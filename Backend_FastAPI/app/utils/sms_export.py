# app/utils/sms_export.py
"""
SMS export (PR-4) helpers: sanitize tên file theo mẫu nhà mạng, render TIN
CUỐI thật (thay sentinel link bằng URL chứa raw code giải mã từ Fernet
ciphertext), dựng + ghi workbook Excel atomic (staging→fsync→os.replace) +
sha256, và verify sha256 theo chunk.

Xem SMS_MARKETING_MODULE_DESIGN.md §8.1 / §8.2 / §8.3.
"""
import hashlib
import io
import os
import re
import unicodedata
import uuid
from typing import Optional, Sequence, Tuple

from openpyxl import Workbook

from app.utils.sms_token import LINK_SENTINEL, build_link, decrypt_code

XLSX_MEDIA = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
# Tiền tố file staging (ghi dở) — DÙNG CHUNG service (ghi) + cleanup task (quét
# orphan) để không lệch quy ước. §8.3.
STAGING_PREFIX = ".staging."
# Ký tự không hợp lệ trong tên file (Windows + *nix) → thay '-'.
_BAD_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|\r\n\t]+')
_FILENAME_MAX = 120
_HASH_CHUNK = 65536


def _slug_label(text: str) -> str:
    """Bỏ dấu tiếng Việt + ÉP ASCII (loại emoji/CJK/Cyrillic còn lại) + gọn
    khoảng trắng → nhãn an toàn cho hệ thống nhà mạng VÀ cho header
    Content-Disposition (Starlette encode latin-1; ký tự >U+00FF sẽ 500)."""
    s = unicodedata.normalize("NFD", text or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d").replace("Đ", "D")
    # Ép ASCII: bỏ MỌI ký tự non-ASCII còn lại (emoji/CJK/…) → filename
    # luôn latin-1-safe, không crash header lúc download.
    s = s.encode("ascii", "ignore").decode("ascii")
    s = _BAD_FILENAME_CHARS.sub("-", s)
    s = re.sub(r"\s+", "-", s.strip())
    # strip cả '.' đầu/cuối: filename KHÔNG bao giờ mở đầu bằng '.' → tránh
    # đụng tiền tố STAGING_PREFIX ('.staging.') khiến orphan-sweep xoá nhầm
    # file final; cũng tránh tạo file ẩn.
    s = re.sub(r"-{2,}", "-", s).strip(".-")
    return s


def sanitize_export_filename(
    group_label: Optional[str], campaign_name: str, carrier_label: str
) -> str:
    """`{Nhóm}-{Campaign}-{NhàMạng}.xlsx` ≤120 ký tự, ASCII. ĐẢM BẢO:
    (1) luôn kết thúc `.xlsx`; (2) phần nhà mạng LUÔN hiện diện đầy đủ → 2 nhà
    mạng KHÔNG bao giờ trùng tên (cùng storage_path → ghi đè). Chỉ cắt phần
    nhóm/campaign (đầu) khi quá dài."""
    grp = _slug_label(group_label or "Nhom") or "Nhom"
    camp = _slug_label(campaign_name) or "Campaign"
    carrier = _slug_label(carrier_label) or "NhaMang"
    ext = ".xlsx"
    # Carrier dài bất thường (gần như không xảy ra) — cắt nhẹ nhưng vẫn giữ +
    # chừa tối thiểu 1 ký tự cho head.
    max_carrier = _FILENAME_MAX - len(ext) - 2
    if len(carrier) > max_carrier:
        carrier = carrier[:max_carrier] or "NhaMang"
    head_budget = _FILENAME_MAX - len(ext) - len(carrier) - 1  # 1 dấu '-'
    head = f"{grp}-{camp}"[:head_budget].rstrip("-")
    if not head:
        head = (grp or "Nhom")[:head_budget] or "Nhom"
    return f"{head}-{carrier}{ext}"


def render_export_message(
    skeleton: Optional[str],
    *,
    has_link: bool,
    token_ciphertext: Optional[str],
    token_key_version: Optional[str],
) -> Optional[str]:
    """TIN CUỐI THẬT cho file nhà mạng. Campaign có {link}: giải mã code từ
    ciphertext (Fernet) → URL thật → thay sentinel. Trả **None** nếu KHÔNG
    giải mã được (key-ring sai) hoặc tin cuối còn sót sentinel — caller fail
    cả batch thay vì ghi tin hỏng/thiếu link cho nhà mạng."""
    msg = skeleton or ""
    if has_link:
        if not (token_ciphertext and token_key_version):
            return None
        code = decrypt_code(token_ciphertext, token_key_version)
        if not code:
            return None
        msg = msg.replace(LINK_SENTINEL, build_link(code))
    # Bất biến: tin cuối KHÔNG còn sentinel nội bộ.
    if LINK_SENTINEL in msg:
        return None
    return msg


def build_carrier_workbook(rows: Sequence[Tuple[str, str]]) -> bytes:
    """Excel đúng mẫu nhà mạng: Sheet1, KHÔNG header, data từ row 2; cột A =
    84xxxxxxxxx (text), cột B = nội dung SMS (text). rows = [(phone_intl, msg)].

    Nội dung cột B GIỮ NGUYÊN tin cuối (KHÔNG escape CSV-injection): đây là văn
    bản nhà mạng GỬI thật — thêm prefix sẽ sai tin. An toàn vì mọi tin luôn bắt
    đầu bằng `[QC]` (assemble_skeleton, NĐ91) nên không thể mở đầu bằng
    `=`/`+`/`-`/`@` → Excel không diễn dịch thành formula; cột số chỉ chứa
    digit; danh bạ do admin tự nhập (không phải public untrusted)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    r = 2  # row 1 để trống theo mẫu nhà mạng
    for phone_intl, msg in rows:
        a = ws.cell(row=r, column=1, value=phone_intl)
        a.number_format = "@"
        b = ws.cell(row=r, column=2, value=msg)
        b.number_format = "@"
        r += 1
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def write_export_atomic(
    rows: Sequence[Tuple[str, str]], final_path: str
) -> Tuple[str, int]:
    """Dựng workbook + ghi ATOMIC ra `final_path`: staging cùng thư mục →
    fsync → os.replace. Trả (sha256_hex, size_bytes). SYNC + blocking →
    caller chạy qua `asyncio.to_thread` để KHÔNG chặn event loop (§8.3)."""
    content = build_carrier_workbook(rows)
    sha = hashlib.sha256(content).hexdigest()
    os.makedirs(os.path.dirname(final_path), exist_ok=True)
    staging = os.path.join(
        os.path.dirname(final_path),
        f"{STAGING_PREFIX}{uuid.uuid4().hex[:12]}.xlsx",
    )
    try:
        with open(staging, "wb") as f:
            f.write(content)
            f.flush()
            os.fsync(f.fileno())
        os.replace(staging, final_path)  # atomic kể cả khi đích đã tồn tại
    except Exception:
        if os.path.exists(staging):
            try:
                os.remove(staging)
            except OSError:
                pass
        raise
    # Best-effort fsync thư mục cha → bền vững entry rename qua crash/mất điện
    # (Linux/prod; Windows không fsync directory được → bỏ qua sạch).
    try:
        dir_fd = os.open(os.path.dirname(final_path), os.O_RDONLY)
        try:
            os.fsync(dir_fd)
        finally:
            os.close(dir_fd)
    except (OSError, AttributeError):
        pass
    return sha, len(content)


def sha256_file(path: str) -> str:
    """sha256 đọc theo chunk (bộ nhớ chặn) — verify integrity lúc download mà
    KHÔNG nạp cả file PII vào RAM. SYNC → caller dùng `asyncio.to_thread`."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(_HASH_CHUNK), b""):
            h.update(chunk)
    return h.hexdigest()
