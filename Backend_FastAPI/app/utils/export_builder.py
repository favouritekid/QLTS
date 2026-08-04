# app/utils/export_builder.py
"""Dựng tệp xuất (csv/xlsx) — MỘT nguồn cho mọi báo cáo dạng bảng.

Có helper này vì các quy tắc dưới đây rất dễ làm sai và sai thì **im lặng**:

* **Ô tiền không được sanitize.** ``DANGEROUS_PREFIXES`` có cả ``-`` nên số âm
  sẽ bị thêm dấu nháy → Excel đọc thành text, người dùng bôi đen cột không ra
  tổng. Chỉ ô TEXT do người dùng nhập mới cần sanitize.
* **CSV phải có BOM.** Không BOM thì Excel bản tiếng Việt đọc theo ANSI và
  tiếng Việt thành mojibake.
* **Cột mã/CCCD phải ép TEXT** (``@``) kẻo mất số 0 đầu.
* **Sheet phụ ghi bộ lọc đã áp.** Người mở file sau vài tuần phải biết đây là
  số của phạm vi nào, nếu không họ sẽ đối chiếu nhầm với màn hình.

Ai thêm báo cáo xuất mới thì khai cột + chỉ số nhóm ô rồi gọi hàm này, đừng
chép lại vòng lặp openpyxl.
"""

import csv
import io
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo

from app.constants.export_formats import (
    CSV_MEDIA_TYPE,
    CSV_UTF8_BOM,
    MONEY_NUMBER_FORMAT,
    TEXT_NUMBER_FORMAT,
    XLSX_MEDIA_TYPE,
)
from app.utils.csv_helpers import sanitize_csv_cell

VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh")

SHEET_META = "Thong tin xuat"


def _csv_money(value: Any) -> str:
    """Ô tiền cho CSV: số nguyên đồng, không phần thập phân.

    ``Decimal('9000000.00')`` in ra "9000000.00"; dưới locale vi-VN dấu ``.``
    là phân tách nghìn nên nhóm 2 chữ số cuối là sai định dạng và Excel nhập ô
    đó thành TEXT (lệch trái, không vào SUM). VND không có phần lẻ nên bỏ hẳn
    đuôi là an toàn và đọc được ở mọi locale.
    """
    if value is None or value == "":
        return ""
    if isinstance(value, Decimal):
        return str(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if isinstance(value, float):
        return str(int(round(value)))
    return str(value)


def _csv_force_text(value: Any) -> str:
    """Ô buộc TEXT cho CSV: bọc ``="..."`` để Excel không cắt số 0 đầu."""
    if value is None or value == "":
        return ""
    raw = sanitize_csv_cell(value)
    # Escape dấu " bên trong để không thoát khỏi chuỗi của công thức.
    return '="' + raw.replace('"', '""') + '"'


def _meta_rows(
    *,
    exporter_name: str,
    applied_filters: Dict[str, Any],
    row_count: int,
    notes: Sequence[Tuple[str, str]],
) -> List[List[str]]:
    now_vn = datetime.now(VN_TZ).strftime("%d/%m/%Y %H:%M")
    rows: List[List[str]] = [
        ["Thông tin lần xuất", ""],
        ["Thời điểm xuất", now_vn],
        ["Người xuất", exporter_name or ""],
        ["Số dòng", str(row_count)],
        ["", ""],
        ["Bộ lọc đã áp dụng", ""],
    ]
    if applied_filters:
        for label, value in applied_filters.items():
            rows.append([str(label), "" if value is None else str(value)])
    else:
        rows.append(["(không lọc)", "toàn bộ phạm vi được phép xem"])

    if notes:
        rows.append(["", ""])
        rows.append(["Lưu ý khi đối chiếu số liệu", ""])
        for label, text in notes:
            rows.append([label, text])
    return rows


def build_simple_export(
    *,
    columns: Sequence[str],
    rows: Iterable[Sequence[Any]],
    money_indexes: set,
    text_indexes: set,
    fmt: str,
    filename_stem: str,
    sheet_title: str,
    exporter_name: str = "",
    applied_filters: Optional[Dict[str, Any]] = None,
    force_text_indexes: Optional[set] = None,
    notes: Sequence[Tuple[str, str]] = (),
    column_widths: Optional[Sequence[int]] = None,
) -> Tuple[bytes, str, str]:
    """Dựng tệp xuất → ``(content, media_type, filename)``.

    ``money_indexes`` / ``text_indexes`` / ``force_text_indexes`` là chỉ số cột
    (0-based). Ô KHÔNG thuộc nhóm nào được ghi thẳng (giá trị do hệ thống sinh
    như mã hồ sơ, ngày tháng, nhãn trạng thái — không cần sanitize).
    """
    rows = [list(r) for r in rows]
    force_text_indexes = force_text_indexes or set()
    applied_filters = applied_filters or {}
    ts = datetime.now(VN_TZ).strftime("%Y%m%d_%H%M%S")

    if (fmt or "").lower() == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([sanitize_csv_cell(c) for c in columns])
        for row in rows:
            out: List[str] = []
            for idx, value in enumerate(row):
                if idx in money_indexes:
                    # KHÔNG sanitize: dấu '-' của số âm sẽ bị thêm dấu nháy.
                    # Ghi số NGUYÊN đồng: "9000000.00" bị Excel bản tiếng Việt
                    # coi là nhóm nghìn sai (dấu '.' là phân tách nghìn) nên
                    # nhập thành TEXT — đúng lỗi "ô tiền không cộng được" mà
                    # bản xuất này sinh ra để dẹp.
                    out.append(_csv_money(value))
                elif idx in force_text_indexes:
                    # Ép TEXT cho mã hồ sơ / CCCD / năm học. Không có bước này,
                    # CCCD "001234567890" vào Excel thành 1234567890 — mất số 0
                    # đầu ở mọi tỉnh 001-096, trong chính tệp dùng để đối chiếu
                    # người nộp. Dạng ="..." là quy ước Excel/LibreOffice đều
                    # hiểu; giá trị được sanitize trước nên không mở đường
                    # chèn công thức.
                    out.append(_csv_force_text(value))
                elif idx in text_indexes:
                    out.append(sanitize_csv_cell(value))
                else:
                    out.append("" if value is None else str(value))
            writer.writerow(out)

        # Ghi chú + bộ lọc: XLSX có sheet phụ, CSV thì không có chỗ nào khác.
        # Đặt SAU dữ liệu và chỉ ở cột đầu → Excel vẫn đọc đúng tiêu đề, và
        # bôi đen cột tiền không dính (các ô này rỗng).
        meta = _meta_rows(
            exporter_name=exporter_name,
            applied_filters=applied_filters,
            row_count=len(rows),
            notes=notes,
        )
        writer.writerow([])
        for meta_row in meta:
            writer.writerow([sanitize_csv_cell(c) for c in meta_row])

        content = (CSV_UTF8_BOM + buf.getvalue()).encode("utf-8")
        return content, CSV_MEDIA_TYPE, f"{filename_stem}_{ts}.csv"

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append([sanitize_csv_cell(c) for c in columns])
    head_fill = PatternFill("solid", fgColor="2F5496")
    head_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for row in rows:
        ws.append(
            [
                sanitize_csv_cell(v) if idx in text_indexes else v
                for idx, v in enumerate(row)
            ]
        )

    last_row = ws.max_row
    for idx in money_indexes:
        letter = get_column_letter(idx + 1)
        for r in range(2, last_row + 1):
            ws[f"{letter}{r}"].number_format = MONEY_NUMBER_FORMAT
    for idx in force_text_indexes:
        letter = get_column_letter(idx + 1)
        for r in range(2, last_row + 1):
            ws[f"{letter}{r}"].number_format = TEXT_NUMBER_FORMAT

    widths = column_widths or [18] * len(columns)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = ws.cell(row=2, column=1).coordinate
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_row}"

    meta = wb.create_sheet(SHEET_META)
    for meta_row in _meta_rows(
        exporter_name=exporter_name,
        applied_filters=applied_filters,
        row_count=len(rows),
        notes=notes,
    ):
        meta.append([sanitize_csv_cell(c) for c in meta_row])
    meta.column_dimensions["A"].width = 30
    meta.column_dimensions["B"].width = 80
    for r in range(1, meta.max_row + 1):
        meta.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), XLSX_MEDIA_TYPE, f"{filename_stem}_{ts}.xlsx"
