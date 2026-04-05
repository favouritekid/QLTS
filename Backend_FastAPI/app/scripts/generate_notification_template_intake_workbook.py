from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app import models
from app.core.event_catalog import get_event_by_key, get_notifiable_events
from app.database import AsyncSessionLocal
from app.scripts.seed_notification_template_library import _INTERNAL_OVERRIDES, _ZNS_DRAFTS


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
READY_FILL = PatternFill("solid", fgColor="E2F0D9")
DRAFT_FILL = PatternFill("solid", fgColor="FCE4D6")
DEFER_FILL = PatternFill("solid", fgColor="EDEDED")

MAIN_HEADERS = [
    "seed_status",
    "template_track",
    "rollout_phase",
    "module",
    "event_key",
    "scenario_key",
    "notification_class",
    "current_rule_enabled",
    "template_scope",
    "actor_type",
    "recipient_primary",
    "recipient_additional",
    "primary_channel",
    "browser_enabled",
    "email_enabled",
    "zalo_enabled",
    "content_mode",
    "template_code",
    "title_sample",
    "message_sample",
    "email_subject_sample",
    "email_body_sample",
    "zalo_template_id",
    "required_variables",
    "link_strategy",
    "dedup_needed",
    "dedup_key_sample",
    "business_trigger",
    "desired_action",
    "review_decision",
    "review_notes",
]

CURRENT_RULE_HEADERS = [
    "event",
    "class",
    "enabled",
    "channels",
    "resolvers",
    "steps",
    "title_template",
    "message_template",
]

GAP_ROWS = [
    ["admission", "override dispatch", "reuse_event", "P1", "Add APPLICATION_STATUS_CHANGED + LEAD_STATUS_CHANGED", "app/routers/admissions.py", "Do not create a new SystemEvent"],
    ["consultation", "status cascade", "reuse_event", "P1", "Dispatch LEAD_STATUS_CHANGED when consultation create/update changes lead pipeline state", "app/routers/leads.py", "Needed for pipeline visibility"],
    ["finance", "payment_overdue", "planned_event", "P2", "Beat task + promote catalog + sync rule + dedup", "app/services/invoice_service.py", "Feature, not a simple patch"],
    ["admission", "withdraw flow", "out_of_scope", "N/A", "Do not implement notification now", "app/services/admission_service.py", "Requires router endpoint + state machine review"],
    ["finance", "fee_fully_paid", "out_of_scope", "N/A", "Do not implement notification now", "app/services/payment_service.py", "Reconsider only if enrollment automation or external applicant notification requires it"],
    ["finance", "payment_rejected / invoice_issued / refund_processed / fee_calculated", "out_of_scope", "N/A", "Do not implement notification now", "app/services/payment_service.py", "Internal workflow is sufficient in current scope"],
]

BUSINESS_GAP_HEADERS = [
    "gap_no",
    "event_or_case",
    "business_flow",
    "current_state",
    "required_work",
    "needs_new_system_event",
    "needs_new_router",
    "needs_beat_task",
    "effort",
]

BUSINESS_GAP_ROWS = [
    [
        1,
        "application_status_changed / overridden",
        "Admin or manager overrides admission decision.",
        "Override endpoint exists and commits, but no notification is dispatched.",
        "Add APPLICATION_STATUS_CHANGED + LEAD_STATUS_CHANGED after commit.",
        "no",
        "no",
        "no",
        "small",
    ],
    [
        2,
        "application_status_changed / withdrawn",
        "Applicant or admin withdraws an admission profile.",
        "withdraw_profile() exists in service but has no router caller and no notification dispatch.",
        "Out of scope for notification now. Revisit only with router endpoint + state machine review.",
        "no",
        "yes",
        "no",
        "medium",
    ],
    [
        3,
        "consultation_created when status_updated=True",
        "Creating consultation can also move the lead pipeline status.",
        "CONSULTATION_CREATED is dispatched, but LEAD_STATUS_CHANGED is missing when status_updated=True.",
        "Conditionally dispatch LEAD_STATUS_CHANGED after CONSULTATION_CREATED.",
        "no",
        "no",
        "no",
        "small",
    ],
    [
        4,
        "payment_overdue",
        "Invoice or fee is overdue and should notify applicant + finance.",
        "SystemEvent exists and overdue repository exists, but no periodic task dispatches it. Catalog is still internal_future.",
        "Add beat task, dispatch PAYMENT_OVERDUE, promote to user, sync DB rule.",
        "no",
        "no",
        "yes",
        "large",
    ],
    [
        5,
        "fee_fully_paid",
        "Fee is fully settled after verification or waiver.",
        "Service already detects fee_remaining <= 0 and syncs lead, but there is no event or dispatch.",
        "Out of scope for notification now. Reconsider only when enrollment automation or external applicant notification requires it.",
        "yes",
        "no",
        "no",
        "medium",
    ],
    [
        6,
        "payment_rejected",
        "Checker rejects a pending payment.",
        "reject_payment() exists but post-commit callback is still TODO.",
        "Out of scope for notification now.",
        "yes",
        "no",
        "no",
        "small",
    ],
    [
        7,
        "invoice_issued",
        "Invoice changes from draft to issued and becomes payable.",
        "issue_invoice() exists but post-commit callback is still TODO.",
        "Out of scope for notification now.",
        "yes",
        "no",
        "no",
        "small",
    ],
    [
        8,
        "refund_processed",
        "Approved refund is processed and balances are updated.",
        "process_approved_refund() returns (refund, None), so no notification is dispatched.",
        "Out of scope for notification now.",
        "yes",
        "no",
        "no",
        "small",
    ],
]


def _module_name(category: str) -> str:
    return {
        "application": "admission",
        "lead": "lead",
        "consultation": "consultation",
        "finance": "finance",
        "ctv": "ctv",
        "system": "system",
        "security": "security",
        "pipeline": "pipeline",
    }.get(category, category)


def _infer_actor(event_key: str, module: str, scenario_key: str) -> str:
    if event_key == "consultation_reminder":
        return "scheduler"
    if event_key in {"system_alert", "suspicious_login"}:
        return "system"
    if event_key == "ctv_claim_submitted":
        return "collaborator"
    if event_key in {"ctv_claim_approved", "ctv_claim_rejected", "ctv_approved", "ctv_suspended"}:
        return "manager"
    if module == "finance" and "overdue" in event_key:
        return "scheduler"
    if module == "finance":
        return "accountant"
    if module == "admission" and scenario_key in {"overridden", "withdrawn"}:
        return "manager"
    if module in {"lead", "consultation", "admission"}:
        return "current_user"
    return "system"


def _email_body(message: str) -> str:
    return f"Kinh gui,\n\n{message}\n\nVui long kiem tra chi tiet tren he thong.\n"


def _preferred_rule_content(rule, definition):
    if rule and not str(rule.title_template).startswith("[LEGACY DISABLED]"):
        return str(rule.title_template), str(rule.message_template)
    override = _INTERNAL_OVERRIDES.get(definition.event.value, {})
    return (
        str(override.get("title_template", definition.display_name)),
        str(override.get("message_template", definition.description or definition.display_name)),
    )


def _primary_channel(channels: list[str]) -> str:
    for value in ("browser", "email", "zalo"):
        if value in channels:
            return value
    return "browser"


def _enabled_flag(channels: list[str], target: str) -> str:
    return "yes" if target in channels else "no"


def _style_status(row, seed_status: str) -> None:
    fill = {"READY": READY_FILL, "DRAFT": DRAFT_FILL, "DEFERRED": DEFER_FILL}.get(seed_status)
    if fill:
        for cell in row:
            cell.fill = fill


def _add_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _set_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _scenario_rows() -> list[dict[str, str]]:
    return [
        {
            "seed_status": "READY",
            "template_track": "scenario_variant",
            "rollout_phase": "phase_1a",
            "module": "admission",
            "event_key": "application_status_changed",
            "scenario_key": "overridden",
            "notification_class": "user",
            "current_rule_enabled": "yes",
            "template_scope": "template_only",
            "actor_type": "manager",
            "recipient_primary": "lead_owner",
            "recipient_additional": "all_admins",
            "primary_channel": "browser",
            "browser_enabled": "yes",
            "email_enabled": "yes",
            "zalo_enabled": "no",
            "content_mode": "template_override",
            "template_code": "TPL_APPLICATION_STATUS_CHANGED_OVERRIDDEN_V1",
            "title_sample": "Ho so duoc duyet dac biet: #$application_id",
            "message_sample": "Ho so #$application_id da duoc chuyen sang trang thai overridden theo phe duyet dac biet.",
            "email_subject_sample": "Ho so duoc duyet dac biet: #$application_id",
            "email_body_sample": _email_body("Ho so #$application_id da duoc chuyen sang trang thai overridden theo phe duyet dac biet."),
            "zalo_template_id": "",
            "required_variables": "application_id, lead_id, old_status, new_status, actor_id",
            "link_strategy": "/admissions/${application_id}",
            "dedup_needed": "yes",
            "dedup_key_sample": "app:${application_id}:status:overridden",
            "business_trigger": "Admin override ho so, can wording rieng de phan biet voi status thong thuong.",
            "desired_action": "Seed template variant de admin gan qua template_override.",
            "review_decision": "",
            "review_notes": "",
        },
        {
            "seed_status": "READY",
            "template_track": "scenario_variant",
            "rollout_phase": "phase_1b",
            "module": "lead",
            "event_key": "lead_status_changed",
            "scenario_key": "consultation_auto_update",
            "notification_class": "user",
            "current_rule_enabled": "yes",
            "template_scope": "template_only",
            "actor_type": "system",
            "recipient_primary": "lead_owner",
            "recipient_additional": "",
            "primary_channel": "browser",
            "browser_enabled": "yes",
            "email_enabled": "no",
            "zalo_enabled": "no",
            "content_mode": "template_override",
            "template_code": "TPL_LEAD_STATUS_CHANGED_CONSULTATION_AUTO_V1",
            "title_sample": "Lead tu dong doi trang thai sau cap nhat tu van",
            "message_sample": "Lead $lead_name da duoc he thong cap nhat trang thai theo dien bien tu van moi nhat.",
            "email_subject_sample": "Lead tu dong doi trang thai sau cap nhat tu van",
            "email_body_sample": _email_body("Lead $lead_name da duoc he thong cap nhat trang thai theo dien bien tu van moi nhat."),
            "zalo_template_id": "",
            "required_variables": "lead_id, lead_name, old_status, new_status, actor_id",
            "link_strategy": "/leads/${lead_id}",
            "dedup_needed": "yes",
            "dedup_key_sample": "lead:${lead_id}:status:${new_status}",
            "business_trigger": "Tao consultation co the lam thay doi pipeline status.",
            "desired_action": "Seed variant cho consultation cascade.",
            "review_decision": "",
            "review_notes": "",
        },
    ]


def _planned_finance_rows() -> list[dict[str, str]]:
    base = [
        ("phase_2", "payment_overdue", "scheduler", "lead_contact", "finance_staff", "TPL_PAYMENT_OVERDUE_V1", "Khoan thanh toan da qua han", "Khoan phi #$fee_id cua ban da qua han $days_overdue ngay voi so tien $amount.", "fee_id, user_id, amount, days_overdue, fee_type", "/finance/fees/${fee_id}", "payment_overdue:${fee_id}:${days_overdue}", "Beat task phat hien invoice/fee qua han."),
    ]
    rows = []
    for phase, event_key, actor, recipient, additional, code, title, message, vars_, link, dedup, trigger in base:
        rows.append(
            {
                "seed_status": "DEFERRED" if phase == "deferred" else "DRAFT",
                "template_track": "planned_event",
                "rollout_phase": phase,
                "module": "finance",
                "event_key": event_key,
                "scenario_key": "default",
                "notification_class": "planned",
                "current_rule_enabled": "no",
                "template_scope": "template_plus_disabled_rule",
                "actor_type": actor,
                "recipient_primary": recipient,
                "recipient_additional": additional,
                "primary_channel": "email" if recipient == "lead_contact" else "browser",
                "browser_enabled": "yes",
                "email_enabled": "yes" if recipient == "lead_contact" else "no",
                "zalo_enabled": "yes" if event_key in {"payment_overdue", "fee_fully_paid"} else "no",
                "content_mode": "template_override",
                "template_code": code,
                "title_sample": title,
                "message_sample": message,
                "email_subject_sample": title,
                "email_body_sample": _email_body(message),
                "zalo_template_id": "DRAFT_ONLY" if event_key in {"payment_overdue", "fee_fully_paid"} else "",
                "required_variables": vars_,
                "link_strategy": link,
                "dedup_needed": "yes",
                "dedup_key_sample": dedup,
                "business_trigger": trigger,
                "desired_action": "Chuan bi template cho rollout tai runtime khi event duoc wire.",
                "review_decision": "",
                "review_notes": "",
            }
        )
    return rows


async def build_workbook(output_path: Path) -> None:
    async with AsyncSessionLocal() as db:
        rules = (
            await db.execute(
                select(models.NotificationRule).options(selectinload(models.NotificationRule.actions))
            )
        ).scalars().all()

    rules_by_event = {rule.event: rule for rule in rules}
    enabled_by_event = {rule.event: rule.enabled for rule in rules}

    wb = Workbook()
    intake = wb.active
    intake.title = "Template_Intake"
    _add_headers(intake, MAIN_HEADERS)

    intake_rows = []
    current_rows = []

    for definition in get_notifiable_events():
        event_key = definition.event.value
        rule = rules_by_event.get(event_key)
        title, message = _preferred_rule_content(rule, definition)
        channels = [a.channel for a in sorted(getattr(rule, "actions", []), key=lambda a: (a.step, a.channel))]
        if not channels:
            channels = list(definition.default_channels)
        resolvers = []
        if rule:
            for action in sorted(rule.actions, key=lambda a: (a.step, a.channel)):
                if isinstance(action.recipient_config, dict) and action.recipient_config.get("resolver_type"):
                    resolvers.append(action.recipient_config["resolver_type"])

        intake_rows.append(
            {
                "seed_status": "READY" if enabled_by_event.get(event_key) else "DRAFT",
                "template_track": "internal_library",
                "rollout_phase": "baseline",
                "module": _module_name(definition.category),
                "event_key": event_key,
                "scenario_key": "default",
                "notification_class": definition.notification_class,
                "current_rule_enabled": "yes" if enabled_by_event.get(event_key) else "no",
                "template_scope": "template_only",
                "actor_type": _infer_actor(event_key, _module_name(definition.category), "default"),
                "recipient_primary": resolvers[0] if resolvers else definition.default_resolver,
                "recipient_additional": ", ".join(sorted(set(resolvers[1:]))) if len(resolvers) > 1 else "",
                "primary_channel": _primary_channel(channels),
                "browser_enabled": _enabled_flag(channels, "browser"),
                "email_enabled": _enabled_flag(channels, "email"),
                "zalo_enabled": "no",
                "content_mode": "template_override",
                "template_code": f"TPL_{event_key.upper()}_V1",
                "title_sample": title,
                "message_sample": message,
                "email_subject_sample": title,
                "email_body_sample": _email_body(message),
                "zalo_template_id": "",
                "required_variables": ", ".join(v.name for v in definition.variables),
                "link_strategy": definition.link_strategy or "",
                "dedup_needed": "yes" if definition.dedup_key_template else "no",
                "dedup_key_sample": definition.dedup_key_template or "",
                "business_trigger": definition.description,
                "desired_action": "Seed template library de admin review va gan vao rule khi can.",
                "review_decision": "",
                "review_notes": "",
            }
        )

        current_rows.append(
            {
                "event": event_key,
                "class": definition.notification_class,
                "enabled": "yes" if enabled_by_event.get(event_key) else "no",
                "channels": ", ".join(channels),
                "resolvers": ", ".join(resolvers or [definition.default_resolver]),
                "steps": ", ".join(str(a.step) for a in sorted(getattr(rule, "actions", []), key=lambda a: (a.step, a.channel))),
                "title_template": rule.title_template if rule else "",
                "message_template": rule.message_template if rule else "",
            }
        )

    intake_rows.extend(_scenario_rows())
    intake_rows.extend(_planned_finance_rows())

    allowed_zns_events = {"payment_verified", "application_status_changed", "consultation_reminder"}
    for draft in _ZNS_DRAFTS:
        event_key = draft.allowed_events[0]
        if event_key not in allowed_zns_events:
            continue
        definition = get_event_by_key(event_key)
        intake_rows.append(
            {
                "seed_status": "DRAFT",
                "template_track": "zns_draft",
                "rollout_phase": "phase_2" if event_key == "payment_verified" else "baseline",
                "module": _module_name(definition.category if definition else draft.category),
                "event_key": event_key,
                "scenario_key": "zns_external",
                "notification_class": definition.notification_class if definition else "planned",
                "current_rule_enabled": "yes" if enabled_by_event.get(event_key) else "no",
                "template_scope": "template_only",
                "actor_type": _infer_actor(event_key, _module_name(definition.category if definition else draft.category), "zns_external"),
                "recipient_primary": "lead_contact" if event_key in {"payment_verified", "application_status_changed", "consultation_reminder"} else "collaborator_user",
                "recipient_additional": "",
                "primary_channel": "zalo",
                "browser_enabled": "no",
                "email_enabled": "no",
                "zalo_enabled": "yes",
                "content_mode": "channel_native",
                "template_code": draft.template_code,
                "title_sample": draft.title_template,
                "message_sample": draft.message_template,
                "email_subject_sample": "",
                "email_body_sample": "",
                "zalo_template_id": "",
                "required_variables": ", ".join(draft.variables),
                "link_strategy": "",
                "dedup_needed": "yes" if definition and definition.dedup_key_template else "no",
                "dedup_key_sample": definition.dedup_key_template if definition and definition.dedup_key_template else "",
                "business_trigger": draft.description,
                "desired_action": "Dien ma template OA that neu quyet dinh rollout ZNS.",
                "review_decision": "",
                "review_notes": "",
            }
        )

    for row_data in intake_rows:
        intake.append([row_data.get(header, "") for header in MAIN_HEADERS])
        _style_status(intake[intake.max_row], row_data["seed_status"])

    _set_widths(intake, {"A": 12, "B": 18, "C": 12, "D": 14, "E": 28, "F": 24, "G": 18, "H": 12, "I": 28, "J": 14, "K": 18, "L": 22, "M": 14, "N": 12, "O": 12, "P": 12, "Q": 16, "R": 34, "S": 38, "T": 58, "U": 38, "V": 58, "W": 18, "X": 32, "Y": 30, "Z": 12, "AA": 30, "AB": 42, "AC": 42, "AD": 18, "AE": 42})
    for row in intake.iter_rows(min_row=2, max_row=intake.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    guide = wb.create_sheet("HuongDan")
    guide["A1"] = "Notification template intake workbook"
    guide["A1"].font = Font(bold=True, size=14)
    guide["A3"] = "Cach dung"
    guide["A3"].fill = SUBHEADER_FILL
    for index, text in enumerate(
        [
            "1. Sheet Template_Intake da co san mau noi dung theo baseline hien tai + draft de xuat.",
            "2. Cac cot co dropdown: seed_status, template_track, rollout_phase, module, event_key, notification_class, template_scope, actor_type, recipient_primary, primary_channel, yes/no, content_mode, review_decision.",
            "3. Moi dong la mot ung vien template. Ban sua title/message/email/zalo_template_id/review_notes de chot nghiep vu.",
            "4. override va withdraw cua admission la variant cua application_status_changed, khong phai event moi.",
            "4a. Theo scope decision hien tai, chi override la immediate notification fix; withdraw dang out_of_scope_by_design.",
            "5. Current_DB_Rules la snapshot cau hinh rule hien tai trong database de doi chieu.",
            "6. Coverage_Gaps phan biet ro immediate fixes, planned backlog va out_of_scope_by_design.",
        ],
        start=4,
    ):
        guide[f"A{index}"] = text
    guide.column_dimensions["A"].width = 120

    current_ws = wb.create_sheet("Current_DB_Rules")
    _add_headers(current_ws, CURRENT_RULE_HEADERS)
    for row_data in current_rows:
        current_ws.append([row_data[h] for h in CURRENT_RULE_HEADERS])
    _set_widths(current_ws, {"A": 28, "B": 18, "C": 12, "D": 20, "E": 24, "F": 12, "G": 42, "H": 72})
    for row in current_ws.iter_rows(min_row=2, max_row=current_ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    gap_ws = wb.create_sheet("Coverage_Gaps")
    _add_headers(gap_ws, ["module", "item", "type", "priority", "recommended_action", "hotspot", "notes"])
    for row in GAP_ROWS:
        gap_ws.append(row)
    _set_widths(gap_ws, {"A": 14, "B": 32, "C": 16, "D": 10, "E": 44, "F": 36, "G": 44})
    for row in gap_ws.iter_rows(min_row=2, max_row=gap_ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    business_ws = wb.create_sheet("Business_Gaps_8")
    _add_headers(business_ws, BUSINESS_GAP_HEADERS)
    for row in BUSINESS_GAP_ROWS:
        business_ws.append(row)
    _set_widths(
        business_ws,
        {
            "A": 8,
            "B": 32,
            "C": 36,
            "D": 48,
            "E": 54,
            "F": 18,
            "G": 16,
            "H": 16,
            "I": 12,
        },
    )
    for row in business_ws.iter_rows(min_row=2, max_row=business_ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    options_ws = wb.create_sheet("Options")
    option_columns = {
        "A": ["READY", "DRAFT", "DEFERRED"],
        "B": ["internal_library", "scenario_variant", "planned_event", "zns_draft"],
        "C": ["baseline", "phase_1a", "phase_1b", "phase_1c", "phase_2", "phase_3", "deferred"],
        "D": ["lead", "consultation", "admission", "finance", "ctv", "system", "security", "pipeline"],
        "E": sorted({row["event_key"] for row in intake_rows}),
        "F": ["user", "broadcast_only", "internal_future", "planned"],
        "G": ["template_only", "template_plus_disabled_rule", "template_plus_enabled_rule"],
        "H": ["system", "scheduler", "current_user", "admin", "manager", "officer", "accountant", "collaborator", "applicant", "mixed_human"],
        "I": ["lead_owner", "unit_managers", "all_admins", "all_users", "specific_users", "collaborator_user", "lead_contact", "admission_contact", "finance_staff", "applicant", "mixed_internal", "mixed_external"],
        "J": ["browser", "email", "zalo"],
        "K": ["yes", "no"],
        "L": ["inherit_default", "template_override", "inline_override", "channel_native"],
        "M": ["approve", "revise", "defer", "drop"],
    }
    for column, values in option_columns.items():
        for idx, value in enumerate(values, start=1):
            options_ws[f"{column}{idx}"] = value
    options_ws.sheet_state = "hidden"

    validations = {
        "A": "=Options!$A$1:$A$3",
        "B": "=Options!$B$1:$B$4",
        "C": "=Options!$C$1:$C$7",
        "D": "=Options!$D$1:$D$8",
        "E": f"=Options!$E$1:$E${len(option_columns['E'])}",
        "G": "=Options!$F$1:$F$4",
        "I": "=Options!$G$1:$G$3",
        "J": "=Options!$H$1:$H$10",
        "K": "=Options!$I$1:$I$12",
        "M": "=Options!$J$1:$J$3",
        "N": "=Options!$K$1:$K$2",
        "O": "=Options!$K$1:$K$2",
        "P": "=Options!$K$1:$K$2",
        "Q": "=Options!$L$1:$L$4",
        "Z": "=Options!$K$1:$K$2",
        "AD": "=Options!$M$1:$M$4",
    }
    for column, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        intake.add_data_validation(dv)
        dv.add(f"{column}2:{column}400")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="/tmp/Notification_Template_Seed_Intake.xlsx")
    args = parser.parse_args()

    import asyncio

    asyncio.run(build_workbook(Path(args.output)))
    print(f"Workbook created at {args.output}")


if __name__ == "__main__":
    main()
