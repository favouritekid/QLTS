# app/services/tuition_export_service.py
"""Xuất danh sách học phí theo bộ lọc màn hình "Thu học phí".

Vì sao có service này: trước đây muốn lấy bảng "hồ sơ × học phí × đã đóng ×
còn lại" thì phải SSH vào máy chủ chạy SQL tay rồi ghép file — không audit,
không phân quyền, chỉ người giữ khoá SSH làm được.

Hai quyết định định hình toàn bộ file, đừng đổi mà không đọc lý do:

1. **Lọc ở mức hoá đơn, xuất ở mức KHOẢN PHÍ.** Bộ lọc của workspace là
   invoice-centric, nhưng ba cột tiền kế toán cần (học phí / đã đóng / còn lại)
   nằm ở mức ``Fee``. Một khoản phí nhiều đợt có nhiều hoá đơn; xuất theo hoá
   đơn là lặp lại cùng số tiền ⇒ bôi đen cột cộng ra gấp đôi mà không có dấu
   hiệu gì. Nên tập dòng = ``DISTINCT invoice.fee_id``.
   Hệ quả ANY-match (một fee lọt vào nếu ≥1 hoá đơn khớp, nhưng tiền là của cả
   khoản phí) được nói thẳng trong sheet "Thong tin xuat".

2. **Ô tiền là SỐ, không đi qua ``sanitize_csv_cell``.** Danh sách ký tự nguy
   hiểm của helper đó có cả dấu ``-``, nên số âm (hồ sơ đóng dư) sẽ bị thêm dấu
   nháy thành text và Excel không cộng được. Sanitize chỉ áp cho ô TEXT do
   người dùng nhập — xem ``_TEXT_COLUMN_INDEXES``.
"""

import csv
import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

import structlog
from sqlalchemy.ext.asyncio import AsyncSession

from app.constants.export_formats import (
    CSV_MEDIA_TYPE,
    CSV_UTF8_BOM,
    MONEY_NUMBER_FORMAT,
    TEXT_NUMBER_FORMAT,
    XLSX_MEDIA_TYPE,
)
from app.repositories.fee_repository import FeeRepository, InvoiceRepository
from app.utils.csv_helpers import sanitize_csv_cell
from app.utils.exceptions import BadRequest
from app.utils.id_helpers import format_profile_code

log = structlog.get_logger(__name__)

VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh")

# Trần dòng mỗi lần xuất. Prod hiện ~400 dòng nên đây là HÀNG RÀO chống truy vấn
# nuốt hết bộ nhớ, không phải chính sách nghiệp vụ.
MAX_EXPORT_ROWS = 10_000

SHEET_DATA = "Danh sach hoc phi"
SHEET_META = "Thong tin xuat"

COLUMNS: List[str] = [
    "Mã hồ sơ",                 # 0
    "Ngày nhập hồ sơ",          # 1
    "Officer phụ trách",        # 2
    "Họ và tên",                # 3
    "Số CCCD",                  # 4
    "Trình độ ngành đăng ký",   # 5
    "Ngành đăng ký",            # 6
    "Loại phí",                 # 7
    "Năm học",                  # 8
    "Học kỳ",                   # 9
    "Trạng thái khoản phí",     # 10
    "Học phí ngành học",        # 11
    "Tổng học phí đã đóng",     # 12
    "Đã miễn giảm",             # 13
    "Số tiền còn lại",          # 14
    "Đơn vị",                   # 15
]

# Ô TEXT do người dùng nhập → phải sanitize (chống chèn công thức vào Excel).
# CỐ Ý không gồm cột tiền: sanitize sẽ biến số âm thành text.
_TEXT_COLUMN_INDEXES = {2, 3, 4, 5, 6, 15}

# Ô cần ép TEXT để giữ số 0 đầu / không bị Excel hiểu thành số.
_FORCE_TEXT_INDEXES = {0, 4, 8}

# Ô tiền — ghi kiểu số + number_format nghìn.
_MONEY_INDEXES = {11, 12, 13, 14}

_FEE_TYPE_LABELS = {
    "tuition": "Học phí",
    "application": "Lệ phí xét tuyển",
    "enrollment": "Phí nhập học",
    "insurance": "Bảo hiểm",
    "dormitory": "Ký túc xá",
    "other": "Khác",
}

_FEE_STATUS_LABELS = {
    "pending": "Chờ tính",
    "calculated": "Đã tính",
    "invoiced": "Đã phát hành hoá đơn",
    "partial": "Đã thu một phần",
    "paid": "Đã thu đủ",
    "overdue": "Quá hạn",
    "waived": "Được miễn",
    "cancelled": "Đã huỷ",
}

_NOT_DECIDED = "(chưa chốt ngành)"
_NO_OFFICER = "(chưa phân công)"


def _fmt_dt(value: Optional[datetime]) -> str:
    if value is None:
        return ""
    return value.astimezone(VN_TZ).strftime("%d/%m/%Y %H:%M")


def _row_for(fee: Any) -> List[Any]:
    """Một khoản phí → một dòng. Giữ nguyên kiểu dữ liệu cho ô tiền."""
    profile = fee.admission_profile
    lead = profile.lead if profile else None
    officer = getattr(lead, "assigned_officer", None) if lead else None
    unit = getattr(lead, "unit", None) if lead else None

    degree = fee.resolved_degree_level or _NOT_DECIDED
    major = fee.resolved_major.name if fee.resolved_major else _NOT_DECIDED

    remaining = fee.final_amount - fee.paid_amount - fee.waived_amount

    return [
        format_profile_code(profile.id) if profile else "",
        _fmt_dt(profile.created_at) if profile else "",
        (officer.full_name if officer else "") or _NO_OFFICER,
        (lead.full_name if lead else "") or "",
        profile.citizen_id if profile and profile.citizen_id else "",
        degree,
        major,
        _FEE_TYPE_LABELS.get(fee.fee_type, fee.fee_type),
        f"{fee.academic_year}-{fee.academic_year + 1}",
        fee.semester_no if fee.semester_no is not None else "",
        _FEE_STATUS_LABELS.get(fee.status, fee.status),
        fee.final_amount,
        fee.paid_amount,
        fee.waived_amount,
        remaining,
        (unit.name if unit else "") or "",
    ]


def _meta_rows(
    *,
    exporter_name: str,
    applied_filters: Dict[str, Any],
    row_count: int,
) -> List[List[str]]:
    """Nội dung sheet phụ — người mở file sau vài tuần phải biết đây là gì."""
    now_vn = datetime.now(VN_TZ).strftime("%d/%m/%Y %H:%M")
    rows: List[List[str]] = [
        ["Thông tin lần xuất", ""],
        ["Thời điểm xuất", now_vn],
        ["Người xuất", exporter_name or ""],
        ["Số dòng", str(row_count)],
        ["", ""],
        ["Bộ lọc đã áp dụng", ""],
    ]
    if applied_filters:
        for label, value in applied_filters.items():
            rows.append([str(label), "" if value is None else str(value)])
    else:
        rows.append(["(không lọc)", "toàn bộ phạm vi được phép xem"])

    rows += [
        ["", ""],
        ["Lưu ý khi đối chiếu số liệu", ""],
        [
            "Mỗi dòng là MỘT KHOẢN PHÍ",
            "Ba cột tiền là của cả khoản phí, không phải của riêng đợt thu.",
        ],
        [
            "Bộ lọc áp ở mức hoá đơn",
            "Một khoản phí có mặt nếu ít nhất một đợt của nó khớp bộ lọc. "
            "Vì vậy tổng cột 'Số tiền còn lại' ở đây có thể KHÁC ô 'Còn phải "
            "thu' trên màn hình (ô đó cộng theo từng đợt).",
        ],
        [
            "Số tiền âm",
            "Là hồ sơ đã đóng dư, giữ nguyên dấu âm chứ không làm tròn về 0.",
        ],
    ]
    return rows


def _build_csv(rows: List[List[Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([sanitize_csv_cell(c) for c in COLUMNS])
    for row in rows:
        out: List[str] = []
        for idx, value in enumerate(row):
            if idx in _MONEY_INDEXES:
                # Ô tiền: ghi thẳng, KHÔNG sanitize (dấu '-' của số âm sẽ bị
                # helper thêm dấu nháy → Excel đọc thành text).
                out.append("" if value is None else str(value))
            elif idx in _TEXT_COLUMN_INDEXES:
                out.append(sanitize_csv_cell(value))
            else:
                out.append("" if value is None else str(value))
        writer.writerow(out)
    return (CSV_UTF8_BOM + buf.getvalue()).encode("utf-8")


def _build_xlsx(
    rows: List[List[Any]],
    *,
    exporter_name: str,
    applied_filters: Dict[str, Any],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA

    ws.append([sanitize_csv_cell(c) for c in COLUMNS])
    head_fill = PatternFill("solid", fgColor="2F5496")
    head_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for row in rows:
        ws.append(
            [
                sanitize_csv_cell(v) if idx in _TEXT_COLUMN_INDEXES else v
                for idx, v in enumerate(row)
            ]
        )

    last_row = ws.max_row
    for idx in _MONEY_INDEXES:
        letter = get_column_letter(idx + 1)
        for r in range(2, last_row + 1):
            ws[f"{letter}{r}"].number_format = MONEY_NUMBER_FORMAT
    for idx in _FORCE_TEXT_INDEXES:
        letter = get_column_letter(idx + 1)
        for r in range(2, last_row + 1):
            ws[f"{letter}{r}"].number_format = TEXT_NUMBER_FORMAT

    widths = [12, 18, 22, 26, 16, 20, 30, 18, 12, 10, 22, 18, 20, 16, 18, 20]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = ws.cell(row=2, column=1).coordinate
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"

    meta = wb.create_sheet(SHEET_META)
    for meta_row in _meta_rows(
        exporter_name=exporter_name,
        applied_filters=applied_filters,
        row_count=len(rows),
    ):
        meta.append([sanitize_csv_cell(c) for c in meta_row])
    meta.column_dimensions["A"].width = 30
    meta.column_dimensions["B"].width = 80
    for r in range(1, meta.max_row + 1):
        meta.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def build_tuition_export(
    db: AsyncSession,
    *,
    fmt: str,
    unit_id: Optional[int],
    exporter_name: str,
    applied_filters: Optional[Dict[str, Any]] = None,
    **filters: Any,
) -> Tuple[bytes, str, str]:
    """Dựng file xuất → ``(content, media_type, filename)``.

    ``filters`` nhận đúng bộ tham số lọc của danh sách hoá đơn (xem
    ``InvoiceRepository.get_distinct_fee_ids_for_filter``).

    Vượt trần thì **từ chối**, không cắt bớt: một file bị cắt âm thầm trông y
    hệt một file đủ, người nhận không có cách nào biết.
    """
    invoice_repo = InvoiceRepository(db)
    fee_repo = FeeRepository(db)

    # Lấy MAX+1 trong MỘT truy vấn: đủ để biết có vượt trần không mà không cần
    # cặp count-rồi-fetch (hai câu có khe TOCTOU giữa chúng).
    fee_ids = await invoice_repo.get_distinct_fee_ids_for_filter(
        limit=MAX_EXPORT_ROWS + 1, unit_id=unit_id, **filters
    )
    if len(fee_ids) > MAX_EXPORT_ROWS:
        cap_vn = f"{MAX_EXPORT_ROWS:,}".replace(",", ".")  # 10.000 kiểu VN
        raise BadRequest(
            f"Kết quả lọc vượt quá {cap_vn} dòng cho một lần xuất. "
            "Hãy thu hẹp bộ lọc (năm học / học kỳ / đơn vị / ngành) rồi xuất lại."
        )

    fees = await fee_repo.get_many_for_export(fee_ids)
    rows = [_row_for(fee) for fee in fees]

    log.info(
        "tuition_export_built",
        row_count=len(rows),
        fmt=fmt,
        unit_id=unit_id,
    )

    ts = datetime.now(VN_TZ).strftime("%Y%m%d_%H%M%S")
    if (fmt or "").lower() == "csv":
        return _build_csv(rows), CSV_MEDIA_TYPE, f"danh_sach_hoc_phi_{ts}.csv"

    content = _build_xlsx(
        rows,
        exporter_name=exporter_name,
        applied_filters=applied_filters or {},
    )
    return content, XLSX_MEDIA_TYPE, f"danh_sach_hoc_phi_{ts}.xlsx"
