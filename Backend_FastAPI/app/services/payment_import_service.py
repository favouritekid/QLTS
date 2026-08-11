# app/services/payment_import_service.py
"""
Payment Import Service (BV-2) - parse + resolve/validate + lưu batch preview.

Kế toán thu offline → import file mẫu → hệ thống đối chiếu từng dòng theo CCCD,
phân bổ FIFO (chỉ GỐC học phí) và phân loại MATCHED / WARNING / ERROR. BV-2 KHÔNG
ghi Payment — chỉ đọc + tính + lưu batch 'preview' để xem trước (pha commit ghi
tiền nằm ở BV-3).

Ref: Documents/BULK_PAYMENT_IMPORT_VERIFY_PLAN.md (DESIGN v2). Bất biến bắt buộc:
  1. đọc ``dtype=str`` CCCD/ref/amount (giữ số 0 đầu, không float)
  2. method 'cash'/'bank_transfer' (seed thực, KHÔNG 'bank')
  3. fee tuition active-only (NOT IN cancelled/waived) + đúng năm qua hồ sơ
  4. FIFO principal-first: principal_remaining = invoice.amount − paid_amount
     (KHÔNG remaining_amount vì gồm penalty)
  5. sổ phân bổ in-batch chống double-alloc khi 2 dòng cùng CCCD+kỳ
  6. IDOR unit-scope + bỏ lead xóa mềm + CCCD lỗi sạch
  7. resolve read-only thuần (không ghi gì trong resolve_and_validate)
"""
import csv
import hashlib
import io
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Callable, Dict, List, Optional, Set, Tuple

import pandas as pd
import structlog
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app import models
from app.models.finance import (
    Fee,
    Invoice,
    FeeStatusEnum,
    InvoiceStatusEnum,
    PAYABLE_INVOICE_STATUSES,
    Payment,
    PaymentImportBatch,
    PaymentImportRow,
    PaymentImportBatchStatusEnum,
    PaymentImportRowStatusEnum,
    PaymentImportCommitStatusEnum,
    PaymentMethod,
    PaymentStatusEnum,
    PaymentTransaction,
    RefundRequest,
    RefundStatusEnum,
    TransactionTypeEnum,
)
from app.constants.export_formats import (
    CSV_MEDIA_TYPE,
    CSV_UTF8_BOM,
    TEXT_NUMBER_FORMAT,
    XLSX_MEDIA_TYPE,
)
from app.repositories.fee_repository import FeeRepository, InvoiceRepository
from app.services.duplicate_review_token import RangBuoc, cap_phieu, soat_phieu
from app.repositories.payment_repository import (
    MAX_DUPLICATE_CANDIDATES,
    WINDOW_DO_TRUNG_NGAY,
    PaymentRepository,
)
from app.utils.admission_status import NON_PAYABLE_PROFILE_STATUSES
from app.utils.csv_helpers import sanitize_csv_row
from app.utils.exceptions import (
    BadRequest,
    BusinessRuleViolation,
    ConflictError,
    ResourceNotFoundError,
)

log = structlog.get_logger(__name__)

# ---------------------------------------------------------------------------
# Template columns (tiếng Việt — chốt 06-24) + mapping
# ---------------------------------------------------------------------------
COL_CCCD = "Số CCCD"
COL_NAME = "Họ và tên học sinh"
COL_AMOUNT = "Số tiền thu (VNĐ)"
COL_DATE = "Ngày thu"
COL_METHOD = "Hình thức"
COL_REF = "Mã tham chiếu"
COL_NOTE = "Ghi chú"

REQUIRED_COLS = [COL_CCCD, COL_AMOUNT, COL_DATE, COL_METHOD]
TEMPLATE_COLS = [
    COL_CCCD,
    COL_NAME,
    COL_AMOUNT,
    COL_DATE,
    COL_METHOD,
    COL_REF,
    COL_NOTE,
]

# Hình thức → PaymentMethod.code thực (seed fin20260131002): cash / bank_transfer.
# Input đã qua _norm_name (bỏ dấu + lower) TRƯỚC khi tra map → chỉ cần key KHÔNG dấu
# (key có dấu sẽ không bao giờ khớp).
METHOD_MAP = {
    "tien mat": "cash",
    "tm": "cash",
    "cash": "cash",
    "chuyen khoan": "bank_transfer",
    "ck": "bank_transfer",
    "bank_transfer": "bank_transfer",
}

CCCD_RE = re.compile(r"^\d{12}$")
MAX_IMPORT_ROWS = 5000
# Trần 1 khoản = max cột Numeric(15,2). Chặn ở parser để số quá lớn KHÔNG lọt
# xuống INSERT (payment_import_row.amount / Payment.amount) → DataError → 500.
MAX_AMOUNT = Decimal("9999999999999.99")
# Mã tham chiếu lưu vào Payment.reference_code / PaymentTransaction.external_reference
# = String(100). Ref dài hơn → DataError ở flush → 500 không bắt → chặn sớm thành
# lỗi dòng sạch.
MAX_REF_LEN = 100


class _NghiTrungBiChan(BusinessRuleViolation):
    """Dòng bị hàng rào nghi trùng giữ lại — KHÔNG phải dòng hỏng.

    Tách riêng khỏi ``BusinessRuleViolation`` vì hai thứ này cần kết cục khác
    nhau: dòng hỏng thật (số dư đổi, hồ sơ bị xoá) thành ``error`` và dừng ở
    đó, còn dòng này chỉ đang chờ kế toán soát — nó phải giữ trạng thái
    commit-được để lượt gửi lại kèm xác nhận còn chọn tới nó.
    """


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------
@dataclass
class RowDraft:
    """Một dòng file đã parse (chưa resolve)."""

    row_no: int
    citizen_id: str
    name: str
    amount: Optional[Decimal]
    payment_date: Optional[date]
    method_code: Optional[str]
    reference: str
    note: str
    raw: dict
    parse_error: Optional[str] = None


@dataclass
class Allocation:
    """Một phần phân bổ vào 1 invoice (đợt)."""

    invoice_id: int
    installment_no: int
    amount: Decimal


@dataclass
class RowResult:
    row_no: int
    #: Trục KIỂM. Không có trục GHI ở đây: bước xem trước KHÔNG ghi
    #: tiền, nên mọi câu hỏi về số phận lúc ghi đều chưa có nghĩa.
    validation_status: str  # matched | warned | error
    message: Optional[str] = None
    citizen_id: Optional[str] = None
    profile_id: Optional[int] = None
    fee_id: Optional[int] = None
    amount: Optional[Decimal] = None
    method_code: Optional[str] = None
    payment_date: Optional[date] = None
    reference: Optional[str] = None
    allocations: List[Allocation] = field(default_factory=list)
    raw: dict = field(default_factory=dict)


@dataclass
class PreviewResult:
    rows: List[RowResult]
    matched_count: int
    warned_count: int
    failed_count: int
    total_amount: Decimal


# ---------------------------------------------------------------------------
# Parse helpers
# ---------------------------------------------------------------------------
def _norm(s: Optional[str]) -> str:
    return (s or "").strip()


def _norm_name(s: Optional[str]) -> str:
    """Bỏ dấu + lower + gộp khoảng trắng để cross-check tên."""
    s = unicodedata.normalize("NFD", _norm(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.lower().split())


def _check_thousands_groups(int_groups: List[str], raw: str) -> None:
    """Xác thực nhóm phân cách nghìn của PHẦN NGUYÊN: nhóm đầu 1-3 chữ số, các nhóm
    sau ĐÚNG 3. Chặn '1.23.456' / '1,23,456' (nhóm giữa sai) bị nối thầm thành số."""
    if not (
        int_groups
        and 1 <= len(int_groups[0]) <= 3
        and all(len(g) == 3 for g in int_groups[1:])
    ):
        raise ValueError(f"số tiền không hợp lệ: '{raw}'")


def parse_amount_vn(raw: str) -> Decimal:
    """Parse số tiền kiểu VN. '.' = phân cách nghìn, ',' = thập phân.

    "7.200.000" -> 7200000 · "7.200.000,50" -> 7200000.50 · "7200000.0" (float
    string từ Excel, nhóm cuối 1 chữ số) -> 7200000. Heuristic: dấu phân cách ĐƠN
    ('.' hoặc ',') với nhóm cuối ĐÚNG 3 chữ số = phân cách nghìn ("500.000" và
    "500,000" đều = 500000 vì VND nguyên); nhóm cuối 1-2 chữ số = thập phân
    ("7,50" -> 7.50). Nhiều dấu = phân cách nghìn (US "7,200,000").
    """
    # Bỏ MỌI khoảng trắng kể cả NBSP   / narrow   — file ngân hàng/Excel
    # hay ghi '7 200 000' với non-breaking space (re \s khớp Unicode whitespace).
    s = re.sub(r"\s", "", _norm(raw))
    if not s:
        raise ValueError("trống")
    has_dot = "." in s
    has_comma = "," in s
    if has_dot and has_comma:
        if s.rfind(",") > s.rfind("."):  # ',' bên phải = thập phân (VN 7.200.000,50)
            _check_thousands_groups(s.split(",")[0].split("."), raw)
            s = s.replace(".", "").replace(",", ".")
        else:  # '.' bên phải = thập phân (US 7,200,000.00)
            _check_thousands_groups(s.split(".")[0].split(","), raw)
            s = s.replace(",", "")
    elif has_comma:
        if s.count(",") > 1:  # nhiều ',' = phân cách nghìn (US 7,200,000)
            _check_thousands_groups(s.split(","), raw)
            s = s.replace(",", "")
        elif len(s.rsplit(",", 1)[1]) == 3:  # 1 ',' + nhóm cuối 3 số -> nghìn
            # Đối xứng nhánh '.': VND nguyên nên '500,000' = 500000, KHÔNG phải thập
            # phân 500,000 -> 500.00 (P1: ghi nhầm 500đ thay vì 500.000đ).
            _check_thousands_groups(s.split(","), raw)
            s = s.replace(",", "")
        else:  # 1 ',' + nhóm cuối 1-2 số = thập phân VN (7,50 -> 7.50)
            s = s.replace(",", ".")
    elif has_dot:  # '.' nhập nhằng
        if len(s.rsplit(".", 1)[1]) == 3:  # nhóm cuối 3 chữ số -> nghìn
            _check_thousands_groups(s.split("."), raw)
            s = s.replace(".", "")
        # else: '.' là thập phân (float string) -> giữ nguyên
    try:
        val = Decimal(s)
    except InvalidOperation:
        raise ValueError(f"số tiền không hợp lệ: '{raw}'")
    # Cột Numeric(15,2) → chuẩn hóa về 2 chữ số thập phân (VND thực tế là số nguyên).
    # Tránh '…,505' (3 chữ số thập phân) lọt xuống BV-3 → lỗi/làm tròn ngầm khi INSERT.
    val = val.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if val <= 0:
        raise ValueError("số tiền phải > 0")
    if val > MAX_AMOUNT:  # chặn tràn cột Numeric(15,2) → DataError/500 ở pha persist
        raise ValueError(f"số tiền vượt giới hạn cho phép ({_money(MAX_AMOUNT)} đ)")
    return val


def parse_date_vn(raw: str) -> date:
    """Parse ngày thu. Chấp nhận:
    - text VN: ``dd/mm/yyyy`` / ``dd-mm-yyyy``
    - ISO: ``yyyy-mm-dd``
    - chuỗi datetime của Excel khi đọc ``dtype=str``: ``2026-09-05 00:00:00``
      (openpyxl trả ô Date thật → pandas str-hóa kèm giờ). Kế toán rất hay để
      Excel tự định dạng ô ngày → PHẢI cắt phần giờ, nếu không HÀNG LOẠT dòng
      bị false-error "ngày không hợp lệ".
    """
    s = _norm(raw)
    if not s:
        raise ValueError("thiếu ngày thu")
    s = re.split(r"[ T]", s, maxsplit=1)[0]  # bỏ phần giờ '... 00:00:00' / '...T...'
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(s, fmt).date()
        except ValueError:
            continue
        # strptime %Y nuốt năm 1–4 chữ số: '05/09/26' → năm 0026 âm thầm. BV-2 KHÔNG
        # lộ (row không có cột ngày → vẫn MATCHED) nhưng BV-3 ghi Payment năm 0026.
        if not (2000 <= d.year <= 2100):
            raise ValueError(f"năm phải đủ 4 chữ số (2000–2100): '{raw}'")
        return d
    raise ValueError(f"ngày không hợp lệ (cần dd/mm/yyyy): '{raw}'")


def file_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def parse_template(content: bytes, filename: str) -> List[RowDraft]:
    """Đọc file mẫu → list RowDraft. CCCD/ref/amount đọc **dtype=str** (giữ số 0
    đầu / không float). Lỗi từng dòng ghi vào ``parse_error`` (không raise)."""
    ext = (filename.rsplit(".", 1)[-1] or "").lower()
    if ext not in ("csv", "xlsx"):
        raise BadRequest("Chỉ hỗ trợ file .xlsx hoặc .csv")
    if not content:
        raise BadRequest("File rỗng")
    try:
        if ext == "csv":
            df = pd.read_csv(io.BytesIO(content), dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
        else:
            # Đọc MỌI sheet rồi chọn sheet ĐẦU TIÊN đủ cột bắt buộc — tránh chỉ đọc
            # sheet 0 (im lặng bỏ data khi nằm ở sheet sau, sau sheet "Hướng dẫn").
            sheets = pd.read_excel(
                io.BytesIO(content), engine="openpyxl", dtype=str, sheet_name=None
            )
            df = None
            for sheet_df in sheets.values():
                sheet_df.columns = [str(c).strip() for c in sheet_df.columns]
                if all(c in sheet_df.columns for c in REQUIRED_COLS):
                    df = sheet_df
                    break
            if df is None:  # không sheet nào đủ cột → lấy sheet đầu để báo lỗi cột
                df = next(iter(sheets.values()), pd.DataFrame())
    except Exception as exc:  # noqa: BLE001
        raise BadRequest(f"Không đọc được file: {exc}")

    df = df.fillna("")
    # Header hay thừa khoảng trắng (' '/' ') → strip để không trượt so khớp
    # cột bắt buộc rồi từ chối nhầm cả file (CSV đã strip ở trên; lặp lại vô hại).
    df.columns = [str(c).strip() for c in df.columns]
    # Strip có thể làm 2 cột khác nhau (vd 'Số CCCD' và 'Số CCCD ') TRÙNG tên →
    # row.get(col) trả pandas Series → rác vào amount/CCCD/raw. Từ chối rõ ràng thay
    # vì âm thầm parse sai (file cột trùng = nhập nhằng, kế toán phải sửa).
    _cols = list(df.columns)
    _dups = sorted({c for c in _cols if _cols.count(c) > 1})
    if _dups:
        raise BadRequest(
            f"File có cột trùng tên (sau khi bỏ khoảng trắng): {', '.join(_dups)}"
        )
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise BadRequest(f"File thiếu cột bắt buộc: {', '.join(missing)}")

    if len(df) > MAX_IMPORT_ROWS:
        raise BadRequest(
            f"File {len(df)} dòng vượt giới hạn {MAX_IMPORT_ROWS} dòng/lần import."
        )

    drafts: List[RowDraft] = []
    for i, row in df.iterrows():
        # row_no = SỐ DÒNG trên bảng tính kế toán nhìn thấy: header = dòng 1, nên
        # dòng dữ liệu đầu (index 0) = dòng 2 → báo lỗi khớp đúng vị trí trong Excel.
        row_no = int(i) + 2
        cells = {c: _norm(str(row.get(c, ""))) for c in df.columns}
        # Bỏ qua dòng TRỐNG HOÀN TOÀN (Excel hay để dòng trống ở cuối) → tránh
        # false-error "thiếu CCCD" hàng loạt.
        if not any(cells.values()):
            continue
        # raw = giá trị gốc (đã chuẩn hóa chuỗi) cho audit + pha commit re-parse.
        # KHÔNG sanitize-CSV ở ingest: sanitize là việc của tầng EXPORT — làm ở đây
        # sẽ chèn dấu ' vào reference/method → lệch preview vs tiền THỰC ghi (BV-3).
        raw = dict(cells)
        cccd = cells.get(COL_CCCD, "")
        name = cells.get(COL_NAME, "")
        ref = cells.get(COL_REF, "")
        note = cells.get(COL_NOTE, "")

        err = None
        amount = None
        try:
            amount = parse_amount_vn(str(row.get(COL_AMOUNT, "")))
        except ValueError as e:
            err = str(e)
        pay_date = None
        try:
            pay_date = parse_date_vn(str(row.get(COL_DATE, "")))
        except ValueError as e:
            err = err or str(e)
        method_raw = _norm_name(str(row.get(COL_METHOD, "")))
        method_code = METHOD_MAP.get(method_raw)
        if method_code is None:
            err = err or f"hình thức không hợp lệ: '{row.get(COL_METHOD, '')}'"
        # Ref dài hơn cột String(100) → DataError ở commit → 500. Chặn sớm = lỗi dòng.
        if len(ref) > MAX_REF_LEN:
            err = err or f"mã tham chiếu quá dài (tối đa {MAX_REF_LEN} ký tự)"

        drafts.append(
            RowDraft(
                row_no=row_no,
                citizen_id=cccd,
                name=name,
                amount=amount,
                payment_date=pay_date,
                method_code=method_code,
                reference=ref,
                note=note,
                raw=raw,
                parse_error=err,
            )
        )
    return drafts


# ---------------------------------------------------------------------------
# Resolve + validate (READ-ONLY — không ghi Payment)
# ---------------------------------------------------------------------------
async def resolve_and_validate(
    db: AsyncSession,
    drafts: List[RowDraft],
    academic_year: int,
    semester_no: int,
    unit_id: Optional[int],
) -> PreviewResult:
    """Đối chiếu từng dòng → MATCHED/WARNING/ERROR. KHÔNG ghi gì.

    - khóa = CCCD + academic_year (unique) + IDOR unit-scope + bỏ lead xóa mềm
    - fee tuition kỳ S, status NOT IN (cancelled, waived)
    - FIFO principal-first: principal_remaining = invoice.amount - paid (KHÔNG
      remaining_amount vì gồm penalty)
    - sổ phân bổ in-batch per-invoice → chống double-alloc khi 2 dòng cùng CCCD+kỳ
    """
    # --- Prefetch theo lô: 3 query thay vì 3N (tránh N+1 giữ 1 connection lâu +
    # cạn pool; dedup CCCD trùng miễn phí). (citizen_id, năm) & (profile, tuition,
    # kỳ) đều UNIQUE nên map 1-1 an toàn. ---
    cccds = {
        d.citizen_id
        for d in drafts
        if not d.parse_error and CCCD_RE.match(d.citizen_id or "")
    }
    profiles = await _fetch_profiles(db, cccds, academic_year, unit_id)
    fees = await _fetch_tuition_fees(db, [p.id for p in profiles.values()], semester_no)
    fee_ids = [f.id for f in fees.values()]
    invoices_by_fee = await _fetch_payable_invoices(db, fee_ids)
    # Chỉ cần status-set cho fee KHÔNG có hóa đơn payable (nhánh hiếm "đợt nháp / đã thu
    # đủ / chưa có đợt"). File bình thường mọi fee đều payable → list rỗng → helper
    # return {} ngay, KHÔNG tốn query trên đường happy-path.
    no_payable_fee_ids = [fid for fid in fee_ids if fid not in invoices_by_fee]
    invoice_statuses = await _fetch_invoice_status_sets(db, no_payable_fee_ids)

    # #6 — nghi TRÙNG GIAO DỊCH (re-import sao kê chồng ngày → tạo Payment thứ 2
    # cho cùng giao dịch khi hồ sơ còn nợ). reference_code là FREE-TEXT (phiếu thu
    # PT-/mã hồ sơ HS-/cash), TÁI DÙNG hợp lệ được ở thu góp KHÁC số tiền → KHÔNG
    # đặt unique. Tín hiệu re-import = tổng payment verified đã ghi cho (hồ sơ, ref)
    # KHỚP số tiền dòng. Prefetch 1 query group; bỏ ref rỗng/synthetic. CẢNH BÁO
    # mềm (không chặn — kế toán quyết).
    existing_ref_sums: Dict[Tuple[int, str], Decimal] = {}
    _prof_ids = [p.id for p in profiles.values()]
    if _prof_ids:
        # Scope ĐÚNG fee mà dòng import sẽ áp (tuition + ĐÚNG học kỳ S): nếu gộp mọi
        # loại phí/kỳ, ref chung (vd lệ phí 500k cùng ref 'TIENMAT' với học phí 500k)
        # sẽ false-positive "nghi trùng". sum-on-collision phòng 2 ref raw _norm về
        # cùng key.
        for pid, ref, total in (
            await db.execute(
                select(
                    models.AdmissionProfile.id,
                    Payment.reference_code,
                    func.sum(Payment.amount),
                )
                .join(Invoice, Invoice.id == Payment.invoice_id)
                .join(Fee, Fee.id == Invoice.fee_id)
                .join(
                    models.AdmissionProfile,
                    models.AdmissionProfile.id == Fee.admission_profile_id,
                )
                .where(
                    models.AdmissionProfile.id.in_(_prof_ids),
                    Fee.fee_type == "tuition",
                    Fee.semester_no == semester_no,
                    Payment.status == PaymentStatusEnum.verified.value,
                    Payment.reference_code.isnot(None),
                    Payment.reference_code != "",
                )
                .group_by(models.AdmissionProfile.id, Payment.reference_code)
            )
        ).all():
            _k = (pid, _norm(ref))
            existing_ref_sums[_k] = existing_ref_sums.get(_k, Decimal("0")) + total

    # G2 — code hình thức ĐANG hoạt động (mirror commit:1017). Parser đã loại method
    # text lạ (:346); đây bắt method map-OK nhưng PaymentMethod inactive/missing → ERROR
    # ngay ở preview (đối xứng commit, hết "khớp giả" rồi commit fail).
    active_methods = {
        c
        for (c,) in (
            await db.execute(
                select(PaymentMethod.code).where(
                    PaymentMethod.code.in_(["cash", "bank_transfer"]),
                    PaymentMethod.is_active.is_(True),
                )
            )
        ).all()
    }
    # G1 — đếm CCCD trùng (và CCCD+tiền trùng) trong file. Cái chốt money DUY NHẤT
    # ("không thu vượt nợ gốc") KHÔNG bắt được trùng-còn-trong-nợ → thu khống. Cảnh báo
    # (WARNING, không chặn — thu nhiều đợt là hợp lệ; kế toán quyết).
    _valid = [
        d for d in drafts
        if not d.parse_error and CCCD_RE.match(d.citizen_id or "")
    ]
    cccd_counts = Counter(d.citizen_id for d in _valid)
    cccd_amount_counts = Counter((d.citizen_id, d.amount) for d in _valid)

    # Sổ phân bổ trong batch: invoice_id -> tổng đã phân bổ ở các dòng TRƯỚC.
    batch_alloc: Dict[int, Decimal] = {}
    results: List[RowResult] = []
    # (B3) hàng đợi dò trùng với phiếu ĐÃ GHI — hỏi một lượt sau vòng lặp.
    # Khoá theo ``row_no`` chứ không theo (khoản phí, tiền, ngày): một tệp thu
    # hai lần cùng số tiền cho cùng hồ sơ trong cùng ngày là ca có thật, gom
    # theo bộ ba đó thì hai dòng dính làm một.
    khoa_do_trung: List[Tuple[int, int, Decimal, datetime]] = []
    res_theo_dong: Dict[int, RowResult] = {}

    for d in drafts:
        res = RowResult(
            row_no=d.row_no,
            validation_status=PaymentImportRowStatusEnum.error.value,
            citizen_id=d.citizen_id,
            amount=d.amount,
            method_code=d.method_code,
            payment_date=d.payment_date,
            reference=d.reference or None,
            raw=d.raw,
        )

        # (0) lỗi parse
        if d.parse_error:
            res.message = d.parse_error
            results.append(res)
            continue
        # (6) CCCD định dạng
        if not CCCD_RE.match(d.citizen_id):
            res.message = "CCCD phải đúng 12 chữ số" if d.citizen_id else "thiếu CCCD"
            results.append(res)
            continue
        # (G2) hình thức đã map (parser) nhưng PaymentMethod inactive/missing → ERROR
        # sớm (commit:1017 cũng chặn → tránh "khớp giả" ở preview).
        if d.method_code not in active_methods:
            res.message = "hình thức chưa được kích hoạt trong hệ thống"
            results.append(res)
            continue

        # (2/6) hồ sơ theo CCCD + năm + IDOR unit-scope (bỏ lead xóa mềm) — prefetch
        profile = profiles.get(d.citizen_id)
        if profile is None:
            res.message = (
                f"không tìm thấy hồ sơ CCCD {d.citizen_id} (năm {academic_year})"
            )
            results.append(res)
            continue
        res.profile_id = profile.id

        # (3) học phí kỳ S, active-only — prefetch
        fee = fees.get(profile.id)
        if fee is None:
            res.message = f"hồ sơ chưa được thiết lập học phí HK{semester_no}"
            results.append(res)
            continue
        res.fee_id = fee.id

        # (4/5) phân bổ FIFO principal-first, trừ sổ in-batch — prefetch
        warnings: List[str] = []
        invoices = invoices_by_fee.get(fee.id, [])
        if not invoices:
            # Tách rõ hành động cho kế toán (thay vì gộp 'nháp hoặc đã thu đủ').
            # Ưu tiên 'nháp' khi vừa có đợt nháp vừa có đợt đã thu đủ — vì còn việc
            # phải làm: phát hành đợt nháp để thu tiếp.
            statuses = invoice_statuses.get(fee.id, set())
            if InvoiceStatusEnum.draft.value in statuses:
                res.message = "đợt còn nháp — kế toán cần phát hành hóa đơn trước"
            elif InvoiceStatusEnum.paid.value in statuses:
                res.message = "học phí đã thu đủ"
            else:
                res.message = "chưa có đợt hóa đơn để thu"
            results.append(res)
            continue

        remaining_total = Decimal("0")
        avail: List[Tuple[Invoice, Decimal]] = []  # (invoice, available_principal)
        for inv in invoices:
            principal_rem = (inv.amount or Decimal("0")) - (
                inv.paid_amount or Decimal("0")
            )
            principal_rem -= batch_alloc.get(inv.id, Decimal("0"))
            if principal_rem < 0:
                principal_rem = Decimal("0")
            avail.append((inv, principal_rem))
            remaining_total += principal_rem

        if d.amount > remaining_total:
            res.message = (
                f"thu {_money(d.amount)} vượt tổng còn nợ gốc HK{semester_no} "
                f"({_money(remaining_total)})"
            )
            results.append(res)
            continue

        # cross-check tên (warning)
        if d.name and _norm_name(d.name) != _norm_name(
            profile.lead.full_name if profile.lead else ""
        ):
            warnings.append(
                f"tên lệch: file '{d.name}' vs hồ sơ "
                f"'{profile.lead.full_name if profile.lead else ''}'"
            )

        # cảnh báo ngày thu lệch xa năm học (gõ nhầm năm '2027' thay '2026', hoặc
        # nhầm dd/mm vs mm/dd) — surface ở preview để kế toán soát, KHÔNG chặn cứng
        # (thu sớm/muộn trong khoảng [năm học, +1] là hợp lệ).
        if d.payment_date and abs(d.payment_date.year - academic_year) > 1:
            warnings.append(
                f"ngày thu {d.payment_date:%d/%m/%Y} lệch xa năm học {academic_year}"
            )

        # (G1) CCCD xuất hiện nhiều dòng trong file → cảnh báo (không chặn). Nhấn
        # mạnh khi CÙNG số tiền (nghi copy nhầm → thu khống), vì chốt "không vượt
        # nợ" không bắt được.
        dup_n = cccd_counts.get(d.citizen_id, 0)
        if dup_n > 1:
            if cccd_amount_counts.get((d.citizen_id, d.amount), 0) > 1:
                warnings.append(
                    f"CCCD xuất hiện {dup_n} dòng — CÙNG số tiền, nghi copy nhầm"
                )
            else:
                warnings.append(
                    f"CCCD xuất hiện {dup_n} dòng trong file — kiểm tra trùng "
                    "(nếu thu nhiều đợt thì bỏ qua)"
                )

        # (#6) nghi TRÙNG GIAO DỊCH với payment ĐÃ GHI (re-import) — prefetch ở trên.
        # Khớp (hồ sơ, mã tham chiếu, tổng đã ghi == số tiền dòng): thu góp hợp lệ
        # khác số tiền nên KHÔNG dính. Bỏ ref rỗng/synthetic (không dedup được).
        _ref = _norm(d.reference or "")
        _da_bao_trung_theo_ref = False
        if _ref and not _ref.startswith(("BULK-", "PAY-")):
            _prior = existing_ref_sums.get((profile.id, _ref))
            if _prior is not None and _prior == d.amount:
                _da_bao_trung_theo_ref = True
                warnings.append(
                    f"nghi trùng giao dịch: mã tham chiếu '{_ref}' + số tiền "
                    f"{_money(d.amount)} đã ghi nhận cho hồ sơ này — kiểm tra "
                    "trước khi thu lại (re-import?)"
                )

        # (B3) Góc mà luật theo mã tham chiếu ở trên KHÔNG thấy: cùng khoản phí,
        # cùng số tiền, gần ngày, nhưng mã tham chiếu khác nhau — hoặc rỗng, hoặc
        # tự sinh 'BULK-'. Đây chính là hình dạng của một tệp bị nhập hai lần sau
        # khi ai đó sửa cột mã.
        #
        # Xếp hàng lại để hỏi MỘT lượt cho cả tệp sau vòng lặp: một tệp có vài
        # trăm dòng, hỏi từng dòng là N+1 mà cả hàm này dựng lên để tránh.
        #
        # Chỉ xếp hàng khi luật theo mã CHƯA bắn cho dòng này: hai câu cảnh báo
        # cho cùng một nghi ngờ chỉ làm kế toán đọc lướt cả hai.
        if d.payment_date and not _da_bao_trung_theo_ref:
            khoa_do_trung.append(
                (
                    d.row_no,
                    fee.id,
                    d.amount,
                    datetime(
                        d.payment_date.year,
                        d.payment_date.month,
                        d.payment_date.day,
                        tzinfo=timezone.utc,
                    ),
                )
            )
            res_theo_dong[d.row_no] = res

        # FIFO allocate
        left = d.amount
        for inv, principal_rem in avail:
            if left <= 0:
                break
            if principal_rem <= 0:
                continue
            take = min(left, principal_rem)
            res.allocations.append(
                Allocation(
                    invoice_id=inv.id, installment_no=inv.installment_no, amount=take
                )
            )
            batch_alloc[inv.id] = batch_alloc.get(inv.id, Decimal("0")) + take
            left -= take

        if len(res.allocations) > 1:
            warnings.append(f"thu phân bổ {len(res.allocations)} đợt")

        if warnings:
            res.validation_status = PaymentImportRowStatusEnum.warned.value
            res.message = " · ".join(warnings)
        else:
            res.validation_status = PaymentImportRowStatusEnum.matched.value
        results.append(res)

    # (B3) Một lượt hỏi cho cả tệp, dùng chung luật dò trùng với đường ghi tay.
    if khoa_do_trung:
        ung_vien = await PaymentRepository(db).find_duplicate_candidates_bulk(
            keys=khoa_do_trung,
            window_days=WINDOW_DO_TRUNG_NGAY,
        )
        for row_no, payment_ids in ung_vien.items():
            r = res_theo_dong.get(row_no)
            if r is None or not payment_ids:
                continue
            # Nêu vài mã phiếu để kế toán tra được, nhưng không đổ cả danh sách:
            # câu cảnh báo dài quá thì không ai đọc, mà số lượng mới là thứ
            # quyết định có nên dừng lại hay không.
            ma_phieu = ", ".join(f"#{i}" for i in payment_ids[:3])
            neu_them = "…" if len(payment_ids) > 3 else ""
            cau = (
                f"nghi trùng với {len(payment_ids)} phiếu đã ghi cho cùng khoản "
                f"phí — cùng số tiền {_money(r.amount)}, lệch không quá "
                f"{WINDOW_DO_TRUNG_NGAY} ngày ({ma_phieu}{neu_them})"
            )
            r.message = f"{r.message} · {cau}" if r.message else cau
            # Chỉ NÂNG matched → warned. Dòng đang ở trạng thái lỗi thì lỗi đó
            # mới là việc kế toán phải xử lý trước, đừng che nó bằng cảnh báo.
            if r.validation_status == PaymentImportRowStatusEnum.matched.value:
                r.validation_status = PaymentImportRowStatusEnum.warned.value

    matched = sum(
        1 for r in results if r.validation_status == PaymentImportRowStatusEnum.matched.value
    )
    warned = sum(
        1 for r in results if r.validation_status == PaymentImportRowStatusEnum.warned.value
    )
    failed = sum(
        1 for r in results if r.validation_status == PaymentImportRowStatusEnum.error.value
    )
    total = sum(
        (r.amount or Decimal("0"))
        for r in results
        if r.validation_status != PaymentImportRowStatusEnum.error.value
    )
    return PreviewResult(
        rows=results,
        matched_count=matched,
        warned_count=warned,
        failed_count=failed,
        total_amount=total,
    )


async def _fetch_profiles(
    db: AsyncSession,
    citizen_ids: set,
    academic_year: int,
    unit_id: Optional[int],
) -> Dict[str, models.AdmissionProfile]:
    """Lô CCCD → {citizen_id: hồ sơ} trong 1 query (thay N). IDOR unit-scope, bỏ
    lead xóa mềm, eager lead cho cross-check tên. (citizen_id, academic_year) UNIQUE
    → mỗi CCCD tối đa 1 hồ sơ/năm nên map 1-1 an toàn.

    Lọc ``Lead.deleted_at IS NULL`` — nếu không, hồ sơ của lead đã xóa mềm vẫn khớp
    CCCD → false MATCH → BV-3 ghi tiền vào hồ sơ "ma".
    """
    from sqlalchemy.orm import selectinload

    if not citizen_ids:
        return {}
    stmt = (
        select(models.AdmissionProfile)
        .join(models.Lead)
        .options(selectinload(models.AdmissionProfile.lead))
        .where(
            models.AdmissionProfile.citizen_id.in_(list(citizen_ids)),
            models.AdmissionProfile.academic_year == academic_year,
            models.Lead.deleted_at.is_(None),
        )
    )
    if unit_id is not None:
        stmt = stmt.where(models.Lead.unit_id == unit_id)
    rows = (await db.execute(stmt)).scalars().all()
    return {p.citizen_id: p for p in rows}


async def _fetch_tuition_fees(
    db: AsyncSession,
    profile_ids: List[int],
    semester_no: int,
) -> Dict[int, Fee]:
    """Lô profile_id → {profile_id: fee} (tuition kỳ S, active-only) trong 1 query.
    (profile, fee_type=tuition, semester_no) UNIQUE → mỗi profile tối đa 1 fee kỳ S.

    Không cần lọc unit: hồ sơ đã resolve dưới unit-scope nên fee đã trong phạm vi.
    """
    if not profile_ids:
        return {}
    stmt = select(Fee).where(
        Fee.admission_profile_id.in_(profile_ids),
        Fee.fee_type == "tuition",
        Fee.semester_no == semester_no,
        Fee.status.notin_([FeeStatusEnum.cancelled.value, FeeStatusEnum.waived.value]),
    )
    rows = (await db.execute(stmt)).scalars().all()
    return {f.admission_profile_id: f for f in rows}


async def _fetch_payable_invoices(
    db: AsyncSession,
    fee_ids: List[int],
) -> Dict[int, List[Invoice]]:
    """Lô fee_id → {fee_id: [invoice PAYABLE]} sắp installment_no, trong 1 query."""
    if not fee_ids:
        return {}
    stmt = (
        select(Invoice)
        .where(
            Invoice.fee_id.in_(fee_ids),
            Invoice.status.in_(list(PAYABLE_INVOICE_STATUSES)),
        )
        .order_by(Invoice.fee_id, Invoice.installment_no)
    )
    by_fee: Dict[int, List[Invoice]] = {}
    for inv in (await db.execute(stmt)).scalars().all():
        by_fee.setdefault(inv.fee_id, []).append(inv)
    return by_fee


async def _id_invoice_cua_fee(
    db: AsyncSession,
    fee_ids: List[int],
    unit_id: Optional[int],
) -> Set[int]:
    """Tập id invoice của các fee — CÙNG bộ lọc IDOR với hàm khoá.

    Dùng để đối chiếu tập đã khoá với tập hiện tại. Phải cùng bộ lọc, nếu không
    một invoice thuộc đơn vị khác sẽ hiện ra như "đợt mới phát hành" và mọi lượt
    commit đều chết oan.
    """
    if not fee_ids:
        return set()
    stmt = (
        select(Invoice.id)
        .join(Fee, Invoice.fee_id == Fee.id)
        .join(models.AdmissionProfile)
        .join(models.Lead)
        .where(Invoice.fee_id.in_(fee_ids))
    )
    if unit_id is not None:
        stmt = stmt.where(models.Lead.unit_id == unit_id)
    return set((await db.execute(stmt)).scalars().all())


async def _fetch_invoice_status_sets(
    db: AsyncSession,
    fee_ids: List[int],
) -> Dict[int, set]:
    """Lô fee_id → {fee_id: set(status MỌI invoice)} (1 query). Phân biệt nhánh
    KHÔNG-payable: đợt còn nháp (draft → cần phát hành) vs đã thu đủ (paid) vs chưa
    có đợt → message preview rõ hành động thay vì gộp 'nháp hoặc đã thu đủ'."""
    if not fee_ids:
        return {}
    stmt = select(Invoice.fee_id, Invoice.status).where(Invoice.fee_id.in_(fee_ids))
    out: Dict[int, set] = {}
    for fee_id, status in (await db.execute(stmt)).all():
        out.setdefault(fee_id, set()).add(status)
    return out


def _money(v: Decimal) -> str:
    return f"{v:,.0f}".replace(",", ".")


# ---------------------------------------------------------------------------
# Persist preview batch + orchestrator (BV-2)
# ---------------------------------------------------------------------------
async def preview_import(
    db: AsyncSession,
    *,
    content: bytes,
    filename: str,
    academic_year: int,
    semester_no: int,
    created_by_id: int,
    unit_id: Optional[int],
) -> Tuple[PaymentImportBatch, PreviewResult]:
    """Pha 1 (dry-run): parse → resolve/validate READ-ONLY → lưu batch 'preview'.

    KHÔNG ghi Payment. Trả ``(batch, preview)`` để router build response giàu thông
    tin (kèm phân bổ FIFO dự kiến — chỉ in-memory, KHÔNG persist vào row).
    """
    drafts = parse_template(content, filename)
    preview = await resolve_and_validate(
        db, drafts, academic_year, semester_no, unit_id
    )
    batch = await create_preview_batch(
        db,
        preview=preview,
        academic_year=academic_year,
        semester_no=semester_no,
        file_name=filename,
        file_sha256_hex=file_sha256(content),
        created_by_id=created_by_id,
    )
    return batch, preview


async def create_preview_batch(
    db: AsyncSession,
    *,
    preview: PreviewResult,
    academic_year: int,
    semester_no: int,
    file_name: str,
    file_sha256_hex: str,
    created_by_id: int,
) -> PaymentImportBatch:
    """Lưu kết quả preview thành 1 ``PaymentImportBatch`` 'preview' + các
    ``PaymentImportRow`` (audit + để pha commit tham chiếu ``batch_id``).

    Tôn trọng partial-unique ``uq_payment_import_batch_active_file`` (tối đa 1 batch
    còn-hiệu-lực / file):
    - cùng sha256 đã ``committed`` → ``ConflictError`` (chống double-count; phải đảo
      (void) lô đó trước khi import lại).
    - ``preview`` cũ cùng sha256 → xóa & tạo lại (DB có thể đã đổi giữa 2 lần xem).
    """
    existing = (
        await db.execute(
            select(PaymentImportBatch).where(
                PaymentImportBatch.file_sha256 == file_sha256_hex,
                PaymentImportBatch.status != PaymentImportBatchStatusEnum.void.value,
            )
        )
    ).scalar_one_or_none()
    if existing is not None:
        if existing.status == PaymentImportBatchStatusEnum.committed.value:
            ngay = (
                f" ngày {existing.committed_at:%d/%m/%Y}"
                if existing.committed_at
                else ""
            )
            raise ConflictError(
                f"File này đã được import và ghi nhận ở lô #{existing.id}{ngay}. "
                "Hãy đảo (void) lô đó trước nếu muốn import lại."
            )
        # ⚠️ `status` MỘT MÌNH không đủ để kết luận "chưa ghi đồng nào".
        #
        # Lô còn dòng bị hàng rào nghi trùng giữ lại vẫn mang `preview` trong
        # khi các dòng khác ĐÃ ghi tiền và đang giữ `payment_ids`. Đó là trạng
        # thái bình thường của mọi lượt commit có `duplicate_review_required`,
        # không phải ngoại lệ hiếm.
        #
        # Xoá lô ở trạng thái đó không phải "làm mới ảnh chụp" mà là xoá bằng
        # chứng: `Payment` vẫn nằm trong sổ, tiền vẫn ở invoice/fee, nhưng hàng
        # nối chúng với dòng file biến mất — hết đường lần ngược "khoản này vào
        # sổ từ đâu", và lượt import sau không thấy dòng cũ nên có thể ghi lần
        # hai.
        so_dong_da_ghi = (
            await db.execute(
                select(func.count())
                .select_from(PaymentImportRow)
                .where(
                    PaymentImportRow.batch_id == existing.id,
                    PaymentImportRow.commit_status
                    == PaymentImportCommitStatusEnum.committed.value,
                )
            )
        ).scalar_one()
        if so_dong_da_ghi:
            # ⚠️ KHÔNG mời "void lô" ở đây: `void_batch` chỉ nhận lô đã
            # `committed`. Lô này vẫn `preview`, nên void là một lối ra KHÔNG
            # tồn tại — chỉ đường vào một thông báo lỗi thứ hai.
            raise ConflictError(
                f"Lô #{existing.id} của file này đã ghi tiền {so_dong_da_ghi} "
                "dòng và còn dòng chờ soát. Hãy mở lô này để xử lý các dòng "
                "còn chờ soát. Sau khi hoàn tất, bạn có thể đảo lô trước khi "
                "nhập lại nếu cần."
            )

        # preview cũ CHƯA ghi đồng nào → thay thế (cascade xóa rows) rồi tạo lại
        await db.delete(existing)
        await db.flush()

    batch = PaymentImportBatch(
        academic_year=academic_year,
        semester_no=semester_no,
        file_name=(file_name or "")[:255],
        file_sha256=file_sha256_hex,
        row_count=len(preview.rows),
        matched_count=preview.matched_count,
        warned_count=preview.warned_count,
        failed_count=preview.failed_count,
        total_amount=preview.total_amount,
        status=PaymentImportBatchStatusEnum.preview.value,
        created_by_id=created_by_id,
    )
    db.add(batch)
    try:
        await db.flush()
    except IntegrityError as exc:
        # Race: 2 upload cùng file đồng thời lọt qua pre-check → partial-unique chặn.
        await db.rollback()
        raise ConflictError(
            "File đang được xử lý bởi một thao tác khác. Vui lòng thử lại."
        ) from exc

    for r in preview.rows:
        # citizen_id cột String(12): chỉ lưu khi đúng 12 số, ngược lại None (CMND 9
        # số / rác dài hơn sẽ tràn cột) — giá trị gốc vẫn nằm trong ``raw`` để đối soát.
        cid = r.citizen_id if (r.citizen_id and CCCD_RE.match(r.citizen_id)) else None
        db.add(
            PaymentImportRow(
                batch_id=batch.id,
                row_no=r.row_no,
                citizen_id=cid,
                raw=r.raw or {},
                resolved_profile_id=r.profile_id,
                resolved_fee_id=r.fee_id,
                validation_status=r.validation_status,
                # Trục GHI đặt ngay từ bước xem trước, theo đúng bảng chân trị:
                # dòng hỏng từ khâu đọc KHÔNG có gì để ghi (`not_applicable`),
                # còn lại là đang chờ lượt ghi. Để mặc định `pending` cho cả
                # dòng lỗi thì nó nằm mãi trong nhóm "chờ ghi" — một hàng đợi
                # không bao giờ vơi, và ràng buộc hai trục ở cơ sở dữ liệu cũng
                # từ chối ngay.
                commit_status=(
                    PaymentImportCommitStatusEnum.not_applicable.value
                    if r.validation_status == PaymentImportRowStatusEnum.error.value
                    else PaymentImportCommitStatusEnum.pending.value
                ),
                message=r.message,
                amount=r.amount,
                payment_ids=None,
            )
        )
    await db.flush()
    return batch


# ---------------------------------------------------------------------------
# Template generation (BV-2) — ở SERVICE để router chỉ stream bytes (CLAUDE.md:
# "Router: Input/Output ONLY") + test được generator.
# ---------------------------------------------------------------------------
_TEMPLATE_EXAMPLE = [
    "001234567890",  # Số CCCD (text, giữ số 0 đầu)
    "Nguyễn Văn An",  # Họ và tên học sinh
    "7.200.000",  # Số tiền thu (VNĐ) — GỐC học phí
    "05/09/2026",  # Ngày thu dd/mm/yyyy
    "TM",  # Hình thức: TM=tiền mặt / CK=chuyển khoản
    "PT-2026-0001",  # Mã tham chiếu (tùy chọn)
    "Thu học phí HK1",  # Ghi chú (tùy chọn)
]
_TEMPLATE_DESCRIPTIONS = [
    "Số CCCD 12 chữ số — ĐỊNH DẠNG TEXT (giữ số 0 đầu). Bắt buộc.",
    "Họ và tên học sinh (để đối chiếu — lệch chỉ cảnh báo).",
    "Số tiền đã thu (VNĐ), GỐC học phí. VD: 7.200.000. Bắt buộc.",
    "Ngày thu, định dạng dd/mm/yyyy. Bắt buộc.",
    "Hình thức: TM = tiền mặt, CK = chuyển khoản. Bắt buộc.",
    "Mã tham chiếu phiếu thu/UNC (tùy chọn) — định dạng TEXT.",
    "Ghi chú (tùy chọn).",
]
# Giữ tên cũ làm alias để không phải sửa mọi call site trong file này;
# nguồn sự thật là app/constants/export_formats.py.
_XLSX_MEDIA = XLSX_MEDIA_TYPE


def build_template(fmt: str) -> Tuple[bytes, str, str]:
    """Sinh file mẫu import → ``(content, media_type, filename)``.

    - CSV: prepend BOM (``\\ufeff``) → Excel locale VN không đọc mojibake header
      tiếng Việt (không BOM → ANSI → kế toán điền & upload lại → lệch cột → lỗi).
    - XLSX: cột CCCD + Mã tham chiếu ``number_format='@'`` (TEXT) tới HẾT vùng nhập
      (``MAX_IMPORT_ROWS``) — không chỉ vài dòng đầu, nếu không file > N dòng sẽ mất
      số 0 đầu CCCD từ vùng chưa định dạng → khớp nhầm người.
    """
    if (fmt or "").lower() == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(TEMPLATE_COLS)
        writer.writerow(_TEMPLATE_EXAMPLE)
        # BOM (utf-8-sig) để Excel locale VN không đọc mojibake header tiếng Việt.
        content = (CSV_UTF8_BOM + buf.getvalue()).encode("utf-8")
        return content, CSV_MEDIA_TYPE, "mau_import_thu_hoc_phi.csv"

    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Thu hoc phi"
    ws.append(TEMPLATE_COLS)

    header_fill = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    for idx, cell in enumerate(ws[1]):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment(_TEMPLATE_DESCRIPTIONS[idx], "QLTS")

    ws.append(_TEMPLATE_EXAMPLE)

    # number_format='@' (TEXT) cho cột CCCD + Mã tham chiếu tới HẾT vùng nhập.
    cccd_col = TEMPLATE_COLS.index(COL_CCCD) + 1
    ref_col = TEMPLATE_COLS.index(COL_REF) + 1
    for col_idx in (cccd_col, ref_col):
        letter = ws.cell(row=1, column=col_idx).column_letter
        for r in range(2, MAX_IMPORT_ROWS + 2):
            ws[f"{letter}{r}"].number_format = TEXT_NUMBER_FORMAT

    widths = [16, 22, 18, 14, 12, 18, 24]
    for i, w in enumerate(widths):
        ws.column_dimensions[ws.cell(row=1, column=i + 1).column_letter].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), _XLSX_MEDIA, "mau_import_thu_hoc_phi.xlsx"


# ---------------------------------------------------------------------------
# BV-3 — Ghi tiền (auto-verify) + commit lô
#
# ``get_system_user`` + ``auto_verify_payment`` là GENERIC (đặt tạm ở đây; tách sang
# payment_service.py khi refactor lệ-phí dùng chung — plan §8). auto_verify KHÔNG
# dispatch notification / lead-sync → ``commit_batch`` GỘP 1 lần/hồ-sơ (tránh fan-out
# 3N của verify_payment per-row).
# ---------------------------------------------------------------------------
async def get_system_user(db: AsyncSession) -> "models.User":
    """Tài khoản kỹ thuật 'system' (checker cho auto-verify by policy).

    DÙNG CHUNG resolver canonical ``_get_system_application_fee_user``
    (admission_service) → **1 NGUỒN** fingerprint cho mọi luồng auto-verify by-policy
    (lệ phí + bulk import), hết drift. Canonical chặt hơn (thêm check không có
    UserUnitAssignment active + reuse ``_is_bcrypt_hash``). Maker-checker thỏa vì
    importer (kế toán) ≠ system_user (``chk_payment_no_self_approval``).

    (Eventual home: ``payment_service`` — plan §8.)
    """
    from app.services.admission_service import _get_system_application_fee_user

    return await _get_system_application_fee_user(db)


async def auto_verify_payment(
    db: AsyncSession,
    *,
    invoice: Invoice,
    fee: Fee,
    method_id: int,
    amount: Decimal,
    payment_date: datetime,
    reference: Optional[str],
    importer_id: int,
    system_user: "models.User",
    idempotency_key: str,
) -> Optional[Payment]:
    """Tạo Payment 'verified' (maker=importer, checker=system_user) áp vào ``invoice``
    ĐÃ TỒN TẠI + cập nhật invoice/fee + PaymentTransaction (audit, neo idempotency).

    Replicate money-math của ``verify_payment`` (payment_service.py:296-335) nhưng tạo
    Payment verified NGAY (không từ pending) và KHÔNG dispatch/lead-sync. Idempotent:
    idempotency_key đã có → trả ``None`` (re-commit an toàn).

    ⚠️ Caller PHẢI đã get_for_update + refresh ``invoice`` rồi ``fee``.

    Ở đây chỉ có ĐUÔI ``invoice → fee`` của thứ tự khoá chung, và điều đó đúng: hàm
    này TẠO một Payment mới ở trạng thái 'verified' ngay tại chỗ, nên không có hàng
    payment nào đang tồn tại để tranh chấp — không có bậc "payment" để khoá trước.
    Thứ tự đầy đủ toàn hệ là batch → payment → invoice → fee; đường bulk vào từ bậc
    thứ ba trở đi nên vẫn nhất quán, không đảo chiều với ``verify_payment``.
    (Đừng đọc dòng này thành "verify cũng chỉ khoá invoice→fee" — verify khoá payment
    trước, xem payment_service.verify_payment.)
    """
    dup = (
        await db.execute(
            select(PaymentTransaction.id).where(
                PaymentTransaction.idempotency_key == idempotency_key
            )
        )
    ).scalar_one_or_none()
    if dup is not None:
        return None

    now = datetime.now(timezone.utc)
    # Đổi ngành: snapshot ngành ghi nhận doanh thu (bất biến) — bulk auto-verify
    # stamp ngay (tuition-only). fee đã được caller lock/refresh.
    from app.services.fee_calculation_service import recognized_major_id_for_fee
    payment = Payment(
        invoice_id=invoice.id,
        method_id=method_id,
        amount=amount,
        reference_code=reference,
        status=PaymentStatusEnum.verified.value,
        payment_date=payment_date,
        verified_at=now,
        created_by_id=importer_id,
        verified_by_id=system_user.id,
        intent_id=None,
        notes="Bulk import thu học phí — auto-verify (system_user)",
        recognized_major_id=recognized_major_id_for_fee(fee),
    )
    db.add(payment)
    await db.flush()

    # Money-math GỐC = verify_payment (1 nguồn sự thật chung, tránh 2 đường ghi tiền
    # trôi dạt). is_fully_paid GỒM penalty → trả đủ GỐC nhưng còn phạt thì 'partial'.
    from app.services.payment_service import apply_verified_payment_balances

    fee_balance_before, fee_remaining = apply_verified_payment_balances(
        invoice=invoice, fee=fee, amount=amount, now=now
    )

    db.add(
        PaymentTransaction(
            payment_id=payment.id,
            fee_id=fee.id,
            transaction_type=TransactionTypeEnum.payment.value,
            amount=amount,
            balance_before=fee_balance_before,
            balance_after=fee_remaining,
            external_reference=reference,
            gateway_response={
                "verification_mode": "bulk_import",
                "idempotency_key": idempotency_key,
                "importer_id": importer_id,
                "verified_by_id": system_user.id,
            },
            idempotency_key=idempotency_key,
            performed_by_id=system_user.id,
            notes=f"Bulk import payment. Invoice: {invoice.invoice_number}",
        )
    )
    await db.flush()
    return payment


async def _load_profile_with_lead(
    db: AsyncSession, profile_id: int
) -> "Optional[models.AdmissionProfile]":
    from sqlalchemy.orm import selectinload

    return (
        await db.execute(
            select(models.AdmissionProfile)
            .options(selectinload(models.AdmissionProfile.lead))
            .where(models.AdmissionProfile.id == profile_id)
        )
    ).scalar_one_or_none()


async def _lead_has_other_settled_hk1(
    db: AsyncSession, profile_id: int, exclude_fee_ids: "set[int]"
) -> bool:
    """LEAD này (qua MỌI hồ sơ) còn HK1 tuition fee đã SETTLED nào KHÁC không
    (loại trừ các fee đang void trong lô)?

    Void 1 fee HK1 chỉ nên lùi lead khỏi sts10 khi lead KHÔNG còn HK1 settled ở
    bất kỳ hồ sơ/năm nào khác — nếu không sẽ tụt nhãn oan (lead multi-profile/
    multi-year, forward sync chỉ fire 1 lần). Settled = mirror ``is_hk1_settled``
    (paid/waived HOẶC remaining<=0), trừ cancelled.
    """
    lead_id = (
        await db.execute(
            select(models.AdmissionProfile.lead_id).where(
                models.AdmissionProfile.id == profile_id
            )
        )
    ).scalar_one_or_none()
    if lead_id is None:
        return False

    conditions = [
        models.AdmissionProfile.lead_id == lead_id,
        Fee.fee_type == "tuition",
        Fee.semester_no == 1,
        Fee.status != FeeStatusEnum.cancelled.value,
        or_(
            Fee.status.in_(
                [FeeStatusEnum.paid.value, FeeStatusEnum.waived.value]
            ),
            (Fee.final_amount - Fee.paid_amount - Fee.waived_amount) <= 0,
        ),
    ]
    if exclude_fee_ids:
        conditions.append(~Fee.id.in_(exclude_fee_ids))

    stmt = (
        select(Fee.id)
        .join(
            models.AdmissionProfile,
            models.AdmissionProfile.id == Fee.admission_profile_id,
        )
        .where(*conditions)
        .limit(1)
    )
    return (await db.execute(stmt)).first() is not None


@dataclass
class CommitResult:
    batch_id: int
    committed_count: int  # số dòng ghi được ở LƯỢT này
    #: Số dòng THỬ ghi và hỏng (TOCTOU, số dư đổi, lỗi hệ thống). KHÔNG gồm dòng
    #: bị hàng rào nghi trùng giữ lại — hai thứ đó đòi hai hành động khác hẳn
    #: nhau (một cái phải sửa dữ liệu, một cái chỉ cần soát rồi xác nhận), nên
    #: gộp chúng vào một con số là buộc người đọc đoán.
    failed_count: int
    #: Số dòng đang chờ kế toán xác nhận — mỗi dòng có một phiếu riêng.
    review_required_count: int
    payment_count: int  # tổng Payment tạo
    total_amount: Decimal  # tổng tiền của CẢ LÔ (không phải riêng lượt này)


async def _assert_batch_creator_in_unit(
    db: AsyncSession, batch: PaymentImportBatch, unit_id: Optional[int]
) -> None:
    """IDOR unit-scope theo ĐƠN VỊ người tạo lô: manager (unit_id!=None) chỉ thao tác lô
    do user CÙNG đơn vị tạo; admin/accountant (unit_id=None) → mọi lô. Ngoài scope → 404
    (không lộ tồn tại). 1 NGUỒN quy tắc dùng chung commit/void/detail (tránh drift
    quyền)."""
    if unit_id is None:
        return
    creator_unit = (
        await db.execute(
            select(models.User.unit_id).where(models.User.id == batch.created_by_id)
        )
    ).scalar_one_or_none()
    if creator_unit != unit_id:
        raise ResourceNotFoundError("Không tìm thấy lô import")


async def commit_batch(
    db: AsyncSession,
    *,
    batch_id: int,
    importer_id: int,
    unit_id: Optional[int],
    confirmed_tokens: Optional[Dict[int, str]] = None,
) -> Tuple[CommitResult, Callable]:
    """Pha 2 — GHI TIỀN. RE-VALIDATE TOCTOU dưới khóa + savepoint per-row + idempotency
    + GỘP lead-sync 1 lần/hồ-sơ. Trả ``(CommitResult, post_commit)``.

    - Khóa lô (FOR UPDATE) → serialize commit đồng thời cùng lô.
    - **Thứ tự khoá: MỌI invoice của các fee liên quan (theo ``fee_id, id``) →
      MỌI fee (theo id) → đọc version → soát phiếu → ghi.** Sau khi chạm fee
      đầu tiên, đường này không xin thêm khoá invoice nào; đợt hoá đơn mới xuất
      hiện giữa hai pha thì dừng sạch (``ConflictError``), không khoá bù.
    - KHÔNG dùng snapshot preview để ghi: re-fetch fee/invoice HIỆN TẠI (số dư có thể
      đổi giữa preview→commit). Cuối vòng còn ``left>0`` = vượt nợ hiện tại → raise →
      savepoint rollback CẢ dòng (không ghi nửa vời).
    - 1 dòng lỗi không abort cả lô (savepoint per-row).
    """
    from app.services.fee_calculation_service import is_hk1_settled_fee
    from app.services.lead_admission_sync import sync_lead_tuition_paid

    batch = (
        await db.execute(
            select(PaymentImportBatch)
            .where(PaymentImportBatch.id == batch_id)
            .with_for_update()
        )
    ).scalar_one_or_none()
    if batch is None:
        raise ResourceNotFoundError("Không tìm thấy lô import")
    # P1 IDOR: committer unit-scope — chặn manager commit/poison lô đơn vị khác.
    await _assert_batch_creator_in_unit(db, batch, unit_id)
    if batch.status != PaymentImportBatchStatusEnum.preview.value:
        raise ConflictError(
            f"Lô #{batch_id} không ở trạng thái preview (hiện: {batch.status})"
        )

    system_user = await get_system_user(db)
    if system_user.id == importer_id:
        raise BusinessRuleViolation(
            "Tài khoản import trùng system_user — vi phạm maker-checker"
        )

    method_ids = {
        code: mid
        for code, mid in (
            await db.execute(
                select(PaymentMethod.code, PaymentMethod.id).where(
                    PaymentMethod.code.in_(["cash", "bank_transfer"]),
                    PaymentMethod.is_active.is_(True),
                )
            )
        ).all()
    }

    fee_repo = FeeRepository(db)
    inv_repo = InvoiceRepository(db)

    # Chọn dòng theo TRỤC GHI, không theo trục kiểm. `committed` bị loại bằng
    # CẤU TRÚC DỮ LIỆU, không bằng một phép kiểm rải rác trong vòng lặp: một
    # dòng đã có tiền mà lọt vào đây là mở lại đường ghi hai lần — khoá
    # idempotency của Payment chỉ chặn theo (lô, dòng, HOÁ ĐƠN) nên không cứu
    # được phần tiền rơi sang đợt khác khi đợt cũ đã hết dư.
    CO_THE_GHI = (
        PaymentImportCommitStatusEnum.pending.value,
        PaymentImportCommitStatusEnum.duplicate_review_required.value,
        PaymentImportCommitStatusEnum.failed.value,
    )
    rows = list(
        (
            await db.execute(
                select(PaymentImportRow)
                .where(
                    PaymentImportRow.batch_id == batch_id,
                    PaymentImportRow.commit_status.in_(CO_THE_GHI),
                )
                .order_by(PaymentImportRow.row_no)
            )
        )
        .scalars()
        .all()
    )

    committed_count = 0
    failed_count = 0
    payment_count = 0
    total_amount = Decimal("0")
    cleared_profiles: Dict[int, str] = {}
    # (B3) Phiếu do CHÍNH lô này ghi — để dòng sau không tự tố dòng trước.
    #
    # Nạp sẵn phiếu của các lượt commit TRƯỚC, không bắt đầu từ tập rỗng. `rows`
    # cố tình chỉ gồm dòng CÒN ghi được, nên dòng `committed` ở lượt trước nằm
    # ngoài — và mã phiếu của chúng sẽ không bao giờ được nạp. Khi một dòng
    # `failed` được thử lại (hoặc phiếu hết hạn rồi xác nhận lại), nó đụng phải
    # phiếu mà chính lô này vừa ghi ở lượt trước và bị tố là nghi trùng. Người
    # ghi soát mãi một cảnh báo về chính mình.
    payment_ids_cua_lo: Set[int] = {
        pid
        for (ids,) in (
            await db.execute(
                select(PaymentImportRow.payment_ids).where(
                    PaymentImportRow.batch_id == batch_id,
                    PaymentImportRow.commit_status
                    == PaymentImportCommitStatusEnum.committed.value,
                )
            )
        ).all()
        if isinstance(ids, list)
        for pid in ids
        if isinstance(pid, int)
    }
    # Đếm riêng dòng bị hàng rào trùng chặn: chúng là thứ DUY NHẤT commit lại
    # với cờ xác nhận sẽ cứu được, nên lô còn dòng như vậy thì chưa được đóng.
    so_dong_chan_trung = 0

    # ────────────────────────────────────────────────────────────────────────
    # PHA SOÁT PHIẾU — chạy TRỌN VẸN trước khi ghi dòng đầu tiên.
    #
    # Thứ tự ở đây là cả vấn đề. Soát tuần tự (dòng nào tới lượt thì soát dòng
    # đó) thì dòng đầu tiên ghi xong sẽ làm `fee.duplicate_guard_version` nhích
    # — trigger ở tầng cơ sở dữ liệu — và phiếu HỢP LỆ của dòng thứ hai cùng
    # khoản phí lập tức hết hiệu lực. Kế toán xác nhận cả năm dòng, hệ thống
    # ghi một dòng rồi bắt xác nhận lại bốn dòng còn lại, mỗi lượt được đúng
    # một dòng. Vòng 409 mà cả đợt này sinh ra để xoá, mọc lại ở chỗ khác.
    #
    # Nên: khoá hết Fee liên quan, đọc version BAN ĐẦU dưới khoá, soát mọi
    # phiếu theo version ấy, rồi mới ghi.
    #
    # ── THỨ TỰ KHOÁ TOÀN CỤC ────────────────────────────────────────────────
    # Khoá Fee trước là bản vá của bẫy trên, nhưng nó dựng ra một chiều
    # Fee → Invoice, ngược với ghi tay (Invoice → Fee). Hai chiều ngược nhau
    # trên cùng cặp hàng = kẹt chéo thật, đã tái hiện được bằng hai session
    # PostgreSQL (``tests/services/test_import_vs_manual_deadlock.py``).
    #
    # Invariant kể từ đây — **sau khi bắt đầu khoá Fee, đường nhập lô tuyệt
    # đối không xin thêm bất kỳ khoá Invoice nào.** Vì vậy pha này cầm TRỌN
    # tập invoice trước, kể cả đợt đã thu đủ mà lượt ghi sẽ không đụng tới:
    # cầm phần "dự đoán sẽ dùng" là để hở đúng cái khe cần đóng.
    #
    # ``khoa_moi_invoice_cua_fee`` khoá ``FOR UPDATE OF invoice`` — KHÔNG kéo
    # ``fee`` theo như ``get_for_update``. Ngữ nghĩa khoá-kèm-fee của hàm kia
    # là hàng rào của đường ghi tay và phải giữ nguyên; xem ghi chú ở
    # ``InvoiceRepository.khoa_moi_invoice_cua_fee``.
    ma_fee_lien_quan = sorted(
        {r.resolved_fee_id for r in rows if r.resolved_fee_id is not None}
    )
    invoice_da_khoa = await inv_repo.khoa_moi_invoice_cua_fee(
        ma_fee_lien_quan, unit_id
    )
    invoice_theo_fee: Dict[int, List[Invoice]] = {}
    for _inv in invoice_da_khoa:
        invoice_theo_fee.setdefault(_inv.fee_id, []).append(_inv)

    version_ban_dau: Dict[int, int] = {}
    for _ma_fee in ma_fee_lien_quan:
        _fee = await fee_repo.get_for_update(_ma_fee, unit_id)
        if _fee is not None:
            version_ban_dau[_ma_fee] = _fee.duplicate_guard_version

    # Cửa sổ giữa hai pha: một đợt hoá đơn MỚI có thể được phát hành sau khi ta
    # chụp tập invoice và trước khi ta cầm Fee. Khoá bổ sung nó ở đây là phá
    # đúng invariant vừa dựng, nên đường duy nhất còn lại là dừng sạch —
    # fail-closed. Lượt commit sau bắt đầu từ đầu và cầm trọn tập mới.
    #
    # Không tự thử lại tại chỗ: giao dịch này đang giữ khoá lô và một tập hàng
    # invoice/fee: "thử lại" mà không nhả khoá thì chỉ lặp lại đúng ảnh chụp cũ.
    _id_hien_tai = await _id_invoice_cua_fee(db, ma_fee_lien_quan, unit_id)
    _id_da_khoa = {i.id for i in invoice_da_khoa}
    _id_moi = _id_hien_tai - _id_da_khoa
    if _id_moi:
        raise ConflictError(
            "Có đợt hoá đơn mới phát hành giữa lúc khoá — lô chưa ghi gì, "
            f"vui lòng commit lại (hoá đơn mới: {sorted(_id_moi)})"
        )

    dong_da_soat: Set[int] = set()
    for row in rows:
        phieu = (confirmed_tokens or {}).get(row.row_no)
        if not phieu or row.resolved_fee_id is None or row.amount is None:
            continue
        gv = version_ban_dau.get(row.resolved_fee_id)
        if gv is None:
            continue
        try:
            _ngay = parse_date_vn(str((row.raw or {}).get(COL_DATE, "")))
        except Exception:  # noqa: BLE001 — dòng hỏng ngày sẽ chết ở vòng ghi
            continue
        if soat_phieu(
            phieu,
            RangBuoc(
                flow="import",
                user_id=importer_id,
                unit_id=unit_id,
                fee_id=row.resolved_fee_id,
                # Nhập lô phân bổ sang nhiều đợt nên không có MỘT hoá đơn để
                # ràng buộc; `batch_id` + `row_no` mới là thứ định danh ở đây.
                invoice_id=None,
                amount=row.amount,
                payment_date=datetime(
                    _ngay.year, _ngay.month, _ngay.day, tzinfo=timezone.utc
                ),
                guard_version=gv,
                batch_id=batch_id,
                row_no=row.row_no,
            ),
        ):
            dong_da_soat.add(row.row_no)

    for row in rows:
        row_payment_ids: List[int] = []
        row_total = Decimal("0")  # row-local (Bug 1: cộng tổng SAU khi savepoint OK)
        row_cleared: Optional[tuple] = None
        try:
            async with db.begin_nested():
                raw = row.raw or {}
                method_code = METHOD_MAP.get(_norm_name(str(raw.get(COL_METHOD, ""))))
                method_id = method_ids.get(method_code)
                if method_id is None:
                    raise BusinessRuleViolation("hình thức không hợp lệ")
                pay_date = parse_date_vn(str(raw.get(COL_DATE, "")))
                # date → datetime tz-aware (cột Payment.payment_date DateTime(tz)).
                pay_dt = datetime(
                    pay_date.year, pay_date.month, pay_date.day, tzinfo=timezone.utc
                )
                reference = _norm(str(raw.get(COL_REF, ""))) or None
                # Defense: ref > String(100) → DataError ở flush không bắt được → 500
                # cả lô. Bắt thành lỗi DÒNG (preview cũng đã chặn ở parse_template).
                if reference is not None and len(reference) > MAX_REF_LEN:
                    raise BusinessRuleViolation(
                        f"mã tham chiếu quá dài (tối đa {MAX_REF_LEN} ký tự)"
                    )
                amount = row.amount
                if amount is None or amount <= 0:
                    raise BusinessRuleViolation("số tiền không hợp lệ")
                if row.resolved_fee_id is None:
                    raise BusinessRuleViolation("thiếu học phí đã resolve")

                # Tập invoice của khoản phí này ĐÃ được khoá trọn ở pha đầu.
                # Không gọi `inv_repo.get_for_update` ở đây: đó vừa là xin thêm
                # khoá Invoice sau khi đã cầm Fee (phá invariant thứ tự khoá),
                # vừa kéo `fee` vào theo thứ tự invoice vì `FOR UPDATE` của hàm
                # ấy không có `OF`.
                cua_fee = invoice_theo_fee.get(row.resolved_fee_id, [])
                if not cua_fee:
                    raise BusinessRuleViolation(
                        "không tìm thấy hóa đơn (IDOR / đổi giữa 2 pha)"
                    )

                # `refresh` đọc lại hàng ĐANG bị chính giao dịch này khoá — một
                # câu SELECT thường, không xin thêm khoá nào.
                locked = []
                for li in sorted(cua_fee, key=lambda i: i.installment_no):
                    await db.refresh(li)
                    if li.status in PAYABLE_INVOICE_STATUSES:
                        locked.append(li)
                if not locked:
                    raise BusinessRuleViolation(
                        "không còn đợt hóa đơn payable (đã thu đủ / đổi giữa 2 pha)"
                    )
                fee = await fee_repo.get_for_update(locked[0].fee_id, unit_id)
                if fee is None:
                    raise BusinessRuleViolation("không tìm thấy học phí")
                await db.refresh(fee)
                if fee.status in (
                    FeeStatusEnum.cancelled.value,
                    FeeStatusEnum.waived.value,
                ):
                    raise BusinessRuleViolation(
                        f"học phí đã {fee.status} (đổi giữa 2 pha)"
                    )
                # Re-check lead chưa xóa mềm: preview lọc Lead.deleted_at IS NULL nhưng
                # get_for_update KHÔNG lọc → chặn ghi tiền vào hồ sơ bị xóa mềm giữa
                # preview→commit (mirror filter của _fetch_profiles).
                _pd_row = (
                    await db.execute(
                        select(
                            models.Lead.deleted_at,
                            models.AdmissionProfile.status,
                        )
                        .join(
                            models.AdmissionProfile,
                            models.AdmissionProfile.lead_id == models.Lead.id,
                        )
                        .where(models.AdmissionProfile.id == fee.admission_profile_id)
                    )
                ).one_or_none()
                lead_deleted = _pd_row[0] if _pd_row else None
                profile_status = _pd_row[1] if _pd_row else None
                if lead_deleted is not None:
                    raise BusinessRuleViolation(
                        "hồ sơ đã bị xóa giữa preview→commit"
                    )
                # P0: bulk import does NOT pass through assert_payable_target, so
                # inline the profile guard here (mirrors the fee cancelled/waived
                # + lead-soft-deleted checks above). Refuse auto-verifying money
                # onto a withdrawn/rejected/refund-pending profile — the invoice
                # can still be `issued` because withdraw does not cancel the fee.
                if profile_status in NON_PAYABLE_PROFILE_STATUSES:
                    raise BusinessRuleViolation(
                        f"hồ sơ đã {profile_status} — không thể thu tiền vào "
                        f"hồ sơ đã rút/từ chối/đang chờ hoàn"
                    )
                # Đồng bộ với ``assert_payable_target``: học phí vừa định giá lại
                # do đổi ngành, đang chờ kế toán xác nhận thì KHÔNG được auto-verify
                # tiền vào. Import hàng loạt là đường ghi tiền im lặng nhất — bỏ
                # sót ở đây thì cả một lô tiền đáp lên mức giá chưa ai duyệt.
                if getattr(fee, "awaiting_accountant_confirmation", False):
                    raise BusinessRuleViolation(
                        "học phí đang chờ kế toán xác nhận đổi ngành — "
                        "xác nhận trước khi thu tiền"
                    )
                # (B3) Hàng rào chống ghi trùng, đặt SAU khi đã khoá `fee`: luật
                # đọc mọi hoá đơn của khoản phí, nên không có điểm gặp chung thì
                # hai lô chạy song song đều thấy "chưa trùng" rồi cùng ghi.
                #
                # Bỏ qua DÒNG chứ không chặn cả lô: raise ở đây rơi vào savepoint
                # của riêng dòng này, phần còn lại của tệp vẫn vào bình thường.
                # Chặn cả lô vì một dòng nghi ngờ sẽ biến mọi tệp lớn thành ngõ cụt.
                if row.row_no not in dong_da_soat:
                    # Loại trừ phiếu do CHÍNH LÔ NÀY vừa ghi. Không có vế này thì
                    # dòng 2 của một tệp có hai dòng giống hệt nhau sẽ đụng phải
                    # phiếu mà dòng 1 vừa flush và bị từ chối — trong khi xem
                    # trước đã cố ý coi hai dòng đó là hai khoản thu riêng (map
                    # dò trùng khoá theo `row_no` chính vì lẽ đó). Hai pha nói
                    # ngược nhau thì pha ghi là pha thắng, và tiền rơi mất.
                    ket_do = await PaymentRepository(db).find_duplicate_candidates_bulk(
                        keys=[(row.row_no, fee.id, amount, pay_dt)],
                        exclude_payment_ids=payment_ids_cua_lo or None,
                    )
                    ung_vien_ids = ket_do.get(row.row_no, [])
                    if ung_vien_ids:
                        bi_cat = len(ung_vien_ids) > MAX_DUPLICATE_CANDIDATES
                        so_luong = (
                            f"hơn {MAX_DUPLICATE_CANDIDATES}"
                            if bi_cat
                            else str(len(ung_vien_ids))
                        )
                        ma_phieu = ", ".join(f"#{i}" for i in ung_vien_ids[:3])
                        them = "…" if len(ung_vien_ids) > 3 else ""
                        raise _NghiTrungBiChan(
                            f"nghi trùng với {so_luong} phiếu đã ghi cho cùng "
                            f"khoản phí — cùng số tiền, lệch không quá "
                            f"{WINDOW_DO_TRUNG_NGAY} ngày ({ma_phieu}{them}). "
                            f"Soát lại; nếu đúng là khoản thu riêng thì commit lại "
                            f"với xác nhận bỏ qua cảnh báo trùng."
                        )

                was_hk1 = is_hk1_settled_fee(fee)

                left = amount
                for li in locked:
                    if left <= 0:
                        break
                    if li.status not in PAYABLE_INVOICE_STATUSES:
                        continue  # đổi trạng thái sau khi khóa
                    principal_rem = (li.amount or Decimal("0")) - (
                        li.paid_amount or Decimal("0")
                    )
                    if principal_rem <= 0:
                        continue
                    take = min(left, principal_rem)
                    idem = f"bulkimport:{batch_id}:{row.row_no}:{li.id}"
                    payment = await auto_verify_payment(
                        db,
                        invoice=li,
                        fee=fee,
                        method_id=method_id,
                        amount=take,
                        payment_date=pay_dt,
                        reference=reference,
                        importer_id=importer_id,
                        system_user=system_user,
                        idempotency_key=idem,
                    )
                    if payment is not None:
                        row_payment_ids.append(payment.id)
                        row_total += take
                    left -= take

                if left > 0:
                    raise BusinessRuleViolation(
                        f"thu {_money(amount)} vượt còn nợ gốc HIỆN TẠI "
                        "(số dư đổi giữa preview→commit)"
                    )

                if not was_hk1:
                    now_hk1 = is_hk1_settled_fee(fee)
                    if now_hk1:
                        row_cleared = (
                            fee.admission_profile_id,
                            reference or f"BULK-{batch_id}-{row.row_no}",
                        )
                row.payment_ids = row_payment_ids
                # Trạng thái GHI và `payment_ids` đặt trong CÙNG savepoint: một
                # dòng "đã ghi" mà không có mã phiếu, hoặc ngược lại, là hai nửa
                # của cùng một sự thật lệch nhau — và ràng buộc "committed bắt
                # buộc có payment_ids" sẽ bắt được ngay tại đây thay vì để lộ ra
                # ở một báo cáo nào đó sau này.
                row.commit_status = (
                    PaymentImportCommitStatusEnum.committed.value
                )
            # savepoint committed → giờ mới cộng tổng (raise giữa chừng đã rollback DB).
            committed_count += 1
            # Chỉ ghi nhận SAU khi savepoint qua: dòng bị rollback không để lại
            # phiếu nào, đưa id của nó vào tập loại trừ là loại trừ thứ không tồn tại.
            payment_ids_cua_lo.update(row_payment_ids)
            payment_count += len(row_payment_ids)
            total_amount += row_total
            if row_cleared:
                cleared_profiles[row_cleared[0]] = row_cleared[1]
        except _NghiTrungBiChan as exc:
            # KHÁC hẳn lỗi thật: dòng này ghi lại được, chỉ cần kế toán soát rồi
            # xác nhận. Nay nó có trạng thái GHI riêng thay vì phải mượn 'warned'
            # — trước đây mượn như vậy nên một dòng đã ghi tiền và một dòng đang
            # chờ xác nhận trông y hệt nhau trong mọi truy vấn.
            so_dong_chan_trung += 1
            row.validation_status = PaymentImportRowStatusEnum.warned.value
            row.commit_status = (
                PaymentImportCommitStatusEnum.duplicate_review_required.value
            )
            row.message = str(exc)[:500]
        except (
            BusinessRuleViolation,
            ResourceNotFoundError,
            ValueError,
            ConflictError,
        ) as exc:
            # Lỗi nghiệp vụ/giá trị → message đã sạch (tiếng Việt) → hiện thẳng cho
            # kế toán. Hỏng ở bước GHI, không phải ở khâu đọc: trục kiểm giữ
            # nguyên kết quả của bước xem trước, đúng như định nghĩa của nó.
            failed_count += 1
            row.commit_status = PaymentImportCommitStatusEnum.failed.value
            row.message = str(exc)[:500]
        except Exception as exc:  # noqa: BLE001 — lỗi KHÔNG lường (DB/IntegrityError)
            # KHÔNG nhét dump SQLAlchemy/asyncpg cho kế toán (xấu + lộ chi tiết kỹ
            # thuật) → message generic + LOG đầy đủ cho dev. Savepoint per-row đã
            # rollback dòng này nên các dòng khác vẫn ghi (lô ROBUST: lỗi-lạ 1 dòng
            # không abort cả lô).
            log.error(
                "bulk_commit_row_unexpected_error",
                batch_id=batch_id,
                row_no=row.row_no,
                error=str(exc),
            )
            failed_count += 1
            row.commit_status = PaymentImportCommitStatusEnum.failed.value
            row.message = "lỗi hệ thống khi ghi dòng này — vui lòng liên hệ kỹ thuật"

    # GỘP lead-sync (projection) 1 lần/hồ-sơ HK1 vừa cleared. Bug 2: bọc savepoint +
    # try/except MỖI hồ-sơ — lỗi projection 1 hồ-sơ KHÔNG được hủy tiền đã ghi của CẢ
    # lô (sync chạy trong body trước router commit; raise → outer rollback → mất tiền).
    for profile_id, ref in cleared_profiles.items():
        try:
            async with db.begin_nested():
                profile = await _load_profile_with_lead(db, profile_id)
                if profile is not None:
                    await sync_lead_tuition_paid(
                        db=db,
                        profile=profile,
                        transaction_id=ref,
                        changed_by_user_id=importer_id,
                        reason=f"Thu học phí qua import lô #{batch_id}",
                    )
        except Exception as exc:  # noqa: BLE001
            # KHÔNG hủy tiền đã ghi vì lỗi projection 1 hồ-sơ, NHƯNG log ERROR + stack
            # trace (backend chưa có Sentry) để lỗi projection HỆ THỐNG không ẩn mình:
            # tiền đã ghi mà lead đứng yên sts cũ phải lộ ra để truy.
            log.error(
                "bulk_commit_lead_sync_failed",
                batch_id=batch_id,
                profile_id=profile_id,
                error=str(exc),
                exc_info=True,
            )

    # ── Cấp PHIẾU MỚI cho các dòng còn bị chặn, sau khi đã ghi xong mọi dòng
    # ghi được. Cuối lượt chứ không phải giữa chừng: phiếu mang theo
    # `guard_version`, và mỗi lần ghi lại làm nó nhích, nên phiếu cấp sớm sẽ
    # chết trước khi kế toán kịp nhìn thấy.
    await db.flush()
    if so_dong_chan_trung:
        version_hien_tai = {
            ma: gv
            for ma, gv in (
                await db.execute(
                    select(Fee.id, Fee.duplicate_guard_version).where(
                        Fee.id.in_(ma_fee_lien_quan)
                    )
                )
            ).all()
        }
        for row in rows:
            if row.commit_status != (
                PaymentImportCommitStatusEnum.duplicate_review_required.value
            ):
                continue
            gv = version_hien_tai.get(row.resolved_fee_id)
            if gv is None or row.amount is None:
                continue
            _ngay = parse_date_vn(str((row.raw or {}).get(COL_DATE, "")))
            row.duplicate_review_token = cap_phieu(
                RangBuoc(
                    flow="import",
                    user_id=importer_id,
                    unit_id=unit_id,
                    fee_id=row.resolved_fee_id,
                    invoice_id=None,
                    amount=row.amount,
                    payment_date=datetime(
                        _ngay.year, _ngay.month, _ngay.day, tzinfo=timezone.utc
                    ),
                    guard_version=gv,
                    batch_id=batch_id,
                    row_no=row.row_no,
                )
            )

    # ── Đếm lại CẢ HAI HỌ từ trạng thái dòng THỰC TẾ của TOÀN LÔ. Không cộng
    # dồn gì hết: `batch.failed_count += failed_count` của bản trước đếm một
    # dòng hai lần (nó vừa làm tăng `failed_count` vừa giữ `warned`) và không ai
    # trừ lại khi lượt sau ghi được. Lô #5 trên máy dev: đúng một dòng, sổ ghi
    # thành hai.
    #
    # Đọc từ cơ sở dữ liệu chứ không đếm trên `rows`: `rows` chỉ gồm những dòng
    # CÓ THỂ ghi ở lượt này, nên dòng hỏng từ khâu đọc và dòng đã ghi ở lượt
    # trước đều nằm ngoài.
    await db.flush()

    async def _dem(cot) -> Dict[str, int]:
        return {
            gt: so
            for gt, so in (
                await db.execute(
                    select(cot, func.count())
                    .where(PaymentImportRow.batch_id == batch_id)
                    .group_by(cot)
                )
            ).all()
        }

    dem_kiem = await _dem(PaymentImportRow.validation_status)
    dem_ghi = await _dem(PaymentImportRow.commit_status)
    E = PaymentImportRowStatusEnum
    C = PaymentImportCommitStatusEnum
    batch.matched_count = dem_kiem.get(E.matched.value, 0)
    batch.warned_count = dem_kiem.get(E.warned.value, 0)
    batch.failed_count = dem_kiem.get(E.error.value, 0)
    batch.committed_row_count = dem_ghi.get(C.committed.value, 0)
    batch.review_required_count = dem_ghi.get(C.duplicate_review_required.value, 0)
    batch.commit_failed_count = dem_ghi.get(C.failed.value, 0)
    batch.not_applicable_count = dem_ghi.get(C.not_applicable.value, 0)

    # Tổng tiền của CẢ LÔ, không phải của riêng lượt này: gán `total_amount` là
    # đè, nên lượt xác nhận lại (chỉ ghi vài dòng còn sót) sẽ làm sổ lô tụt
    # xuống bằng đúng phần vừa ghi và mất phần đã vào ở lượt trước. Cộng từ
    # PHIẾU THẬT, không từ số dự kiến.
    _ma_phieu_toan_lo = [
        pid
        for (ids,) in (
            await db.execute(
                select(PaymentImportRow.payment_ids).where(
                    PaymentImportRow.batch_id == batch_id
                )
            )
        ).all()
        if isinstance(ids, list)
        for pid in ids
        if isinstance(pid, int)
    ]
    batch.total_amount = (
        (
            await db.execute(
                select(func.coalesce(func.sum(Payment.amount), 0)).where(
                    Payment.id.in_(_ma_phieu_toan_lo)
                )
            )
        ).scalar_one()
        if _ma_phieu_toan_lo
        else Decimal("0")
    )
    # Chỉ đánh dấu 'committed' khi THỰC SỰ ghi được tiền. 0 dòng ghi (tất cả fail
    # TOCTOU) → GIỮ 'preview' để re-import được (chưa có endpoint void; tránh khóa
    # file vĩnh viễn qua partial-unique).
    #
    # (B3) Còn dòng bị hàng rào trùng chặn thì CŨNG giữ 'preview'. Đóng lô ở đây
    # là dựng ngõ cụt: lô 'committed' từ chối commit lại (409), nên những dòng
    # kia vĩnh viễn không vào được — mà chúng là tiền thật kế toán đã thu. Giữ
    # 'preview' cho phép soát rồi commit lại với `confirm_duplicates`; các dòng
    # đã ghi ở lượt này không bị ghi hai lần vì idempotency key
    # `bulkimport:{lô}:{dòng}:{hoá đơn}` đã chặn sẵn.
    # Trạng thái lô SUY RA từ tập trục ghi, không từ biến đếm của riêng lượt
    # này: một lượt xác nhận lại có `committed_count` nhỏ vẫn có thể là lượt
    # khép được lô, còn một lượt ghi được nhiều dòng nhưng còn dòng chờ xác
    # nhận thì không.
    con_cho = batch.review_required_count > 0
    if batch.committed_row_count > 0 and not con_cho:
        batch.status = PaymentImportBatchStatusEnum.committed.value
        batch.committed_at = datetime.now(timezone.utc)
    await db.flush()

    result = CommitResult(
        batch_id=batch_id,
        committed_count=committed_count,
        failed_count=failed_count,
        review_required_count=so_dong_chan_trung,
        payment_count=payment_count,
        # Tổng của CẢ LÔ, đọc lại từ phiếu thật — không phải tổng của riêng
        # lượt này, vì lượt xác nhận lại chỉ ghi vài dòng còn sót.
        total_amount=batch.total_amount,
    )

    async def post_commit() -> None:
        # QUYẾT ĐỊNH SẢN PHẨM (đã chốt): bulk import KHÔNG gửi notification per-row
        # (tránh fan-out 3N) — KHÁC verify tay (fire PAYMENT_RECEIVED mỗi payment).
        # Lead-sync (projection) đã làm trong body. Nếu sau cần báo phụ huynh: thêm 1
        # notif GỘP per-hồ-sơ + PHẢI seed notification_rule (nếu không sẽ fire-zero
        # theo fail-closed dispatcher). Để trống là CỐ Ý, không phải thiếu sót.
        return None

    return result, post_commit


# ---------------------------------------------------------------------------
# BV-3.5 — Void (đảo) lô đã committed: rút lại tiền + mở lại file để re-import
# ---------------------------------------------------------------------------
@dataclass
class VoidResult:
    batch_id: int
    reversed_count: int  # số Payment đã đảo (rút lại)
    reversed_amount: Decimal  # tổng tiền đã rút


async def void_batch(
    db: AsyncSession,
    *,
    batch_id: int,
    user_id: int,
    unit_id: Optional[int],
    reason: str,
) -> Tuple[VoidResult, Callable]:
    """Đảo (void) 1 lô đã ``committed`` — rút lại MỌI Payment bulk đã ghi.

    Mỗi Payment 'verified' của lô → reverse invoice/fee.paid_amount (về 'issued' nếu
    hết, 'partial' nếu còn) + ``PaymentTransaction(type=reversal, amount âm)`` +
    Payment → 'refunded' (constraint chỉ có refunded; ``type=reversal`` phân biệt với
    customer-refund ở ledger). Batch → 'void' → file_sha256 thoát partial-unique →
    re-import được.

    ATOMIC (KHÔNG savepoint per-row): void phải đảo TRỌN lô — 1 lỗi → rollback cả
    (router get_db không commit).

    Thứ tự khoá: **batch → payments → invoices (asc id) → fees (asc id)**. Khớp thứ
    tự thống nhất toàn hệ (batch → payment → invoice → fee); ``verify_payment`` và
    ``reject_payment`` nay cũng khoá payment TRƯỚC invoice và fee.
    (Trước đây dòng này chỉ ghi "invoice → fee" — thiếu hai bậc đầu khiến người đọc
    tưởng đường void chỉ bắt đầu từ invoice. Lưu ý hàm này KHÔNG tranh chấp với
    verify: nó chỉ đảo phiếu 'verified', còn verify chỉ nhận 'pending'.)

    Lead-status: void TỰ lùi lead khỏi sts10 ("Đã hoàn tất học phí") về status TRƯỚC
    (đối xứng forward sync) cho hồ sơ HK1 KHÔNG còn cleared sau đảo — qua
    ``revert_lead_tuition_paid`` (đọc LeadStatusHistory). Best-effort: chỉ khi lead
    ĐANG ở sts10 (đã chuyển tiếp/nhập học → giữ nguyên, không kéo lùi). KHÔNG đẩy
    sts18 (đó là refund THẬT — ``sync_lead_tuition_refunded``).
    """
    from app.services.payment_service import reverse_payment_balances

    batch = (
        await db.execute(
            select(PaymentImportBatch)
            .where(PaymentImportBatch.id == batch_id)
            .with_for_update()
        )
    ).scalar_one_or_none()
    if batch is None:
        raise ResourceNotFoundError("Không tìm thấy lô import")
    # IDOR: void unit-scope theo đơn vị NGƯỜI TẠO (accountant đã bị Casbin chặn).
    await _assert_batch_creator_in_unit(db, batch, unit_id)
    if batch.status != PaymentImportBatchStatusEnum.committed.value:
        raise ConflictError(
            f"Chỉ đảo (void) được lô đã committed (hiện: {batch.status})"
        )

    fee_repo = FeeRepository(db)
    inv_repo = InvoiceRepository(db)

    rows = list(
        (
            await db.execute(
                select(PaymentImportRow).where(PaymentImportRow.batch_id == batch_id)
            )
        )
        .scalars()
        .all()
    )
    payment_ids = sorted({pid for r in rows for pid in (r.payment_ids or [])})
    # Lock Payment rows FOR UPDATE (asc id) TRƯỚC khi check refund + đảo — serialize
    # với refund-creation (request_refund cũng khóa Payment, of=Payment): refund tạo
    # đồng thời phải CHỜ → sau void thấy status='refunded' → bị chặn (chỉ refund
    # 'verified'). order_by(id) cho thứ tự khóa xác định (chống void↔void deadlock).
    payments = (
        list(
            (
                await db.execute(
                    select(Payment)
                    .where(Payment.id.in_(payment_ids))
                    .order_by(Payment.id)
                    .with_for_update()
                )
            )
            .scalars()
            .all()
        )
        if payment_ids
        else []
    )

    # Lock TẤT CẢ invoice (asc id) RỒI TẤT CẢ fee (asc id) — deadlock-safe (không giữ
    # khóa fee khi đi xin khóa invoice). KHÔNG db.refresh: void không pre-load (Payment.
    # invoice lazy='select') nên get_for_update là lần nạp ĐẦU → cột đã tươi.
    invoice_ids = sorted({p.invoice_id for p in payments})
    inv_by_id: Dict[int, Invoice] = {}
    for iid in invoice_ids:
        li = await inv_repo.get_for_update(iid, unit_id)
        if li is None:
            raise BusinessRuleViolation(
                f"không khóa được hóa đơn #{iid} (IDOR / đã đổi)"
            )
        inv_by_id[iid] = li
    fee_ids = sorted({inv.fee_id for inv in inv_by_id.values()})
    fee_by_id: Dict[int, Fee] = {}
    for fid in fee_ids:
        lf = await fee_repo.get_for_update(fid, unit_id)
        if lf is None:
            raise BusinessRuleViolation(f"không khóa được học phí #{fid}")
        fee_by_id[fid] = lf

    # (#3) Guard out-of-band: invoice/fee bị HỦY/MIỄN giữa commit→void (vd luồng gộp
    # đợt SQL) → KHÔNG resurrect (reverse_payment_balances set 'issued'/'partial' vô
    # điều kiện → hồi sinh hóa đơn đã hủy / IntegrityError). Refuse → xử lý thủ công.
    for inv in inv_by_id.values():
        if inv.status == InvoiceStatusEnum.cancelled.value:
            raise BusinessRuleViolation(
                f"hóa đơn #{inv.id} đã bị hủy giữa commit→void — xử lý thủ công"
            )
    for fee in fee_by_id.values():
        if fee.status in (
            FeeStatusEnum.cancelled.value,
            FeeStatusEnum.waived.value,
        ):
            raise BusinessRuleViolation(
                f"học phí #{fee.id} đã {fee.status} giữa commit→void — xử lý thủ công"
            )

    # (P1) Guard refund: payment trong lô đã/đang được hoàn lẻ (RefundRequest non-
    # rejected) → KHÔNG đảo (trừ tiền 2 LẦN — refund đã rút phần đó). Dưới khóa fee nên
    # refund-processing đồng thời (cũng khóa fee) bị serialize. Khe tạo refund SAU check
    # bị chặn bởi guard payment.status='verified' ở process_approved_refund.
    if payment_ids:
        live_refund = (
            await db.execute(
                select(RefundRequest.id)
                .where(
                    RefundRequest.payment_id.in_(payment_ids),
                    RefundRequest.status != RefundStatusEnum.rejected.value,
                )
                .limit(1)
            )
        ).scalar_one_or_none()
        if live_refund is not None:
            raise ConflictError(
                "Lô có payment đã/đang được hoàn tiền (refund) — xử lý refund "
                "trước khi đảo lô"
            )

    now = datetime.now(timezone.utc)
    reversed_count = 0
    reversed_amount = Decimal("0")
    for p in payments:
        # Idempotent: Payment đã 'refunded' (đã đảo/hoàn lẻ trước) → bỏ qua (không
        # đảo 2 lần).
        if p.status != PaymentStatusEnum.verified.value:
            continue
        inv = inv_by_id.get(p.invoice_id)
        fee = fee_by_id.get(inv.fee_id) if inv is not None else None
        if inv is None or fee is None:
            raise BusinessRuleViolation(f"thiếu hóa đơn/học phí cho payment #{p.id}")
        bal_before, fee_remaining = reverse_payment_balances(
            invoice=inv, fee=fee, amount=p.amount
        )
        p.status = PaymentStatusEnum.refunded.value
        db.add(
            PaymentTransaction(
                payment_id=p.id,
                fee_id=fee.id,
                transaction_type=TransactionTypeEnum.reversal.value,
                amount=-p.amount,
                balance_before=bal_before,
                balance_after=fee_remaining,
                external_reference=p.reference_code,
                gateway_response={
                    "verification_mode": "bulk_void",
                    "batch_id": batch_id,
                    "voided_by_id": user_id,
                },
                idempotency_key=f"bulkvoid:{batch_id}:{p.id}",
                performed_by_id=user_id,
                notes=f"Void lô import #{batch_id}. Lý do: {reason}"[:1000],
            )
        )
        reversed_count += 1
        reversed_amount += p.amount

    # ⚠️ Flush việc đảo tiền xuống DB TRƯỚC vòng lùi-lead. autoflush=False (database.py)
    # → các UPDATE/INSERT đảo tiền ở trên còn PENDING; nếu không flush ở đây, lần flush
    # ĐẦU TIÊN là `db.flush()` BÊN TRONG savepoint của revert → revert lỗi → ROLLBACK TO
    # SAVEPOINT cuốn theo cả đảo tiền (mất tiền đã đảo dù lô vẫn chuyển 'void'). Flush
    # NGOÀI savepoint = đảo tiền bền trong outer txn; revert lỗi chỉ mất projection.
    await db.flush()

    # Lùi lead (projection) cho hồ sơ HK1 KHÔNG còn cleared sau khi đảo tiền — đối
    # xứng forward sync ở commit. Void = SỬA NHẦM ghi nhận (KHÔNG phải học sinh rút)
    # → lùi về status TRƯỚC sts10, KHÔNG đẩy sts18 (đó là refund thật). Bọc savepoint
    # + try/except MỖI hồ-sơ: lỗi projection KHÔNG hủy đảo tiền cả lô (tiền đã đảo giữ).
    from app.services.fee_calculation_service import is_hk1_settled_fee
    from app.services.lead_admission_sync import revert_lead_tuition_paid

    reverted_leads = 0
    _voided_fee_ids = set(fee_by_id.keys())
    for fee in fee_by_id.values():
        if fee.fee_type != "tuition" or fee.semester_no != 1:
            continue  # chỉ HK1 đụng pipeline lead
        if is_hk1_settled_fee(fee):
            continue  # HK1 fee NÀY vẫn settled (nguồn thu khác) → giữ lead sts10
        # (#4) Lead-level: lead còn HK1 settled ở hồ-sơ/năm KHÁC (ngoài lô đang
        # void) → KHÔNG lùi nhãn (tránh tụt sts10 oan khi multi-profile/year).
        if await _lead_has_other_settled_hk1(
            db, fee.admission_profile_id, _voided_fee_ids
        ):
            continue
        try:
            async with db.begin_nested():
                profile = await _load_profile_with_lead(db, fee.admission_profile_id)
                if profile is not None and await revert_lead_tuition_paid(
                    db=db,
                    profile=profile,
                    changed_by_user_id=user_id,
                    reason=f"Đảo (void) lô import #{batch_id}",
                ):
                    reverted_leads += 1
        except Exception as exc:  # noqa: BLE001
            log.error(
                "bulk_void_lead_revert_failed",
                batch_id=batch_id,
                profile_id=fee.admission_profile_id,
                error=str(exc),
                exc_info=True,
            )
    if reverted_leads:
        log.info(
            "bulk_void_leads_reverted", batch_id=batch_id, count=reverted_leads
        )

    batch.status = PaymentImportBatchStatusEnum.void.value
    batch.voided_at = now
    batch.void_reason = (reason or "")[:1000]
    # (#4) Lịch sử lô phản ánh tiền THỰC còn lại = đã thu − đã đảo (full void → 0),
    # tránh batch 'void' vẫn hiện total như còn thu đủ.
    batch.total_amount = batch.total_amount - reversed_amount
    if batch.total_amount < Decimal("0"):
        batch.total_amount = Decimal("0")
    await db.flush()

    result = VoidResult(
        batch_id=batch_id,
        reversed_count=reversed_count,
        reversed_amount=reversed_amount,
    )

    async def post_commit() -> None:
        return None

    return result, post_commit


# ---------------------------------------------------------------------------
# Read helpers (response build + lịch sử lô)
# ---------------------------------------------------------------------------
async def load_batch_with_rows(
    db: AsyncSession, batch_id: int
) -> Tuple[Optional[PaymentImportBatch], List[PaymentImportRow]]:
    """Lô + các dòng (sắp row_no) cho response build. ``(None, [])`` nếu không có."""
    batch = (
        await db.execute(
            select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
        )
    ).scalar_one_or_none()
    if batch is None:
        return None, []
    rows = list(
        (
            await db.execute(
                select(PaymentImportRow)
                .where(PaymentImportRow.batch_id == batch_id)
                .order_by(PaymentImportRow.row_no)
            )
        )
        .scalars()
        .all()
    )
    return batch, rows


async def list_batches(
    db: AsyncSession, *, unit_id: Optional[int], skip: int = 0, limit: int = 50
) -> Tuple[List[PaymentImportBatch], int]:
    """Lịch sử lô import (mới nhất trước), phân trang.

    P1 IDOR: unit-scope (manager) chỉ thấy lô do user CÙNG ĐƠN VỊ tạo; admin/accountant
    (unit_id=None) → toàn hệ.
    """
    count_q = select(func.count()).select_from(PaymentImportBatch)
    list_q = select(PaymentImportBatch)
    if unit_id is not None:
        count_q = count_q.join(
            models.User, PaymentImportBatch.created_by_id == models.User.id
        ).where(models.User.unit_id == unit_id)
        list_q = list_q.join(
            models.User, PaymentImportBatch.created_by_id == models.User.id
        ).where(models.User.unit_id == unit_id)
    total = (await db.execute(count_q)).scalar_one()
    items = list(
        (
            await db.execute(
                list_q.order_by(
                    PaymentImportBatch.created_at.desc(),
                    PaymentImportBatch.id.desc(),
                )
                .offset(skip)
                .limit(limit)
            )
        )
        .scalars()
        .all()
    )
    return items, total


# ---------------------------------------------------------------------------
# BV-5 R2/R1 — xem lại per-row + xuất file kết quả
# ---------------------------------------------------------------------------
async def get_batch_detail_scoped(
    db: AsyncSession, batch_id: int, unit_id: Optional[int]
) -> Tuple[PaymentImportBatch, List[PaymentImportRow]]:
    """Lô + dòng, có IDOR unit-scope (mirror commit:998). Ngoài scope / không tồn tại
    → ``ResourceNotFoundError`` (404, không lộ tồn tại)."""
    batch, rows = await load_batch_with_rows(db, batch_id)
    if batch is None:
        raise ResourceNotFoundError("Không tìm thấy lô import")
    await _assert_batch_creator_in_unit(db, batch, unit_id)
    return batch, rows


def _result_status_label(batch_status: str, row: PaymentImportRow) -> str:
    """Nhãn Trạng thái cho file kết quả kế toán đọc.

    Đọc thẳng trục GHI thay vì suy từ (trạng thái lô × trạng thái kiểm). Bản
    trước phải suy vì chỉ có một trục, và phép suy ấy nói sai đúng ở ca quan
    trọng nhất: một dòng bị hàng rào giữ lại nằm trong lô đã đóng vẫn hiện "Đã
    ghi", trong khi không có đồng nào của nó vào sổ.
    """
    cs = row.commit_status
    if cs == PaymentImportCommitStatusEnum.not_applicable.value:
        return "Lỗi (không ghi)" if batch_status == (
            PaymentImportBatchStatusEnum.committed.value
        ) else "Lỗi"
    if batch_status == PaymentImportBatchStatusEnum.void.value:
        return "Đã đảo"
    if cs == PaymentImportCommitStatusEnum.committed.value:
        return "Đã ghi"
    if cs == PaymentImportCommitStatusEnum.duplicate_review_required.value:
        return "Chờ xác nhận trùng"
    if cs == PaymentImportCommitStatusEnum.failed.value:
        return "Ghi hỏng"
    return "Dự kiến ghi"  # pending


# "(hệ thống)" trong nhãn để tránh trùng nếu file có sẵn cột tên "Tên hồ sơ".
COL_PROFILE_NAME = "Tên hồ sơ (hệ thống)"  # tên hệ thống (Lead.full_name)
_RESULT_EXTRA_COLS = ["Trạng thái", "Lý do", "Mã Payment", "Đã ghi (đồng)"]


async def build_result_file(
    db: AsyncSession, batch_id: int, fmt: str, unit_id: Optional[int]
) -> Tuple[bytes, str, str]:
    """File kết quả = NGUYÊN dòng gốc (`raw`) + cột Trạng thái/Lý do/Mã Payment/Đã ghi
    → ``(content, media_type, filename)``. IDOR scope.

    🔴 P1 chống formula injection: MỌI ô (raw = user nhập + header cột do file quyết) qua
    ``sanitize_csv_cell`` (CSV + XLSX); XLSX thêm ``number_format='@'`` (text) phòng
    openpyxl diễn giải chuỗi mở đầu '=' thành công thức.
    """
    batch, rows = await get_batch_detail_scoped(db, batch_id, unit_id)
    committed_v = PaymentImportBatchStatusEnum.committed.value
    # Cột "Tên hồ sơ" = tên hệ thống để kế toán đối chiếu cạnh "Họ và tên học sinh"
    # của file (file có thể ghi lệch). Batch-query 1 lần (anti-N+1); dòng không
    # resolve được hồ sơ → "(không có hồ sơ)".
    profile_ids = {
        r.resolved_profile_id for r in rows if r.resolved_profile_id is not None
    }
    profile_names: Dict[int, str] = {}
    if profile_ids:
        name_rows = await db.execute(
            select(models.AdmissionProfile.id, models.Lead.full_name)
            .join(models.Lead, models.AdmissionProfile.lead_id == models.Lead.id)
            .where(models.AdmissionProfile.id.in_(profile_ids))
        )
        profile_names = {pid: (name or "") for pid, name in name_rows.all()}

    def _profile_name(r: PaymentImportRow) -> str:
        pid = r.resolved_profile_id
        if pid is None:
            return "(không có hồ sơ)"
        return profile_names.get(pid) or "(không có hồ sơ)"

    # Header = cột gốc của file + cột kết quả. JSONB KHÔNG giữ thứ tự key → sắp theo
    # TEMPLATE_COLS (cột template, đúng thứ tự file mẫu) rồi cột lạ (nếu có) ở cuối.
    # Cột gốc do FILE quyết → cũng phải sanitize.
    first_raw = rows[0].raw if rows and rows[0].raw else {}
    if first_raw:
        raw_keys = [c for c in TEMPLATE_COLS if c in first_raw] + [
            k for k in first_raw if k not in TEMPLATE_COLS
        ]
    else:
        raw_keys = list(TEMPLATE_COLS)
    # Chèn "Tên hồ sơ" NGAY SAU "Họ và tên học sinh" (đối chiếu cạnh nhau); file thiếu
    # cột tên → đặt cuối phần cột gốc, ngay trước cột kết quả.
    name_pos = raw_keys.index(COL_NAME) + 1 if COL_NAME in raw_keys else len(raw_keys)
    display_cols = raw_keys[:name_pos] + [COL_PROFILE_NAME] + raw_keys[name_pos:]
    header = display_cols + _RESULT_EXTRA_COLS

    def _row_cells(r: PaymentImportRow) -> List[str]:
        raw = r.raw or {}
        label = _result_status_label(batch.status, r)
        pay = ", ".join(str(p) for p in (r.payment_ids or []))
        written = (
            f"{r.amount:.0f}"  # VND nguyên — bỏ đuôi '.00' của Numeric(15,2)
            # Số tiền ĐÃ ghi lấy theo trục GHI của chính dòng, không theo
            # trạng thái lô: lô đóng không có nghĩa mọi dòng đều vào sổ.
            if (
                r.commit_status == PaymentImportCommitStatusEnum.committed.value
                and r.amount is not None
            )
            else ""
        )
        base = [raw.get(k, "") for k in raw_keys]
        base = base[:name_pos] + [_profile_name(r)] + base[name_pos:]
        cells = base + [label, r.message or "", pay, written]
        return sanitize_csv_row(cells)  # = [sanitize_csv_cell(c) for c in cells]

    fname = f"ket_qua_import_lo_{batch_id}"
    if (fmt or "").lower() == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(sanitize_csv_row(header))
        for r in rows:
            writer.writerow(_row_cells(r))
        content = (CSV_UTF8_BOM + buf.getvalue()).encode("utf-8")  # BOM
        return content, CSV_MEDIA_TYPE, f"{fname}.csv"

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Ket qua"
    ws.append(sanitize_csv_row(header))
    for r in rows:
        ws.append(_row_cells(r))
    # Force TEXT mọi ô (chống injection + giữ số 0 đầu CCCD).
    for ws_row in ws.iter_rows():
        for cell in ws_row:
            cell.number_format = TEXT_NUMBER_FORMAT
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), _XLSX_MEDIA, f"{fname}.xlsx"
