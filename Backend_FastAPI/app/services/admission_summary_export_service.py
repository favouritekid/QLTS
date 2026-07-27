"""Xuất báo cáo tuyển sinh (snapshot) ra Excel — orchestration only, no FastAPI.

Báo cáo "ảnh chụp" phân bố TẠI THỜI ĐIỂM xuất, KHÁC cockpit tuần (funnel theo
thời gian). Là một PHỄU: mỗi lead xếp vào đúng GIAI ĐOẠN CAO NHẤT đạt được, nên
9 cột đếm của phễu (5 tư vấn + Hồ sơ chưa đóng phí + Lệ phí + Đóng một phần +
Đóng đủ HK1) CỘNG = "Tổng lead". Ưu tiên cao→thấp:
  Học phí > Lệ phí > Hồ sơ > Đang tư vấn.
(Các cột "Tổng học phí" (số đếm = tổng con) và "Doanh thu" (tiền) KHÔNG thuộc 9
cột phễu; nhóm "Chi tiết hồ sơ" cũng ngoài phễu.)

- HỌC PHÍ (đã đóng học phí HK1 = fee tuition ``semester_no=1``, partial hoặc đủ):
    Đóng một phần (partial) | Đóng đủ HK1 (đã thu đủ/miễn) | Tổng học phí (SỐ HỒ
    SƠ đã đóng học phí = hp_1p + hp_dhk1; là cột tổng con, KHÔNG cộng vào phễu) |
    Doanh thu (SỐ TIỀN học phí, VND). ⚠️ ĐỔI NGÀNH: Doanh thu gán theo NGÀNH-LÚC-
    THU (``Payment.recognized_major_id``, ledger payment+refund+reversal đã net)
    — KHÁC đếm hồ sơ (theo NV1 hiện tại). Tiền đã thu cho ngành A vẫn ở A dù hồ
    sơ đã đổi sang B → ngành A có thể là dòng "chỉ còn doanh thu, không hồ sơ";
    recognized NULL → dòng "(Đã thu, chưa xác định ngành)" tách riêng.
- LỆ PHÍ: đã đóng/miễn lệ phí xét tuyển (application: paid_amount>0 OR waived)
    NHƯNG chưa đóng học phí.
- HỒ SƠ (đã có hồ sơ nhưng CHƯA đóng phí nào) — 1 cột. Chi tiết nháp/nợ/đủ xem
    khối "Chi tiết hồ sơ" (tránh cột 0/0 vô nghĩa vì hồ sơ đã nộp thường đã đóng
    phí nên chỉ còn hồ sơ nháp ở đây).
- ĐANG TƯ VẤN (CHƯA có hồ sơ), theo ``consultation_status`` hiện tại: Chưa tiếp
    cận (sts00) | Đã kết nối (sts02) | Có nhu cầu tìm hiểu (sts03) | Đồng ý tư
    vấn/Hẹn liên hệ (sts06+sts05) | Từ chối tư vấn/Đã ngừng liên hệ (sts04+sts20).
    Lead chưa hồ sơ mà trạng thái khác (sts01/15/19 phổ quát/NULL) → gộp "Chưa
    tiếp cận". ⚠️ Vì phễu, cột này KHÁC filter lead: lead đã có hồ sơ (dù nháp)
    nằm ở nhóm Hồ sơ, còn filter đếm chúng theo consultation_status.
- THAM CHIẾU (ngoài phễu, KHÔNG cộng tổng): "Tổng hồ sơ đã tạo" = MỌI hồ sơ ở
    mọi giai đoạn, phân Chưa hoàn thiện/Hoàn thiện 1 phần/Đủ điều kiện. Giải
    thích vì sao nhóm Hồ sơ (phễu) < tổng hồ sơ thật: hồ sơ đã đóng phí đã tiến
    sang nhóm Lệ phí/Học phí.

Scope: admin = toàn trường (hoặc đơn vị chọn); manager = ép đơn vị của mình.
Read-only → service không commit. Raises domain exceptions, never HTTPException.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app import models
from app.constants.cash_ledger import CASH_TRANSACTION_TYPES, sql_in_list
from app.services.admission_report_service import AdmissionReportService
from app.utils.csv_helpers import sanitize_csv_cell

VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh")

# ---- Quy ước cột (xem docstring) --------------------------------------------
# Mỗi cột = (key, nhãn). Nhóm = (tên nhóm, cột, màu nền).
TV_COLS = [
    ("tv_ctc", "Chưa tiếp cận"),
    ("tv_kn", "Đã kết nối"),
    ("tv_nc", "Có nhu cầu tìm hiểu"),
    ("tv_dyh", "Đồng ý tư vấn/Hẹn liên hệ"),
    ("tv_tcn", "Từ chối tư vấn/Đã ngừng liên hệ"),
]
# Hồ sơ (PHỄU) = lead ĐÃ có hồ sơ nhưng CHƯA đóng phí nào. 1 cột (chi tiết
# nháp/nợ/đủ ở khối "Chi tiết hồ sơ" bên dưới, tránh cột 0/0 vô nghĩa).
HS_COLS = [("hs_cd", "Hồ sơ chưa đóng phí")]
LP_COLS = [("le_phi", "Lệ phí")]
# Học phí — 3 cột ĐẾM hồ sơ (Đóng một phần / Đóng đủ HK1 / Tổng học phí = số hồ
# sơ đã đóng học phí = hp_1p + hp_dhk1) + 1 cột TIỀN (Doanh thu = học phí HK1 đã
# thu, VND). Chỉ hp_1p + hp_dhk1 thuộc phễu; hp_tong là tổng con, hp_dt là tiền.
HP_COLS = [
    ("hp_1p", "Đóng một phần"),
    ("hp_dhk1", "Đóng đủ HK1"),
    ("hp_tong", "Tổng học phí"),
    ("hp_dt", "Doanh thu"),
]
# CHI TIẾT HỒ SƠ (KHÔNG thuộc phễu, KHÔNG cộng vào Tổng lead): MỌI hồ sơ đã tạo
# ở mọi giai đoạn (kể cả đã đóng phí) phân theo trạng thái. Nhóm 'Hồ sơ' của phễu
# chỉ gồm hồ sơ CHƯA đóng phí; hồ sơ đã đóng phí đã tiến sang Lệ phí/Học phí →
# Tổng hồ sơ đã tạo = Hồ sơ (phễu) + Lệ phí + Học phí.
REF_COLS = [
    ("ref_ct", "Chưa hoàn thiện"),
    ("ref_1p", "Hoàn thiện 1 phần"),
    ("ref_dk", "Đủ điều kiện"),
    ("ref_tong", "Tổng hồ sơ"),
]

# Consultation status HIỆN TẠI → cột "Đang tư vấn" (chỉ áp cho lead CHƯA có hồ
# sơ). Lead chưa hồ sơ mà trạng thái không nằm ở đây (sts01/sts15/sts19 phổ quát,
# hoặc NULL) → gộp "Chưa tiếp cận" để phễu vẫn đủ (mỗi lead đúng 1 cột).
_TV_FALLBACK = "tv_ctc"
_TV_OF_CS = {
    "sts00": "tv_ctc",
    "sts02": "tv_kn",
    "sts03": "tv_nc",
    "sts06": "tv_dyh",
    "sts05": "tv_dyh",
    "sts04": "tv_tcn",
    "sts20": "tv_tcn",
}

_UNASSIGNED = 0  # sentinel cột "Chưa PC" cho lead không có nhân viên phụ trách
_TL_KEY = "_tl"  # pseudo-key: Tổng lead (dùng cho khối per-nhân-viên sheet 2)

# ---- Styling ----------------------------------------------------------------
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor="4472C4")
_HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
_GRPA_FILL = PatternFill("solid", fgColor="D9E1F2")  # Đang tư vấn
_GRPB_FILL = PatternFill("solid", fgColor="E2EFDA")  # Hồ sơ
_GRPC_FILL = PatternFill("solid", fgColor="FCE4D6")  # Lệ phí
_GRPD_FILL = PatternFill("solid", fgColor="FFF2CC")  # Học phí
_GRPE_FILL = PatternFill("solid", fgColor="EDEDED")  # Tham chiếu (xám)
_SUB_FONT = Font(bold=True, color="1F3864", size=9)
_TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
_TOTAL_FONT = Font(bold=True, size=10)
_TITLE_FONT = Font(bold=True, size=14, color="1F3864")
_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_MONEY_FMT = "#,##0"

# Nhóm cột (sheet 1). 4 nhóm đầu = PHỄU (mỗi lead đúng 1 cột đếm, cộng = Tổng
# lead); nhóm cuối = THAM CHIẾU (không cộng vào tổng). Sheet 2 chèn khối "Tổng
# lead" ở đầu.
GROUPS = [
    ("Đang tư vấn (chưa có hồ sơ)", TV_COLS, _GRPA_FILL),
    ("Hồ sơ (chưa đóng phí)", HS_COLS, _GRPB_FILL),
    ("Lệ phí", LP_COLS, _GRPC_FILL),
    ("Học phí", HP_COLS, _GRPD_FILL),
    ("Chi tiết hồ sơ đã tạo (mọi giai đoạn — KHÔNG thuộc phễu)", REF_COLS, _GRPE_FILL),
]
ALL_COLS = TV_COLS + HS_COLS + LP_COLS + HP_COLS + REF_COLS
ALL_KEYS = [k for k, _ in ALL_COLS]
MONEY_KEYS = {"hp_dt"}  # chỉ Doanh thu là VND; hp_tong là số ĐẾM hồ sơ
# Các key được cộng dồn theo lead (gồm _tl để sheet 2 có khối Tổng lead).
STORE_KEYS = [_TL_KEY] + ALL_KEYS

# Lead is year-scoped via its offering's academic year (lead has no year column).
# A lead WITHOUT an offering has no year of its own → scope it by the VN-tz year of
# its created_at (a brand-new lead that hasn't picked a ngành still counts in the
# year it arrived). Without this guard an offering-less lead from ANY prior year
# would show up in EVERY year's report and inflate "Tổng lead".
# Shared by _LEAD_SQL and _OFFICER_SQL, so it may reference only lead/po columns
# (the officer query doesn't join admission_profile) → both sheets stay reconciled.
_YEAR_SCOPE = """
      AND (
          EXISTS (
              SELECT 1 FROM offering_academic_info oai
              WHERE oai.offering_id = po.id AND oai.academic_year = :year
          )
          OR (
              po.id IS NULL
              AND EXTRACT(
                  YEAR FROM l.created_at AT TIME ZONE 'Asia/Ho_Chi_Minh'
              )::int = :year
          )
      )"""

# Một dòng / lead. Các CTE gom fee + hồ sơ về mức profile TRƯỚC khi nối vào lead
# nên KHÔNG có row-multiplication (không cần GROUP BY). Mỗi lead lấy 1 hồ sơ đại
# diện của năm (mới nhất) qua DISTINCT ON — nếu 1 lead lỡ có 2 hồ sơ cùng năm thì
# "Tổng lead" vẫn khớp filter (đếm theo lead).
#   - has_doc_debt: hồ sơ đã nộp còn NỢ giấy tờ = có mã trong document_debt->codes
#     chưa được duyệt (verified/paper_submitted). Khớp outstanding_debt_codes.
#   - has_app: đã đóng/miễn lệ phí xét tuyển.
#   - hk1_partial / hk1_settled: học phí HK1 (fee tuition semester_no=1) đóng một
#     phần / đã đủ (paid|waived hoặc remaining<=0), bỏ cancelled.
#   - hk1_paid: TIỀN học phí HK1 đã thu (Σ paid_amount trên hồ sơ đã đóng) — cho
#     cột "Doanh thu". ("Tổng học phí" là SỐ ĐẾM hồ sơ, không dùng final_amount.)
# Ngành (pid): nguyện vọng 1 của hồ sơ năm (display_order=1) nếu có, else ngành
# quan tâm (offering) của lead.
_LEAD_SQL = text(
    """
    WITH fee_agg AS (
        SELECT f.admission_profile_id AS ap_id,
               bool_or(f.fee_type='application' AND f.status<>'cancelled'
                       AND (f.paid_amount > 0 OR f.status='waived')) AS has_app,
               bool_or(f.fee_type='tuition' AND f.semester_no=1
                       AND f.status='partial') AS hk1_partial,
               bool_or(f.fee_type='tuition' AND f.semester_no=1
                       AND f.status<>'cancelled'
                       AND (f.status IN ('paid','waived')
                            OR (f.final_amount - f.paid_amount
                                - f.waived_amount) <= 0)) AS hk1_settled,
               COALESCE(sum(f.paid_amount) FILTER (
                   WHERE f.fee_type='tuition' AND f.semester_no=1
                     AND f.status<>'cancelled'
                     AND (f.status='partial' OR f.status IN ('paid','waived')
                          OR (f.final_amount - f.paid_amount
                              - f.waived_amount) <= 0)), 0) AS hk1_paid
        FROM fee f
        GROUP BY f.admission_profile_id
    ),
    prof AS (
        SELECT DISTINCT ON (ap.lead_id)
               ap.id AS ap_id, ap.lead_id, ap.status AS pstatus,
               (ap.document_debt IS NOT NULL AND EXISTS (
                   SELECT 1
                   FROM jsonb_array_elements_text(
                            ap.document_debt->'codes') AS d(code)
                   WHERE NOT EXISTS (
                       SELECT 1 FROM profile_document pd
                       JOIN config_document_type cdt
                            ON cdt.id = pd.document_type_id
                       WHERE pd.profile_id = ap.id AND pd.category='path'
                         AND cdt.code = d.code
                         AND pd.status IN ('verified','paper_submitted')
                   )
               )) AS has_doc_debt,
               COALESCE(fa.has_app, false) AS has_app,
               COALESCE(fa.hk1_partial, false) AS hk1_partial,
               COALESCE(fa.hk1_settled, false) AS hk1_settled,
               COALESCE(fa.hk1_paid, 0) AS hk1_paid
        FROM admission_profile ap
        LEFT JOIN fee_agg fa ON fa.ap_id = ap.id
        WHERE ap.academic_year = :year
        ORDER BY ap.lead_id, ap.id DESC
    ),
    nv1maj AS (
        SELECT DISTINCT ON (nv1.admission_profile_id)
               nv1.admission_profile_id AS ap_id, nv1_po.program_id
        FROM admission_profile_choice nv1
        JOIN admission_path nv1_path ON nv1.admission_path_id = nv1_path.id
        JOIN offering_academic_info nv1_oai
             ON nv1_path.academic_info_id = nv1_oai.id
        JOIN program_offering nv1_po ON nv1_oai.offering_id = nv1_po.id
        WHERE nv1.display_order = 1
        ORDER BY nv1.admission_profile_id, nv1.id
    )
    SELECT l.id,
           COALESCE(nm.program_id, po.program_id) AS pid,
           l.consultation_status_id AS cs,
           l.assigned_officer_id AS off,
           p.pstatus AS pstatus,
           COALESCE(p.has_doc_debt, false) AS has_doc_debt,
           COALESCE(p.has_app, false) AS has_app,
           -- F13: a withdrawn / awaiting-refund profile's HK1 tuition is
           -- refunded (or being refunded), so it is NOT realized revenue —
           -- zero the tuition flags/amount so it drops out of 'Doanh thu' and
           -- the Học phí funnel, consistent with a finalized withdrawn profile
           -- (whose cancelled fees already drop via the status<>'cancelled'
           -- filter). Lệ phí xét tuyển (has_app) is non-refundable and stays.
           CASE WHEN p.pstatus IN ('withdrawn', 'withdrawal_pending')
                THEN false ELSE COALESCE(p.hk1_partial, false) END AS hk1_partial,
           CASE WHEN p.pstatus IN ('withdrawn', 'withdrawal_pending')
                THEN false ELSE COALESCE(p.hk1_settled, false) END AS hk1_settled,
           CASE WHEN p.pstatus IN ('withdrawn', 'withdrawal_pending')
                THEN 0 ELSE COALESCE(p.hk1_paid, 0) END AS hk1_paid
    FROM lead l
    LEFT JOIN program_offering po ON l.offering_id = po.id
    LEFT JOIN prof p ON p.lead_id = l.id
    LEFT JOIN nv1maj nm ON nm.ap_id = p.ap_id
    WHERE l.deleted_at IS NULL
      AND (CAST(:unit AS INTEGER) IS NULL OR l.unit_id = CAST(:unit AS INTEGER))"""
    + _YEAR_SCOPE
)
_OFFICER_SQL = text(
    """
    SELECT u.id, COALESCE(u.full_name, u.username) AS nm, count(l.id) AS c
    FROM lead l
    JOIN "user" u ON l.assigned_officer_id = u.id
    LEFT JOIN program_offering po ON l.offering_id = po.id
    WHERE l.deleted_at IS NULL
      AND (CAST(:unit AS INTEGER) IS NULL OR l.unit_id = CAST(:unit AS INTEGER))"""
    + _YEAR_SCOPE
    + """
    GROUP BY 1, 2 ORDER BY c DESC, u.id
    """
)
# Include inactive majors too: a lead/profile may still point to a major archived
# mid-year; filtering by is_active would drop it into '(Chưa xác định ngành)' instead
# of its real name. Majors with no leads are dropped later by the ``ordered`` filter,
# so this never adds empty rows.
_MAJOR_SQL = text(
    """
    SELECT id, code, name, degree_level FROM major_program
    ORDER BY CASE degree_level WHEN 'Cao đẳng' THEN 0
                               WHEN 'Trung cấp' THEN 1 ELSE 2 END,
             name, code
    """
)

# Doanh thu học phí (cột hp_dt) theo NGÀNH-LÚC-THU — đổi ngành: tiền đã thu cho
# ngành A vẫn thuộc A dù hồ sơ đã chuyển sang B. Aggregate ledger THẬT (payment +
# refund + reversal, amount signed → net) theo Payment.recognized_major_id, KHÔNG
# theo fee.paid_amount của NV1 hiện tại. GỒM reversal (void lô) để net đúng ngành
# gốc. CHỈ HK1 (semester_no=1) — khớp cột đếm hồ sơ. recognized NULL → sentinel
# ``:rev_null``. Per-officer để khớp sheet 2.
#  • F13 (khớp _LEAD_SQL): LOẠI hồ sơ withdrawn/withdrawal_pending — học phí đang
#    hoàn/đã hoàn KHÔNG phải doanh thu thực (cột đếm cũng zero chúng); nếu để lại,
#    withdrawal_pending (refund duyệt nhưng CHƯA post ledger) sẽ thổi doanh thu.
#  • YEAR-SCOPE (khớp _LEAD_SQL/_OFFICER_SQL qua _YEAR_SCOPE): revenue chỉ của
#    lead TRONG cùng phạm vi năm với cột officer → officer luôn ∈ officer_id_set
#    (không rơi 'Chưa PC' oan) + tổng Excel reconcile với báo cáo tuần.
_REVENUE_SQL = text(
    """
    SELECT COALESCE(pay.recognized_major_id, CAST(:rev_null AS INTEGER)) AS rmaj,
           l.assigned_officer_id AS off,
           COALESCE(SUM(pt.amount), 0) AS revenue
    FROM payment_transaction pt
    -- LEFT JOIN, KHÔNG phải INNER: ``payment_transaction.payment_id`` là nullable
    -- (điều chỉnh sổ không gắn payment nào). Với INNER, những giao dịch đó bị
    -- Excel BỎ HẲN trong khi báo cáo tuần vẫn tính chúng (xếp vào "chưa xác định
    -- ngành") ⇒ hai artifact người dùng đối chiếu lệch nhau đúng số đó, và lệch
    -- theo hướng THIẾU tiền nên không ai thấy. Nay ``recognized_major_id`` NULL
    -- rơi vào sentinel ``:rev_null`` giống hệt quy tắc của báo cáo tuần.
    LEFT JOIN payment pay ON pt.payment_id = pay.id
    JOIN fee f ON pt.fee_id = f.id
              AND f.fee_type = 'tuition' AND f.semester_no = 1
    JOIN admission_profile ap ON f.admission_profile_id = ap.id
                              AND ap.academic_year = :year
                              AND ap.status NOT IN
                                  ('withdrawn', 'withdrawal_pending')
    JOIN lead l ON ap.lead_id = l.id AND l.deleted_at IS NULL
    LEFT JOIN program_offering po ON l.offering_id = po.id
    WHERE pt.transaction_type IN """
    + sql_in_list(CASH_TRANSACTION_TYPES)
    + """
      AND (CAST(:unit AS INTEGER) IS NULL OR l.unit_id = CAST(:unit AS INTEGER))
    """
    + _YEAR_SCOPE
    + """
    GROUP BY 1, 2
    """
)

# Sentinel doanh thu chưa xác định ngành (tuition thu khi fee chưa chốt ngành).
# KHÁC ``NONE`` (-1, ngành HỒ SƠ chưa xác định) — 2 nghĩa không gộp.
_REVENUE_NULL = -2


def _H(cell, fill=_HEAD_FILL, font=_HEAD_FONT):
    cell.fill = fill
    cell.font = font
    cell.alignment = _CTR
    cell.border = _BORDER


def _officer_short_names(oname: dict, officer_ids: list) -> dict:
    """Short header label (last name token) per officer.

    Whitespace-safe ('   '.split() → [] → '#id', avoids IndexError),
    CSV-injection-safe (sanitize_csv_cell), and disambiguated so two officers
    sharing a last name fall back to the full name instead of colliding.
    """
    base = {}
    for oid in officer_ids:
        parts = (oname.get(oid) or "").split()
        base[oid] = parts[-1] if parts else f"#{oid}"
    counts: dict = {}
    for v in base.values():
        counts[v] = counts.get(v, 0) + 1
    out = {}
    for oid in officer_ids:
        label = base[oid]
        if counts[label] > 1:  # collision → fuller name to keep columns distinct
            label = (oname.get(oid) or "").strip() or f"#{oid}"
        out[oid] = sanitize_csv_cell(label)
    return out


class AdmissionSummaryExportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def build_xlsx(
        self,
        current_user: models.User,
        academic_year: int,
        unit_id: Optional[int] = None,
    ) -> tuple[bytes, str]:
        # Reuse the weekly report's scope rule (admin all/chosen unit, manager
        # forced own unit) so viewing and exporting enforce IDOR identically.
        scope_unit = AdmissionReportService._resolve_scope(current_user, unit_id)
        params = {"year": academic_year, "unit": scope_unit}
        leads = (await self.db.execute(_LEAD_SQL, params)).mappings().all()
        officers = (await self.db.execute(_OFFICER_SQL, params)).mappings().all()
        majors = (await self.db.execute(_MAJOR_SQL)).mappings().all()
        revenue = (
            await self.db.execute(
                _REVENUE_SQL, {**params, "rev_null": _REVENUE_NULL}
            )
        ).mappings().all()

        wb = self._build_workbook(academic_year, leads, officers, majors, revenue)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        ts = datetime.now(VN_TZ).strftime("%Y%m%d_%H%M%S")
        filename = f"bao_cao_tuyen_sinh_{academic_year}_{ts}.xlsx"
        return bio.getvalue(), filename

    # ----------------------------------------------------------- workbook build
    def _build_workbook(self, year, leads, officers, majors, revenue) -> Workbook:
        officer_ids = [r["id"] for r in officers]
        officer_id_set = set(officer_ids)
        oname = {r["id"]: r["nm"] for r in officers}
        oshort = _officer_short_names(oname, officer_ids)
        # "Chưa PC" cần cho lead-không-officer VÀ doanh thu-không-officer (phiếu
        # thu của hồ sơ chưa gán cán bộ vẫn phải vào cột nào đó ở sheet 2).
        # ⚠️ KHÔNG guard `and r["revenue"]`: revenue loop bump MỌI row (kể cả net
        # 0 do void sạch) vào cột theo r["off"] → nếu off ∉ set mà cột _UNASSIGNED
        # chưa add (revenue=0 falsy làm rev_unassigned False) → bump KeyError → 500.
        rev_unassigned = any(
            r["off"] not in officer_id_set for r in revenue
        )
        n_unassigned = sum(1 for r in leads if r["off"] not in officer_id_set)
        # Sheet-2 columns = officers (+ "Chưa PC" if any lead has no officer) so
        # EVERY lead is attributed → sheet 2 reconciles exactly with sheet 1.
        officer_cols = list(officer_ids)
        if n_unassigned or rev_unassigned:
            officer_cols.append(_UNASSIGNED)
            oshort[_UNASSIGNED] = "Chưa PC"

        major_ids = [m["id"] for m in majors]
        minfo = {m["id"]: m for m in majors}
        NONE = -1
        # _REVENUE_NULL (-2): doanh thu tuition thu khi fee chưa chốt ngành.
        all_keys = major_ids + [NONE, _REVENUE_NULL]

        # s1[pid][key] = giá trị tổng (sheet 1); s2[pid][key][officer] (sheet 2).
        s1 = {p: {k: 0 for k in STORE_KEYS} for p in all_keys}
        s2 = {
            p: {k: {o: 0 for o in officer_cols} for k in STORE_KEYS} for p in all_keys
        }

        def bump(pid, col, key, val):
            s1[pid][key] += val
            s2[pid][key][col] += val

        for r in leads:
            pid = r["pid"] if r["pid"] in s1 else NONE
            off = r["off"]
            # col is always a valid key: an unassigned lead implies _UNASSIGNED ∈ cols
            col = off if off in officer_id_set else _UNASSIGNED

            bump(pid, col, _TL_KEY, 1)  # Tổng lead

            ps = r["pstatus"]
            has_app = r["has_app"]
            partial = r["hk1_partial"]
            settled = r["hk1_settled"]
            has_hp = partial or settled  # đã đóng học phí HK1 (một phần hoặc đủ)

            # ---- PHỄU: mỗi lead vào GIAI ĐOẠN CAO NHẤT (đúng 1 cột) ----
            if has_hp:  # Học phí (cao nhất)
                # partial & settled loại trừ: HK1 đã đủ (remaining<=0) → "Đóng đủ
                # HK1", KHÔNG đếm lại "Đóng một phần" dù status='partial'.
                if partial and not settled:
                    bump(pid, col, "hp_1p", 1)
                else:
                    bump(pid, col, "hp_dhk1", 1)
                bump(pid, col, "hp_tong", 1)  # tổng con: số hồ sơ đã đóng học phí
                # Doanh thu (hp_dt) KHÔNG lấy từ đây nữa — gán riêng theo ledger
                # recognized_major bên dưới (đổi ngành: tiền theo ngành-lúc-thu).
            elif has_app:  # Lệ phí (đã đóng lệ phí, chưa học phí)
                bump(pid, col, "le_phi", 1)
            elif ps is not None:  # Hồ sơ (có hồ sơ, chưa đóng phí nào) — 1 cột
                bump(pid, col, "hs_cd", 1)
            else:  # Đang tư vấn (chưa có hồ sơ)
                bump(pid, col, _TV_OF_CS.get(r["cs"], _TV_FALLBACK), 1)

            # ---- THAM CHIẾU (ngoài phễu): mọi hồ sơ đã tạo, phân theo trạng thái
            if ps is not None:
                if ps == "draft":
                    rk = "ref_ct"
                elif r["has_doc_debt"]:
                    rk = "ref_1p"
                else:
                    rk = "ref_dk"
                bump(pid, col, rk, 1)
                bump(pid, col, "ref_tong", 1)

        # ---- DOANH THU (hp_dt) theo NGÀNH-LÚC-THU (recognized_major) — TÁCH khỏi
        # đếm hồ sơ. Đổi ngành: tiền đã thu cho ngành A vẫn ở A dù hồ sơ đã sang B.
        # Ledger đã net payment+refund+reversal (SUM signed). rmaj có thể là ngành
        # KHÔNG còn lead (chỉ còn tiền — vẫn hiện) hoặc _REVENUE_NULL (chưa xác
        # định ngành). Guard: rmaj lạ (major hard-deleted) rơi về _REVENUE_NULL.
        for r in revenue:
            rmaj = r["rmaj"]
            if rmaj not in s1:
                rmaj = _REVENUE_NULL
            off = r["off"]
            col = off if off in officer_id_set else _UNASSIGNED
            bump(rmaj, col, "hp_dt", int(r["revenue"] or 0))

        # Giữ ngành có lead HOẶC có doanh thu (đổi ngành: ngành chỉ còn tiền vẫn
        # phải xuất hiện — không bị lọc mất theo lead).
        def _has_data(p):
            return s1[p][_TL_KEY] > 0 or s1[p]["hp_dt"] != 0

        ordered = [p for p in major_ids if _has_data(p)]
        # 2 dòng sentinel cuối, chỉ khi có dữ liệu: NONE (ngành hồ sơ chưa xác
        # định) + _REVENUE_NULL (doanh thu chưa xác định ngành) — tách bạch.
        if _has_data(NONE):
            ordered.append(NONE)
        if _has_data(_REVENUE_NULL):
            ordered.append(_REVENUE_NULL)

        def desc_of(pid):
            if pid == NONE:
                return "", "(Chưa xác định ngành)", ""
            if pid == _REVENUE_NULL:
                return "", "(Đã thu, chưa xác định ngành)", ""
            m = minfo[pid]
            return (
                sanitize_csv_cell(m["code"]),
                sanitize_csv_cell(m["name"]),
                sanitize_csv_cell(m["degree_level"]),
            )

        ts_label = datetime.now(VN_TZ).strftime("%d/%m/%Y %H:%M")
        wb = Workbook()
        DESC = ["TT", "Mã ngành", "Ngành, nghề", "Trình độ", "Liên thông/VB2"]
        ND = len(DESC)

        self._sheet_summary(wb, year, ts_label, DESC, ND, ordered, desc_of, s1)
        self._sheet_by_officer(
            wb,
            year,
            ts_label,
            DESC,
            ND,
            ordered,
            desc_of,
            officer_cols,
            oshort,
            n_unassigned,
            s2,
        )
        self._sheet_notes(wb, year, n_unassigned, len(leads))
        return wb

    # --------------------------------------------------------- sheet 1: chung
    def _sheet_summary(self, wb, year, ts_label, DESC, ND, ordered, desc_of, s1):
        ws = wb.active
        ws.title = "Số liệu chung"
        c_tong = ND + 1  # cột "Tổng lead"
        c_metric = c_tong + 1  # cột đầu tiên của khối PHÂN LOẠI
        total = ND + 1 + len(ALL_KEYS)

        ws.cell(1, 1, f"SỐ LIỆU TUYỂN SINH NĂM {year}").font = _TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total)
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(2, 1, "Thời điểm báo cáo:").font = Font(italic=True, size=9)
        ws.cell(2, 2, ts_label).font = Font(italic=True, size=9)

        # Header 3 hàng (4,5,6): DESC + Tổng lead merge dọc; PHÂN LOẠI → nhóm → cột.
        for i, h in enumerate(DESC, 1):
            ws.merge_cells(start_row=4, start_column=i, end_row=6, end_column=i)
            _H(ws.cell(4, i, h))
        ws.merge_cells(start_row=4, start_column=c_tong, end_row=6, end_column=c_tong)
        _H(ws.cell(4, c_tong, "Tổng lead"))
        ws.merge_cells(start_row=4, start_column=c_metric, end_row=4, end_column=total)
        _H(ws.cell(4, c_metric, "PHÂN LOẠI"))

        col = c_metric
        for gname, gcols, gfill in GROUPS:
            n = len(gcols)
            if n == 1:  # nhóm 1 cột (Lệ phí) → gộp nhãn nhóm + cột theo chiều dọc
                ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
                _H(ws.cell(5, col, gname), fill=gfill, font=_SUB_FONT)
            else:
                ws.merge_cells(
                    start_row=5, start_column=col, end_row=5, end_column=col + n - 1
                )
                _H(ws.cell(5, col, gname), fill=gfill, font=_SUB_FONT)
                for j, (k, label) in enumerate(gcols):
                    _H(ws.cell(6, col + j, label), fill=gfill, font=_SUB_FONT)
            col += n

        # TỔNG CỘNG (row 7)
        ws.cell(7, 1, "TỔNG CỘNG")
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=ND)
        ws.cell(7, 1).font = _TOTAL_FONT
        ws.cell(7, 1).alignment = _CTR
        for i in range(1, ND + 1):
            ws.cell(7, i).fill = _TOTAL_FILL
            ws.cell(7, i).border = _BORDER
        vals = [sum(s1[p][_TL_KEY] for p in ordered)] + [
            sum(s1[p][k] for p in ordered) for k in ALL_KEYS
        ]
        for j, v in enumerate(vals):
            c = ws.cell(7, c_tong + j, v)
            c.fill = _TOTAL_FILL
            c.font = _TOTAL_FONT
            c.alignment = _CTR
            c.border = _BORDER
            if j >= 1 and ALL_KEYS[j - 1] in MONEY_KEYS:
                c.number_format = _MONEY_FMT

        r = 8
        for n, pid in enumerate(ordered, 1):
            code, name, deg = desc_of(pid)
            ws.cell(r, 1, n).alignment = _CTR
            ws.cell(r, 2, code).alignment = _CTR
            ws.cell(r, 3, name).alignment = _LEFT
            ws.cell(r, 4, deg).alignment = _CTR
            ws.cell(r, 5, "Không").alignment = _CTR
            row_vals = [s1[pid][_TL_KEY]] + [s1[pid][k] for k in ALL_KEYS]
            for j, v in enumerate(row_vals):
                c = ws.cell(r, c_tong + j, v)
                c.alignment = _CTR
                if j == 0:
                    c.font = Font(bold=True)
                elif ALL_KEYS[j - 1] in MONEY_KEYS:
                    c.number_format = _MONEY_FMT
            for i in range(1, total + 1):
                ws.cell(r, i).border = _BORDER
            r += 1

        for col_letter, w in zip("ABCDE", (4.5, 11, 34, 10, 9)):
            ws.column_dimensions[col_letter].width = w
        ws.column_dimensions[get_column_letter(c_tong)].width = 8
        for j, k in enumerate(ALL_KEYS):
            ws.column_dimensions[get_column_letter(c_metric + j)].width = (
                14 if k in MONEY_KEYS else 11
            )
        # Freeze header rows 1-7 + descriptor cols A-E + Tổng lead.
        ws.freeze_panes = ws.cell(8, c_metric).coordinate

    # ----------------------------------------------------- sheet 2: nhân viên
    def _sheet_by_officer(
        self,
        wb,
        year,
        ts_label,
        DESC,
        ND,
        ordered,
        desc_of,
        officer_cols,
        oshort,
        n_unassigned,
        s2,
    ):
        ws = wb.create_sheet("Chia theo nhân viên")
        noff = len(officer_cols)
        block_w = 1 + noff  # cột "Tổng" + mỗi nhân viên (+ "Chưa PC" nếu có)
        # Khối đầu = Tổng lead; sau đó 4 nhóm PHÂN LOẠI.
        blocks = [("Tổng lead", [(_TL_KEY, "Tổng lead")], _TOTAL_FILL)] + GROUPS
        n_metric = len(STORE_KEYS)  # 1 (_tl) + 15 (ALL_KEYS: 5+1+1+4+4)
        aStart = ND + 1
        total2 = ND + n_metric * block_w

        ws.cell(1, 1, f"SỐ LIỆU TUYỂN SINH NĂM {year} — CHIA THEO NHÂN VIÊN").font = (
            _TITLE_FONT
        )
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=max(total2, 1)
        )
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(2, 1, "Thời điểm báo cáo:").font = Font(italic=True, size=9)
        ws.cell(2, 2, ts_label).font = Font(italic=True, size=9)
        if n_unassigned:
            ws.cell(
                2,
                6,
                f"{n_unassigned} lead chưa phân công nằm ở cột 'Chưa PC' (mỗi "
                "khối) → cột 'Tổng' khớp sheet 'Số liệu chung'.",
            ).font = Font(italic=True, color="C00000", size=9)

        if not officer_cols:
            return  # không có lead trong phạm vi → sheet chỉ có tiêu đề

        for i, h in enumerate(DESC, 1):
            ws.merge_cells(start_row=3, start_column=i, end_row=5, end_column=i)
            _H(ws.cell(3, i, h))

        # Header: row3 nhóm, row4 nhãn cột (merge block_w), row5 Tổng + nhân viên.
        col = aStart
        for gname, gcols, gfill in blocks:
            span = len(gcols) * block_w
            ws.merge_cells(
                start_row=3, start_column=col, end_row=3, end_column=col + span - 1
            )
            _H(ws.cell(3, col, gname))
            col += span
        col = aStart
        for gname, gcols, gfill in blocks:
            for k, label in gcols:
                ws.merge_cells(
                    start_row=4,
                    start_column=col,
                    end_row=4,
                    end_column=col + block_w - 1,
                )
                _H(ws.cell(4, col, label), fill=gfill, font=_SUB_FONT)
                _H(ws.cell(5, col, "Tổng"), fill=gfill, font=_SUB_FONT)
                for t, oid in enumerate(officer_cols):
                    _H(ws.cell(5, col + 1 + t, oshort[oid]), fill=gfill, font=_SUB_FONT)
                col += block_w

        # TỔNG CỘNG row 6
        ws.cell(6, 1, "TỔNG CỘNG")
        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=ND)
        ws.cell(6, 1).font = _TOTAL_FONT
        ws.cell(6, 1).alignment = _CTR
        for i in range(1, ND + 1):
            ws.cell(6, i).fill = _TOTAL_FILL
            ws.cell(6, i).border = _BORDER

        def emit(row, pid, is_total):
            col = aStart
            for gname, gcols, gfill in blocks:
                for k, _label in gcols:
                    vals = [
                        (
                            sum(s2[p][k][oid] for p in ordered)
                            if is_total
                            else s2[pid][k][oid]
                        )
                        for oid in officer_cols
                    ]
                    # cột "Tổng" của khối = cộng nhân viên (khớp sheet Số liệu chung)
                    cells = [(col, sum(vals), True)]
                    cells += [(col + 1 + t, v, False) for t, v in enumerate(vals)]
                    for cc, v, is_block_total in cells:
                        c = ws.cell(row, cc, v)
                        c.alignment = _CTR
                        c.border = _BORDER
                        if k in MONEY_KEYS:
                            c.number_format = _MONEY_FMT
                        if is_total:
                            c.fill = _TOTAL_FILL
                            c.font = _TOTAL_FONT
                        elif is_block_total:
                            c.font = Font(bold=True)
                    col += block_w

        emit(6, None, True)
        r = 7
        for n, pid in enumerate(ordered, 1):
            code, name, deg = desc_of(pid)
            ws.cell(r, 1, n).alignment = _CTR
            ws.cell(r, 2, code).alignment = _CTR
            ws.cell(r, 3, name).alignment = _LEFT
            ws.cell(r, 4, deg).alignment = _CTR
            ws.cell(r, 5, "Không").alignment = _CTR
            emit(r, pid, False)
            for i in range(1, ND + 1):
                ws.cell(r, i).border = _BORDER
            r += 1

        for col_letter, w in zip("ABCDE", (4.5, 11, 28, 9, 8)):
            ws.column_dimensions[col_letter].width = w
        # Khối tiền (Tổng học phí / Doanh thu) cần rộng hơn — cột "Tổng" của khối
        # cộng mọi nhân viên có thể tới hàng tỷ; width 8 sẽ hiện "########".
        cc = aStart
        for gname, gcols, gfill in blocks:
            for k, _label in gcols:
                w = 14 if k in MONEY_KEYS else 8
                for t in range(block_w):
                    ws.column_dimensions[get_column_letter(cc + t)].width = w
                cc += block_w
        ws.freeze_panes = ws.cell(7, aStart).coordinate

    # ----------------------------------------------------- sheet 3: quy ước
    def _sheet_notes(self, wb, year, n_unassigned, n_leads):
        ws = wb.create_sheet("Quy ước & ghi chú")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 82
        ws.cell(1, 1, "QUY ƯỚC ĐẾM & GHI CHÚ").font = _TITLE_FONT
        ws.merge_cells("A1:B1")
        rows = [
            (
                "MÔ HÌNH: PHỄU",
                "Mỗi lead xếp vào ĐÚNG 1 cột theo GIAI ĐOẠN CAO NHẤT: Học phí > "
                "Lệ phí > Hồ sơ > Đang tư vấn. 9 cột đếm của phễu (5 tư vấn + Hồ "
                "sơ + Lệ phí + Đóng một phần + Đóng đủ HK1) CỘNG = Tổng lead. Cột "
                "'Tổng học phí', 'Doanh thu' và nhóm 'Chi tiết hồ sơ' KHÔNG cộng.",
            ),
            (
                "Tổng lead",
                "Tổng số lead theo ngành = nguyện vọng 1 của hồ sơ; nếu chưa "
                "có hồ sơ thì theo ngành quan tâm (offering).",
            ),
            (
                "ĐANG TƯ VẤN — 5 cột",
                "Lead CHƯA có hồ sơ, theo consultation_status hiện tại. ⚠️ KHÁC "
                "bộ lọc lead: lead đã có hồ sơ (dù nháp) nằm ở nhóm Hồ sơ, còn "
                "bộ lọc vẫn đếm chúng theo trạng thái tư vấn.",
            ),
            ("• Chưa tiếp cận", "sts00; kèm lead chưa-hồ-sơ có trạng thái khác/rỗng."),
            ("• Đã kết nối", "sts02 (Đã kết nối liên hệ)."),
            ("• Có nhu cầu tìm hiểu", "sts03."),
            ("• Đồng ý tư vấn/Hẹn liên hệ", "sts06 (Đồng ý tư vấn) hoặc sts05 (Hẹn)."),
            (
                "• Từ chối tư vấn/Đã ngừng liên hệ",
                "sts04 (Từ chối) hoặc sts20 (Ngừng).",
            ),
            (
                "Hồ sơ (chưa đóng phí) — 1 cột",
                "Lead ĐÃ có hồ sơ nhưng CHƯA đóng phí nào (chưa lệ phí, chưa học "
                "phí). Chi tiết nháp/nợ/đủ xem nhóm 'Chi tiết hồ sơ' (thường chỉ "
                "còn hồ sơ nháp vì hồ sơ đã nộp phần lớn đã đóng phí).",
            ),
            (
                "Lệ phí",
                "Đã ĐÓNG/miễn lệ phí xét tuyển (application: paid_amount>0 hoặc "
                "miễn) NHƯNG chưa đóng học phí.",
            ),
            (
                "HỌC PHÍ — 4 cột",
                "Đã đóng học phí HK1 (fee tuition semester_no=1). 3 cột ĐẾM hồ sơ "
                "+ 1 cột TIỀN (Doanh thu).",
            ),
            ("• Đóng một phần", "SỐ hồ sơ HK1 = partial (đã đóng chưa đủ)."),
            ("• Đóng đủ HK1", "SỐ hồ sơ HK1 đã thu đủ hoặc được miễn."),
            (
                "• Tổng học phí",
                "SỐ hồ sơ đã đóng học phí = Đóng một phần + Đóng đủ HK1 (KHÔNG "
                "phải tiền; là tổng con, không cộng vào phễu).",
            ),
            (
                "• Doanh thu",
                "TIỀN học phí HK1 ĐÃ THU (Σ paid_amount). Đơn vị VND (miễn học "
                "phí không tạo doanh thu; không gồm lệ phí).",
            ),
            (
                "CHI TIẾT HỒ SƠ ĐÃ TẠO — 4 cột (ngoài phễu)",
                "KHÔNG thuộc phễu, KHÔNG cộng vào Tổng lead. Đếm MỌI hồ sơ đã tạo "
                "(mọi giai đoạn, kể cả đã đóng phí), phân Chưa hoàn thiện (nháp)/"
                "Hoàn thiện 1 phần (nộp+nợ giấy tờ)/Đủ điều kiện (nộp+đủ) + Tổng. "
                "= Hồ sơ(phễu) + Lệ phí + Học phí.",
            ),
            (
                "• Vì sao ≠ nhóm 'Hồ sơ'",
                "Nhóm 'Hồ sơ' (phễu) chỉ gồm hồ sơ CHƯA đóng phí; hồ sơ đã đóng "
                "phí đã tiến sang Lệ phí/Học phí. Tổng hồ sơ đã tạo = nhóm Hồ sơ "
                "+ Lệ phí + Học phí.",
            ),
        ]
        _H(ws.cell(3, 1, "Cột"))
        _H(ws.cell(3, 2, "Cách đếm"))
        r = 4
        for a, b in rows:
            ws.cell(r, 1, a).font = Font(bold=True)
            ws.cell(r, 1).alignment = _LEFT
            ws.cell(r, 2, b).alignment = _LEFT
            for i in (1, 2):
                ws.cell(r, i).border = _BORDER
            r += 1
        r += 1
        notes = [
            "GHI CHÚ:",
            "• PHỄU: mỗi lead đúng 1 cột (trong 9 cột đếm) → 5 cột Tư vấn + Hồ "
            "sơ chưa đóng phí + Lệ phí + Đóng một phần + Đóng đủ HK1 CỘNG = Tổng "
            "lead. 'Tổng học phí' (số đếm) và 'Doanh thu' (tiền) KHÔNG cộng.",
            "• Nhóm 'Chi tiết hồ sơ đã tạo' (xám) NGOÀI phễu — KHÔNG cộng vào "
            "Tổng lead; dùng để tra tổng số hồ sơ thật ở mọi giai đoạn.",
            "• Cột 'Đang tư vấn' KHÁC bộ lọc lead ở cột 'Đồng ý tư vấn/Hẹn': phễu "
            "chỉ tính lead CHƯA có hồ sơ; lead đã có hồ sơ nháp xếp sang nhóm Hồ "
            "sơ (4/5 cột tư vấn còn lại KHỚP bộ lọc).",
            "• Chỉ 'Doanh thu' là SỐ TIỀN (VND); tất cả cột còn lại là SỐ LƯỢNG "
            "(lead / hồ sơ) — kể cả 'Tổng học phí'.",
            "• Sheet 'Chia theo nhân viên': mỗi khối có cột 'Tổng' (= cộng các "
            "nhân viên, KHỚP sheet 'Số liệu chung') và cột 'Chưa PC' (lead chưa "
            "phân công nhân viên) nên sheet 2 đối chiếu khớp sheet 1.",
        ]
        if n_unassigned:
            notes.append(
                f"• Có {n_unassigned} lead chưa phân công nhân viên (cột 'Chưa PC')."
            )
        for t in notes:
            ws.cell(r, 1, t).alignment = _LEFT
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            if t.endswith(":"):
                ws.cell(r, 1).font = Font(bold=True, size=11)
            r += 1
