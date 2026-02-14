import csv
import io
from datetime import datetime
from typing import List, Literal, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
import structlog

from .. import database, models, schemas
from ..core.deps import CasbinAuth, LeadListFilter, get_lead_for_user, get_lead_list_filter, require_admin_or_manager  # ✅ Phase 2.2
from ..schemas.collaborator import LeadValidityUpdate
from ..services import distribution_service, insights_service, lead_service
from ..utils.csv_helpers import sanitize_csv_cell
from ..services.notification_dispatcher import dispatch  # ✅ NOTIFICATION 2.0
from ..core.events import SystemEvents  # ✅ NOTIFICATION 2.0
from ..core.constants import UserRole
from app.core.rate_limits import limiter, RateLimits

log = structlog.get_logger(__name__)

router = APIRouter(tags=["Leads"])

LeadAccessDep = Depends(get_lead_for_user)


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.post("", response_model=schemas.Lead, status_code=status.HTTP_201_CREATED)
async def create_new_lead(
    request: Request,
    lead_in: schemas.LeadCreate,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
):
    """
    Tạo một Lead mới.

    Role-based behavior:
    - Admin: Can set any unit_id, can assign to any officer or use auto-assignment
    - Manager: Can assign to officers in their unit or use auto-assignment
    - Officer: Auto-assigned to themselves, unit forced to their unit
    """
    result, callback = await lead_service.create_lead(db, lead_in, created_by=current_user)
    await db.commit()
    await callback()

    return result


@limiter.limit(RateLimits.DATA_READ)  # 1000/hour
@router.get("/my/reassign-quota")
async def get_my_reassign_quota(
    request: Request,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
):
    """
    Get current user's remaining reassign quota.
    
    Returns:
        - allowed: True if user can still reassign
        - used: Number of reassigns used this week
        - limit: Maximum reassigns allowed per week
        - remaining: Reassigns left this week
    """
    quota = await lead_service.check_reassign_quota(db, current_user.id)
    return quota


@limiter.limit(RateLimits.DATA_READ)  # 1000/hour
@router.get("/distribution-preview")
async def get_distribution_preview(
    request: Request,
    offering_id: int = Query(..., description="Offering ID to preview distribution"),
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
):
    """
    Preview which unit will receive the next lead based on distribution config.

    This does NOT increment the distribution cursor - it's read-only.

    Returns:
    - offering_id: The offering being queried
    - has_config: Whether distribution config exists for this offering
    - next_unit_id: Unit that will receive the next lead
    - next_unit_name: Name of the unit
    - configs: List of active distribution configs with weights
    - total_slots: Total weighted slots in distribution cycle
    """
    stats = await distribution_service.get_distribution_stats(db, offering_id)

    # ✅ PHASE 8: next_unit_name now comes from service (no db.get() in router)
    return {
        "offering_id": offering_id,
        "has_config": len(stats.get("configs", [])) > 0,
        "next_unit_id": stats.get("next_unit_id"),
        "next_unit_name": stats.get("next_unit_name"),  # ✅ From service
        "configs": stats.get("configs", []),
        "total_slots": stats.get("total_slots", 0),
        "cursor_position": stats.get("cursor_value"),
    }


@limiter.limit(RateLimits.DATA_READ)  # 1000/hour
@router.get("/check-duplicate")
async def check_duplicate(
    request: Request,
    phone: Optional[str] = Query(None, description="Phone number to check"),
    email: Optional[str] = Query(None, description="Email to check"),
    unit_id: Optional[int] = Query(None, description="Unit ID for email check scope"),
    exclude_id: Optional[int] = Query(None, description="Lead ID to exclude (for edit mode)"),
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
):
    """
    Real-time duplicate check for phone/email.

    - Phone: Checked globally (must be unique across ALL units)
    - Email: Checked within same unit only

    Returns:
        - phone_available: True if phone is available (or not provided)
        - email_available: True if email is available (or not provided)
        - phone_conflict: Info about conflicting lead (if phone taken)
        - email_conflict: Info about conflicting lead (if email taken)
    """
    from ..repositories import LeadRepository

    repo = LeadRepository(db)
    result = {
        "phone_available": True,
        "email_available": True,
        "phone_conflict": None,
        "email_conflict": None,
    }

    # Check phone (global scope)
    if phone:
        # Normalize phone using the same function as Pydantic schema validators
        from ..utils.phone_helpers import normalize_vietnam_phone
        normalized_phone = normalize_vietnam_phone(phone) or phone.replace(" ", "").replace("-", "").replace(".", "")
        existing = await repo.check_phone_conflict(
            phone=normalized_phone,
            phone2=None,
            exclude_id=exclude_id
        )
        if existing:
            officer_name = "Chưa phân công"
            if existing.assigned_officer:
                officer_name = existing.assigned_officer.full_name or "N/A"

            # Get unit name
            unit_name = "N/A"
            if existing.unit_id:
                await db.refresh(existing, ["unit"])
                if existing.unit:
                    unit_name = existing.unit.name

            result["phone_available"] = False
            result["phone_conflict"] = (
                f"SĐT đã tồn tại: {existing.full_name} ({existing.phone}) - "
                f"Đơn vị: {unit_name} - Quản lý: {officer_name}"
            )

    # Check email (unit scope)
    if email and unit_id:
        existing = await repo.check_email_conflict(
            email=email,
            unit_id=unit_id,
            exclude_id=exclude_id
        )
        if existing:
            officer_name = "Chưa phân công"
            if existing.assigned_officer:
                officer_name = existing.assigned_officer.full_name or "N/A"

            result["email_available"] = False
            result["email_conflict"] = (
                f"Email đã tồn tại trong đơn vị: {existing.full_name} ({existing.phone}) - "
                f"Quản lý: {officer_name}"
            )

    return result


@limiter.limit(RateLimits.DATA_READ)  # 1000/hour
@router.get("", response_model=schemas.LeadsPage)
async def get_all_leads(
    request: Request,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    # === ⭐️ THÊM CÁC THAM SỐ QUERY ===
    status: Optional[str] = Query(
        None, description="Filter by status (comma-separated)"
    ),
    assigned_officer_id: Optional[str] = Query(
        None, description="Filter by assigned officer ID(s) (comma-separated, e.g. '1,2,3'). Note: Officers can only see their own leads."
    ),
    # ✅ Phase 2: Inject LeadListFilter for role-based filtering
    lead_filter: LeadListFilter = Depends(get_lead_list_filter),
    offering_id: Optional[str] = Query(
        None, description="Filter by program offering ID(s) (comma-separated, e.g. '1,2,3')"
    ),
    source: Optional[str] = Query(
        None, description="Filter by source (comma-separated)"
    ),
    search: Optional[str] = Query(
        None, description="Search term for name, email, phone"
    ),
    sort_by: str = Query("created_at", description="Field to sort by"),
    order: str = Query("desc", description="Sort order (asc or desc)"),
    # === PIPELINE STAGE FILTER ===
    pipeline_stage_id: Optional[str] = Query(
        None, description="Filter by pipeline stage ID(s) (comma-separated, e.g. 'stg01,stg02')"
    ),
    # === DATE RANGE FILTER ===
    date_from: Optional[datetime] = Query(
        None, description="Filter leads created/updated from this date (ISO format)"
    ),
    date_to: Optional[datetime] = Query(
        None, description="Filter leads created/updated until this date (ISO format)"
    ),
    date_field: str = Query(
        "created_at", description="Date field to filter on (created_at or updated_at)"
    ),
    # === KẾT THÚC THÊM THAM SỐ ===
):
    """
    Lấy danh sách Leads (có phân trang, filter, search, sort).

    **Role-based filtering:**
    - Admin/Manager: Xem tất cả leads
    - Officer: Chỉ xem leads được gán cho mình
    """
    skip = (page - 1) * page_size

    # ✅ Phase 2: Use role-enforced filter from dependency
    # Officers are forced to see only their leads, managers see only their unit
    effective_officer_id = lead_filter.assigned_officer_id
    effective_unit_id = lead_filter.unit_id

    total, leads = await lead_service.get_leads(
        db,
        skip=skip,
        limit=page_size,
        # === ⭐️ TRUYỀN THAM SỐ VÀO SERVICE ===
        status=status,
        assigned_officer_id=effective_officer_id,  # Now role-enforced via dependency
        unit_id=effective_unit_id,  # ✅ Phase 2: Also role-enforced
        offering_id=offering_id,  # Now a string (comma-separated for multi)
        source=source,
        search=search,
        sort_by=sort_by,
        order=order,
        # === PIPELINE STAGE FILTER ===
        pipeline_stage_id=pipeline_stage_id,  # Already string (comma-separated for multi)
        # === DATE RANGE FILTER ===
        date_from=date_from,
        date_to=date_to,
        date_field=date_field,
        # === KẾT THÚC TRUYỀN THAM SỐ ===
    )
    return {"total_count": total, "leads": leads}


# ==============================================================================
# EXPORT (must be defined BEFORE /{lead_id} to avoid route conflict)
# ==============================================================================

@limiter.limit(RateLimits.DATA_EXPORT)  # 20/hour - Export operation
@router.get("/export")
async def export_leads(
    request: Request,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
    format: Literal["csv", "xlsx"] = Query("csv", description="Export format (csv or xlsx)"),
    # ✅ RBAC enforcement via LeadListFilter dependency
    lead_filter: LeadListFilter = Depends(get_lead_list_filter),
    # Apply same filters as get_all_leads (aligned param types)
    status: Optional[str] = Query(
        None, description="Filter by status (comma-separated)"
    ),
    assigned_officer_id: Optional[str] = Query(
        None, description="Filter by assigned officer ID(s) (comma-separated)"
    ),
    offering_id: Optional[str] = Query(
        None, description="Filter by program offering ID(s) (comma-separated)"
    ),
    source: Optional[str] = Query(
        None, description="Filter by source (comma-separated)"
    ),
    search: Optional[str] = Query(
        None, description="Search term for name, email, phone"
    ),
    sort_by: str = Query("created_at", description="Field to sort by"),
    order: str = Query("desc", description="Sort order (asc or desc)"),
    pipeline_stage_id: Optional[str] = Query(
        None, description="Filter by pipeline stage ID(s) (comma-separated)"
    ),
    date_from: Optional[datetime] = Query(
        None, description="Filter leads from this date (ISO format)"
    ),
    date_to: Optional[datetime] = Query(
        None, description="Filter leads until this date (ISO format)"
    ),
    date_field: str = Query(
        "created_at", description="Date field to filter on (created_at or updated_at)"
    ),
):
    """
    Export leads to CSV or Excel file.

    Apply same filters as the list endpoint to allow exporting filtered results.
    Maximum 10,000 leads per export to prevent performance issues.

    **RBAC:** Officers only export their own leads, Managers only their unit.
    """
    # ✅ Use role-enforced filter from dependency (same as list endpoint)
    effective_officer_id = lead_filter.assigned_officer_id
    effective_unit_id = lead_filter.unit_id

    # Get filtered leads (no pagination, but limit to 10,000)
    total, leads = await lead_service.get_leads(
        db,
        skip=0,
        limit=10000,  # Export limit
        status=status,
        assigned_officer_id=effective_officer_id,
        unit_id=effective_unit_id,
        offering_id=offering_id,
        source=source,
        search=search,
        sort_by=sort_by,
        order=order,
        pipeline_stage_id=pipeline_stage_id,
        date_from=date_from,
        date_to=date_to,
        date_field=date_field,
    )

    # Helper: extract display names from relationships (already eager-loaded)
    def _lead_row(lead):
        # Split offering into 3 columns: Trình độ ĐT, Ngành, Hình thức
        degree_level = ""
        program_name = ""
        offering_type = ""
        if lead.offering and lead.offering.program:
            degree_level = lead.offering.program.degree_level or ""
            full_name = lead.offering.program.name or ""
            # Strip degree_level prefix from name to get pure program name
            # e.g. "Cao đẳng Công nghệ thông tin" → "Công nghệ thông tin"
            if degree_level and full_name.startswith(degree_level):
                program_name = full_name[len(degree_level):].strip()
            else:
                program_name = full_name
        if lead.offering:
            offering_type = lead.offering.offering_type or ""

        # ✅ SECURITY: Sanitize all user-data strings to prevent CSV/Excel formula injection
        return {
            "ID": lead.id,
            "Họ tên": sanitize_csv_cell(lead.full_name),
            "Email": sanitize_csv_cell(lead.email),
            "SĐT": sanitize_csv_cell(lead.phone),
            "SĐT 2": sanitize_csv_cell(lead.phone2),
            "Trạng thái": lead.status,
            "Điểm lead": lead.lead_score,
            "Nguồn": sanitize_csv_cell(lead.source),
            "Trình độ học vấn": sanitize_csv_cell(lead.education_level),
            "GPA": lead.gpa or "",
            "Địa chỉ": sanitize_csv_cell(lead.location),
            "Trình độ ĐT": degree_level,
            "Ngành": program_name,
            "Hình thức": offering_type,
            "Cán bộ phụ trách": sanitize_csv_cell(lead.assigned_officer.full_name if lead.assigned_officer else ""),
            "Giai đoạn pipeline": lead.pipeline_stage.name if lead.pipeline_stage else "",
            "Trạng thái tư vấn": lead.consultation_status.name if lead.consultation_status else "",
            "Đơn vị": lead.unit.name if lead.unit else "",
            "Ngày tạo": lead.created_at.isoformat() if lead.created_at else "",
            "Ngày cập nhật": lead.updated_at.isoformat() if lead.updated_at else "",
        }

    fieldnames = [
        "ID", "Họ tên", "Email", "SĐT", "SĐT 2",
        "Trạng thái", "Điểm lead", "Nguồn", "Trình độ học vấn", "GPA", "Địa chỉ",
        "Trình độ ĐT", "Ngành", "Hình thức",
        "Cán bộ phụ trách", "Giai đoạn pipeline",
        "Trạng thái tư vấn", "Đơn vị", "Ngày tạo", "Ngày cập nhật",
    ]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if format == "csv":
        # Generate CSV with UTF-8 BOM for Vietnamese encoding in Excel
        output = io.StringIO()
        output.write("\ufeff")  # UTF-8 BOM
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for lead in leads:
            writer.writerow(_lead_row(lead))

        output.seek(0)
        filename = f"leads_export_{timestamp}.csv"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    else:  # xlsx — validated by Literal type
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Export"

        # Header row with styling
        ws.append(fieldnames)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for lead in leads:
            row = _lead_row(lead)
            ws.append([row[col] for col in fieldnames])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError):
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"leads_export_{timestamp}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )


@limiter.limit(RateLimits.DATA_READ)  # 1000/hour
@router.get("/{lead_id}", response_model=schemas.Lead)
async def get_lead_details(
    request: Request,
    lead: models.Lead = LeadAccessDep,
):
    """Lấy thông tin chi tiết của một Lead."""
    return lead


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.put("/{lead_id}", response_model=schemas.Lead)
async def update_existing_lead(
    request: Request,
    lead_in: schemas.LeadUpdate,
    lead: models.Lead = LeadAccessDep,
    # Lấy current_user từ Casbin check hoặc get_current_user
    current_user: models.User = CasbinAuth,  # <<< LẤY USER TỪ DEPENDENCY
    db: AsyncSession = Depends(database.get_db),
):
    """Cập nhật một Lead (chỉ Admin/Manager)."""
    # <<< SỬA Ở ĐÂY: Truyền current_user vào service >>>

    # Track which fields are being updated
    update_data = lead_in.model_dump(exclude_unset=True)
    updated_fields = list(update_data.keys())
    status_changed = "consultation_status_id" in updated_fields or "pipeline_stage_id" in updated_fields
    unit_changed = "unit_id" in updated_fields and update_data.get("unit_id") != lead.unit_id

    # Store old values before update
    old_unit_id = lead.unit_id
    old_officer_id = lead.assigned_officer_id

    result = await lead_service.update_lead(db, lead.id, lead_in, updated_by=current_user)
    await db.commit()

    # ✅ NOTIFICATION 2.0: Dispatch LEAD_REASSIGNED if unit changed
    if unit_changed:
        try:
            _, notif_cb = await dispatch(
                db=db,
                event=SystemEvents.LEAD_REASSIGNED,
                payload={
                    "lead_id": result.id,
                    "old_officer_id": old_officer_id,
                    "new_officer_id": result.assigned_officer_id,
                    "old_unit_id": old_unit_id,
                    "new_unit_id": result.unit_id,
                    "actor_id": current_user.id,
                    "actor_name": current_user.full_name or current_user.username,
                    "reason": "Unit transfer by admin",
                },
                dedupe_key=f"lead_reassigned:{result.id}:{result.unit_id}",
            )
            await db.commit()
            if notif_cb:
                await notif_cb()
        except Exception as e:
            log.warning("Failed to dispatch LEAD_REASSIGNED notification", error=str(e))

    # ✅ NOTIFICATION 2.0: Dispatch notification if status changed
    if status_changed:
        try:
            _, notif_cb = await dispatch(
                db=db,
                event=SystemEvents.LEAD_STATUS_CHANGED,
                payload={
                    "lead_id": result.id,
                    "lead_name": result.full_name or result.email or f"Lead #{result.id}",
                    "officer_id": result.assigned_officer_id,
                    "officer_name": result.assigned_officer.full_name if result.assigned_officer else "Unknown",
                    "old_status": lead.consultation_status_id or "none",
                    "new_status": result.consultation_status_id or "none",
                    "old_stage": lead.pipeline_stage_id,
                    "new_stage": result.pipeline_stage_id,
                    "actor_id": current_user.id,
                    "actor_name": current_user.full_name or current_user.username,
                    "updated_fields": updated_fields,
                },
                dedupe_key=f"lead_status_changed:{result.id}:{result.consultation_status_id}",
            )
            await db.commit()
            if notif_cb:
                await notif_cb()
        except Exception as e:
            log.warning(
                "Failed to dispatch lead status changed notification",
                lead_id=lead.id,
                error=str(e)
            )

    # ✅ REAL-TIME SYNC: Always dispatch LEAD_UPDATED for UI refresh
    try:
        _, notif_cb = await dispatch(
            db=db,
            event=SystemEvents.LEAD_UPDATED,
            payload={
                "lead_id": result.id,
                "updated_fields": updated_fields,
                "status_changed": status_changed,
                "actor_id": current_user.id,
                "updated_by": current_user.full_name or current_user.username,
                "actor_name": current_user.full_name or current_user.username,
                "updated_summary": ", ".join(updated_fields[:3]) + ("..." if len(updated_fields) > 3 else ""),
                "updated_at": datetime.now().isoformat(),
                "message": f"Lead updated by {current_user.full_name or current_user.username}",
            },
            dedupe_key=f"lead_updated:{result.id}:{int(datetime.now().timestamp())}",
        )
        await db.commit()
        if notif_cb:
            await notif_cb()
    except Exception as e:
        log.warning("Failed to dispatch LEAD_UPDATED notification", lead_id=result.id, error=str(e))

    return result


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.delete("/{lead_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_lead(
    request: Request,
    lead_id: int,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
):
    """
    (Admin only) Soft delete a Lead.

    Sets deleted_at timestamp instead of physically deleting the lead.
    Preserves all historical data (consultations, applications, logs).
    Deleted leads are filtered out from normal queries.

    **Permission:** Admin only (enforced by Casbin)

    **Business Rules:**
    - Cannot delete leads that have an admission profile
    - Must cancel/delete admission profile first before deleting lead

    **Status Code:** 204 No Content on success

    **Raises:**
    - 400 Bad Request: If lead has an admission profile (cannot be deleted)
    - 404 Not Found: If lead doesn't exist or already deleted
    - 403 Forbidden: If user doesn't have admin permission
    """
    deleted_lead = await lead_service.delete_lead(db, lead_id, deleted_by=current_user)
    await db.commit()

    # Dispatch notification for lead deletion
    try:
        _, notif_cb = await dispatch(
            db=db,
            event=SystemEvents.LEAD_DELETED,
            payload={
                "lead_id": deleted_lead.id,
                "lead_name": deleted_lead.full_name or "Unknown",
                "unit_id": deleted_lead.unit_id,
                "officer_id": deleted_lead.assigned_officer_id,
                "actor_id": current_user.id,
                "actor_name": current_user.full_name or current_user.username,
            },
            dedupe_key=f"lead_deleted:{deleted_lead.id}",
        )
        await db.commit()
        if notif_cb:
            await notif_cb()
    except Exception as e:
        log.warning("Failed to dispatch lead deletion notification", error=str(e))

    return None


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.post("/{lead_id}/restore", response_model=schemas.Lead)
async def restore_lead(
    request: Request,
    lead_id: int,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
):
    """
    (Admin only) Restore a soft-deleted Lead and its related consultations.

    Clears deleted_at timestamp to restore the lead and all consultations
    that were deleted at the same time.

    **Permission:** Admin only (enforced by Casbin)

    **Business Rules:**
    - Only restores leads that were previously soft-deleted
    - Automatically restores related consultations deleted at the same time
    - Restored lead status is reset to 'new'

    **Raises:**
    - 400 Bad Request: If lead is not deleted (cannot restore active lead)
    - 404 Not Found: If lead doesn't exist
    - 403 Forbidden: If user doesn't have admin permission
    """
    restored_lead = await lead_service.restore_lead(db, lead_id, restored_by=current_user)
    await db.commit()

    log.info(
        "Lead restored via API",
        lead_id=lead_id,
        restored_by_user_id=current_user.id
    )

    return restored_lead


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.post(
    "/{lead_id}/consultations",
    response_model=schemas.Consultation,
    status_code=status.HTTP_201_CREATED,
)
async def add_new_consultation(
    request: Request,
    consultation_in: schemas.ConsultationCreate,
    lead: models.Lead = LeadAccessDep,  # <-- THAY ĐỔI (IDOR Check)
    current_user: models.User = CasbinAuth,  # <-- THAY ĐỔI (Casbin Check)
    db: AsyncSession = Depends(database.get_db),
):
    """Thêm một ghi chú tư vấn mới cho Lead (Đã xác thực 2 lớp)."""
    # Service 'add_consultation' có logic check quyền sở hữu
    # nhưng check ở đây vẫn an toàn hơn
    result = await lead_service.add_consultation(
        db, lead.id, current_user.id, consultation_in
    )
    await db.commit()

    # Dispatch notification for consultation created
    try:
        _, notif_cb = await dispatch(
            db=db,
            event=SystemEvents.CONSULTATION_CREATED,
            payload={
                "consultation_id": result.id,
                "lead_id": lead.id,
                "officer_id": lead.assigned_officer_id,
                "status_id": result.consultation_status_id or "",
                "actor_id": current_user.id,
                "actor_name": current_user.full_name or current_user.username,
                "unit_id": lead.unit_id,
            },
        )
        await db.commit()
        if notif_cb:
            await notif_cb()
    except Exception as e:
        log.warning(
            "Failed to dispatch consultation created notification",
            lead_id=lead.id,
            consultation_id=result.id,
            error=str(e)
        )

    return result


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.post("/{lead_id}/assign", response_model=schemas.Lead)
async def assign_lead_manually(
    request: Request,
    assign_data: schemas.AssignLead,
    lead: models.Lead = LeadAccessDep,  # <-- THAY ĐỔI (IDOR Check)
    current_user: models.User = CasbinAuth,  # <-- THAY ĐỔI (Casbin Check)
    db: AsyncSession = Depends(database.get_db),
):
    """(Admin/Manager only) Gán thủ công một Lead (Đã xác thực 2 lớp)."""
    result = await lead_service.assign_lead_manually(
        db, lead.id, assign_data.officer_id, current_user
    )
    await db.commit()

    # ❌ REDUNDANT: lead_service.assign_lead_manually already dispatches LEAD_ASSIGNED with richer payload
    # Removed to prevent duplicate notifications.

    return result


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.post("/{lead_id}/action", response_model=schemas.Lead)
async def perform_lead_action(
    request: Request,
    action_data: schemas.LeadAction,
    lead: models.Lead = LeadAccessDep,  # <-- THAY ĐỔI (IDOR Check)
    current_user: models.User = CasbinAuth,  # <-- THAY ĐỔI (Casbin Check)
    db: AsyncSession = Depends(database.get_db),
):
    """Xử lý hành động (reject/reassign) của Officer (Đã xác thực 2 lớp)."""
    result, callback = await lead_service.process_officer_action(
        db, lead.id, current_user, action_data.action, action_data.reason
    )
    await db.commit()
    # ✅ FIX: Call callback AFTER commit to ensure Celery sees committed data
    await callback()
    return result


@limiter.limit(RateLimits.DATA_READ)  # 1000/hour
@router.get("/{lead_id}/timeline", response_model=List[schemas.TimelineItem])
async def get_lead_timeline(
    request: Request,
    lead: models.Lead = LeadAccessDep,  # <-- THAY ĐỔI (IDOR Check)
    db: AsyncSession = Depends(database.get_db),
):
    """Lấy lịch sử tổng hợp (timeline) của một Lead (Đã xác thực quyền)."""
    return await lead_service.get_lead_timeline(db, lead.id)


@limiter.limit(RateLimits.DATA_READ)  # 1000/hour
@router.get("/{lead_id}/insights", response_model=schemas.LeadInsights)
async def get_lead_insights(
    request: Request,
    lead: models.Lead = LeadAccessDep,  # <-- THAY ĐỔI (IDOR Check)
    db: AsyncSession = Depends(database.get_db),
):
    """Lấy các chỉ số insight 360 độ của một Lead (Đã xác thực quyền)."""
    timeline = await lead_service.get_lead_timeline(db, lead.id)
    return await insights_service.get_lead_insights(db, lead, timeline)


@limiter.limit(RateLimits.DATA_READ)  # 1000/hour
@router.get("/{lead_id}/workflow-context", response_model=schemas.WorkflowContext)
async def get_workflow_context(
    request: Request,
    lead: models.Lead = LeadAccessDep,
    current_user: models.User = CasbinAuth,
    db: AsyncSession = Depends(database.get_db),
):
    """
    Get workflow context for a lead - Phase-Based Workflow.
    
    Provides frontend with:
    - Current phase (consultation, admission, fee, enrolled)
    - Allowed statuses user can select
    - Phase constraints and locked states
    
    Used to dynamically filter status dropdowns based on lead's current phase.
    """
    from app.services.phase_manager import (
        derive_phase_from_admission,
        get_allowed_statuses_for_phase,
        is_terminal_phase,
        UNIVERSAL_STATUSES,
    )
    from app.repositories import AdmissionRepository, PipelineRepository
    
    # Get admission profile to derive phase
    admission_repo = AdmissionRepository(db)
    pipeline_repo = PipelineRepository(db)
    
    admission_profile = await admission_repo.get_profile_by_lead_id(lead.id)
    current_phase = derive_phase_from_admission(admission_profile)
    
    # Get allowed status IDs for this phase and user role
    user_role = current_user.role if isinstance(current_user.role, str) else current_user.role.value
    allowed_status_ids = get_allowed_statuses_for_phase(
        current_phase, 
        user_role
    )
    
    # Get full status objects for allowed statuses
    all_statuses = await pipeline_repo.get_all_statuses_with_stage()
    
    allowed_statuses = []
    for status in all_statuses:
        if status.id in allowed_status_ids:
            allowed_statuses.append(schemas.WorkflowAllowedStatus(
                id=status.id,
                name=status.name,
                phase=status.phase if hasattr(status, 'phase') else "consultation",
                color_code=status.color_code,
                outcome_type=status.outcome_type.value if hasattr(status.outcome_type, 'value') else str(status.outcome_type),
                is_universal=status.is_universal,
            ))
    
    return schemas.WorkflowContext(
        lead_id=lead.id,
        current_phase=current_phase.value,
        current_status_id=lead.consultation_status_id,
        current_stage_id=lead.pipeline_stage_id,
        allowed_statuses=allowed_statuses,
        is_terminal_phase=is_terminal_phase(current_phase),
        can_change_status=not is_terminal_phase(current_phase),
        has_admission_profile=admission_profile is not None,
        admission_status=admission_profile.status if admission_profile else None,
    )

@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.put(
    "/{lead_id}/consultations/{consultation_id}", response_model=schemas.Consultation
)
async def update_a_consultation(
    request: Request,
    consultation_id: int,
    consultation_in: schemas.ConsultationUpdate,
    lead: models.Lead = LeadAccessDep,  # <-- IDOR Check
    current_user: models.User = CasbinAuth,  # <-- Casbin Check
    db: AsyncSession = Depends(database.get_db),
):
    """
    Cập nhật một ghi chú tư vấn (Admin: any consultation, Officer: most recent only).

    Permission Rules:
    - Admin: Can update any consultation
    - Officer: Can only update the most recent consultation to maintain consultation chain integrity
    - Other roles: Cannot update consultations

    Business Logic:
    - If status_id is changed and this is the most recent consultation, lead status will be updated
    - Changes are logged in LeadStatusHistory
    - Real-time Socket.IO event is emitted to all connected clients

    This prevents Officers from breaking the consultation chain by editing historical consultations.
    """
    # ✅ PHASE 8: Use LeadRepository instead of db.get()
    from ..repositories import LeadRepository
    repo = LeadRepository(db)
    consultation = await repo.get_consultation_with_status_stage(consultation_id)
    old_status_id = consultation.consultation_status_id if consultation else None

    result = await lead_service.update_consultation(
        db, lead.id, consultation_id, consultation_in, current_user
    )

    # ✅ TRANSACTION PATTERN: Router commits the transaction
    await db.commit()

    # ✅ NOTIFICATION 2.0: Dispatch CONSULTATION_UPDATED if status changed
    if old_status_id and result.consultation_status_id != old_status_id:
        try:
            _, notif_cb = await dispatch(
                db=db,
                event=SystemEvents.CONSULTATION_UPDATED,
                payload={
                    "consultation_id": result.id,
                    "lead_id": lead.id,
                    "officer_id": lead.assigned_officer_id,
                    "old_status_id": old_status_id,
                    "new_status_id": result.consultation_status_id,
                    "actor_id": current_user.id,
                    "actor_name": current_user.full_name or current_user.username,
                },
                dedupe_key=f"consultation_updated:{result.id}:{result.consultation_status_id}",
            )
            await db.commit()
            if notif_cb:
                await notif_cb()
        except Exception as e:
            log.warning("Failed to dispatch CONSULTATION_UPDATED notification", error=str(e))

    return result


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.delete(
    "/{lead_id}/consultations/{consultation_id}", status_code=status.HTTP_204_NO_CONTENT
)
async def delete_a_consultation(
    request: Request,
    consultation_id: int,
    lead: models.Lead = LeadAccessDep,  # <-- THAY ĐỔI (IDOR Check)
    current_user: models.User = CasbinAuth,  # <-- THAY ĐỔI (Casbin Check)
    db: AsyncSession = Depends(database.get_db),
):
    """
    Xóa một ghi chú tư vấn (Admin: any consultation, Officer: most recent only).

    Permission Rules:
    - Admin: Can delete any consultation
    - Officer: Can only delete the most recent consultation to maintain consultation chain integrity
    - Other roles: Cannot delete consultations

    This prevents Officers from breaking the consultation chain by deleting historical consultations.
    """
    await lead_service.delete_consultation(db, lead.id, consultation_id, current_user)

    # ✅ TRANSACTION PATTERN: Router commits the transaction
    await db.commit()

    # ✅ NOTIFICATION 2.0: Dispatch CONSULTATION_DELETED
    try:
        _, notif_cb = await dispatch(
            db=db,
            event=SystemEvents.CONSULTATION_DELETED,
            payload={
                "consultation_id": consultation_id,
                "lead_id": lead.id,
                "officer_id": lead.assigned_officer_id,
                "actor_id": current_user.id,
                "actor_name": current_user.full_name or current_user.username,
            },
            dedupe_key=f"consultation_deleted:{consultation_id}",
        )
        await db.commit()
        if notif_cb:
            await notif_cb()
    except Exception as e:
        log.warning("Failed to dispatch CONSULTATION_DELETED notification", error=str(e))

    return None


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.post(
    "/{lead_id}/consultations/{consultation_id}/restore",
    response_model=schemas.Consultation,
)
async def restore_a_consultation(
    request: Request,
    consultation_id: int,
    lead: models.Lead = LeadAccessDep,
    current_user: models.User = CasbinAuth,
    db: AsyncSession = Depends(database.get_db),
):
    """
    Restore a soft-deleted consultation.

    Permission Rules:
    - Admin/Manager: Can restore any consultation
    - Officer: Can only restore if assigned to the lead

    Business Logic:
    - Restores the consultation by clearing deleted_at
    - If restored consultation is the most recent, updates lead status accordingly
    """
    await lead_service.restore_consultation(
        db, lead.id, consultation_id, current_user
    )

    # Commit the transaction
    await db.commit()

    # Re-fetch with eager loading to avoid lazy load issues
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    result = await db.execute(
        select(models.Consultation)
        .where(models.Consultation.id == consultation_id)
        .options(
            selectinload(models.Consultation.consultation_status)
            .selectinload(models.ConsultationStatus.stage),
            selectinload(models.Consultation.lead),
        )
    )
    consultation = result.scalar_one()

    log.info(
        "Consultation restored via API",
        consultation_id=consultation_id,
        lead_id=lead.id,
        restored_by=current_user.id,
    )

    return consultation


@limiter.limit(RateLimits.DATA_READ)  # 1000/hour
@router.get("/import/template")
async def download_import_template(
    request: Request,
    format: str = Query("csv", description="Template format (csv or xlsx)"),
    current_user: models.User = CasbinAuth,
):
    """
    Download CSV/Excel template for lead import.

    Returns a pre-formatted template file with:
    - Correct column headers (required + optional)
    - Example data row with Vietnamese sample
    - Column descriptions as comments (Excel only)

    **Query Parameters:**
    - `format`: "csv" or "xlsx" (default: csv)

    **Template Columns:**
    - Required: full_name, email, phone, source, unit_id
    - Optional: offering_id, education_level, gpa, location

    **Usage:**
    1. Download template
    2. Fill in lead data
    3. Upload via POST /api/admin/users/leads/import
    """
    # Define template data
    headers = ["full_name", "email", "phone", "source", "unit_id", "offering_id", "education_level", "gpa", "location"]
    example_row = [
        "Nguyễn Văn An",
        "nguyenvanan@gmail.com",
        "0901234567",
        "website",
        "1",  # unit_id - Replace with your actual unit ID
        "",   # offering_id - Optional
        "bachelor",  # high_school, bachelor, master, phd
        "3.5",  # 0.0-4.0
        "Hà Nội"
    ]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if format.lower() == "csv":
        # Generate CSV template
        output = io.StringIO()
        writer = csv.writer(output)

        # Add header comment
        output.write("# Lead Import Template\n")
        output.write("# Required columns: full_name, email, phone, source, unit_id\n")
        output.write("# Optional columns: offering_id, education_level, gpa, location\n")
        output.write("# Education levels: high_school, bachelor, master, phd\n")
        output.write("# Sources: website, referral, social_media, walk_in, email, phone, event, other\n")
        output.write("#\n")

        # Write headers and example
        writer.writerow(headers)
        writer.writerow(example_row)

        output.seek(0)
        filename = f"lead_import_template_{timestamp}.csv"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    elif format.lower() in ["xlsx", "excel"]:
        # Generate Excel template with styling
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.comments import Comment

        wb = Workbook()
        ws = wb.active
        ws.title = "Lead Import Template"

        # Header row with styling
        ws.append(headers)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        column_descriptions = [
            "Full name (required)",
            "Email address (required, unique per unit)",
            "Phone number (required)",
            "Lead source (required): website, referral, social_media, walk_in, email, phone, event, other",
            "Organization Unit ID (required): Get from /api/organization-units",
            "Program Offering ID (optional): Get from /api/offerings",
            "Education level (optional): high_school, bachelor, master, phd",
            "GPA (optional): 0.0-4.0 scale",
            "Location (optional): City/Province"
        ]

        for idx, cell in enumerate(ws[1], start=0):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            # Add comment with description
            cell.comment = Comment(column_descriptions[idx], "System")

        # Example row
        ws.append(example_row)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError):
                    pass
            adjusted_width = max(max_length + 2, 15)  # Min 15 chars
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"lead_import_template_{timestamp}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    else:
        return {"error": f"Unsupported format '{format}'. Supported: csv, xlsx"}


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.post("/bulk-assign", status_code=status.HTTP_200_OK)
async def bulk_assign_leads(
    request: Request,
    bulk_assign_data: schemas.BulkAssignLeadsSchema,
    officer_id: int = Query(..., description="Officer ID to assign leads to"),
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
):
    """
    (Admin/Manager only) Bulk assign multiple leads to a single officer.

    Assigns multiple leads to one officer in a single operation.
    Continues on errors (doesn't fail fast) and returns detailed results.

    **Request Body:**
    ```json
    {
        "lead_ids": [1, 2, 3, 4, 5]
    }
    ```

    **Query Parameters:**
    - `officer_id`: Target officer ID to assign all leads to

    **Response:**
    ```json
    {
        "total": 5,
        "successful": 4,
        "failed": 1,
        "assigned_lead_ids": [1, 2, 3, 4],
        "errors": [
            {"lead_id": 5, "error": "Lead with id 5 not found"}
        ]
    }
    ```

    **Permission:** Admin or Manager only (enforced by Casbin)

    **Business Rules:**
    - Officer must exist, be active, and have role="officer"
    - Creates AssignmentLog for each successful assignment
    - Logs state change in LeadStatusHistory
    - Updates lead status to "assigned"
    - Updates officer's last_assigned_at timestamp
    """
    result = await lead_service.bulk_assign_leads(
        db,
        lead_ids=bulk_assign_data.lead_ids,
        officer_id=officer_id,
        assigner=current_user
    )
    await db.commit()
    return result


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.post("/bulk-update-stage", status_code=status.HTTP_200_OK)
async def bulk_update_leads_stage(
    request: Request,
    bulk_data: schemas.BulkUpdateStageSchema,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
):
    """
    (Admin only) Bulk update pipeline stage for multiple leads.
    
    ✅ PHASE 8: Uses lead_service.bulk_update_pipeline_stage (Repository pattern).

    **Request Body:**
    ```json
    {"lead_ids": [1, 2, 3], "pipeline_stage_id": "stage_2"}
    ```
    """
    result = await lead_service.bulk_update_pipeline_stage(
        db,
        lead_ids=bulk_data.lead_ids,
        pipeline_stage_id=bulk_data.pipeline_stage_id,
        updated_by=current_user
    )
    await db.commit()
    return result


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.post("/bulk-delete", status_code=status.HTTP_200_OK)
async def bulk_delete_leads(
    request: Request,
    bulk_data: schemas.BulkDeleteSchema,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
):
    """
    (Admin only) Bulk soft delete multiple leads.

    **Request Body:**
    ```json
    {"lead_ids": [1, 2, 3]}
    ```

    **Response:**
    - deleted_count: Number of successfully deleted leads
    - skipped: List of lead IDs that couldn't be deleted (with reasons)
    """
    deleted_count = 0
    skipped = []

    for lead_id in bulk_data.lead_ids:
        try:
            deleted_lead = await lead_service.delete_lead(db, lead_id, deleted_by=current_user)
            if deleted_lead:
                deleted_count += 1
        except Exception as e:
            log.warning(f"Failed to delete lead {lead_id}: {e}")
            skipped.append({"lead_id": lead_id, "reason": str(e)})

    await db.commit()
    return {
        "message": f"Deleted {deleted_count} leads",
        "deleted_count": deleted_count,
        "skipped": skipped
    }


@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.post("/import", response_model=schemas.LeadImportResult)
async def officer_import_leads(
    request: Request,
    file: UploadFile = File(
        ..., description="CSV or Excel file containing lead data (.csv, .xlsx)"
    ),
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
):
    """
    Import leads từ file CSV hoặc Excel (cho Officer/Admin/Manager).

    **Khác với Admin import:**
    - Leads được tự động gán cho officer đang import
    - Officer chỉ có thể import vào unit của mình

    **Required columns:** full_name, email, phone, source
    **Optional columns:** phone2, offering_id, education_level, gpa, location

    **Note:** unit_id sẽ được tự động set thành unit của officer.
    """
    log.info(
        "Received officer lead import request",
        user_id=current_user.id,
        role=current_user.role,
        filename=file.filename,
    )

    # Validate officer has a unit
    if not current_user.unit_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You must be assigned to a unit before importing leads."
        )

    # Read file content
    try:
        content = await file.read()
    except Exception as e:
        log.error(
            "Failed to read uploaded file",
            filename=file.filename,
            error=str(e),
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to read uploaded file. Please check the file format and try again."
        )

    # Call service with auto-assign parameters
    try:
        result, callback = await lead_service.import_leads_from_file_content(
            file_content=content,
            filename=file.filename or "unknown",
            db=db,
            default_unit_id=current_user.unit_id,  # Force unit to officer's unit
            auto_assign_officer_id=current_user.id,  # Auto-assign to officer
        )
        await db.commit()
        await callback()
        return result

    except ValueError as e:
        log.warning(
            "Lead import validation failed",
            user_id=current_user.id,
            filename=file.filename,
            error=str(e),
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e),
        )

# ==============================================================================
# FSM-COMPLIANT STATUS UPDATE (Spec v3.0)
# ==============================================================================

@limiter.limit(RateLimits.DATA_WRITE)  # 200/hour
@router.patch("/{lead_id}/status", response_model=schemas.Lead)
async def update_lead_consultation_status(
    request: Request,
    lead_id: int,
    status_update: schemas.LeadStatusUpdate,
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = CasbinAuth,
    # ✅ SMART DEPENDENCY: Validates FSM transition BEFORE service
    # This dependency enforces RULE #12: Backend must validate transitions
    # Returns validated ConsultationStatus or raises BusinessRuleViolation (400)
):
    """
    Update lead consultation status with FSM validation (Spec v3.0 compliant).
    
    This endpoint enforces:
    - Rule #11: NULL status → only NOT_CONTACTED
    - Rule #12: Backend FSM validation (not just UI)
    - Phase guards (user/role cannot cross phases)
    - Trigger type enforcement (cannot set system-only statuses)
    
    **Architecture Compliance:**
    - Smart Dependency: validate_status_transition() does ALL validation
    - Dumb Router: Just coordinates service + commit
    - Service: Pure Python, raises domain exceptions
    
    **Request Body:**
    ```json
    {
        "consultation_status_id": "sts02"
    }
    ```
    
    **Error Responses:**
    - 400: Invalid FSM transition (BusinessRuleViolation)
    - 404: Lead or status not found (ResourceNotFoundError)
    """
    from ..core.deps import validate_status_transition
    from ..services.status_helper import StatusHelper
    
    # ✅ SMART DEPENDENCY: Validate transition using FSM engine
    validated_status = await validate_status_transition(
        to_status_id=status_update.consultation_status_id,
        lead_id=lead_id,
        db=db,
        current_user=current_user
    )
    
    # Get lead (already validated by dependency, but need for update)
    from ..repositories.lead_repository import LeadRepository
    repo = LeadRepository(db)
    lead = await repo.get_lead_by_id_and_unit(lead_id, current_user.unit_id)
    
    if not lead:
        from ..utils.exceptions import ResourceNotFoundError
        raise ResourceNotFoundError(f"Lead {lead_id} not found")
    
    # Store old status for history/notification
    old_status_id = lead.consultation_status_id
    
    # ✅ UPDATE STATUS: Use StatusHelper to sync all fields
    await StatusHelper.sync_lead_status(lead, validated_status)
    
    # Create history record
    history = models.LeadHistory(
        lead_id=lead.id,
        consultation_status_id=validated_status.id,
        changed_by_user_id=current_user.id,
        notes=f"Status updated via FSM-validated endpoint"
    )
    db.add(history)
    
    # Flush before commit to get updated timestamp
    await db.flush()
    
    # Commit transaction
    await db.commit()
    await db.refresh(lead)
    
    log.info(
        "Lead status updated via FSM endpoint",
        lead_id=lead_id,
        old_status=old_status_id,
        new_status=validated_status.id,
        user_id=current_user.id
    )
    
    # TODO: Dispatch status change notification if needed
    # await dispatch(db, SystemEvents.LEAD_STATUS_CHANGED, {...})

    return lead


# =============================================================================
# LEAD VALIDITY STATUS (Collaborator System Phase 1)
# =============================================================================


@router.post("/{lead_id}/validity", response_model=schemas.Lead)
async def set_lead_validity(
    validity_data: LeadValidityUpdate,
    lead: models.Lead = Depends(get_lead_for_user),
    db: AsyncSession = Depends(database.get_db),
    current_user: models.User = Depends(require_admin_or_manager),
):
    """
    Update lead validity status (admin/manager only).

    Validity status is a quality control field separate from lead lifecycle status.
    Used by CTV system to track lead quality: raw, duplicate, invalid, valid, qualified.
    """
    from app.services.collaborator_service import update_lead_validity
    from app.repositories.lead_repository import LeadRepository

    updated_lead, _ = await update_lead_validity(db, lead, validity_data)
    await db.commit()

    repo = LeadRepository(db)
    return await repo.get_by_id_shallow(updated_lead.id)
