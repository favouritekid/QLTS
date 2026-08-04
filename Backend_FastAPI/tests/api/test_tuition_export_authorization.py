"""Authz + route-order cho `GET /api/invoices/export` (PR-A / H1).

Ma trận: admin (wildcard) · accountant · manager → QUA gate; officer · user → 403.

Ca quan trọng nhất là ``test_export_not_shadowed_by_invoice_id``: nếu ai đó dời
khai báo route xuống dưới ``/{invoice_id}`` thì FastAPI ép "export" thành int và
trả **422**, không phải 404 — triệu chứng không ai đoán ra. Cùng bẫy đã có tiền
lệ ở payments.py (/import/batches).
"""

from __future__ import annotations

import pytest_asyncio

from app import models
from app.database import AsyncSessionLocal
from app.main import fastapi_app
from app.security import get_password_hash
from tests.fixtures.constants import AuthURLs
from tests.fixtures.users import create_user_with_role, get_auth_headers

EXPORT_URL = "/api/invoices/export"
DEBT_EXPORT_URL = "/api/finance/debt-report/export"


@pytest_asyncio.fixture
async def accountant_token_headers(client) -> dict:
    """Accountant + headers (conftest KHÔNG có sẵn fixture này)."""
    try:
        from casbin_async_sqlalchemy_adapter import CasbinRule
    except ImportError:  # pragma: no cover
        CasbinRule = None
    info = await create_user_with_role(
        session_factory=AsyncSessionLocal,
        user_data={
            "username": "acct_export_test",
            "email": "acct_export_test@test.com",
            "password": "AcctPass123!",
            "role": "accountant",
            "status": "active",
        },
        casbin_role="role:accountant",
        unit_id=None,
        models=models,
        get_password_hash=get_password_hash,
        CasbinRule=CasbinRule,
        app=fastapi_app,
    )
    return await get_auth_headers(client, info, AuthURLs.LOGIN)


class TestExportAuthz:
    async def test_officer_denied(self, client, officer_token_headers):
        r = await client.get(EXPORT_URL, headers=officer_token_headers)
        assert r.status_code == 403

    async def test_regular_user_denied(self, client, regular_user_token_headers):
        r = await client.get(EXPORT_URL, headers=regular_user_token_headers)
        assert r.status_code == 403

    async def test_accountant_passes_gate(self, client, accountant_token_headers):
        r = await client.get(EXPORT_URL, headers=accountant_token_headers)
        assert r.status_code != 403, r.text
        assert r.status_code == 200

    async def test_manager_passes_gate(self, client, manager_token_headers):
        r = await client.get(EXPORT_URL, headers=manager_token_headers)
        assert r.status_code != 403, r.text
        assert r.status_code == 200

    async def test_admin_passes_gate(self, client, admin_token_headers):
        r = await client.get(EXPORT_URL, headers=admin_token_headers)
        assert r.status_code == 200, r.text


class TestExportRouteShape:
    async def test_export_not_shadowed_by_invoice_id(
        self, client, admin_token_headers
    ):
        """[N] /export KHÔNG bị /{invoice_id} nuốt.

        Nếu bị nuốt: FastAPI ép "export" thành int → 422. Test này đỏ đúng lúc
        ai đó dời thứ tự khai báo route.
        """
        r = await client.get(EXPORT_URL, headers=admin_token_headers)
        assert r.status_code != 422, "route /export đang bị /{invoice_id} nuốt"

    async def test_invalid_format_rejected(self, client, admin_token_headers):
        r = await client.get(
            EXPORT_URL, params={"format": "pdf"}, headers=admin_token_headers
        )
        assert r.status_code == 422

    async def test_default_format_is_xlsx(self, client, admin_token_headers):
        r = await client.get(EXPORT_URL, headers=admin_token_headers)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        assert "danh_sach_khoan_phi_" in r.headers["content-disposition"]

    async def test_csv_format_has_bom(self, client, admin_token_headers):
        r = await client.get(
            EXPORT_URL, params={"format": "csv"}, headers=admin_token_headers
        )
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("text/csv")
        assert r.content.startswith("﻿".encode("utf-8"))


class TestDebtReportExportAuthz:
    """`GET /api/finance/debt-report/export` — quyền qua Casbin THẬT.

    🔴 Route này KHÁC /api/invoices/export: policy sẵn có cho báo cáo công nợ là
    literal '/api/finance/debt-report', thêm một segment là **hết khớp
    keyMatch4** ⇒ không có va chạm nào che, thiếu grant là accountant/manager
    403 thật. Vì vậy phải kiểm bằng request thật qua ASGI, không chỉ test
    service.
    """

    async def test_officer_denied(self, client, officer_token_headers):
        r = await client.get(DEBT_EXPORT_URL, headers=officer_token_headers)
        assert r.status_code == 403

    async def test_regular_user_denied(self, client, regular_user_token_headers):
        r = await client.get(DEBT_EXPORT_URL, headers=regular_user_token_headers)
        assert r.status_code == 403

    async def test_accountant_passes_gate(self, client, accountant_token_headers):
        """[N] Bỏ grant accountant khỏi template là ca này đỏ."""
        r = await client.get(DEBT_EXPORT_URL, headers=accountant_token_headers)
        assert r.status_code != 403, r.text
        assert r.status_code == 200

    async def test_manager_passes_gate(self, client, manager_token_headers):
        r = await client.get(DEBT_EXPORT_URL, headers=manager_token_headers)
        assert r.status_code != 403, r.text
        assert r.status_code == 200

    async def test_admin_passes_gate(self, client, admin_token_headers):
        r = await client.get(DEBT_EXPORT_URL, headers=admin_token_headers)
        assert r.status_code == 200, r.text

    async def test_not_shadowed_by_debt_report_route(
        self, client, admin_token_headers
    ):
        """/debt-report/export không bị /debt-report nuốt (trả JSON thay vì tệp)."""
        r = await client.get(DEBT_EXPORT_URL, headers=admin_token_headers)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        assert "bao_cao_cong_no_" in r.headers["content-disposition"]

    async def test_csv_has_bom_and_vietnamese_header(
        self, client, admin_token_headers
    ):
        r = await client.get(
            DEBT_EXPORT_URL, params={"format": "csv"}, headers=admin_token_headers
        )
        assert r.status_code == 200
        assert r.content.startswith("﻿".encode("utf-8"))
        head_line = r.content.decode("utf-8").splitlines()[0]
        assert "Mã hồ sơ" in head_line
        assert "total_outstanding" not in head_line  # hết khoá kỹ thuật

    async def test_invalid_format_rejected(self, client, admin_token_headers):
        r = await client.get(
            DEBT_EXPORT_URL,
            params={"format": "pdf"},
            headers=admin_token_headers,
        )
        assert r.status_code == 422


class TestExportCacheHeaders:
    """Tệp xuất mang PII (họ tên + CCCD) nên không được để cache lưu.

    Không có Cache-Control thì RFC 7234 §4.2.2 cho phép cache theo suy nghiệm;
    hệ này xác thực bằng cookie nên cache dùng chung không bị cấm lưu. Trên máy
    dùng chung ở quầy, người sau mở lại từ lịch sử là có tệp mà không cần đăng
    nhập lại. Cùng quy ước với sms_export / enrollment_letters / admissions.
    """

    async def test_tuition_export_sets_no_store(self, client, admin_token_headers):
        r = await client.get(EXPORT_URL, headers=admin_token_headers)
        assert r.status_code == 200
        assert r.headers.get("cache-control") == "private, no-store"
        assert r.headers.get("x-content-type-options") == "nosniff"

    async def test_debt_export_sets_no_store(self, client, admin_token_headers):
        r = await client.get(DEBT_EXPORT_URL, headers=admin_token_headers)
        assert r.status_code == 200
        assert r.headers.get("cache-control") == "private, no-store"
        assert r.headers.get("x-content-type-options") == "nosniff"


class TestExportMetadataNoExistenceOracle:
    """Sheet "Bộ lọc đã áp dụng" không được thành máy dò tồn tại.

    Nếu ID có thật (nhưng ngoài phạm vi) cho ra chuỗi KHÁC với ID không tồn
    tại, người xem dò được danh mục người dùng / đơn vị của đơn vị khác mà
    không cần quyền xem dữ liệu của họ. Hai trường hợp phải KHÔNG phân biệt được.
    """

    @staticmethod
    def _meta_text(content: bytes) -> str:
        import io as _io

        from openpyxl import load_workbook

        wb = load_workbook(_io.BytesIO(content))
        return chr(10).join(
            str(c.value or "")
            for row in wb["Thong tin xuat"].iter_rows()
            for c in row
        )

    async def test_officer_out_of_scope_and_nonexistent_look_same(
        self, client, manager_token_headers, manager_other_unit_user_in_db
    ):
        """[N] officer CÓ THẬT ở đơn vị khác ≡ officer không tồn tại.

        Dùng user thuộc ĐƠN VỊ KHÁC tường minh (fixture), không đoán ID —
        ID nhỏ trong DB test rất dễ trúng chính người đang gọi.
        """
        other_user_id = manager_other_unit_user_in_db["id"]

        r_other = await client.get(
            EXPORT_URL,
            params={"format": "xlsx", "officer_id": other_user_id},
            headers=manager_token_headers,
        )
        assert r_other.status_code == 200
        meta_other = self._meta_text(r_other.content)

        r_missing = await client.get(
            EXPORT_URL,
            params={"format": "xlsx", "officer_id": 999999},
            headers=manager_token_headers,
        )
        assert r_missing.status_code == 200
        meta_missing = self._meta_text(r_missing.content)

        def _officer_line(meta: str) -> str:
            for ln in meta.splitlines():
                if "ngoài phạm vi hoặc không tồn tại" in ln:
                    return ln
            return ""

        line_other = _officer_line(meta_other)
        line_missing = _officer_line(meta_missing)
        assert line_other, f"user ngoài phạm vi phải hiện dạng mờ: {meta_other}"
        assert line_missing, f"user không tồn tại phải hiện dạng mờ: {meta_missing}"

        # Tên thật KHÔNG được lọt ra.
        other_username = manager_other_unit_user_in_db["username"]
        assert other_username not in meta_other, "lộ tên người ngoài phạm vi"

    async def test_scope_and_user_filter_are_separate_rows(
        self, client, manager_token_headers
    ):
        """Sheet phải tách "Phạm vi quyền" (vai trò áp) khỏi "Bộ lọc đơn vị"."""
        r = await client.get(
            EXPORT_URL, params={"format": "xlsx"}, headers=manager_token_headers
        )
        assert r.status_code == 200
        meta = self._meta_text(r.content)
        assert "Phạm vi quyền" in meta
