"""Test xuất báo cáo công nợ (PR-A / H2).

Lỗi cũ mà các ca ở đây khoá lại: CSV dựng ở TRÌNH DUYỆT nên header là khoá kỹ
thuật tiếng Anh, không BOM (Excel tiếng Việt mojibake), ô tiền bọc thành chuỗi
(không cộng được), tên tệp cố định (xuất 2 lần đè nhau).
"""

from __future__ import annotations

import csv
import io
import itertools
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app import models
from app.models.finance import Fee, Invoice
from app.services import finance_report_service as frs
from app.services.finance_report_service import FinanceReportService
from app.utils.exceptions import BadRequest
import pytest

_phone_seq = itertools.count(1)


async def _seed_debtor(
    db: AsyncSession,
    deps: dict,
    *,
    amount: str,
    paid: str,
    lead_name: str = "Trần Thị Nợ",
    citizen_id: str = "012345678901",
    due_date: date = date(2026, 7, 1),
):
    lead = models.Lead(
        full_name=lead_name,
        phone=f"09{next(_phone_seq):08d}",
        source="debt_export_test",
        unit_id=deps["unit_id"],
        consultation_status_id=deps["initial_status_id"],
    )
    db.add(lead)
    await db.flush()

    profile = models.AdmissionProfile(
        lead_id=lead.id,
        status="submitted",
        academic_year=2026,
        applied_rules={},
        citizen_id=citizen_id,
    )
    db.add(profile)
    await db.flush()

    fee = Fee(
        admission_profile_id=profile.id,
        fee_type="tuition",
        academic_year=2026,
        semester_no=1,
        base_amount=Decimal(amount),
        final_amount=Decimal(amount),
        paid_amount=Decimal(paid),
        status="partial",
    )
    db.add(fee)
    await db.flush()

    db.add(
        Invoice(
            fee_id=fee.id,
            invoice_number=f"INV-DEBT-{profile.id}",
            installment_no=1,
            amount=Decimal(amount),
            paid_amount=Decimal(paid),
            penalty_amount=Decimal("0"),
            status="partial",
            due_date=due_date,
        )
    )
    await db.flush()
    return profile, fee


def _csv_rows(content: bytes) -> list:
    """Chỉ phần BẢNG — CSV nay nối khối "Thong tin xuat" sau một dòng trống."""
    text = content.decode("utf-8")
    assert text.startswith("﻿"), "CSV phải mở đầu bằng BOM UTF-8"
    out = []
    for r in csv.reader(io.StringIO(text[1:])):
        if not any(c.strip() for c in r):
            break
        out.append(r)
    return out


def _unquote_forced(value: str) -> str:
    if value.startswith('="') and value.endswith('"'):
        return value[2:-1].replace('""', '"')
    return value


class TestDebtExportContent:
    async def test_export_matches_report_row_for_row(self, db, seeded_dependencies):
        """[N] File xuất khớp TỪNG DÒNG với get_debt_report.

        Đỏ nếu ai đó dựng lại truy vấn riêng cho export thay vì gọi báo cáo.
        """
        await _seed_debtor(db, seeded_dependencies, amount="9000000", paid="2000000")
        await _seed_debtor(
            db, seeded_dependencies, amount="7000000", paid="1000000",
            citizen_id="012345678902", lead_name="Lê Văn Nợ",
        )

        service = FinanceReportService(db)
        report = await service.get_debt_report()
        content, _, _ = await service.build_debt_report_export(
            fmt="csv", exporter_name="Kế toán A"
        )
        rows = _csv_rows(content)[1:]

        assert len(rows) == len(report.items)
        for row, item in zip(rows, report.items):
            assert _unquote_forced(row[0]) == item.profile_code
            assert row[1] == item.profile_name
            assert Decimal(row[9]) == Decimal(item.total_outstanding)

    async def test_header_is_vietnamese_not_technical_keys(
        self, db, seeded_dependencies
    ):
        """[N] Tiêu đề tiếng Việt — không còn khoá kỹ thuật như bản dựng ở client."""
        await _seed_debtor(db, seeded_dependencies, amount="9000000", paid="2000000")
        service = FinanceReportService(db)
        content, _, _ = await service.build_debt_report_export(
            fmt="csv", exporter_name="Kế toán A"
        )
        header = _csv_rows(content)[0]
        assert header == frs.DEBT_REPORT_COLUMNS
        joined = " ".join(header)
        for technical in ("total_outstanding", "profile_code", "aging_bucket"):
            assert technical not in joined

    async def test_money_cells_are_numbers_in_xlsx(self, db, seeded_dependencies):
        """[N] Ô tiền là SỐ (Excel cộng được) — bản cũ bọc mọi ô thành chuỗi."""
        await _seed_debtor(db, seeded_dependencies, amount="9000000", paid="2000000")
        service = FinanceReportService(db)
        content, _, _ = await service.build_debt_report_export(
            fmt="xlsx", exporter_name="Kế toán A"
        )
        ws = load_workbook(io.BytesIO(content)).worksheets[0]
        cell = ws.cell(row=2, column=10)  # "Còn nợ"
        assert isinstance(cell.value, (int, float, Decimal)), type(cell.value)
        assert cell.number_format == "#,##0"

    async def test_filename_has_timestamp(self, db, seeded_dependencies):
        """Tên tệp có mốc thời gian — xuất nhiều lần không đè lên nhau."""
        await _seed_debtor(db, seeded_dependencies, amount="9000000", paid="2000000")
        service = FinanceReportService(db)
        _, media, filename = await service.build_debt_report_export(
            fmt="csv", exporter_name="Kế toán A"
        )
        assert filename.startswith("bao_cao_cong_no_")
        assert filename != "bao_cao_cong_no.csv"
        assert media.startswith("text/csv")

    async def test_meta_sheet_warns_about_scope(self, db, seeded_dependencies):
        """Sheet phụ phải cảnh báo 'Đã thu' chỉ tính trên đợt CÒN NỢ.

        Đây là điểm dễ đọc nhầm nhất của báo cáo này: hồ sơ trả xong đợt 1 và
        còn nợ đợt 2 thì tiền đợt 1 KHÔNG có trong cột 'Đã thu'.
        """
        await _seed_debtor(db, seeded_dependencies, amount="9000000", paid="2000000")
        service = FinanceReportService(db)
        content, _, _ = await service.build_debt_report_export(
            fmt="xlsx", exporter_name="Kế toán A"
        )
        wb = load_workbook(io.BytesIO(content))
        meta_text = "\n".join(
            str(c.value or "")
            for row in wb["Thong tin xuat"].iter_rows()
            for c in row
        )
        assert "ĐỢT THU CÒN NỢ" in meta_text
        assert "KHÔNG phải tổng đã thu" in meta_text

    async def test_formula_injection_neutralised(self, db, seeded_dependencies):
        """[N] Tên thí sinh chứa công thức → bị vô hiệu."""
        await _seed_debtor(
            db, seeded_dependencies, amount="9000000", paid="2000000",
            lead_name="=HYPERLINK(\"http://x\")",
        )
        service = FinanceReportService(db)
        content, _, _ = await service.build_debt_report_export(
            fmt="csv", exporter_name="Kế toán A"
        )
        row = _csv_rows(content)[1]
        assert row[1].startswith("'"), row[1]


class TestDebtExportCap:
    """Trần dòng: phải chặn TRƯỚC khi nạp, nhưng KHÔNG được từ chối oan.

    Ba hợp đồng ở đây tương ứng ba cách hỏng khác nhau:
      * bỏ ``aging`` khỏi preflight → từ chối oan người lọc nhóm nhỏ;
      * kiểm sau khi nạp → cap không chặn được truy vấn, chỉ chặn dựng workbook;
      * bỏ hậu kiểm → khe TOCTOU giữa đếm và nạp cho lọt quá trần.
    """

    async def _seed_n(self, db, deps, n: int, *, days_overdue: int, tag: str):
        from datetime import timedelta
        due = date.today() - timedelta(days=days_overdue)
        for i in range(n):
            await _seed_debtor(
                db, deps, amount="9000000", paid="2000000",
                citizen_id=f"{tag}{i:05d}"[:12], due_date=due,
            )

    async def test_total_over_cap_but_selected_aging_under_cap_still_exports(
        self, db, seeded_dependencies, monkeypatch
    ):
        """[N] Bỏ aging khỏi preflight là ca này đỏ (từ chối oan).

        Tổng vượt trần, nhưng nhóm tuổi nợ người dùng chọn thì dưới trần.
        """
        monkeypatch.setattr(frs, "MAX_EXPORT_ROWS", 3)
        # 4 hồ sơ quá hạn >60 ngày (vượt trần nếu tính TỔNG)
        await self._seed_n(db, seeded_dependencies, 4, days_overdue=90, tag="0711")
        # 2 hồ sơ quá hạn 40 ngày → bucket 31_60, dưới trần
        await self._seed_n(db, seeded_dependencies, 2, days_overdue=40, tag="0722")

        service = FinanceReportService(db)
        content, _, _ = await service.build_debt_report_export(
            fmt="csv", exporter_name="Kế toán A", aging="31_60",
        )
        rows = _csv_rows(content)[1:]
        assert len(rows) == 2, f"phải xuất đúng nhóm đã chọn, thực tế {len(rows)}"

    async def test_selected_aging_over_cap_refuses_before_fetch(
        self, db, seeded_dependencies, monkeypatch
    ):
        """[N] Từ chối phải xảy ra TRƯỚC khi gọi get_debt_report.

        Kiểm sau khi nạp thì cap chỉ chặn bước RẺ NHẤT (dựng workbook) mà
        không chặn truy vấn/hydrate — đúng thứ cap sinh ra để chặn.
        """
        monkeypatch.setattr(frs, "MAX_EXPORT_ROWS", 2)
        await self._seed_n(db, seeded_dependencies, 4, days_overdue=90, tag="0733")

        service = FinanceReportService(db)
        called = {"n": 0}
        original = service.get_debt_report

        async def _spy(*a, **kw):
            called["n"] += 1
            return await original(*a, **kw)

        monkeypatch.setattr(service, "get_debt_report", _spy)

        with pytest.raises(BadRequest) as exc:
            await service.build_debt_report_export(
                fmt="csv", exporter_name="Kế toán A", aging="over_60",
            )
        assert "thu hẹp bộ lọc" in str(exc.value.detail)
        assert called["n"] == 0, "get_debt_report bị gọi dù đã vượt trần"

    async def test_toctou_postcheck_refuses_when_fetch_exceeds_cap(
        self, db, seeded_dependencies, monkeypatch
    ):
        """[N] Preflight dưới trần nhưng dữ liệu thật vượt → hậu kiểm chặn.

        Mô phỏng khe TOCTOU: giữa lúc đếm và lúc nạp, kế toán khác phát hành
        thêm hoá đơn. Bỏ hậu kiểm là ca này đỏ.
        """
        monkeypatch.setattr(frs, "MAX_EXPORT_ROWS", 2)
        await self._seed_n(db, seeded_dependencies, 4, days_overdue=90, tag="0744")

        service = FinanceReportService(db)

        async def _stale_count(**kw):
            return 1  # số đếm CŨ, dưới trần

        monkeypatch.setattr(service, "_count_debt_report_rows", _stale_count)

        with pytest.raises(BadRequest) as exc:
            await service.build_debt_report_export(
                fmt="csv", exporter_name="Kế toán A",
            )
        assert "thu hẹp bộ lọc" in str(exc.value.detail)
