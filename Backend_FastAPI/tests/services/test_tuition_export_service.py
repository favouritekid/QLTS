"""Test xuất danh sách học phí (PR-A / H1).

Phần lớn các ca ở đây là **kiểm ngược**: bỏ một bước trong service thì test
phải đỏ. Đánh dấu [N] trong docstring từng ca.

Ba ca quan trọng nhất, đừng xoá khi refactor:
* ``test_distinct_fee_one_row_per_fee`` — khoá quyết định lớn nhất của H1
  (grain = khoản phí, không phải hoá đơn).
* ``test_negative_remaining_not_sanitized_in_csv`` — số âm phải giữ dấu, không
  bị ``sanitize_csv_cell`` biến thành text.
* ``test_over_cap_refuses_instead_of_truncating`` — vượt trần thì TỪ CHỐI, cắt
  im lặng là file thiếu trông y hệt file đủ.
"""

from __future__ import annotations

import csv
import io
import itertools
from datetime import date, datetime, timezone
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app import models
from app.models.finance import Fee, Invoice
from app.repositories.fee_repository import InvoiceRepository
from app.services import tuition_export_service as tes
from app.utils.exceptions import BadRequest

_phone_seq = itertools.count(1)


async def _seed_fee(
    db: AsyncSession,
    deps: dict,
    *,
    invoices: list,           # [(installment_no, amount, status, paid)]
    final_amount=None,
    paid_amount="0",
    waived_amount="0",
    lead_name: str = "Nguyễn Văn An",
    citizen_id: str = "012345678901",
    unit_id: int = None,
    fee_status: str = "partial",
    officer_id: int = None,
):
    """Lead → AdmissionProfile → Fee(tuition) → Invoice(s)."""
    lead = models.Lead(
        full_name=lead_name,
        phone=f"09{next(_phone_seq):08d}",
        source="export_test",
        unit_id=unit_id or deps["unit_id"],
        consultation_status_id=deps["initial_status_id"],
        assigned_officer_id=officer_id,
    )
    db.add(lead)
    await db.flush()

    profile = models.AdmissionProfile(
        lead_id=lead.id,
        status="submitted",
        academic_year=2026,
        applied_rules={},
        citizen_id=citizen_id,
        created_at=datetime(2026, 6, 15, 8, 30, tzinfo=timezone.utc),
    )
    db.add(profile)
    await db.flush()

    total = (
        Decimal(str(final_amount))
        if final_amount is not None
        else sum((Decimal(str(a)) for (_n, a, *_r) in invoices), Decimal("0"))
    )
    fee = Fee(
        admission_profile_id=profile.id,
        fee_type="tuition",
        academic_year=2026,
        semester_no=1,
        base_amount=total,
        final_amount=total,
        paid_amount=Decimal(str(paid_amount)),
        waived_amount=Decimal(str(waived_amount)),
        status=fee_status,
        resolved_degree_level="Cao đẳng",
    )
    db.add(fee)
    await db.flush()

    for no, amount, st, paid in invoices:
        db.add(
            Invoice(
                fee_id=fee.id,
                invoice_number=f"INV-EX-{profile.id}-{no}",
                installment_no=no,
                amount=Decimal(str(amount)),
                paid_amount=Decimal(str(paid)),
                penalty_amount=Decimal("0"),
                status=st,
                due_date=date(2026, 9, 5),
            )
        )
    await db.flush()
    return profile, fee


def _csv_rows(content: bytes) -> list:
    text = content.decode("utf-8")
    assert text.startswith("﻿"), "CSV phải mở đầu bằng BOM UTF-8"
    return list(csv.reader(io.StringIO(text[1:])))


async def _export(db, **kw):
    return await tes.build_tuition_export(
        db, fmt=kw.pop("fmt", "csv"), unit_id=kw.pop("unit_id", None),
        exporter_name=kw.pop("exporter_name", "Kế toán A"), **kw
    )


class TestGrain:
    """Grain = khoản phí. Đây là quyết định lớn nhất của H1."""

    async def test_distinct_fee_one_row_per_fee(self, db, seeded_dependencies):
        """[N] 1 khoản phí + 3 hoá đơn cùng khớp lọc → ĐÚNG 1 dòng.

        Đổi sang grain hoá đơn là test này đỏ. Nếu nó đỏ mà bạn định "sửa cho
        xanh", hãy đọc lại docstring service trước: xuất theo hoá đơn làm cột
        tiền cộng ra gấp đôi.
        """
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "3000000", "partial", "1000000"),
                      (2, "3000000", "issued", "0"),
                      (3, "3000000", "issued", "0")],
            final_amount="9000000", paid_amount="1000000",
        )
        content, _, _ = await _export(db)
        rows = _csv_rows(content)
        assert len(rows) == 2, f"1 header + 1 dòng, thực tế {len(rows)}"

    async def test_money_total_matches_sum_of_fees(self, db, seeded_dependencies):
        """Tổng cột 'đã đóng' == Σ fee.paid_amount — chống nhân đôi."""
        for i in range(3):
            await _seed_fee(
                db, seeded_dependencies,
                invoices=[(1, "5000000", "partial", "2000000"),
                          (2, "5000000", "issued", "0")],
                final_amount="10000000", paid_amount="2000000",
                citizen_id=f"01234567890{i}",
            )
        content, _, _ = await _export(db)
        rows = _csv_rows(content)[1:]
        total_paid = sum(Decimal(r[12]) for r in rows)
        assert total_paid == Decimal("6000000")


class TestMoneyCells:
    async def test_negative_remaining_not_sanitized_in_csv(
        self, db, seeded_dependencies
    ):
        """[N] Hồ sơ đóng dư → ô 'Số tiền còn lại' giữ dấu âm, KHÔNG prefix '.

        ``DANGEROUS_PREFIXES`` của sanitize_csv_cell có cả '-', nên nếu ô tiền
        đi qua sanitize thì Excel đọc thành text và không cộng được.
        """
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "2000000", "paid", "2050000")],
            final_amount="2000000", paid_amount="2050000", fee_status="paid",
        )
        content, _, _ = await _export(db)
        row = _csv_rows(content)[1]
        # Điều cần khoá: KHÔNG có dấu nháy đầu (nếu có thì Excel đọc thành text
        # và không cộng được) và giá trị vẫn âm. Số chữ số thập phân do Decimal
        # quyết định nên so bằng Decimal, không so chuỗi.
        assert not row[14].startswith("'"), (
            f"ô tiền bị sanitize thành text: {row[14]!r}"
        )
        assert Decimal(row[14]) == Decimal("-50000"), row[14]

    async def test_money_cells_are_numbers_in_xlsx(self, db, seeded_dependencies):
        """[N] Ô tiền trong xlsx là SỐ + number_format nghìn (Excel cộng được)."""
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "7000000", "partial", "2500000")],
            final_amount="7000000", paid_amount="2500000",
        )
        content, _, _ = await _export(db, fmt="xlsx")
        ws = load_workbook(io.BytesIO(content))[tes.SHEET_DATA]
        cell = ws.cell(row=2, column=12)  # "Học phí ngành học"
        assert isinstance(cell.value, (int, float, Decimal)), type(cell.value)
        assert cell.number_format == "#,##0"

    async def test_cccd_forced_text_keeps_leading_zero(
        self, db, seeded_dependencies
    ):
        """[N] CCCD ép TEXT → không mất số 0 đầu."""
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "1000000", "issued", "0")],
            citizen_id="001234567890",
        )
        content, _, _ = await _export(db, fmt="xlsx")
        ws = load_workbook(io.BytesIO(content))[tes.SHEET_DATA]
        assert ws.cell(row=2, column=5).number_format == "@"
        assert ws.cell(row=2, column=5).value == "001234567890"


class TestSanitize:
    async def test_formula_injection_in_name_is_neutralised(
        self, db, seeded_dependencies
    ):
        """[N] Tên chứa công thức → phải bị vô hiệu bằng dấu nháy."""
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "1000000", "issued", "0")],
            lead_name="=cmd|'/c calc'!A1",
        )
        content, _, _ = await _export(db)
        row = _csv_rows(content)[1]
        assert row[3].startswith("'"), f"tên phải bị sanitize, thực tế {row[3]!r}"


class TestCap:
    async def test_over_cap_refuses_instead_of_truncating(
        self, db, seeded_dependencies, monkeypatch
    ):
        """[N] Vượt trần → BadRequest, KHÔNG trả file bị cắt."""
        monkeypatch.setattr(tes, "MAX_EXPORT_ROWS", 2)
        for i in range(3):
            await _seed_fee(
                db, seeded_dependencies,
                invoices=[(1, "1000000", "issued", "0")],
                citizen_id=f"09876543210{i}",
            )
        with pytest.raises(BadRequest) as exc:
            await _export(db)
        assert "thu hẹp bộ lọc" in str(exc.value.detail)

    async def test_exactly_at_cap_succeeds(
        self, db, seeded_dependencies, monkeypatch
    ):
        """Đúng bằng trần thì vẫn xuất được (biên dưới)."""
        monkeypatch.setattr(tes, "MAX_EXPORT_ROWS", 2)
        for i in range(2):
            await _seed_fee(
                db, seeded_dependencies,
                invoices=[(1, "1000000", "issued", "0")],
                citizen_id=f"09876543211{i}",
            )
        content, _, _ = await _export(db)
        assert len(_csv_rows(content)) == 3  # header + 2


class TestFilterParity:
    async def test_export_fee_set_equals_list_fee_set(
        self, db, seeded_dependencies
    ):
        """[N] Tập fee xuất ra == tập fee DISTINCT của danh sách (đẳng thức).

        Đỏ nếu ai đó khai lại điều kiện lọc thay vì gọi
        ``_build_invoice_list_conditions``.
        """
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "3000000", "partial", "1000000")],
            citizen_id="011111111111",
        )
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "4000000", "issued", "0")],
            citizen_id="022222222222",
        )

        repo = InvoiceRepository(db)
        invoices, _ = await repo.get_filtered_with_count(skip=0, limit=1000)
        from_list = {inv.fee_id for inv in invoices}
        from_export = set(
            await repo.get_distinct_fee_ids_for_filter(limit=1000)
        )
        assert from_export == from_list


class TestScope:
    async def test_unit_scope_excludes_other_unit(
        self, db, seeded_dependencies, seed_other_unit
    ):
        """IDOR: chỉ thấy khoản phí thuộc đơn vị được phép."""
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "1000000", "issued", "0")],
            citizen_id="033333333333",
        )
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "1000000", "issued", "0")],
            citizen_id="044444444444",
            unit_id=seed_other_unit["unit_id"],
        )
        content, _, _ = await _export(db, unit_id=seeded_dependencies["unit_id"])
        assert len(_csv_rows(content)) == 2  # header + 1


class TestFileShape:
    async def test_filename_and_meta_sheet(self, db, seeded_dependencies):
        """Tên file có mốc thời gian + sheet phụ ghi bộ lọc và cảnh báo."""
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "1000000", "issued", "0")],
        )
        content, media, filename = await _export(
            db, fmt="xlsx", applied_filters={"Năm học": 2026},
        )
        assert filename.startswith("danh_sach_khoan_phi_")
        assert filename.endswith(".xlsx")
        assert "spreadsheetml" in media

        wb = load_workbook(io.BytesIO(content))
        assert tes.SHEET_META in wb.sheetnames
        meta_text = "\n".join(
            str(c.value or "")
            for row in wb[tes.SHEET_META].iter_rows()
            for c in row
        )
        assert "Năm học" in meta_text
        assert "MỘT KHOẢN PHÍ" in meta_text  # cảnh báo grain
        assert "đóng dư" in meta_text        # cảnh báo số âm

    async def test_header_matches_columns(self, db, seeded_dependencies):
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "1000000", "issued", "0")],
        )
        content, _, _ = await _export(db)
        assert _csv_rows(content)[0] == tes.COLUMNS


class TestMixedFeeTypes:
    """File có thể trộn nhiều loại phí — nhãn và cảnh báo phải nói đúng.

    Bộ lọc workspace KHÔNG mặc định lọc loại phí (đúng nguyên tắc "xuất đúng
    cái đang xem"), nên file thường có cả học phí lẫn lệ phí xét tuyển. Rủi ro:
    kế toán bôi đen cột tiền mà quên lọc → cộng trộn hai loại.
    """

    async def _seed_two_fee_types(self, db, deps):
        profile, _fee = await _seed_fee(
            db, deps,
            invoices=[(1, "9000000", "partial", "2000000")],
            final_amount="9000000", paid_amount="2000000",
        )
        # Thêm khoản lệ phí xét tuyển cho CÙNG hồ sơ.
        app_fee = Fee(
            admission_profile_id=profile.id,
            fee_type="application",
            academic_year=2026,
            semester_no=None,
            base_amount=Decimal("70000"),
            final_amount=Decimal("70000"),
            paid_amount=Decimal("70000"),
            status="paid",
        )
        db.add(app_fee)
        await db.flush()
        db.add(
            Invoice(
                fee_id=app_fee.id,
                invoice_number=f"APP-{profile.id}",
                installment_no=1,
                amount=Decimal("70000"),
                paid_amount=Decimal("70000"),
                penalty_amount=Decimal("0"),
                status="paid",
                due_date=date(2026, 9, 5),
            )
        )
        await db.flush()

    async def test_csv_header_is_fee_type_neutral(self, db, seeded_dependencies):
        """[N] Tiêu đề CSV KHÔNG được gọi mọi khoản là "học phí".

        CSV không có sheet phụ nên nhãn trung tính là lớp bảo vệ duy nhất —
        đây là lý do hai cột đổi thành "Giá trị khoản phí" / "Tổng đã đóng".
        """
        await self._seed_two_fee_types(db, seeded_dependencies)
        content, _, filename = await _export(db)
        header = _csv_rows(content)[0]
        assert "Giá trị khoản phí" in header
        assert "Tổng đã đóng" in header
        assert "Học phí ngành học" not in header
        assert "Tổng học phí đã đóng" not in header
        assert filename.startswith("danh_sach_khoan_phi_")

    async def test_xlsx_warns_when_multiple_fee_types(
        self, db, seeded_dependencies
    ):
        """[N] Trộn loại phí → sheet phụ phải có cảnh báo động, liệt kê loại."""
        await self._seed_two_fee_types(db, seeded_dependencies)
        content, _, _ = await _export(db, fmt="xlsx")
        wb = load_workbook(io.BytesIO(content))
        meta = "\n".join(
            str(c.value or "")
            for row in wb[tes.SHEET_META].iter_rows()
            for c in row
        )
        assert "NHIỀU LOẠI PHÍ" in meta
        assert "Học phí" in meta
        assert "Lệ phí xét tuyển" in meta

    async def test_xlsx_no_warning_for_single_fee_type(
        self, db, seeded_dependencies
    ):
        """Chỉ một loại phí → KHÔNG cảnh báo (tránh nhiễu, cảnh báo thừa sẽ bị bỏ qua)."""
        await _seed_fee(
            db, seeded_dependencies,
            invoices=[(1, "9000000", "partial", "2000000")],
            final_amount="9000000", paid_amount="2000000",
        )
        content, _, _ = await _export(db, fmt="xlsx")
        wb = load_workbook(io.BytesIO(content))
        meta = "\n".join(
            str(c.value or "")
            for row in wb[tes.SHEET_META].iter_rows()
            for c in row
        )
        assert "NHIỀU LOẠI PHÍ" not in meta
