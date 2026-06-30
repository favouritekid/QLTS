# tests/integration/test_sms_export.py
"""
Integration test PR-4 (SMS Export BE).

Phủ: gate fail-closed (thiếu attestation / over_limit / drift consent), export
sinh file per nhà mạng (idempotent, campaign→exported), nội dung Excel đúng mẫu
(Sheet1, row1 trống, cột A 84xxx text, cột B [QC]+tin+link thật đã giải mã),
download verify sha256, mark-handed-off (recipient.handed_off_at +
contact.last_handed_off_at + campaign closed + chặn rebuild). Chạy one-off
container (CLAUDE.md).
"""
import io

from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import text

from app.database import AsyncSessionLocal

API = "/api/sms"
_GRANT = {
    "event_type": "granted",
    "occurred_at": "2026-05-01T00:00:00+07:00",
    "basis": "signed_form",
    "disclosure_version": "v1",
    "proof_reference": "ho-so/ref",
}
# 098 → viettel, 090 → mobifone → export tách 2 file.
_PHONE_VIETTEL = "0981234567"
_PHONE_MOBI = "0901234588"


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------
async def _mk_group(client, h, name="Nhom Export QA"):
    res = await client.post(
        f"{API}/contact-groups",
        json={"name": name, "group_type": "custom"},
        headers=h,
    )
    assert res.status_code == 201, res.text
    return res.json()["id"]


async def _mk_contact(client, h, name, phone, granted=True):
    res = await client.post(
        f"{API}/contacts", json={"full_name": name, "phone": phone}, headers=h
    )
    assert res.status_code == 201, res.text
    cid = res.json()["id"]
    if granted:
        r = await client.post(
            f"{API}/contacts/{cid}/consent-events", json=_GRANT, headers=h
        )
        assert r.status_code == 201, r.text
    return cid


async def _add(client, h, cid, gid):
    r = await client.post(
        f"{API}/contacts/{cid}/groups", json={"group_id": gid}, headers=h
    )
    assert r.status_code == 201, r.text


async def _seed_carriers():
    """qlts_test dùng create_all() (KHÔNG chạy migration seed) → bảng prefix
    carrier rỗng, mọi số = 'unknown'. Seed 098→viettel, 090→mobifone để test
    tách file per nhà mạng (idempotent qua ON CONFLICT)."""
    async with AsyncSessionLocal() as s:
        await s.execute(
            text(
                "INSERT INTO sms_prefix_carrier_rule "
                "(prefix, carrier_code, carrier_name, is_active) VALUES "
                "('098','viettel','Viettel', true), "
                "('090','mobifone','MobiFone', true) "
                "ON CONFLICT (prefix) DO NOTHING"
            )
        )
        await s.commit()


async def _attest_all(client, h, cid):
    for kind in ("consent", "dnc", "optout_channel"):
        r = await client.post(
            f"{API}/campaigns/{cid}/attestations",
            json={"kind": kind, "reference": f"ref-{kind}"},
            headers=h,
        )
        assert r.status_code == 201, r.text


async def _build_ready(
    client, h, template="Chao {full_name}, xem {link}", phones=None
):
    """group + contacts (granted) + campaign + attach + build → 'ready'. Trả cid."""
    gid = await _mk_group(client, h)
    phones = phones if phones is not None else [_PHONE_VIETTEL, _PHONE_MOBI]
    for i, p in enumerate(phones):
        cc = await _mk_contact(client, h, f"Phu huynh {i}", p)
        await _add(client, h, cc, gid)
    res = await client.post(
        f"{API}/campaigns",
        json={"name": "Export QA", "sms_template": template},
        headers=h,
    )
    assert res.status_code == 201, res.text
    cid = res.json()["id"]
    r = await client.post(
        f"{API}/campaigns/{cid}/groups", json={"group_id": gid}, headers=h
    )
    assert r.status_code == 201, r.text
    r = await client.post(f"{API}/campaigns/{cid}/build", headers=h)
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "ready"
    return cid


# ---------------------------------------------------------------------
# Gate fail-closed
# ---------------------------------------------------------------------
async def test_export_requires_attestation(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    cid = await _build_ready(client, h)
    res = await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    assert res.status_code == 400, res.text
    assert "xác nhận" in res.json()["detail"].lower()


async def test_export_blocks_over_limit(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    long_tpl = "Chao " + ("a" * 200) + " {full_name}"  # tin cuối > 160 GSM7
    cid = await _build_ready(client, h, template=long_tpl, phones=[_PHONE_VIETTEL])
    await _attest_all(client, h, cid)
    res = await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    assert res.status_code == 400, res.text
    assert "over_limit" in res.json()["detail"] or "segment" in res.json()["detail"]


async def test_export_blocks_consent_drift(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    cid = await _build_ready(client, h)
    await _attest_all(client, h, cid)
    # revoke consent 1 người SAU build (chưa rebuild) → drift fail-closed.
    async with AsyncSessionLocal() as s:
        row = (
            await s.execute(
                text("SELECT id FROM sms_contact WHERE phone_normalized=:p"),
                {"p": _PHONE_VIETTEL},
            )
        ).first()
    contact_id = row[0]
    rv = await client.post(
        f"{API}/contacts/{contact_id}/consent-events",
        json={
            "event_type": "revoked",
            "occurred_at": "2026-06-10T00:00:00+07:00",
            "revoke_source": "manual",
        },
        headers=h,
    )
    assert rv.status_code == 201, rv.text
    res = await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    assert res.status_code == 400, res.text
    assert "consent" in res.json()["detail"].lower()


# ---------------------------------------------------------------------
# Happy path + nội dung file + idempotent
# ---------------------------------------------------------------------
async def test_export_happy_path_per_carrier(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    await _seed_carriers()  # 098→viettel, 090→mobifone (trước build)
    cid = await _build_ready(client, h)
    await _attest_all(client, h, cid)
    res = await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    assert res.status_code == 201, res.text
    body = res.json()
    assert body["total_exportable"] == 2
    batches = body["batches"]
    assert len(batches) == 2
    assert {b["carrier_bucket"] for b in batches} == {"viettel", "mobifone"}
    for b in batches:
        assert b["status"] == "generated"
        assert b["recipient_count"] == 1
        assert b["can_download"] is True
        assert b["can_mark_handed_off"] is True
        assert b["file_name"].endswith(".xlsx")
    # campaign → exported
    assert (
        await client.get(f"{API}/campaigns/{cid}", headers=h)
    ).json()["status"] == "exported"

    # download viettel → xlsx đúng mẫu (row1 trống, A=84xxx text, B=[QC]+link)
    vb = next(b for b in batches if b["carrier_bucket"] == "viettel")
    dl = await client.get(
        f"{API}/campaigns/{cid}/exports/{vb['id']}/download", headers=h
    )
    assert dl.status_code == 200, dl.text
    assert dl.headers["content-type"].startswith(
        "application/vnd.openxmlformats"
    )
    ws = load_workbook(io.BytesIO(dl.content))["Sheet1"]
    assert ws.cell(1, 1).value is None  # row 1 trống
    assert ws.cell(2, 1).value == "84981234567"
    msg = ws.cell(2, 2).value
    assert msg.startswith("[QC]")
    assert "/r/" in msg  # sentinel link đã thay bằng URL thật
    assert "__SMS_LINK__" not in msg


async def test_export_idempotent(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    cid = await _build_ready(client, h)
    await _attest_all(client, h, cid)
    r1 = await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    assert r1.status_code == 201
    ids1 = sorted(b["id"] for b in r1.json()["batches"])
    r2 = await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    assert r2.status_code == 201
    ids2 = sorted(b["id"] for b in r2.json()["batches"])
    assert ids1 == ids2  # cùng batch, không tạo file/token mới
    assert all(b["status"] == "generated" for b in r2.json()["batches"])
    # GET list khớp
    lst = await client.get(f"{API}/campaigns/{cid}/exports", headers=h)
    assert lst.status_code == 200
    assert sorted(b["id"] for b in lst.json()["batches"]) == ids1


async def test_export_no_link_campaign(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    cid = await _build_ready(
        client, h, template="Chao {full_name} uu dai thang 6",
        phones=[_PHONE_VIETTEL],
    )
    await _attest_all(client, h, cid)
    res = await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    assert res.status_code == 201, res.text
    b = res.json()["batches"][0]
    dl = await client.get(
        f"{API}/campaigns/{cid}/exports/{b['id']}/download", headers=h
    )
    msg = load_workbook(io.BytesIO(dl.content))["Sheet1"].cell(2, 2).value
    assert msg.startswith("[QC]")
    assert "/r/" not in msg  # campaign không có {link}


# ---------------------------------------------------------------------
# mark-handed-off + chặn rebuild
# ---------------------------------------------------------------------
async def test_mark_handed_off_and_block_rebuild(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    cid = await _build_ready(client, h)
    await _attest_all(client, h, cid)
    batches = (
        await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    ).json()["batches"]
    for b in batches:
        r = await client.post(
            f"{API}/campaigns/{cid}/exports/{b['id']}/mark-handed-off",
            headers=h,
        )
        assert r.status_code == 200, r.text
        assert r.json()["status"] == "handed_off"
        assert r.json()["can_mark_handed_off"] is False
    # tất cả bàn giao → campaign closed
    assert (
        await client.get(f"{API}/campaigns/{cid}", headers=h)
    ).json()["status"] == "closed"
    # recipient.handed_off_at + contact.last_handed_off_at
    async with AsyncSessionLocal() as s:
        rc = (
            await s.execute(
                text(
                    "SELECT count(*) FROM sms_campaign_recipient "
                    "WHERE campaign_id=:c AND handed_off_at IS NOT NULL"
                ),
                {"c": cid},
            )
        ).scalar()
        assert rc == 2
        cc = (
            await s.execute(
                text(
                    "SELECT count(*) FROM sms_contact WHERE "
                    "last_handed_off_at IS NOT NULL AND "
                    "phone_normalized IN (:a, :b)"
                ),
                {"a": _PHONE_VIETTEL, "b": _PHONE_MOBI},
            )
        ).scalar()
        assert cc == 2
    # rebuild sau handoff → 400; export lại (closed) → 400
    assert (
        await client.post(f"{API}/campaigns/{cid}/build", headers=h)
    ).status_code == 400
    assert (
        await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    ).status_code == 400


async def test_download_unknown_batch_404(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    cid = await _build_ready(client, h)
    res = await client.get(
        f"{API}/campaigns/{cid}/exports/999999/download", headers=h
    )
    assert res.status_code == 404


# ---------------------------------------------------------------------
# Gate drift suppression + lifecycle guards (vá lỗ test /code-review)
# ---------------------------------------------------------------------
async def test_export_blocks_suppression_drift(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    cid = await _build_ready(client, h)
    await _attest_all(client, h, cid)
    # Số vào sms_opt_out SAU build (chưa rebuild) → gate count_new_suppression.
    async with AsyncSessionLocal() as s:
        await s.execute(
            text(
                "INSERT INTO sms_opt_out (phone_normalized, source, "
                "observed_at) VALUES (:p, 'manual', now())"
            ),
            {"p": _PHONE_VIETTEL},
        )
        await s.commit()
    res = await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    assert res.status_code == 400, res.text
    assert "opt-out" in res.json()["detail"].lower()


async def test_stale_revision_download_blocked_after_rebuild(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    cid = await _build_ready(client, h, phones=[_PHONE_VIETTEL])
    await _attest_all(client, h, cid)
    exp = (await client.post(f"{API}/campaigns/{cid}/export", headers=h)).json()
    old_batch = exp["batches"][0]["id"]
    # cùng revision → tải được
    assert (
        await client.get(
            f"{API}/campaigns/{cid}/exports/{old_batch}/download", headers=h
        )
    ).status_code == 200
    # rebuild → revision mới; batch cũ bị vô hiệu + chặn download
    assert (
        await client.post(f"{API}/campaigns/{cid}/build", headers=h)
    ).status_code == 200
    dl = await client.get(
        f"{API}/campaigns/{cid}/exports/{old_batch}/download", headers=h
    )
    assert dl.status_code == 400, dl.text
    async with AsyncSessionLocal() as s:
        row = (
            await s.execute(
                text(
                    "SELECT status, invalidated_at FROM "
                    "sms_campaign_export_batch WHERE id=:b"
                ),
                {"b": old_batch},
            )
        ).first()
    assert row[0] == "invalidated"
    assert row[1] is not None


async def test_download_expired_returns_400(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    cid = await _build_ready(client, h, phones=[_PHONE_VIETTEL])
    await _attest_all(client, h, cid)
    bid = (
        await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    ).json()["batches"][0]["id"]
    async with AsyncSessionLocal() as s:
        await s.execute(
            text(
                "UPDATE sms_campaign_export_batch SET expires_at = now() - "
                "interval '1 day' WHERE id=:b"
            ),
            {"b": bid},
        )
        await s.commit()
    dl = await client.get(
        f"{API}/campaigns/{cid}/exports/{bid}/download", headers=h
    )
    assert dl.status_code == 400, dl.text
    lst = (
        await client.get(f"{API}/campaigns/{cid}/exports", headers=h)
    ).json()
    b = next(x for x in lst["batches"] if x["id"] == bid)
    assert b["is_expired"] is True
    assert b["can_download"] is False


async def test_mark_handed_off_guards(
    client: AsyncClient, admin_token_headers: dict
):
    h = admin_token_headers
    cid = await _build_ready(client, h, phones=[_PHONE_VIETTEL])
    await _attest_all(client, h, cid)
    bid = (
        await client.post(f"{API}/campaigns/{cid}/export", headers=h)
    ).json()["batches"][0]["id"]
    base = f"{API}/campaigns/{cid}/exports"
    assert (
        await client.post(f"{base}/{bid}/mark-handed-off", headers=h)
    ).status_code == 200
    # lần 2 → 400 (đã bàn giao)
    assert (
        await client.post(f"{base}/{bid}/mark-handed-off", headers=h)
    ).status_code == 400
    # batch không tồn tại → 404
    assert (
        await client.post(f"{base}/999999/mark-handed-off", headers=h)
    ).status_code == 404
