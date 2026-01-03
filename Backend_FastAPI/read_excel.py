
import openpyxl
import json

file_path = r"d:\QLTS\Documents\Import_Thong_Tin_Hoc_Vien_Data_1767431894263.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if "Danh mục" in wb.sheetnames:
        ws = wb["Danh mục"]
        data = {}
        
        # Helper to read column data
        def read_column(col_letter):
            values = []
            for cell in ws[col_letter]:
                if cell.value:
                    values.append(str(cell.value).strip())
            return values[1:] # Skip header
            
        # Assuming typical structure, looking for headers to identify columns
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[cell.column_letter] = str(cell.value).strip()
        
        print("Headers found:", headers)
        
        for col, header in headers.items():
            data[header] = read_column(col)
            
        print(json.dumps(data, ensure_ascii=False, indent=2))
        
    else:
        print("Sheet 'Danh mục' not found")
        print("Available sheets:", wb.sheetnames)

except Exception as e:
    print(f"Error: {e}")
