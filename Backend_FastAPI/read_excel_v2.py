
import openpyxl
import json

file_path = r"d:\QLTS\Documents\Import_Thong_Tin_Hoc_Vien_Data_1767431894263.xlsx"

try:
    # Load workbook without read_only=True to avoid iter_cols issue if checking loosely
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Danh mục" in wb.sheetnames:
        ws = wb["Danh mục"]
        data = {}
        
        # Read headers from first row
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[cell.column_letter] = str(cell.value).strip()
        
        print("Headers mapped:", headers)
        
        # Create mapping of what we need
        # We need: Giới tính, Quốc tịch, Dân tộc, Tôn giáo, Loại khuyết tật
        target_cols = {
            "Giới tính": None,
            "Quốc tịch": None,
            "Dân tộc": None,
            "Tôn giáo": None,
            "Loại khuyết tật": None
        }
        
        for col_letter, header_name in headers.items():
            if header_name in target_cols:
                target_cols[header_name] = col_letter

        # Read data for targets
        for name, col in target_cols.items():
            if col:
                values = []
                # iterate rows starting from 2
                for row in range(2, ws.max_row + 1):
                    val = ws[f"{col}{row}"].value
                    if val:
                        values.append(str(val).strip())
                data[name] = list(set(values)) # unique
        
        print(json.dumps(data, ensure_ascii=False, indent=2))
        
    else:
        print("Sheet 'Danh mục' not found")

except Exception as e:
    print(f"Error: {e}")
