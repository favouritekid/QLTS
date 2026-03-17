#!/usr/bin/env python3
"""
Seed database from seed_data_template.xlsx

Each sheet has a dedicated parser that understands its structure:
- Row 1: headers
- Row 2: description row (skipped)
- Data rows: until "GHI CHÚ:" or empty
- Some sheets (15_NhomHoSo) have multi-section layout

Usage:
    python -m scripts.seed_from_xlsx                          # Seed all
    python -m scripts.seed_from_xlsx --sheets 12 13 14        # Specific sheets
    python -m scripts.seed_from_xlsx --dry-run                # Preview only
    python -m scripts.seed_from_xlsx --file /path/to/file.xlsx

Idempotent: uses ON CONFLICT DO NOTHING / skip-if-exists.
"""
import argparse
import asyncio
import json
import os
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

# Setup path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))
os.environ.setdefault("APP_ENV", "development")

from app.database import AsyncSessionLocal  # noqa: E402

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

LOAI_MAPPING = {
    "truong": "Trường",
    "phong": "Phòng ban",
    "khoa": "Khoa",
    "trung tam": "Trung tâm",
    "to": "Tổ",
    "bo mon": "Bộ môn",
}



# Sheets where row 2 is a description row (old-style sheets with [BẮT BUỘC]/[BAT BUOC] markers)
SHEETS_WITH_DESC_ROW = {
    "1_ToChuc", "2_TaiKhoan", "3_NganhHoc", "4_HinhThucDaoTao", "5_NamHoc",
    "6_MonThi", "7_ToHopMon", "8_PhuongThucXT", "9_TieuChi",
    "10_Pipeline", "10b_TrangThaiTuVan", "10c_ChuyenTrangThai", "11_DanhMuc",
}


def read_sheet_data(wb, sheet_name: str, *, skip_desc_row=None) -> list[dict]:
    """Read sheet into list of dicts. Auto-detects description row for old sheets."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

    # Auto-detect: if sheet_name in old-style set, or row 2 starts with "[BẮT BUỘC]"/"[BAT BUOC]"
    if skip_desc_row is None:
        if sheet_name in SHEETS_WITH_DESC_ROW:
            skip_desc_row = True
        elif len(rows) > 1 and rows[1][0] and str(rows[1][0]).startswith("["):
            skip_desc_row = True
        else:
            skip_desc_row = False

    start = 2 if skip_desc_row else 1  # skip description row

    result = []
    for row in rows[start:]:
        if row[0] is None:
            continue
        first_val = str(row[0]).strip()
        if first_val.startswith("GHI CHÚ") or first_val.startswith("["):
            continue
        if first_val.startswith("•"):
            continue
        record = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            record[h] = val
        result.append(record)
    return result


def yn_to_bool(val) -> bool:
    """Convert có/không/true/false to bool."""
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower() if val else ""
    return s in ("có", "co", "true", "1", "yes")


def to_decimal(val) -> Decimal | None:
    if val is None or val == "":
        return None
    return Decimal(str(val))


def to_int(val) -> int | None:
    if val is None or val == "":
        return None
    return int(float(val))


def to_str(val) -> str | None:
    if val is None or val == "":
        return None
    return str(val).strip()


def to_date(val) -> date | None:
    """Convert string/datetime to date object for asyncpg."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


async def upsert_sql(db: AsyncSession, table: str, data: dict, conflict_col: str):
    """INSERT ... ON CONFLICT (conflict_col) DO NOTHING. Returns True if inserted."""
    cols = list(data.keys())
    placeholders = ", ".join(f":{c}" for c in cols)
    col_names = ", ".join(f'"{c}"' for c in cols)
    sql = f'INSERT INTO "{table}" ({col_names}) VALUES ({placeholders}) ON CONFLICT ("{conflict_col}") DO NOTHING'
    result = await db.execute(text(sql), data)
    return result.rowcount > 0


async def upsert_multi_conflict(db: AsyncSession, table: str, data: dict, conflict_cols: list[str]):
    """INSERT ... ON CONFLICT (col1, col2) DO NOTHING."""
    cols = list(data.keys())
    placeholders = ", ".join(f":{c}" for c in cols)
    col_names = ", ".join(f'"{c}"' for c in cols)
    conflict = ", ".join(f'"{c}"' for c in conflict_cols)
    sql = f'INSERT INTO "{table}" ({col_names}) VALUES ({placeholders}) ON CONFLICT ({conflict}) DO NOTHING'
    result = await db.execute(text(sql), data)
    return result.rowcount > 0


async def lookup_id(db: AsyncSession, table: str, code_col: str, code_val: str) -> int | None:
    """Lookup ID by code column."""
    result = await db.execute(
        text(f'SELECT id FROM "{table}" WHERE "{code_col}" = :v'), {"v": code_val}
    )
    row = result.fetchone()
    return row[0] if row else None


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET PARSERS (one per sheet)
# ═══════════════════════════════════════════════════════════════════════════════

async def seed_12_config_degree_level(db: AsyncSession, wb, dry_run: bool) -> int:
    """12_CauHinhTrinhDo → config_degree_level"""
    rows = read_sheet_data(wb, "12_CauHinhTrinhDo")
    count = 0
    for r in rows:
        data = {
            "code": to_str(r["code"]),
            "name": to_str(r["name"]),
            "display_order": to_int(r.get("display_order", 0)),
            "is_active": True,
        }
        if not dry_run:
            if await upsert_sql(db, "config_degree_level", data, "code"):
                count += 1
        else:
            count += 1
    return count


async def seed_13_config_offering_type(db: AsyncSession, wb, dry_run: bool) -> int:
    """13_CauHinhLoaiHinh → config_offering_type"""
    rows = read_sheet_data(wb, "13_CauHinhLoaiHinh")
    count = 0
    for r in rows:
        data = {
            "code": to_str(r["code"]),
            "name": to_str(r["name"]),
            "display_order": to_int(r.get("display_order", 0)),
            "is_active": True,
        }
        if not dry_run:
            if await upsert_sql(db, "config_offering_type", data, "code"):
                count += 1
        else:
            count += 1
    return count


async def seed_14_config_document_type(db: AsyncSession, wb, dry_run: bool) -> int:
    """14_CauHinhLoaiTaiLieu → config_document_type"""
    rows = read_sheet_data(wb, "14_CauHinhLoaiTaiLieu")
    count = 0
    for r in rows:
        data = {
            "code": to_str(r["code"]),
            "name": to_str(r["name"]),
            "description": to_str(r.get("description")),
            "display_order": to_int(r.get("display_order", 0)),
            "is_active": True,
        }
        if not dry_run:
            if await upsert_sql(db, "config_document_type", data, "code"):
                count += 1
        else:
            count += 1
    return count


async def seed_11_danh_muc(db: AsyncSession, wb, dry_run: bool) -> int:
    """11_DanhMuc → config_system_category (handles trailing NONE rows after blank)"""
    ws = wb["11_DanhMuc"]
    rows_raw = list(ws.iter_rows(values_only=True))
    if len(rows_raw) < 2:
        return 0

    # Parse ALL rows, not stopping at blank — pick up NONE rows
    count = 0
    for row in rows_raw[2:]:  # skip header + description
        loai = to_str(row[0])
        ma = to_str(row[1])
        ten = to_str(row[2])
        if not loai or not ma or not ten:
            continue
        if loai.startswith("GHI CHÚ") or loai.startswith("•"):
            continue
        valid_types = {"nationality", "ethnicity", "religion", "disability_type", "family_type"}
        if loai not in valid_types:
            continue
        data = {
            "type": loai,
            "code": ma,
            "name": ten,
            "display_order": to_int(row[3]) if len(row) > 3 and row[3] is not None else 0,
            "is_active": True,
        }
        if not dry_run:
            if await upsert_multi_conflict(db, "config_system_category", data, ["type", "code"]):
                count += 1
        else:
            count += 1
    return count


async def seed_10_pipeline(db: AsyncSession, wb, dry_run: bool) -> int:
    """10_Pipeline → pipeline_stage. Skips if table already has data (avoids FK conflicts)."""
    if not dry_run:
        result = await db.execute(text("SELECT COUNT(*) FROM pipeline_stage"))
        existing = result.scalar()
        if existing > 0:
            print(f"    pipeline_stage already has {existing} rows — skipping (delete first to re-seed)")
            return 0
    rows = read_sheet_data(wb, "10_Pipeline")
    count = 0
    for r in rows:
        # Headers from user's custom sheet
        ma = to_str(r.get("ma_giai_doan") or r.get(list(r.keys())[0]))
        ten = to_str(r.get("ten_giai_doan") or r.get(list(r.keys())[1]))
        thu_tu = to_int(r.get("thu_tu") or r.get(list(r.keys())[2]))
        is_final = yn_to_bool(r.get("la_giai_doan_cuoi") or r.get(list(r.keys())[3]))
        color = to_str(r.get("ma_mau") or r.get(list(r.keys())[4])) or "#6B7280"

        if not ma or not ten:
            continue
        if not dry_run:
            # pipeline_stage has unique constraints on both id AND order
            result = await db.execute(
                text('SELECT id FROM pipeline_stage WHERE id = :id OR "order" = :ord'),
                {"id": ma, "ord": thu_tu}
            )
            if result.fetchone():
                continue
            await db.execute(
                text('INSERT INTO pipeline_stage (id, name, "order", is_final_stage, color_code) '
                     'VALUES (:id, :name, :ord, :final, :color)'),
                {"id": ma, "name": ten, "ord": thu_tu, "final": is_final, "color": color}
            )
            count += 1
        else:
            count += 1
    return count


async def seed_01_to_chuc(db: AsyncSession, wb, dry_run: bool) -> int:
    """1_ToChuc → organization_unit (hierarchical, maps loai values)"""
    rows = read_sheet_data(wb, "1_ToChuc")
    # First pass: create all units without parent_id
    name_to_id = {}
    count = 0

    for r in rows:
        ten = to_str(r.get("ten") or r.get(list(r.keys())[0]))
        loai_raw = to_str(r.get("loai") or r.get(list(r.keys())[1]))
        mo_ta = to_str(r.get("mo_ta") or r.get(list(r.keys())[2]))
        if not ten:
            continue
        loai = LOAI_MAPPING.get(loai_raw.lower().strip(), loai_raw) if loai_raw else "Phòng ban"

        # Check if exists
        result = await db.execute(
            text('SELECT id FROM organization_unit WHERE name = :n'), {"n": ten}
        )
        existing = result.fetchone()
        if existing:
            name_to_id[ten] = existing[0]
            continue

        data = {"name": ten, "type": loai, "description": mo_ta, "is_active": True}
        if not dry_run:
            result = await db.execute(
                text('INSERT INTO organization_unit (name, type, description, is_active, created_at, updated_at) '
                     'VALUES (:name, :type, :description, :is_active, NOW(), NOW()) RETURNING id'),
                data
            )
            row_id = result.fetchone()[0]
            name_to_id[ten] = row_id
            count += 1
        else:
            count += 1

    # Second pass: set parent_id
    if not dry_run:
        for r in rows:
            ten = to_str(r.get("ten") or r.get(list(r.keys())[0]))
            cha = to_str(r.get("don_vi_cha") or r.get(list(r.keys())[3]))
            if ten and cha and cha in name_to_id and ten in name_to_id:
                await db.execute(
                    text('UPDATE organization_unit SET parent_id = :pid WHERE id = :id AND parent_id IS NULL'),
                    {"pid": name_to_id[cha], "id": name_to_id[ten]}
                )

    return count


async def seed_02_tai_khoan(db: AsyncSession, wb, dry_run: bool) -> int:
    """2_TaiKhoan → user + user_unit_assignment + casbin_rule"""
    from passlib.context import CryptContext
    pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto", bcrypt__rounds=12)

    rows = read_sheet_data(wb, "2_TaiKhoan")
    count = 0

    for r in rows:
        username = to_str(r.get("username") or r.get(list(r.keys())[0]))
        email = to_str(r.get("email") or r.get(list(r.keys())[1]))
        password = to_str(r.get("mat_khau") or r.get(list(r.keys())[2]))
        full_name = to_str(r.get("ho_ten") or r.get(list(r.keys())[3]))
        role = to_str(r.get("vai_tro") or r.get(list(r.keys())[4]))
        unit_name = to_str(r.get("don_vi") or r.get(list(r.keys())[5]))
        phone = to_str(r.get("so_dien_thoai") or (r.get(list(r.keys())[6]) if len(r) > 6 else None))

        if not username or not email:
            continue

        # Check if exists
        result = await db.execute(text('SELECT id FROM "user" WHERE username = :u'), {"u": username})
        if result.fetchone():
            continue

        # Lookup unit
        unit_id = None
        if unit_name:
            result = await db.execute(
                text("SELECT id FROM organization_unit WHERE name = :n"), {"n": unit_name}
            )
            row = result.fetchone()
            unit_id = row[0] if row else None

        if not dry_run:
            pw_hash = pwd_ctx.hash(password) if password else pwd_ctx.hash("Default@123")
            result = await db.execute(
                text('''INSERT INTO "user" (username, email, password_hash, full_name, role, unit_id, phone_number, status, total_lead_score, max_capacity, availability_status)
                     VALUES (:username, :email, :pw_hash, :full_name, :role, :unit_id, :phone, 'active', 0, 100, 'available')
                     ON CONFLICT (username) DO NOTHING RETURNING id'''),
                {"username": username, "email": email, "pw_hash": pw_hash,
                 "full_name": full_name, "role": role or "user", "unit_id": unit_id, "phone": phone}
            )
            row = result.fetchone()
            if row:
                user_id = row[0]
                # user_unit_assignment
                if unit_id:
                    await db.execute(
                        text('''INSERT INTO user_unit_assignment (user_id, unit_id, role, is_active, start_date, created_at, updated_at)
                             VALUES (:uid, :unit_id, :role, true, NOW(), NOW(), NOW())
                             ON CONFLICT DO NOTHING'''),
                        {"uid": user_id, "unit_id": unit_id, "role": role or "user"}
                    )
                # casbin_rule
                await db.execute(
                    text('''INSERT INTO casbin_rule (ptype, v0, v1)
                         VALUES ('g', :v0, :v1) ON CONFLICT DO NOTHING'''),
                    {"v0": f"user:{user_id}", "v1": f"role:{role or 'user'}"}
                )
                count += 1
        else:
            count += 1

    return count


async def seed_10b_trang_thai(db: AsyncSession, wb, dry_run: bool) -> int:
    """10b_TrangThaiTuVan → consultation_status. Skips if table already has data."""
    if not dry_run:
        result = await db.execute(text("SELECT COUNT(*) FROM consultation_status"))
        existing = result.scalar()
        if existing > 0:
            print(f"    consultation_status already has {existing} rows — skipping")
            return 0
    rows = read_sheet_data(wb, "10b_TrangThaiTuVan")
    count = 0
    for r in rows:
        keys = list(r.keys())
        ma = to_str(r.get(keys[0]))
        code = to_str(r.get(keys[1]))
        ten = to_str(r.get(keys[2]))
        color = to_str(r.get(keys[3])) or "#6B7280"
        stage_id = to_str(r.get(keys[4])) or None
        phase = to_str(r.get(keys[5])) or "consultation"
        status_type = to_str(r.get(keys[6])) or "transition"
        selectable = to_str(r.get(keys[7])) or "user"
        outcome = to_str(r.get(keys[8])) or "neutral"
        is_final = yn_to_bool(r.get(keys[9]))
        is_universal = yn_to_bool(r.get(keys[10]))
        updates_pipeline = yn_to_bool(r.get(keys[11]))
        counts_funnel = yn_to_bool(r.get(keys[12]))
        display_order = to_int(r.get(keys[13])) or 0
        description = to_str(r.get(keys[14])) if len(keys) > 14 else None
        legacy_status = to_str(r.get(keys[15])) if len(keys) > 15 else None

        if not ma or not ten:
            continue

        if not dry_run:
            # consultation_status has unique on both id AND code
            result = await db.execute(
                text('SELECT id FROM consultation_status WHERE id = :id OR code = :code'),
                {"id": ma, "code": code}
            )
            if result.fetchone():
                continue
            await db.execute(
                text('''INSERT INTO consultation_status
                     (id, code, name, color_code, stage_id, phase, status_type,
                      selectable_mode, outcome_type, is_final, is_universal,
                      updates_pipeline, counts_for_funnel, display_order, description,
                      legacy_status)
                     VALUES (:id, :code, :name, :color, :stage_id, :phase, :stype,
                             :smode, :outcome, :final, :univ, :upipe, :cfunnel, :dord, :desc,
                             :legacy)'''),
                {"id": ma, "code": code, "name": ten, "color": color,
                 "stage_id": stage_id, "phase": phase, "stype": status_type,
                 "smode": selectable, "outcome": outcome,
                 "final": is_final, "univ": is_universal,
                 "upipe": updates_pipeline, "cfunnel": counts_funnel,
                 "dord": display_order, "desc": description,
                 "legacy": legacy_status}
            )
            count += 1
        else:
            count += 1
    return count


async def seed_10c_transitions(db: AsyncSession, wb, dry_run: bool) -> int:
    """10c_ChuyenTrangThai → allowed_transitions. Skips if table already has data."""
    if not dry_run:
        result = await db.execute(text("SELECT COUNT(*) FROM allowed_transitions"))
        existing = result.scalar()
        if existing > 0:
            print(f"    allowed_transitions already has {existing} rows — skipping")
            return 0
    rows = read_sheet_data(wb, "10c_ChuyenTrangThai")
    count = 0
    for r in rows:
        keys = list(r.keys())
        from_id = to_str(r.get(keys[0]))
        to_id = to_str(r.get(keys[1]))
        trigger = to_str(r.get(keys[2])) or "user"
        req_phase = to_str(r.get(keys[3])) if len(keys) > 3 else None
        desc = to_str(r.get(keys[4])) if len(keys) > 4 else None

        if not from_id or not to_id:
            continue

        if not dry_run:
            # No unique constraint on (from_status_id, to_status_id) — check manually
            result = await db.execute(
                text('SELECT id FROM allowed_transitions WHERE from_status_id = :f AND to_status_id = :t'),
                {"f": from_id, "t": to_id}
            )
            if result.fetchone():
                continue
            await db.execute(
                text('''INSERT INTO allowed_transitions
                     (from_status_id, to_status_id, trigger_type, required_phase, is_active, description, created_at, updated_at)
                     VALUES (:f, :t, :trigger, :phase, true, :desc, NOW(), NOW())'''),
                {"f": from_id, "t": to_id, "trigger": trigger, "phase": req_phase, "desc": desc}
            )
            count += 1
        else:
            count += 1
    return count


async def seed_06_mon_thi(db: AsyncSession, wb, dry_run: bool) -> int:
    """6_MonThi → subject"""
    rows = read_sheet_data(wb, "6_MonThi")
    count = 0
    for r in rows:
        keys = list(r.keys())
        code = to_str(r.get(keys[0]))
        name = to_str(r.get(keys[1]))
        order = to_int(r.get(keys[2])) or 0
        if not code or not name:
            continue
        data = {"code": code, "name_vi": name, "display_order": order, "is_active": True}
        if not dry_run:
            if await upsert_sql(db, "subject", data, "code"):
                count += 1
        else:
            count += 1
    return count


async def seed_07_to_hop_mon(db: AsyncSession, wb, dry_run: bool) -> int:
    """7_ToHopMon → subject_group + subject_group_subject"""
    rows = read_sheet_data(wb, "7_ToHopMon")
    count = 0
    for idx, r in enumerate(rows):
        keys = list(r.keys())
        code = to_str(r.get(keys[0]))
        name = to_str(r.get(keys[1]))
        mon1 = to_str(r.get(keys[2]))
        mon2 = to_str(r.get(keys[3]))
        mon3 = to_str(r.get(keys[4]))
        if not code or not name:
            continue

        if not dry_run:
            # subject_group
            result = await db.execute(
                text('INSERT INTO subject_group (code, name, display_order, is_active) '
                     'VALUES (:code, :name, :order, true) ON CONFLICT (code) DO NOTHING RETURNING id'),
                {"code": code, "name": name, "order": idx + 1}
            )
            row = result.fetchone()
            if row:
                group_id = row[0]
                count += 1
                # subject_group_subject
                for pos, mon_code in enumerate([mon1, mon2, mon3], 1):
                    if not mon_code:
                        continue
                    subj_id = await lookup_id(db, "subject", "code", mon_code)
                    if subj_id:
                        await db.execute(
                            text('INSERT INTO subject_group_subject (subject_group_id, subject_id, "position") '
                                 'VALUES (:gid, :sid, :pos) ON CONFLICT DO NOTHING'),
                            {"gid": group_id, "sid": subj_id, "pos": pos}
                        )
        else:
            count += 1
    return count


async def seed_08_phuong_thuc(db: AsyncSession, wb, dry_run: bool) -> int:
    """8_PhuongThucXT → admission_method"""
    rows = read_sheet_data(wb, "8_PhuongThucXT")
    count = 0
    for idx, r in enumerate(rows):
        keys = list(r.keys())
        code = to_str(r.get(keys[0]))
        name = to_str(r.get(keys[1]))
        desc = to_str(r.get(keys[2]))
        req_gpa = yn_to_bool(r.get(keys[3]))
        req_scores = yn_to_bool(r.get(keys[4]))
        if not code or not name:
            continue
        data = {
            "code": code, "name": name, "description": desc,
            "requires_gpa": req_gpa, "requires_subject_scores": req_scores,
            "display_order": idx + 1, "is_active": True,
        }
        if not dry_run:
            if await upsert_sql(db, "admission_method", data, "code"):
                count += 1
        else:
            count += 1
    return count


async def seed_09_tieu_chi(db: AsyncSession, wb, dry_run: bool) -> int:
    """9_TieuChi → admission_criteria + criteria_subject_group"""
    rows = read_sheet_data(wb, "9_TieuChi")
    count = 0
    for r in rows:
        keys = list(r.keys())
        code = to_str(r.get(keys[0]))
        name = to_str(r.get(keys[1]))
        method_code = to_str(r.get(keys[2]))
        min_gpa = to_decimal(r.get(keys[3]))
        min_score = to_decimal(r.get(keys[4]))
        subj_count = to_int(r.get(keys[5]))
        scoring = to_str(r.get(keys[6])) or "sum"
        max_score = to_decimal(r.get(keys[7]))
        groups_str = to_str(r.get(keys[8]))

        if not code or not name or not method_code:
            continue

        method_id = await lookup_id(db, "admission_method", "code", method_code) if not dry_run else 1
        if not method_id and not dry_run:
            print(f"    WARNING: method '{method_code}' not found for criteria '{code}'")
            continue

        if not dry_run:
            result = await db.execute(
                text('''INSERT INTO admission_criteria
                     (method_id, code, name, min_gpa, min_score, required_subject_count,
                      subject_selection_mode, scoring_method, max_possible_score, policy_version, is_active)
                     VALUES (:mid, :code, :name, :min_gpa, :min_score, :subj_count,
                             'fixed', :scoring, :max_score, '2026.1', true)
                     ON CONFLICT (code) DO NOTHING RETURNING id'''),
                {"mid": method_id, "code": code, "name": name, "min_gpa": min_gpa,
                 "min_score": min_score, "subj_count": subj_count,
                 "scoring": scoring, "max_score": max_score}
            )
            row = result.fetchone()
            if row:
                criteria_id = row[0]
                count += 1
            else:
                # Already exists — look up ID for backfill
                criteria_id = await lookup_id(db, "admission_criteria", "code", code)

            # Always backfill criteria_subject_group (idempotent via ON CONFLICT)
            if criteria_id and groups_str:
                for g_code in groups_str.split(","):
                    g_code = g_code.strip()
                    if not g_code:
                        continue
                    g_id = await lookup_id(db, "subject_group", "code", g_code)
                    if g_id:
                        await db.execute(
                            text('INSERT INTO criteria_subject_group (criteria_id, subject_group_id) '
                                 'VALUES (:cid, :gid) ON CONFLICT DO NOTHING'),
                            {"cid": criteria_id, "gid": g_id}
                        )
        else:
            count += 1
    return count


async def seed_03_nganh_hoc(db: AsyncSession, wb, dry_run: bool) -> int:
    """3_NganhHoc → major_program"""
    rows = read_sheet_data(wb, "3_NganhHoc")
    count = 0
    for r in rows:
        keys = list(r.keys())
        ma = to_str(r.get(keys[0]))
        ten = to_str(r.get(keys[1]))
        bac = to_str(r.get(keys[2]))
        khoa = to_str(r.get(keys[3]))
        nang = yn_to_bool(r.get(keys[4])) if len(keys) > 4 else False
        if not ma or not ten:
            continue

        unit_id = None
        if khoa and not dry_run:
            result = await db.execute(
                text("SELECT id FROM organization_unit WHERE name = :n"), {"n": khoa}
            )
            row = result.fetchone()
            unit_id = row[0] if row else None
            if not unit_id:
                print(f"    ERROR: unit '{khoa}' not found for major '{ma}' — skipping")
                continue

        data = {
            "code": ma, "name": ten, "degree_level": bac or "Cao đẳng",
            "unit_id": unit_id, "is_active": True, "is_heavy": nang,
        }
        if not dry_run:
            if await upsert_sql(db, "major_program", data, "code"):
                count += 1
        else:
            count += 1
    return count


async def seed_04_hinh_thuc(db: AsyncSession, wb, dry_run: bool) -> int:
    """4_HinhThucDaoTao → program_offering"""
    rows = read_sheet_data(wb, "4_HinhThucDaoTao")
    count = 0
    for r in rows:
        keys = list(r.keys())
        ma_nganh = to_str(r.get(keys[0]))
        hinh_thuc = to_str(r.get(keys[1]))
        so_ky = to_int(r.get(keys[2]))
        so_tc = to_int(r.get(keys[3]))
        if not ma_nganh or not hinh_thuc:
            continue

        program_id = await lookup_id(db, "major_program", "code", ma_nganh) if not dry_run else 1
        if not program_id and not dry_run:
            print(f"    WARNING: major_program '{ma_nganh}' not found")
            continue

        # Map offering_type to config_offering_type
        ot_map = {"Chính quy": "chinh_quy", "Liên thông": "lien_thong",
                  "Vừa làm vừa học": "vua_lam_vua_hoc", "Từ xa": "tu_xa"}
        ot_code = ot_map.get(hinh_thuc, hinh_thuc.lower().replace(" ", "_"))
        ot_id = await lookup_id(db, "config_offering_type", "code", ot_code) if not dry_run else None

        if not dry_run:
            result = await db.execute(
                text('''INSERT INTO program_offering (offering_type, program_id, offering_type_id, is_active, duration_semesters, total_credits)
                     VALUES (:ot, :pid, :otid, true, :dur, :tc)
                     ON CONFLICT (program_id, offering_type) DO NOTHING RETURNING id'''),
                {"ot": hinh_thuc, "pid": program_id, "otid": ot_id, "dur": so_ky, "tc": so_tc}
            )
            if result.fetchone():
                count += 1
        else:
            count += 1
    return count


async def seed_05_nam_hoc(db: AsyncSession, wb, dry_run: bool) -> int:
    """5_NamHoc → offering_academic_info"""
    rows = read_sheet_data(wb, "5_NamHoc")
    count = 0
    for r in rows:
        keys = list(r.keys())
        ma_nganh = to_str(r.get(keys[0]))
        hinh_thuc = to_str(r.get(keys[1]))
        nam = to_int(r.get(keys[2]))
        hoc_phi = to_decimal(r.get(keys[3]))
        chi_tieu = to_int(r.get(keys[4]))
        diem_chuan = to_decimal(r.get(keys[5])) if len(keys) > 5 else None
        doi_tuong = to_str(r.get(keys[6])) if len(keys) > 6 else None
        published = yn_to_bool(r.get(keys[7])) if len(keys) > 7 else True

        if not ma_nganh or not nam:
            continue

        if not dry_run:
            # Find offering_id via program_id + offering_type
            result = await db.execute(
                text('''SELECT po.id FROM program_offering po
                     JOIN major_program mp ON po.program_id = mp.id
                     WHERE mp.code = :code AND po.offering_type = :ot'''),
                {"code": ma_nganh, "ot": hinh_thuc}
            )
            row = result.fetchone()
            if not row:
                continue
            offering_id = row[0]

            result = await db.execute(
                text('''INSERT INTO offering_academic_info
                     (offering_id, academic_year, tuition_fee_per_year, annual_admission_quota,
                      cutoff_score_previous_year, target_audience, is_published, created_at, updated_at)
                     VALUES (:oid, :year, :fee, :quota, :cutoff, :audience, :pub, NOW(), NOW())
                     ON CONFLICT (offering_id, academic_year) DO NOTHING RETURNING id'''),
                {"oid": offering_id, "year": nam, "fee": hoc_phi, "quota": chi_tieu,
                 "cutoff": diem_chuan, "audience": doi_tuong, "pub": published}
            )
            if result.fetchone():
                count += 1
        else:
            count += 1
    return count


async def seed_15_nhom_ho_so(db: AsyncSession, wb, dry_run: bool) -> int:
    """15_NhomHoSo → document_group + document_group_item (multi-section parser)"""
    ws = wb["15_NhomHoSo"]
    rows = list(ws.iter_rows(values_only=True))

    # Parse two sections
    section = None
    groups = []
    items = []

    for row in rows:
        first = to_str(row[0]) if row[0] else ""
        if "PHẦN A" in first:
            section = "A"
            continue
        if "PHẦN B" in first:
            section = "B"
            continue
        if first.startswith("group_code") or first.startswith("GHI") or not first:
            continue

        if section == "A":
            groups.append({
                "code": to_str(row[0]), "name": to_str(row[1]),
                "offering_type_code": to_str(row[2]),
                "admission_method_code": to_str(row[3]) or None,
            })
        elif section == "B":
            items.append({
                "group_code": to_str(row[0]),
                "document_type_code": to_str(row[1]),
                "is_mandatory": yn_to_bool(row[2]) if row[2] is not None else True,
                "display_order": to_int(row[3]) or 0,
                "requires_upload": yn_to_bool(row[4]) if row[4] is not None else True,
                "submission_format": to_str(row[5]),
            })

    count = 0
    if dry_run:
        return len(groups) + len(items)

    group_code_to_id = {}

    for g in groups:
        ot_id = await lookup_id(db, "config_offering_type", "code", g["offering_type_code"])
        method_id = None
        if g["admission_method_code"]:
            method_id = await lookup_id(db, "admission_method", "code", g["admission_method_code"])

        result = await db.execute(
            text('''INSERT INTO document_group (offering_type_id, admission_method_id, code, name, is_active)
                 VALUES (:otid, :mid, :code, :name, true)
                 ON CONFLICT (code) DO NOTHING RETURNING id'''),
            {"otid": ot_id, "mid": method_id, "code": g["code"], "name": g["name"]}
        )
        row = result.fetchone()
        if row:
            group_code_to_id[g["code"]] = row[0]
            count += 1
        else:
            # Already exists, look up ID
            gid = await lookup_id(db, "document_group", "code", g["code"])
            if gid:
                group_code_to_id[g["code"]] = gid

    for item in items:
        gid = group_code_to_id.get(item["group_code"])
        if not gid:
            continue
        dt_id = await lookup_id(db, "config_document_type", "code", item["document_type_code"])
        if not dt_id:
            print(f"    WARNING: document_type '{item['document_type_code']}' not found")
            continue
        result = await db.execute(
            text('''INSERT INTO document_group_item
                 (group_id, document_type_id, is_mandatory, display_order, requires_upload, submission_format)
                 VALUES (:gid, :dtid, :mand, :ord, :upload, :fmt)
                 ON CONFLICT (group_id, document_type_id) DO NOTHING'''),
            {"gid": gid, "dtid": dt_id, "mand": item["is_mandatory"],
             "ord": item["display_order"], "upload": item["requires_upload"],
             "fmt": item["submission_format"]}
        )
        if result.rowcount > 0:
            count += 1
    return count


async def seed_16_duong_dan(db: AsyncSession, wb, dry_run: bool) -> int:
    """16_DuongDanTuyenSinh → admission_path (+ auto-derive offering_admission_config)"""
    rows = read_sheet_data(wb, "16_DuongDanTuyenSinh")
    count = 0

    for r in rows:
        keys = list(r.keys())
        ma_nganh = to_str(r.get(keys[0]))
        hinh_thuc = to_str(r.get(keys[1]))
        nam_hoc = to_int(r.get(keys[2]))
        method_code = to_str(r.get(keys[3]))
        criteria_code = to_str(r.get(keys[4])) or None
        status = to_str(r.get(keys[5])) or "draft"
        display_name = to_str(r.get(keys[6]))
        visibility = to_str(r.get(keys[7])) or "internal"
        app_fee = to_decimal(r.get(keys[8]))

        if not ma_nganh or not method_code:
            continue

        if dry_run:
            count += 1
            continue

        # Resolve academic_info_id
        result = await db.execute(
            text('''SELECT oai.id FROM offering_academic_info oai
                 JOIN program_offering po ON oai.offering_id = po.id
                 JOIN major_program mp ON po.program_id = mp.id
                 WHERE mp.code = :code AND po.offering_type = :ot AND oai.academic_year = :year'''),
            {"code": ma_nganh, "ot": hinh_thuc, "year": nam_hoc}
        )
        ai_row = result.fetchone()
        if not ai_row:
            continue
        academic_info_id = ai_row[0]

        method_id = await lookup_id(db, "admission_method", "code", method_code)
        if not method_id:
            continue

        criteria_id = None
        if criteria_code:
            criteria_id = await lookup_id(db, "admission_criteria", "code", criteria_code)

        result = await db.execute(
            text('''INSERT INTO admission_path
                 (academic_info_id, admission_method_id, criteria_id, status,
                  display_name, display_order, visibility, application_fee, created_at, updated_at)
                 VALUES (:aiid, :mid, :cid, :status, :dn, 0, :vis, :fee, NOW(), NOW())
                 ON CONFLICT (academic_info_id, admission_method_id) DO NOTHING RETURNING id'''),
            {"aiid": academic_info_id, "mid": method_id, "cid": criteria_id,
             "status": status, "dn": display_name, "vis": visibility, "fee": app_fee}
        )
        if result.fetchone():
            count += 1

    # Auto-derive offering_admission_config (same logic as seed20260214001)
    if not dry_run:
        await db.execute(text('''
            INSERT INTO offering_admission_config (academic_info_id, criteria_id, is_active, created_at)
            SELECT DISTINCT ap.academic_info_id, ap.criteria_id, true, NOW()
            FROM admission_path ap
            WHERE ap.criteria_id IS NOT NULL
              AND NOT EXISTS (
                  SELECT 1 FROM offering_admission_config oac
                  WHERE oac.academic_info_id = ap.academic_info_id
                    AND oac.criteria_id = ap.criteria_id
              )
        '''))

    return count


async def seed_17_phuong_thuc_tt(db: AsyncSession, wb, dry_run: bool) -> int:
    """17_PhuongThucTT → payment_method"""
    rows = read_sheet_data(wb, "17_PhuongThucTT")
    count = 0
    for r in rows:
        data = {
            "code": to_str(r["code"]),
            "name": to_str(r["name"]),
            "is_online": yn_to_bool(r.get("is_online", False)),
            "requires_verification": yn_to_bool(r.get("requires_verification", True)),
            "gateway_code": to_str(r.get("gateway_code")) or None,
            "display_order": to_int(r.get("display_order", 0)),
            "is_active": True,
        }
        if not dry_run:
            if await upsert_sql(db, "payment_method", data, "code"):
                count += 1
        else:
            count += 1
    return count


async def seed_18_ke_hoach_tra_gop(db: AsyncSession, wb, dry_run: bool) -> int:
    """18_KeHoachTraGop → installment_plan"""
    rows = read_sheet_data(wb, "18_KeHoachTraGop")
    count = 0
    for r in rows:
        schedule_raw = to_str(r.get("schedule"))
        # Validate JSON
        try:
            schedule = json.loads(schedule_raw) if schedule_raw else []
        except json.JSONDecodeError:
            print(f"    WARNING: invalid JSON in schedule for '{r.get('code')}'")
            continue

        data = {
            "code": to_str(r["code"]),
            "name": to_str(r["name"]),
            "installment_count": to_int(r.get("installment_count", 1)),
            "penalty_type": to_str(r.get("penalty_type")) or "percentage",
            "penalty_rate": to_decimal(r.get("penalty_rate")) or Decimal("0"),
            "grace_period_days": to_int(r.get("grace_period_days")) or 7,
            "schedule": json.dumps(schedule),
            "is_active": True,
        }
        if not dry_run:
            # Cast schedule to jsonb
            result = await db.execute(
                text('''INSERT INTO installment_plan
                     (code, name, installment_count, penalty_type, penalty_rate,
                      grace_period_days, schedule, is_active)
                     VALUES (:code, :name, :installment_count, :penalty_type, :penalty_rate,
                             :grace_period_days, CAST(:schedule AS jsonb), :is_active)
                     ON CONFLICT (code) DO NOTHING'''),
                data
            )
            if result.rowcount > 0:
                count += 1
        else:
            count += 1
    return count


async def seed_19_lich_nghi_le(db: AsyncSession, wb, dry_run: bool) -> int:
    """19_LichNghiLe → holiday_calendar"""
    rows = read_sheet_data(wb, "19_LichNghiLe")
    count = 0
    for r in rows:
        date_val = to_date(r.get("date"))
        if not date_val:
            continue
        data = {
            "date": date_val,
            "name": to_str(r["name"]),
            "year": to_int(r.get("year", 2026)),
            "is_recurring": yn_to_bool(r.get("is_recurring", False)),
        }
        if not dry_run:
            result = await db.execute(
                text('''INSERT INTO holiday_calendar (date, name, year, is_recurring)
                     VALUES (:date, :name, :year, :is_recurring)
                     ON CONFLICT (date) DO NOTHING'''),
                data
            )
            if result.rowcount > 0:
                count += 1
        else:
            count += 1
    return count


async def seed_20_chinh_sach_giam(db: AsyncSession, wb, dry_run: bool) -> int:
    """20_ChinhSachGiam → tuition_discount_policy"""
    rows = read_sheet_data(wb, "20_ChinhSachGiam")
    count = 0
    for r in rows:
        scope_raw = to_str(r.get("applicable_scope")) or "{}"
        criteria_raw = to_str(r.get("target_criteria")) or "{}"
        try:
            json.loads(scope_raw)
            json.loads(criteria_raw)
        except json.JSONDecodeError:
            print(f"    WARNING: invalid JSON in policy '{r.get('code')}'")
            continue

        data = {
            "code": to_str(r["code"]),
            "name": to_str(r["name"]),
            "description": to_str(r.get("description")),
            "discount_type": to_str(r.get("discount_type", "percentage")),
            "discount_value": to_decimal(r.get("discount_value", 0)),
            "valid_from": to_date(r.get("valid_from")),
            "valid_to": to_date(r.get("valid_to")),
            "applicable_scope": scope_raw,
            "target_criteria": criteria_raw,
            "is_stackable": yn_to_bool(r.get("is_stackable", False)),
            "priority": to_int(r.get("priority", 0)),
            "max_usage": to_int(r.get("max_usage")),
            "current_usage": 0,
            "is_active": True,
        }
        if not dry_run:
            result = await db.execute(
                text('''INSERT INTO tuition_discount_policy
                     (code, name, description, discount_type, discount_value,
                      valid_from, valid_to, applicable_scope, target_criteria,
                      is_stackable, priority, max_usage, current_usage, is_active,
                      created_at, updated_at)
                     VALUES (:code, :name, :description, :discount_type, :discount_value,
                             :valid_from, :valid_to,
                             CAST(:applicable_scope AS jsonb), CAST(:target_criteria AS jsonb),
                             :is_stackable, :priority, :max_usage, :current_usage, :is_active,
                             NOW(), NOW())
                     ON CONFLICT (code) DO NOTHING'''),
                data
            )
            if result.rowcount > 0:
                count += 1
        else:
            count += 1
    return count


# ═══════════════════════════════════════════════════════════════════════════════
# REGISTRY: sheet number → (table, parser, dependencies)
# ═══════════════════════════════════════════════════════════════════════════════

SEED_REGISTRY = [
    # (sheet_prefix, display_name, parser_fn, sheet_name_in_workbook)
    ("12", "config_degree_level", seed_12_config_degree_level, "12_CauHinhTrinhDo"),
    ("13", "config_offering_type", seed_13_config_offering_type, "13_CauHinhLoaiHinh"),
    ("14", "config_document_type", seed_14_config_document_type, "14_CauHinhLoaiTaiLieu"),
    ("11", "config_system_category", seed_11_danh_muc, "11_DanhMuc"),
    ("10", "pipeline_stage", seed_10_pipeline, "10_Pipeline"),
    ("01", "organization_unit", seed_01_to_chuc, "1_ToChuc"),
    ("02", "user + casbin", seed_02_tai_khoan, "2_TaiKhoan"),
    ("10b", "consultation_status", seed_10b_trang_thai, "10b_TrangThaiTuVan"),
    ("10c", "allowed_transitions", seed_10c_transitions, "10c_ChuyenTrangThai"),
    ("06", "subject", seed_06_mon_thi, "6_MonThi"),
    ("07", "subject_group", seed_07_to_hop_mon, "7_ToHopMon"),
    ("08", "admission_method", seed_08_phuong_thuc, "8_PhuongThucXT"),
    ("09", "admission_criteria", seed_09_tieu_chi, "9_TieuChi"),
    ("03", "major_program", seed_03_nganh_hoc, "3_NganhHoc"),
    ("04", "program_offering", seed_04_hinh_thuc, "4_HinhThucDaoTao"),
    ("05", "offering_academic_info", seed_05_nam_hoc, "5_NamHoc"),
    ("15", "document_group", seed_15_nhom_ho_so, "15_NhomHoSo"),
    ("16", "admission_path", seed_16_duong_dan, "16_DuongDanTuyenSinh"),
    ("17", "payment_method", seed_17_phuong_thuc_tt, "17_PhuongThucTT"),
    ("18", "installment_plan", seed_18_ke_hoach_tra_gop, "18_KeHoachTraGop"),
    ("19", "holiday_calendar", seed_19_lich_nghi_le, "19_LichNghiLe"),
    ("20", "tuition_discount_policy", seed_20_chinh_sach_giam, "20_ChinhSachGiam"),
]


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

async def main():
    parser = argparse.ArgumentParser(description="Seed DB from seed_data_template.xlsx")
    parser.add_argument("--file", default="/app/seed_data_template.xlsx",
                        help="Path to xlsx file (default: /app/seed_data_template.xlsx)")
    parser.add_argument("--sheets", nargs="+",
                        help="Specific sheet prefixes to seed (e.g. 12 13 14)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview only, no DB changes")
    args = parser.parse_args()

    xlsx_path = Path(args.file)
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    print(f"{'DRY RUN — ' if args.dry_run else ''}Seeding from: {xlsx_path}")
    print(f"Sheets in workbook: {wb.sheetnames}\n")

    # Filter registry
    registry = SEED_REGISTRY
    if args.sheets:
        registry = [r for r in registry if r[0] in args.sheets]

    async with AsyncSessionLocal() as db:
        total = 0
        for prefix, table_name, parser_fn, sheet_name in registry:
            if sheet_name not in wb.sheetnames:
                print(f"  SKIP  {prefix:>4} {table_name:<30} sheet '{sheet_name}' not found")
                continue
            try:
                count = await parser_fn(db, wb, args.dry_run)
                if not args.dry_run:
                    await db.commit()
                status = "DRY" if args.dry_run else "OK"
                print(f"  {status:>4}  {prefix:>4} {table_name:<30} {count:>4} rows")
                total += count
            except Exception as e:
                await db.rollback()
                print(f"  FAIL  {prefix:>4} {table_name:<30} {e}")

        # Reset sequences
        if not args.dry_run:
            print("\nResetting sequences...")
            async with db.begin():
                result = await db.execute(text(
                    "SELECT tablename FROM pg_tables WHERE schemaname='public' AND tablename != 'alembic_version'"
                ))
                for (tbl,) in result.fetchall():
                    try:
                        await db.execute(text(
                            f"SELECT setval(pg_get_serial_sequence('{tbl}', 'id'), "
                            f"COALESCE((SELECT MAX(id) FROM \"{tbl}\"), 1))"
                        ))
                    except Exception:
                        pass

        print(f"\n{'DRY RUN — ' if args.dry_run else ''}Total: {total} rows")
        if not args.dry_run:
            print("\nNext steps:")
            print("  1. Seed notifications:  python -m app.scripts.seed_notification_rules")
            print("  2. Seed admin nodes:    python -m scripts.seeds.seed_administrative_nodes")


if __name__ == "__main__":
    asyncio.run(main())
