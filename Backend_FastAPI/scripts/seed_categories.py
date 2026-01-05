import sys
import os
import asyncio
import openpyxl

# Add parent dir to path to import app
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import AsyncSessionLocal
from app.repositories.config_repository import SystemCategoryRepository

# WSL Path
FILE_PATH = "/mnt/d/QLTS/Documents/Import_Thong_Tin_Hoc_Vien_Data_1767431894263.xlsx"

async def main():
    print(f"Reading file: {FILE_PATH}")
    if not os.path.exists(FILE_PATH):
        print(f"File not found: {FILE_PATH}")
        return

    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    if "Danh mục" not in wb.sheetnames:
        print("Sheet 'Danh mục' not found")
        return

    ws = wb["Danh mục"]
    
    # Headers mapping
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.column_letter] = str(cell.value).strip()
            
    target_map = {
        "Dân tộc": "ethnicity",
        "Tôn giáo": "religion",
        "Quốc tịch": "nationality",
        "Loại khuyết tật": "disability_type"
    }
    
    col_map = {}
    for col, name in headers.items():
        if name in target_map:
            col_map[target_map[name]] = col
            
    print(f"Mapped columns: {col_map}")
    
    async with AsyncSessionLocal() as db:
        repo = SystemCategoryRepository(db)
        
        for type_key, col_letter in col_map.items():
            print(f"Processing {type_key}...")
            # Extract unique values
            values = set()
            for row in range(2, ws.max_row + 1):
                val = ws[f"{col_letter}{row}"].value
                if val:
                    values.add(str(val).strip())
            
            count = 0
            for val in values:
                # Use value as Code AND Name
                await repo.create_or_update(type_key, val, val)
                count += 1
            print(f"Imported {count} items for {type_key}")
        
        await db.commit()
    print("Seeding completed successfully.")

if __name__ == "__main__":
    asyncio.run(main())
