"""
Generate Excel template for seed data input.
Each sheet = 1 database table, columns = fields to fill.
Run: python -m scripts.generate_seed_template
Output: seed_data_template.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
EXAMPLE_FONT = Font(italic=True, color="6B7280")
NOTE_FONT = Font(italic=True, color="DC2626", size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def create_sheet(name, columns, examples=None, notes=None):
    """Create a sheet with headers, examples, and notes."""
    ws = wb.create_sheet(title=name)

    # Header row
    for col_idx, (col_name, col_desc, required) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Description row
    for col_idx, (col_name, col_desc, required) in enumerate(columns, 1):
        label = f"{'[BẮT BUỘC] ' if required else '[Tùy chọn] '}{col_desc}"
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(italic=True, size=9, color="374151")
        if required:
            cell.fill = REQUIRED_FILL
        cell.alignment = Alignment(wrap_text=True)
        cell.border = THIN_BORDER

    # Example rows
    if examples:
        for row_idx, example in enumerate(examples, 3):
            for col_idx, val in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = EXAMPLE_FONT
                cell.border = THIN_BORDER

    # Notes
    if notes:
        note_row = (len(examples) + 4) if examples else 4
        cell = ws.cell(row=note_row, column=1, value="GHI CHÚ:")
        cell.font = Font(bold=True, color="DC2626")
        for i, note in enumerate(notes):
            cell = ws.cell(row=note_row + 1 + i, column=1, value=f"• {note}")
            cell.font = NOTE_FONT

    # Auto-width
    for col_idx in range(1, len(columns) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 15), 50)

    return ws


# =========================================================================
# 1. TỔ CHỨC (organization_unit)
# =========================================================================
create_sheet("1_ToChuc", [
    ("ten", "Tên đơn vị", True),
    ("loai", "Loại: truong / khoa / phong / ban", True),
    ("mo_ta", "Mô tả ngắn", False),
    ("don_vi_cha", "Tên đơn vị cha (để trống nếu là gốc)", False),
], examples=[
    ("Trường Cao Đẳng ABC", "truong", "Đơn vị gốc", ""),
    ("Khoa Công Nghệ Thông Tin", "khoa", "Khoa CNTT", "Trường Cao Đẳng ABC"),
    ("Phòng Tuyển Sinh", "phong", "Phòng tuyển sinh", "Trường Cao Đẳng ABC"),
], notes=[
    "Đơn vị gốc (trường) để trống cột 'đơn vị cha'",
    "Các khoa/phòng phải trỏ về đơn vị cha đã có ở dòng trên",
])

# =========================================================================
# 2. TÀI KHOẢN (user)
# =========================================================================
create_sheet("2_TaiKhoan", [
    ("username", "Tên đăng nhập (unique)", True),
    ("email", "Email (unique)", True),
    ("mat_khau", "Mật khẩu ban đầu", True),
    ("ho_ten", "Họ và tên đầy đủ", True),
    ("vai_tro", "Vai trò: admin / manager / officer / accountant", True),
    ("don_vi", "Tên đơn vị (khớp sheet ToChuc)", True),
    ("so_dien_thoai", "Số điện thoại", False),
], examples=[
    ("admin", "admin@truong.edu.vn", "Admin@123", "Nguyễn Văn A", "admin", "Trường Cao Đẳng ABC", "0901234567"),
    ("manager01", "manager01@truong.edu.vn", "Manager@123", "Trần Thị B", "manager", "Phòng Tuyển Sinh", "0907654321"),
    ("officer01", "officer01@truong.edu.vn", "Officer@123", "Lê Văn C", "officer", "Phòng Tuyển Sinh", "0909876543"),
], notes=[
    "Mật khẩu nên có ít nhất 8 ký tự, bao gồm chữ hoa, chữ thường, số, ký tự đặc biệt",
    "Vai trò: admin (quản trị), manager (quản lý), officer (tư vấn viên), accountant (kế toán)",
    "Đơn vị phải khớp chính xác tên ở sheet 1_ToChuc",
])

# =========================================================================
# 3. NGÀNH HỌC (major_program)
# =========================================================================
create_sheet("3_NganhHoc", [
    ("ma_nganh", "Mã ngành (theo Bộ GD&ĐT)", True),
    ("ten_nganh", "Tên ngành đào tạo", True),
    ("bac_hoc", "Bậc: Trung cấp / Cao đẳng / Đại học / Thạc sĩ / Tiến sĩ", True),
    ("don_vi_quan_ly", "Tên khoa quản lý (khớp sheet ToChuc)", True),
    ("nganh_nang", "Ngành nặng nhọc/độc hại? (có/không)", False),
], examples=[
    ("6480201", "Công nghệ thông tin", "Cao đẳng", "Khoa Công Nghệ Thông Tin", "không"),
    ("6720301", "Điều dưỡng", "Cao đẳng", "Khoa Y Dược", "không"),
    ("6510216", "Công nghệ ô tô", "Cao đẳng", "Khoa Công Nghệ Thông Tin", "có"),
], notes=[
    "Mã ngành theo danh mục của Bộ GD&ĐT",
    "Bậc học phải khớp: Trung cấp, Cao đẳng, Đại học, Thạc sĩ, Tiến sĩ",
])

# =========================================================================
# 4. HÌNH THỨC ĐÀO TẠO (program_offering)
# =========================================================================
create_sheet("4_HinhThucDaoTao", [
    ("ma_nganh", "Mã ngành (khớp sheet NganhHoc)", True),
    ("hinh_thuc", "Hình thức: Chính quy / Liên thông / Vừa làm vừa học / Từ xa", True),
    ("so_ky", "Số kỳ học", True),
    ("so_tin_chi", "Tổng số tín chỉ", False),
], examples=[
    ("6480201", "Chính quy", 6, 90),
    ("6480201", "Liên thông", 4, 60),
    ("6720301", "Chính quy", 6, 95),
], notes=[
    "Mỗi dòng = 1 hình thức đào tạo của 1 ngành",
    "Cùng ngành có thể có nhiều hình thức (Chính quy, Liên thông...)",
])

# =========================================================================
# 5. THÔNG TIN NĂM HỌC (offering_academic_info)
# =========================================================================
create_sheet("5_NamHoc", [
    ("ma_nganh", "Mã ngành (khớp sheet NganhHoc)", True),
    ("hinh_thuc", "Hình thức (khớp sheet HinhThucDaoTao)", True),
    ("nam_hoc", "Năm tuyển sinh (VD: 2026)", True),
    ("hoc_phi_nam", "Học phí/năm (VNĐ)", True),
    ("chi_tieu", "Chỉ tiêu tuyển sinh", True),
    ("diem_chuan_nam_truoc", "Điểm chuẩn năm trước", False),
    ("doi_tuong", "Đối tượng tuyển sinh", False),
    ("da_cong_bo", "Đã công bố? (có/không)", False),
], examples=[
    ("6480201", "Chính quy", 2026, 5500000, 120, None, "Tốt nghiệp THPT", "có"),
    ("6480201", "Liên thông", 2026, 6000000, 60, None, "Có bằng TC/CĐ", "có"),
], notes=[
    "Mỗi dòng = 1 ngành × 1 hình thức × 1 năm tuyển sinh",
])

# =========================================================================
# 6. MÔN THI (subject) — chỉ điền nếu muốn thay đổi danh sách mặc định
# =========================================================================
create_sheet("6_MonThi", [
    ("ma_mon", "Mã môn (VD: math, physics...)", True),
    ("ten_mon", "Tên tiếng Việt", True),
    ("thu_tu", "Thứ tự hiển thị", False),
], examples=[
    ("math", "Toán", 1),
    ("physics", "Vật lý", 2),
    ("chemistry", "Hóa học", 3),
    ("biology", "Sinh học", 4),
    ("literature", "Ngữ văn", 5),
    ("english", "Tiếng Anh", 6),
    ("history", "Lịch sử", 7),
    ("geography", "Địa lý", 8),
    ("civic_education", "Giáo dục công dân", 9),
], notes=[
    "Đã có sẵn 9 môn mặc định ở trên. Chỉ điền thêm nếu cần bổ sung",
    "Xóa các dòng ví dụ và điền lại nếu muốn thay đổi hoàn toàn",
])

# =========================================================================
# 7. TỔ HỢP MÔN (subject_group)
# =========================================================================
create_sheet("7_ToHopMon", [
    ("ma_to_hop", "Mã tổ hợp (VD: A00, D01...)", True),
    ("ten_to_hop", "Tên hiển thị", True),
    ("mon_1", "Mã môn 1 (khớp sheet MonThi)", True),
    ("mon_2", "Mã môn 2", True),
    ("mon_3", "Mã môn 3", True),
], examples=[
    ("A00", "Toán, Vật lý, Hóa học", "math", "physics", "chemistry"),
    ("A01", "Toán, Vật lý, Tiếng Anh", "math", "physics", "english"),
    ("B00", "Toán, Hóa học, Sinh học", "math", "chemistry", "biology"),
    ("D01", "Toán, Ngữ văn, Tiếng Anh", "math", "literature", "english"),
    ("D07", "Toán, Hóa học, Tiếng Anh", "math", "chemistry", "english"),
    ("C00", "Ngữ văn, Lịch sử, Địa lý", "literature", "history", "geography"),
], notes=[
    "Mã môn phải khớp chính xác với cột 'ma_mon' ở sheet 6_MonThi",
    "Mỗi tổ hợp gồm đúng 3 môn",
])

# =========================================================================
# 8. PHƯƠNG THỨC XÉT TUYỂN (admission_method)
# =========================================================================
create_sheet("8_PhuongThucXT", [
    ("ma_pt", "Mã phương thức (VD: hoc_ba, thpt_qg)", True),
    ("ten_pt", "Tên phương thức", True),
    ("mo_ta", "Mô tả chi tiết", False),
    ("can_gpa", "Yêu cầu GPA? (có/không)", True),
    ("can_diem_thi", "Yêu cầu điểm thi? (có/không)", True),
], examples=[
    ("hoc_ba", "Xét học bạ THPT", "Xét tuyển dựa trên điểm trung bình học bạ", "có", "không"),
    ("thpt_qg", "Xét điểm thi THPT Quốc gia", "Xét tuyển dựa trên điểm thi THPT QG", "không", "có"),
    ("xet_tuyen_thang", "Xét tuyển thẳng", "Dành cho đối tượng ưu tiên", "không", "không"),
], notes=[
    "Mã phương thức dùng chữ thường, không dấu, gạch dưới thay khoảng trắng",
])

# =========================================================================
# 9. TIÊU CHÍ XÉT TUYỂN (admission_criteria)
# =========================================================================
create_sheet("9_TieuChi", [
    ("ma_tieu_chi", "Mã tiêu chí (VD: HB2026_CQ)", True),
    ("ten_tieu_chi", "Tên tiêu chí", True),
    ("ma_phuong_thuc", "Mã phương thức (khớp sheet 8)", True),
    ("diem_gpa_toi_thieu", "Điểm GPA tối thiểu (để trống nếu ko áp dụng)", False),
    ("diem_thi_toi_thieu", "Tổng điểm thi tối thiểu", False),
    ("so_mon_xet", "Số môn xét", False),
    ("cach_tinh_diem", "Cách tính: sum / average / weighted", True),
    ("diem_toi_da", "Điểm tối đa có thể đạt", False),
    ("to_hop_ap_dung", "Mã tổ hợp áp dụng (ngăn cách bằng dấu phẩy)", False),
], examples=[
    ("HB2026_CQ", "Xét học bạ 2026 - Chính quy", "hoc_ba", 6.0, None, 3, "average", 10.0, "A00, A01, B00, D01, D07"),
    ("HB2026_LT", "Xét học bạ 2026 - Liên thông", "hoc_ba", 5.5, None, 3, "average", 10.0, "A00, D01"),
    ("THPTQG2026", "Xét THPT QG 2026", "thpt_qg", None, 15.0, 3, "sum", 30.0, "A00, A01, B00, D01, D07"),
], notes=[
    "Cách tính điểm: sum (tổng), average (trung bình), weighted (có trọng số)",
    "Tổ hợp áp dụng: liệt kê mã tổ hợp cách nhau bằng dấu phẩy",
])

# =========================================================================
# 10. PIPELINE / TRẠNG THÁI TƯ VẤN (tùy chọn)
# =========================================================================
create_sheet("10_Pipeline", [
    ("ma_giai_doan", "Mã giai đoạn pipeline (VD: new, contacted)", True),
    ("ten_giai_doan", "Tên hiển thị", True),
    ("thu_tu", "Thứ tự (0, 1, 2...)", True),
    ("la_giai_doan_cuoi", "Giai đoạn cuối? (có/không)", True),
    ("ma_mau", "Mã màu hex (VD: #3B82F6)", False),
], examples=[
    ("new", "Mới", 0, "không", "#3B82F6"),
    ("contacted", "Tiếp cận", 1, "không", "#F59E0B"),
    ("consulting", "Tư vấn", 2, "không", "#8B5CF6"),
    ("evaluation", "Đánh giá", 3, "không", "#10B981"),
    ("result", "Kết quả", 4, "có", "#EF4444"),
], notes=[
    "Sheet này TÙY CHỌN — nếu để trống sẽ dùng pipeline mặc định",
    "Mã giai đoạn dùng chữ thường, không dấu",
    "Nếu muốn tùy chỉnh trạng thái chi tiết (consultation_status), liên hệ để được hướng dẫn thêm",
])

# =========================================================================
# 11. DANH MỤC HỆ THỐNG (config_system_category) — tùy chọn
# =========================================================================
create_sheet("11_DanhMuc", [
    ("loai", "Loại: nationality / ethnicity / religion / disability", True),
    ("ma", "Mã (VD: VN, 01...)", True),
    ("ten", "Tên hiển thị", True),
    ("thu_tu", "Thứ tự hiển thị", False),
], examples=[
    ("nationality", "VN", "Việt Nam", 1),
    ("nationality", "LA", "Lào", 2),
    ("ethnicity", "01", "Kinh", 1),
    ("ethnicity", "02", "Tày", 2),
    ("religion", "00", "Không tôn giáo", 1),
    ("religion", "01", "Phật giáo", 2),
], notes=[
    "Sheet này TÙY CHỌN — nếu để trống sẽ dùng danh mục mặc định",
    "Danh sách đầy đủ 54 dân tộc, 63 tỉnh thành có thể bổ sung sau",
])

# Remove default sheet
del wb["Sheet"]

# Save
output_path = "/app/seed_data_template.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Sheets: {wb.sheetnames}")
