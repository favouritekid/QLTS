"""Build an editable, multi-sheet admission-config workbook from the prod export.

Reads _admission_export_2026.json (marker '@@@SECTION:<name>' + one JSON line),
writes admission_config_2026.xlsx with VN headers, dropdowns, gap highlighting.
Run inside the backend container (openpyxl present). READ-ONLY on source data.
"""
from __future__ import annotations

import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

AY = 2026
SRC = "/app/_admission_export_2026.json"
OUT = "/app/admission_config_2026.xlsx"

# ---- parse marker file -> {section: [rows]} ----
sections: dict[str, list] = {}
cur = None
for line in open(SRC, encoding="utf-8"):
    line = line.rstrip("\n")
    if line.startswith("@@@SECTION:"):
        cur = line.split(":", 1)[1].strip()
        continue
    if cur and line.strip().startswith("["):
        sections[cur] = json.loads(line)
        cur = None
print("sections:", {k: len(v) for k, v in sections.items()})

# ---- styles ----
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
LOCK_FILL = PatternFill("solid", fgColor="D9D9D9")   # non-editable id cols
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")   # needs-decision cells
NOTE_FILL = PatternFill("solid", fgColor="FCE4D6")   # editable note col
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="top")


def write_sheet(wb, title, rows, headers, *, widths=None, dropdowns=None,
                lock_cols=(), warn_rule=None, note_col=None, freeze="A2"):
    ws = wb.create_sheet(title[:31])
    keys = [h[0] for h in headers]
    labels = [h[1] for h in headers]
    if note_col:
        keys.append("__note__")
        labels.append(note_col)
    # header
    for c, label in enumerate(labels, 1):
        cell = ws.cell(1, c, label)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    # body
    for r, row in enumerate(rows, 2):
        for c, key in enumerate(keys, 1):
            val = "" if key == "__note__" else row.get(key)
            if isinstance(val, bool):
                val = "TRUE" if val else "FALSE"
            elif isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            cell = ws.cell(r, c, val)
            cell.alignment = WRAP
            cell.border = BORDER
            if key in lock_cols:
                cell.fill = LOCK_FILL
            if key == "__note__":
                cell.fill = NOTE_FILL
            if warn_rule and warn_rule(key, row):
                cell.fill = WARN_FILL
    # widths
    for c, key in enumerate(keys, 1):
        w = (widths or {}).get(key, 16)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(keys))}1"
    # dropdowns
    for key, options in (dropdowns or {}).items():
        if key not in keys:
            continue
        col = get_column_letter(keys.index(key) + 1)
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(options), allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max(2, len(rows)+1)}")
    return ws


wb = Workbook()
wb.remove(wb.active)

# ===== README =====
ws = wb.create_sheet("HƯỚNG DẪN")
ws.column_dimensions["A"].width = 120
guide = [
    (f"FILE CẤU HÌNH TUYỂN SINH {AY} — XUẤT TỪ PRODUCTION (qlts.tnpc.edu.vn)", True),
    ("", False),
    ("CÁCH DÙNG: Sửa trực tiếp các ô. Cột nền XÁM (Path ID...) là khoá — KHÔNG sửa.", False),
    ("Ô nền VÀNG = mục cần bạn quyết định/bổ sung. Cột 'GHI CHÚ / ĐIỀU CHỈNH' (cam) để bạn ghi yêu cầu thay đổi.", False),
    ("Các cột Trạng thái / Hiển thị / Phương thức có dropdown chọn sẵn.", False),
    ("Gửi lại file này, tôi sẽ sinh lệnh cập nhật prod theo đúng nội dung bạn sửa (sau khi bạn duyệt).", False),
    ("", False),
    ("CÁC SHEET:", True),
    ("• Paths — TRUNG TÂM: mỗi dòng 1 tổ hợp (ngành × đợt × phương thức). Đủ: học phí, chỉ tiêu, tiêu chí, tổ hợp xét tuyển.", False),
    ("• Programs — ngành + offering + học phí/năm + học phí từng kỳ + chỉ tiêu năm.", False),
    ("• Rounds — các đợt tuyển sinh + thời gian mở/đóng.", False),
    ("• Methods_Criteria — phương thức + tiêu chí (GPA/điểm tối thiểu).", False),
    ("• Documents — bộ hồ sơ giấy tờ (theo phương thức / dùng chung) + tài liệu bắt buộc.", False),
    ("• SubjectGroups — tổ hợp xét tuyển + môn.", False),
    ("• DiscountPolicies — chính sách miễn giảm/học bổng (hiện rỗng — thêm nếu có).", False),
    ("• Reference — danh mục bậc đào tạo & loại hình hợp lệ (dùng khi điền).", False),
    ("", False),
    ("GHI CHÚ QUAN TRỌNG (từ rà soát prod 2026):", True),
    ("1. 'Chỉ tiêu (phương thức)' = admit_quota: ĐỂ TRỐNG nghĩa là KHÔNG giới hạn riêng theo phương thức,", False),
    ("   chỉ giới hạn bởi 'Chỉ tiêu năm (ngành)'. Điền số nếu muốn khống chế riêng học bạ / THPT QG.", False),
    ("2. Phương thức 'xet_tuyen_thang' đang ở trạng thái draft/internal + thiếu tiêu chí → đang ẩn.", False),
    ("   Nếu muốn MỞ: điền tiêu chí + đổi Trạng thái=active, Hiển thị=public.", False),
    ("3. Đợt DOT_3 (T7-9), DOT_4 (T10-12) hiện CHƯA có path. DOT_2 đang MỞ và đầy đủ.", False),
    ("4. KV trường THPT: 75/420 trường chưa gán khu vực ưu tiên cho 2026 → thí sinh các trường đó cần gán KV thủ công.", False),
]
for r, (txt, bold) in enumerate(guide, 1):
    cell = ws.cell(r, 1, txt)
    cell.font = Font(bold=bold, size=12 if bold else 11, color="1F4E78" if bold else "000000")
    cell.alignment = WRAP

# ===== Paths (core) =====
paths_headers = [
    ("path_id", "Path ID (khoá)"), ("program_code", "Mã ngành"), ("program_name", "Tên ngành"),
    ("degree_level", "Bậc"), ("offering_type", "Loại hình"), ("round_code", "Đợt"),
    ("method_code", "Mã phương thức"), ("method_name", "Tên phương thức"),
    ("criteria_code", "Mã tiêu chí"), ("min_gpa", "GPA tối thiểu"), ("min_score", "Điểm tối thiểu"),
    ("scoring_method", "Cách tính điểm"), ("admit_quota", "Chỉ tiêu (phương thức)"),
    ("round_quota", "Chỉ tiêu (đợt)"), ("status", "Trạng thái"), ("visibility", "Hiển thị"),
    ("display_order", "Thứ tự"), ("application_fee", "Lệ phí hồ sơ"),
    ("allow_unverified_submission", "Cho nộp khi chưa xác minh"),
    ("tuition_fee_per_year", "Học phí/năm"), ("annual_admission_quota", "Chỉ tiêu năm (ngành)"),
    ("subject_groups", "Tổ hợp xét tuyển"),
]
def paths_warn(key, row):
    if key == "admit_quota" and row.get("admit_quota") is None and row.get("status") == "active":
        return True
    if key == "criteria_code" and not row.get("criteria_code"):
        return True
    return False
write_sheet(wb, "Paths", sections.get("paths", []), paths_headers,
            widths={"program_name": 34, "method_name": 26, "subject_groups": 32, "tuition_fee_per_year": 16,
                    "scoring_method": 14, "allow_unverified_submission": 14, "annual_admission_quota": 14,
                    "admit_quota": 16, "criteria_code": 24, "__note__": 30},
            dropdowns={"status": ["draft", "active", "inactive", "archived"],
                       "visibility": ["public", "internal"],
                       "method_code": ["hoc_ba", "thpt_qg", "xet_tuyen_thang"]},
            lock_cols=("path_id",), warn_rule=paths_warn, note_col="GHI CHÚ / ĐIỀU CHỈNH")

# ===== Programs =====
write_sheet(wb, "Programs", sections.get("programs", []), [
    ("program_code", "Mã ngành"), ("program_name", "Tên ngành"), ("degree_level", "Bậc"),
    ("is_heavy", "Ngành nặng"), ("offering_type", "Loại hình"), ("duration_semesters", "Số kỳ"),
    ("total_credits", "Tín chỉ"), ("academic_year", "Năm"), ("is_published", "Đã publish"),
    ("annual_admission_quota", "Chỉ tiêu năm"), ("tuition_fee_per_year", "Học phí/năm"),
    ("semester_tuitions", "Học phí từng kỳ (kỳ:tiền)"),
], widths={"program_name": 34, "semester_tuitions": 34, "tuition_fee_per_year": 16, "annual_admission_quota": 14},
   note_col="GHI CHÚ / ĐIỀU CHỈNH")

# ===== Rounds =====
write_sheet(wb, "Rounds", sections.get("rounds", []), [
    ("round_code", "Mã đợt"), ("round_name", "Tên đợt"), ("start_date", "Mở từ"),
    ("end_date", "Đóng ngày"), ("is_active", "Đang bật"), ("archived", "Đã lưu trữ"),
    ("allow_multi_nv", "Cho đa nguyện vọng"), ("confirm_expiry_hours", "Hạn xác nhận (giờ)"),
], widths={"round_name": 30}, dropdowns={"allow_multi_nv": ["TRUE", "FALSE"], "is_active": ["TRUE", "FALSE"]},
   note_col="GHI CHÚ / ĐIỀU CHỈNH")

# ===== Methods + Criteria =====
write_sheet(wb, "Methods_Criteria", sections.get("methods", []), [
    ("method_code", "Mã phương thức"), ("method_name", "Tên phương thức"),
    ("requires_gpa", "Cần GPA"), ("requires_subject_scores", "Cần điểm môn"),
    ("method_active", "PT bật"), ("criteria_code", "Mã tiêu chí"), ("criteria_name", "Tên tiêu chí"),
    ("min_gpa", "GPA tối thiểu"), ("min_score", "Điểm tối thiểu"), ("min_subject_score", "Điểm môn tối thiểu"),
    ("scoring_method", "Cách tính"), ("required_subject_count", "Số môn cần"),
    ("subject_selection_mode", "Chế độ chọn môn"), ("criteria_active", "Tiêu chí bật"),
], widths={"method_name": 26, "criteria_name": 34}, note_col="GHI CHÚ / ĐIỀU CHỈNH")

# ===== Documents =====
write_sheet(wb, "Documents", sections.get("documents", []), [
    ("scope", "Phạm vi áp dụng"), ("offering_type", "Loại hình"), ("method_code", "Phương thức"),
    ("group_name", "Bộ hồ sơ"), ("doc_code", "Mã tài liệu"), ("doc_name", "Tên tài liệu"),
    ("is_mandatory", "Bắt buộc"), ("requires_upload", "Cần tải lên"), ("submission_format", "Hình thức nộp"),
    ("display_order", "Thứ tự"), ("is_active", "Bật"),
], widths={"group_name": 28, "doc_name": 34, "scope": 18, "submission_format": 16},
   dropdowns={"is_mandatory": ["TRUE", "FALSE"]}, note_col="GHI CHÚ / ĐIỀU CHỈNH")

# ===== Subject Groups =====
write_sheet(wb, "SubjectGroups", sections.get("subject_groups", []), [
    ("code", "Mã tổ hợp"), ("name", "Tên tổ hợp"), ("subjects", "Các môn"), ("is_active", "Bật"),
], widths={"name": 30, "subjects": 40}, note_col="GHI CHÚ / ĐIỀU CHỈNH")

# ===== Discount Policies =====
disc = sections.get("discount_policies", [])
write_sheet(wb, "DiscountPolicies", disc if disc else [
    {"code": "", "name": "", "discount_type": "", "discount_value": "", "applicable_scope": "",
     "valid_from": "", "valid_to": "", "is_active": ""}
], [
    ("code", "Mã"), ("name", "Tên chính sách"), ("discount_type", "Loại (percent/amount)"),
    ("discount_value", "Giá trị"), ("applicable_scope", "Phạm vi"), ("valid_from", "Hiệu lực từ"),
    ("valid_to", "Đến"), ("is_active", "Bật"),
], widths={"name": 34}, note_col="GHI CHÚ / ĐIỀU CHỈNH")

# ===== Reference =====
ref_rows = ([{"loại": "BẬC ĐÀO TẠO", "code": r["code"], "name": r["name"], "is_active": r["is_active"]}
             for r in sections.get("ref_degree", [])]
            + [{"loại": "LOẠI HÌNH", "code": r["code"], "name": r["name"], "is_active": r["is_active"]}
               for r in sections.get("ref_offering_type", [])])
write_sheet(wb, "Reference", ref_rows, [
    ("loại", "Danh mục"), ("code", "Mã"), ("name", "Tên"), ("is_active", "Bật"),
], widths={"name": 26, "loại": 16})

wb.save(OUT)
print("WROTE", OUT)
